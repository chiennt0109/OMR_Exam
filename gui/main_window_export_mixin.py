from __future__ import annotations

import copy
import csv
import gc
import json
import os
import re
import shutil
import sys
import unicodedata
from collections import deque
from datetime import date, datetime
from pathlib import Path
import time
import uuid
from typing import TYPE_CHECKING

sys.dont_write_bytecode = True

from PySide6.QtCore import Qt, QEvent, QTimer, QSize
from PySide6.QtGui import QAction, QColor, QImage, QKeySequence, QPixmap, QTransform, QPainter, QPen, QIcon
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QCheckBox,
    QCompleter,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QFormLayout,
    QGridLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QMainWindow,
    QMenu,
    QMessageBox,
    QProgressBar,
    QProgressDialog,
    QPushButton,
    QRadioButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QInputDialog,
    QToolBar,
    QStyle,
    QGroupBox,
    QScrollArea,
)

from core.answer_key_importer import ImportedAnswerKey, ImportedAnswerKeyPackage, import_answer_key
from core.omr_engine import OMRProcessor, OMRResult, RecognitionContext
from core.scoring_engine import ScoringEngine
from models.answer_key import AnswerKeyRepository, SubjectKey
from models.database import OMRDatabase, bootstrap_application_db
from models.exam_session import ExamSession, Student
from models.template import Template, ZoneType
from models.template_repository import TemplateRepository
from gui.ui_branding import app_icon, load_theme, TOOLBAR, STATUS, logo_symbol, logo_main, apply_widget_branding, brand_button
from gui.batch_scan_flow import run_batch_scan_from_api_file

if TYPE_CHECKING:
    from editor.template_editor import TemplateEditorWindow

from gui.main_window_dialogs import PreviewImageWidget, SubjectConfigDialog, NewExamDialog, StudentListPreviewDialog


class MainWindowExportMixin:
    """Excel/API/report export workflows and export row builders."""
    def _has_session_context_for_export(self) -> bool:
        if bool(str(getattr(self, "current_session_id", "") or "").strip()):
            return True
        if getattr(self, "session", None) is not None:
            sid = str(getattr(self.session, "session_id", "") or "").strip()
            if sid:
                return True
        if hasattr(self, "exam_list_table"):
            row = self.exam_list_table.currentRow()
            if row >= 0 and bool(str(self._session_id_for_row(row) or "").strip()):
                return True
        return False

    def _ensure_current_session_loaded(self) -> bool:
        if self.session is not None:
            return True
        sid = str(getattr(self, "current_session_id", "") or "").strip()
        if not sid:
            return False
        payload = self.database.fetch_exam_session(sid) or {}
        if not payload:
            return False
        try:
            self.session = ExamSession.from_dict(payload)
            self.current_session_path = self._session_path_from_id(sid)
        except Exception:
            return False
        return self.session is not None

    def _refresh_export_action_states(self, *, has_session: bool | None = None, has_export_data: bool | None = None) -> None:
        if has_session is None:
            has_session = self._has_session_context_for_export()
        if has_export_data is None:
            has_export_data = self._has_exportable_data()
        for attr_name in [
            "act_export_subject_scores",
            "act_export_subject_score_matrix",
            "act_export_class_subject_scores",
            "act_export_all_classes_subject_scores",
            "act_export_all_scores",
            "act_export_return_by_class",
            "act_export_recheck_by_subject",
            "act_export_recheck_by_class",
            "act_export_subject_api",
            "act_export_reports_center",
            "act_export_range_report",
            "act_export_class_report",
            "act_export_management_report",
        ]:
            action = getattr(self, attr_name, None)
            if action is not None:
                action.setEnabled(bool(has_session))

    def _iter_export_subjects(self) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        subject_rank = {str(name).strip().casefold(): idx for idx, name in enumerate(self.subject_catalog or [])}

        for cfg in self._subject_configs_for_scoring():
            key = str(self._subject_key_from_cfg(cfg) or "").strip()
            if not key:
                continue
            label = self._display_subject_label(cfg)
            pairs.append((label, key))
        if not pairs:
            for key in (str(k) for k in (self.scoring_results_by_subject or {}).keys() if str(k).strip()):
                pairs.append((key, key))

        indexed_pairs = list(enumerate(pairs))

        def _subject_rank_from_key(subject_key: str) -> int:
            key = str(subject_key or "").strip()
            if not key:
                return 10**6
            cfg = self._subject_config_by_subject_key(key) or {}
            candidates = [
                str(cfg.get("name", "") or "").strip(),
                str(cfg.get("subject", "") or "").strip(),
                key,
                key.split("_", 1)[0].strip(),
            ]
            for candidate in candidates:
                if not candidate:
                    continue
                rank = subject_rank.get(candidate.casefold())
                if rank is not None:
                    return rank
            return 10**6

        def _sort_key(item: tuple[int, tuple[str, str]]) -> tuple[int, int]:
            original_idx, entry = item
            rank = _subject_rank_from_key(str(entry[1] or ""))
            return (rank, original_idx)

        return [entry for _idx, entry in sorted(indexed_pairs, key=_sort_key)]

    def _has_exportable_data(self) -> bool:
        for _label, key in self._iter_export_subjects():
            if self._load_scoring_results_for_subject_from_storage(key, force_db=True):
                return True
            if self._refresh_scan_results_from_db(key):
                return True
            cfg = self._subject_config_by_subject_key(key) or {}
            if self._subject_uses_direct_score_import(cfg) and self._direct_score_import_rows_for_subject(cfg):
                return True
        return False

    def _pick_subject_for_export(self, title: str, prompt: str) -> str:
        subjects = self._iter_export_subjects()
        if not subjects:
            return ""
        if len(subjects) == 1:
            return str(subjects[0][1])
        labels = [x[0] for x in subjects]
        choice, ok = QInputDialog.getItem(self, title, prompt, labels, 0, False)
        if not ok:
            return ""
        picked = str(choice or "").strip()
        for label, key in subjects:
            if label == picked:
                return key
        return ""

    @staticmethod
    def _safe_sheet_name(sheet_name: str, fallback: str = "Sheet") -> str:
        cleaned = re.sub(r"[\\/*?:\\[\\]]+", "_", str(sheet_name or "").strip())
        cleaned = cleaned[:31].strip()
        return cleaned or fallback

    @staticmethod
    def _format_birth_date_for_export(value: object) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        candidates = [raw]
        if "T" in raw:
            candidates.append(raw.split("T", 1)[0].strip())
        if " " in raw:
            candidates.append(raw.split(" ", 1)[0].strip())
        for text in candidates:
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y"):
                try:
                    return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
                except Exception:
                    continue
        return raw

    def _score_rows_for_subject(self, subject_key: str, *, force_db: bool = False) -> list[dict]:
        payload = self._load_scoring_results_for_subject_from_storage(subject_key, force_db=force_db)
        rows = [dict(v) for v in payload.values() if isinstance(v, dict)]
        rows.sort(key=lambda x: str(x.get("student_id", "") or ""))
        return rows

    def _align_export_rows_with_session_students(self, rows: list[dict], subject_key: str) -> list[dict]:
        subject = str(subject_key or "").strip()
        if not subject:
            return list(rows or [])

        normalized_rows: list[dict] = [dict(item) for item in (rows or []) if isinstance(item, dict)]
        if not self.session or not getattr(self.session, "students", None):
            return normalized_rows

        student_meta = self._student_meta_by_sid()
        row_by_sid: dict[str, dict] = {}
        for row in normalized_rows:
            sid = str(row.get("student_id", "") or "").strip()
            if sid and sid not in row_by_sid:
                row_by_sid[sid] = row

        ordered: list[dict] = []
        seen: set[str] = set()
        for st in (self.session.students or []):
            sid = str(getattr(st, "student_id", "") or "").strip()
            if not sid or sid in seen:
                continue
            seen.add(sid)
            existing = row_by_sid.get(sid, {})
            meta = student_meta.get(sid, {})
            base = {
                "student_id": sid,
                "name": str(getattr(st, "name", "") or meta.get("name", "") or ""),
                "subject": subject,
                "class_name": str(getattr(st, "class_name", "") or meta.get("class_name", "") or ""),
                "birth_date": str(getattr(st, "birth_date", "") or meta.get("birth_date", "") or ""),
                "exam_room": str(getattr(st, "exam_room", "") or meta.get("exam_room", "") or ""),
                "exam_code": "",
                "mcq_correct": "",
                "tf_correct": "",
                "numeric_correct": "",
                "tf_compare": "",
                "numeric_compare": "",
                "correct": "",
                "wrong": "",
                "blank": "",
                "score": "",
                "recheck_score": "",
                "baithiphuctra": "",
                "phase": "",
                "phase_timestamp": "",
                "phase_mode": "",
                "note": "",
                "status": "",
            }
            base.update(existing)
            ordered.append(base)

        for row in normalized_rows:
            sid = str(row.get("student_id", "") or "").strip()
            if not sid or sid in seen:
                continue
            ordered.append(row)
        return ordered

    def _ensure_export_score_rows_for_subject(self, subject_key: str) -> list[dict]:
        subject = str(subject_key or "").strip()
        if not subject:
            return []
        rows = self._score_rows_for_subject(subject, force_db=True)
        if rows:
            return rows
        cfg = self._subject_config_by_subject_key(subject) or {}
        if self._subject_uses_direct_score_import(cfg) and self._direct_score_import_rows_for_subject(cfg):
            self.calculate_scores(subject_key=subject, mode="Nhập điểm trực tiếp", note="auto_prepare_export_essay")
            rows = self._score_rows_for_subject(subject, force_db=True)
        return rows

    @staticmethod
    def _score_value_for_statistics(row: dict) -> float | None:
        if not isinstance(row, dict):
            return None
        status_text = str(row.get("status", "") or "").strip()
        status_fold = status_text.casefold()
        if status_fold.startswith("lỗi"):
            return None
        if status_fold in {"chưa chấm", "cần chấm lại", "chưa có điểm"}:
            return None
        score_raw = row.get("score", "")
        if score_raw in {"", None}:
            return None
        try:
            return float(score_raw)
        except Exception:
            return None

    def _scan_rows_for_subject(self, subject_key: str) -> list[OMRResult]:
        for candidate_key in self._subject_scan_storage_key_candidates(subject_key):
            rows = list(self.scan_results_by_subject.get(candidate_key, []) or [])
            if rows:
                return rows
        return self._refresh_scan_results_from_db(subject_key) or []

    def _scoring_source_student_ids(self, subject_key: str) -> tuple[set[str], int]:
        subject = str(subject_key or "").strip()
        if not subject:
            return set(), 0
        cfg = self._subject_config_by_subject_key(subject) or {}
        if self._subject_uses_direct_score_import(cfg):
            import_rows = self._direct_score_import_rows_for_subject(cfg)
            student_ids = {
                str((row or {}).get("student_id", "") or "").strip()
                for row in import_rows
                if str((row or {}).get("student_id", "") or "").strip()
            }
            return student_ids, len(import_rows)
        scan_rows = self._scan_rows_for_subject(subject)
        student_ids = {
            str(getattr(scan, "student_id", "") or "").strip()
            for scan in scan_rows
            if str(getattr(scan, "student_id", "") or "").strip()
        }
        return student_ids, len(scan_rows)

    def _student_meta_by_sid(self) -> dict[str, dict[str, str]]:
        lookup: dict[str, dict[str, str]] = {}
        for st in (self.session.students if self.session else []):
            sid = str(getattr(st, "student_id", "") or "").strip()
            if not sid:
                continue
            extra = getattr(st, "extra", {}) or {}
            lookup[sid] = {
                "name": str(getattr(st, "name", "") or "").strip(),
                "class_name": str((extra or {}).get("class_name", "") or "").strip(),
                "exam_room": str((extra or {}).get("exam_room", "") or "").strip(),
                "birth_date": str((extra or {}).get("birth_date", "") or "").strip(),
            }
        return lookup

    def action_export_answer_key_sample(self) -> None:
        self.export_answer_key_sample()

    def action_export_results(self) -> None:
        self.export_results()

    def action_export_subject_scores(self) -> None:
        subject_key = self._pick_subject_for_export("Xuất điểm môn", "Chọn môn cần xuất điểm:")
        if not subject_key:
            return
        self._export_subject_scores(subject_key)

    def action_export_subject_score_matrix(self) -> None:
        self._export_subject_score_matrix()

    def action_export_class_subject_scores(self) -> None:
        self._export_class_subject_scores()

    def action_export_all_classes_subject_scores(self) -> None:
        self._export_all_classes_subject_scores()

    def action_export_all_subject_scores(self) -> None:
        self._export_all_subject_scores()

    def action_export_return_by_class(self) -> None:
        self._export_return_by_class()

    def action_export_recheck_by_subject(self) -> None:
        self._export_recheck_package(group_by="subject")

    def action_export_recheck_by_class(self) -> None:
        self._export_recheck_package(group_by="class")

    def action_export_subject_api_payload(self) -> None:
        subject_key = self._pick_subject_for_export("Xuất API bài làm", "Chọn môn cần xuất API bài làm:")
        if not subject_key:
            return
        self._export_subject_api_payload(subject_key)

    def action_export_score_range_report(self) -> None:
        self.action_open_export_reports_center()

    def action_export_class_report(self) -> None:
        self.action_open_export_reports_center()

    def action_export_management_report(self) -> None:
        self.action_open_export_reports_center()

    def export_answer_key_sample(self) -> None:
        save_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Sample Answer Key",
            "answer_key_sample.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)",
        )
        if not save_path:
            return

        import pandas as pd

        data = {
            "Question": list(range(1, 19)) + [1, 2, 3, 4] + [1, 2, 3, 4, 5, 6],
            "0101": ["C", "C", "C", "B", "D", "D", "D", "B", "B", "B", "B", "B", "D", "C", "D", "B", "C", "C", "ĐSĐĐ", "ĐĐĐS", "ĐSĐS", "ĐDDS", "5", "69", "0,61", "58,3", "49,6", "2"],
            "0102": ["B", "C", "C", "C", "D", "D", "B", "D", "B", "B", "B", "C", "C", "B", "D", "C", "D", "B", "ĐĐSĐ", "ĐĐĐS", "ĐĐSĐ", "ĐSDS", "2", "69", "58,3", "5", "49,6", "0,61"],
            "0103": ["C", "C", "C", "B", "C", "B", "D", "D", "C", "B", "B", "C", "B", "D", "C", "D", "B", "B", "SĐĐĐ", "ĐĐĐS", "ĐĐSĐ", "SĐDS", "2", "0,61", "58,3", "49,6", "5", "69"],
            "0104": ["C", "C", "C", "B", "C", "B", "D", "D", "C", "B", "B", "C", "B", "D", "C", "D", "B", "B", "ĐĐĐS", "SĐSĐ", "ĐĐSĐ", "SĐDD", "5", "2", "49,6", "0,61", "58,3", "69"],
        }
        df = pd.DataFrame(data)
        path = Path(save_path)
        if path.suffix.lower() == ".csv" or "CSV" in selected_filter:
            if path.suffix.lower() != ".csv":
                path = path.with_suffix(".csv")
            df.to_csv(path, index=False)
        else:
            if path.suffix.lower() != ".xlsx":
                path = path.with_suffix(".xlsx")
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, startrow=1)
                ws = writer.sheets[next(iter(writer.sheets))]
                ws["A1"] = "Câu hỏi"
                ws["B1"] = "Mã đề thi"
                ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
                ws["B2"] = "0101"
                ws["C2"] = "0102"
                ws["D2"] = "0103"
                ws["E2"] = "0104"
        QMessageBox.information(self, "Sample exported", f"Saved sample answer key file to:\n{path}")

    def _export_student_subject_matrix_excel(self, output_path: Path) -> None:
        from openpyxl import Workbook

        for _label, key in self._iter_export_subjects():
            self._load_scoring_results_for_subject_from_storage(key, force_db=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "scores_by_student"

        subjects = [key for _label, key in self._iter_export_subjects() if str(key or "").strip()]
        if not subjects:
            subjects = sorted(set(str(k or "").strip() for k in (self.scoring_results_by_subject or {}).keys() if str(k or "").strip()))
        header = ["Student ID", "Name"] + subjects
        ws.append(header)

        base_students: list[tuple[str, str]] = []
        seen: set[str] = set()
        if self.session:
            for st in (self.session.students or []):
                sid = str(getattr(st, "student_id", "") or "").strip()
                name = str(getattr(st, "name", "") or "").strip()
                if not sid or sid in seen:
                    continue
                seen.add(sid)
                base_students.append((sid, name))

        for subject in subjects:
            for sid, row in (self.scoring_results_by_subject.get(subject, {}) or {}).items():
                sid_text = str(sid or "").strip()
                if not sid_text or sid_text in seen:
                    continue
                seen.add(sid_text)
                base_students.append((sid_text, str((row or {}).get("name", "") or "")))

        for sid, name in base_students:
            vals = [sid, name]
            for subject in subjects:
                row = (self.scoring_results_by_subject.get(subject, {}) or {}).get(sid, {}) or {}
                vals.append(row.get("score", ""))
            ws.append(vals)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[col_letter].width = min(40, max(12, max_len + 2))

        wb.save(output_path)

    def export_results(self) -> None:
        rows = self.score_rows or self.calculate_scores()
        if not rows:
            return

        output_dir = QFileDialog.getExistingDirectory(self, "Select export folder")
        if not output_dir:
            return
        self.scoring_engine.export_csv(rows, Path(output_dir) / "results.csv")
        self.scoring_engine.export_json(rows, Path(output_dir) / "results.json")
        self.scoring_engine.export_xml(rows, Path(output_dir) / "results.xml")
        self.scoring_engine.export_excel(rows, Path(output_dir) / "results.xlsx")
        try:
            self._export_student_subject_matrix_excel(Path(output_dir) / "scores_by_student.xlsx")
            QMessageBox.information(self, "Export", "Exported CSV, JSON, XML, XLSX, scores_by_student.xlsx.")
        except Exception as exc:
            QMessageBox.warning(self, "Export", f"Đã export CSV/JSON/XML/XLSX nhưng không tạo được scores_by_student.xlsx:\n{exc}")

    def _export_subject_scores(self, subject_key: str) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        rows = self._ensure_export_score_rows_for_subject(subject)
        if not rows:
            QMessageBox.information(self, "Xuất điểm môn", "Môn này chưa có dữ liệu điểm.")
            return
        cfg = self._subject_config_by_subject_key(subject) or {}
        suggested = self._safe_sheet_name(str(cfg.get("name", "") or "subject_scores"), fallback="subject_scores")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Xuất điểm môn",
            f"{suggested}.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)",
        )
        if not path:
            return
        out = Path(path)
        headers = self._subject_score_export_headers()
        normalized = self._build_subject_score_export_rows(subject)
        if out.suffix.lower() == ".csv":
            with out.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(normalized)
        else:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            from openpyxl.styles import Border, Font, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "scores"
            ws.append(headers)
            for row in normalized:
                ws.append([row.get(key, "") for key in headers])
            border_side = Side(style="thin", color="000000")
            base_font = Font(name="Times New Roman", size=12)
            for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row_cells:
                    cell.font = base_font
                    cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            for cell in ws[1]:
                cell.font = Font(name="Times New Roman", size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            name_col = headers.index("Họ tên") + 1 if "Họ tên" in headers else -1
            if name_col > 0:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=name_col).alignment = Alignment(horizontal="left", vertical="center")
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    max_len = max(max_len, len(str(cell.value or "")))
                ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))
            wb.save(out)
        QMessageBox.information(self, "Xuất điểm môn", f"Đã xuất dữ liệu:\n{out}")

    def _subject_score_export_headers(self) -> list[str]:
        return [
            "STT",
            "SBD",
            "Phòng thi",
            "Họ tên",
            "Ngày sinh",
            "Lớp",
            "Mã đề",
            "Đúng",
            "Sai",
            "Bỏ trống",
            "Điểm",
            "Trạng thái",
            "Ghi chú",
        ]

    def _build_subject_score_export_rows(self, subject: str) -> list[dict[str, object]]:
        rows = self._ensure_export_score_rows_for_subject(subject)
        student_meta = self._student_meta_by_sid()
        room_by_sid: dict[str, str] = {}
        for scan_item in self._scan_rows_for_subject(subject):
            sid_scan = str(getattr(scan_item, "student_id", "") or "").strip()
            room_scan = str(getattr(scan_item, "exam_room", "") or "").strip()
            if sid_scan and room_scan and sid_scan not in room_by_sid:
                room_by_sid[sid_scan] = room_scan
        normalized: list[dict[str, object]] = []
        for row in rows:
            sid = str(row.get("student_id", "") or "").strip()
            meta = student_meta.get(sid, {})
            status_text = str(row.get("status", "") or "").strip()
            status_fold = status_text.casefold()
            has_error = bool(status_fold) and status_fold not in {"ok", "đã sửa"}
            score_value = row.get("recheck_score", "")
            if score_value in {"", None}:
                score_value = row.get("score", "")
            normalized.append({
                "_has_error": has_error,
                "SBD": sid,
                "Phòng thi": str(row.get("exam_room", "") or room_by_sid.get(sid, "") or meta.get("exam_room", "")),
                "Họ tên": str(row.get("name", "") or meta.get("name", "")),
                "Ngày sinh": self._format_birth_date_for_export(row.get("birth_date", "") or meta.get("birth_date", "")),
                "Lớp": str(row.get("class_name", "") or meta.get("class_name", "")),
                "Mã đề": str(row.get("exam_code", "") or ""),
                "Đúng": row.get("correct", 0),
                "Sai": row.get("wrong", 0),
                "Bỏ trống": row.get("blank", 0),
                "Điểm": "" if score_value in {"", None} else score_value,
                "Trạng thái": status_text,
                "Ghi chú": str(row.get("note", "") or ""),
            })
        for idx, row in enumerate(normalized, start=1):
            row["STT"] = idx
        return normalized

    def _export_subject_score_matrix(self) -> None:
        subjects = self._iter_export_subjects()
        if not subjects:
            QMessageBox.information(self, "Xuất điểm các môn", "Chưa có môn nào để xuất.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Xuất điểm các môn", "subject_scores_matrix.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "scores"
        base_headers, score_headers, sorted_rows = self._build_subject_score_matrix_rows(subjects)
        self._write_subject_score_matrix_sheet(ws, base_headers, score_headers, sorted_rows)
        wb.save(path)
        QMessageBox.information(self, "Xuất điểm các môn", f"Đã xuất dữ liệu:\n{path}")

    def _short_subject_label_for_export(self, subject_key: str, fallback_label: str) -> str:
        cfg = self._subject_config_by_subject_key(subject_key) or {}
        raw = str(cfg.get("name", "") or fallback_label or subject_key).strip()
        if not raw:
            raw = str(subject_key or "").strip()
        compact = re.sub(r"\s*[-–]\s*kỳ thi\b.*$", "", raw, flags=re.IGNORECASE).strip()
        compact = re.sub(r"\s*[-–]\s*khối\b.*$", "", compact, flags=re.IGNORECASE).strip()
        return compact or raw

    def _build_subject_score_matrix_rows(self, subjects: list[tuple[str, str]]) -> tuple[list[str], list[str], list[dict[str, object]]]:
        base_headers = ["STT", "SBD", "Họ tên", "Ngày sinh", "Lớp"]
        score_headers = [self._short_subject_label_for_export(key, label) for label, key in subjects]
        student_rows: dict[str, dict[str, object]] = {}
        student_meta = self._student_meta_by_sid()

        if self.session:
            for st in (self.session.students or []):
                sid = str(getattr(st, "student_id", "") or "").strip()
                if not sid:
                    continue
                meta = student_meta.get(sid, {})
                student_rows[sid] = {
                    "SBD": sid,
                    "Họ tên": str(getattr(st, "name", "") or meta.get("name", "") or ""),
                    "Ngày sinh": self._format_birth_date_for_export(
                        getattr(st, "birth_date", "") or meta.get("birth_date", "") or ""
                    ),
                    "Lớp": str(getattr(st, "class_name", "") or meta.get("class_name", "") or ""),
                }

        for idx, (label, key) in enumerate(subjects):
            score_header = score_headers[idx]
            for row in self._build_subject_score_export_rows(key):
                sid = str(row.get("SBD", "") or "").strip()
                if not sid:
                    continue
                rec = student_rows.setdefault(sid, {
                    "SBD": sid,
                    "Họ tên": row.get("Họ tên", ""),
                    "Ngày sinh": self._format_birth_date_for_export(row.get("Ngày sinh", "")),
                    "Lớp": row.get("Lớp", ""),
                })
                if not rec.get("Họ tên") and row.get("Họ tên"):
                    rec["Họ tên"] = row.get("Họ tên", "")
                if not rec.get("Ngày sinh") and row.get("Ngày sinh"):
                    rec["Ngày sinh"] = self._format_birth_date_for_export(row.get("Ngày sinh", ""))
                if not rec.get("Lớp") and row.get("Lớp"):
                    rec["Lớp"] = row.get("Lớp", "")
                rec[score_header] = row.get("Điểm", "")
        sorted_rows = sorted(student_rows.values(), key=lambda x: str(x.get("SBD", "")))
        return base_headers, score_headers, sorted_rows

    def _write_subject_score_matrix_sheet(
        self,
        ws,
        base_headers: list[str],
        score_headers: list[str],
        rows: list[dict[str, object]],
    ) -> None:
        from openpyxl.styles import Alignment
        from openpyxl.styles import Border, Font, Side
        from openpyxl.utils import get_column_letter

        ws.append(base_headers + score_headers)
        for idx, row in enumerate(rows, start=1):
            row["STT"] = idx
            ws.append([row.get(col, "") for col in base_headers] + [row.get(col, "") for col in score_headers])

        border_side = Side(style="thin", color="000000")
        base_font = Font(name="Times New Roman", size=12)
        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row_cells:
                cell.font = base_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        for cell in ws[1]:
            cell.font = Font(name="Times New Roman", size=12, bold=True)
        name_col = base_headers.index("Họ tên") + 1 if "Họ tên" in base_headers else -1
        if name_col > 0:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=name_col).alignment = Alignment(horizontal="left", vertical="center")

        score_col_width = 136 / 7
        score_col_indexes = set(range(len(base_headers) + 1, len(base_headers) + len(score_headers) + 1))
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx in score_col_indexes:
                ws.column_dimensions[col_letter].width = score_col_width
                continue
            max_len = 0
            for cell in ws[col_letter]:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

    @staticmethod
    def _trim_subject_headers_without_scores(rows: list[dict[str, object]], score_headers: list[str]) -> list[str]:
        kept_subject_headers: list[str] = []
        for header in score_headers:
            has_score = False
            for row in rows:
                if header not in row:
                    continue
                value = row.get(header, None)
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                has_score = True
                break
            if has_score:
                kept_subject_headers.append(header)
        return kept_subject_headers

    def _build_class_matrix_rows(
        self,
        class_name: str,
        score_rows: list[dict[str, object]],
        score_headers: list[str],
    ) -> list[dict[str, object]]:
        score_by_sid: dict[str, dict[str, object]] = {}
        for row in score_rows:
            sid = str(row.get("SBD", "") or "").strip()
            if sid:
                score_by_sid[sid] = row
        out_rows: list[dict[str, object]] = []
        seen: set[str] = set()
        if self.session:
            for st in (self.session.students or []):
                sid = str(getattr(st, "student_id", "") or "").strip()
                st_class = str(getattr(st, "class_name", "") or "").strip()
                if not sid or st_class != class_name:
                    continue
                rec: dict[str, object] = {
                    "SBD": sid,
                    "Họ tên": str(getattr(st, "name", "") or ""),
                    "Ngày sinh": self._format_birth_date_for_export(getattr(st, "birth_date", "") or ""),
                    "Lớp": st_class,
                }
                score_row = score_by_sid.get(sid, {})
                for header in score_headers:
                    if header in score_row:
                        rec[header] = score_row.get(header, "")
                out_rows.append(rec)
                seen.add(sid)
        for row in score_rows:
            sid = str(row.get("SBD", "") or "").strip()
            st_class = str(row.get("Lớp", "") or "").strip()
            if not sid or sid in seen or st_class != class_name:
                continue
            rec = {
                "SBD": sid,
                "Họ tên": row.get("Họ tên", ""),
                "Ngày sinh": self._format_birth_date_for_export(row.get("Ngày sinh", "")),
                "Lớp": st_class,
            }
            for header in score_headers:
                if header in row:
                    rec[header] = row.get(header, "")
            out_rows.append(rec)
        out_rows.sort(key=lambda x: str(x.get("SBD", "")))
        return out_rows

    def _class_options_for_export(self) -> list[str]:
        class_names: list[str] = []
        seen: set[str] = set()
        if self.session:
            for st in (self.session.students or []):
                class_name = str(getattr(st, "class_name", "") or "").strip()
                if class_name and class_name not in seen:
                    seen.add(class_name)
                    class_names.append(class_name)
        if class_names:
            return sorted(class_names)
        subjects = self._iter_export_subjects()
        _base_headers, _score_headers, rows = self._build_subject_score_matrix_rows(subjects)
        for row in rows:
            class_name = str(row.get("Lớp", "") or "").strip()
            if class_name and class_name not in seen:
                seen.add(class_name)
                class_names.append(class_name)
        return sorted(class_names)

    def _export_class_subject_scores(self) -> None:
        subjects = self._iter_export_subjects()
        if not subjects:
            QMessageBox.information(self, "Xuất điểm lớp", "Chưa có môn nào để xuất.")
            return
        class_options = self._class_options_for_export()
        if not class_options:
            QMessageBox.information(self, "Xuất điểm lớp", "Chưa có dữ liệu lớp để xuất.")
            return
        picked_class, ok = QInputDialog.getItem(self, "Xuất điểm lớp", "Chọn lớp cần xuất:", class_options, 0, False)
        if not ok:
            return
        class_name = str(picked_class or "").strip()
        if not class_name:
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Xuất điểm lớp",
            f"{self._safe_sheet_name(class_name, fallback='class_scores')}.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return
        from openpyxl import Workbook

        base_headers, score_headers, rows = self._build_subject_score_matrix_rows(subjects)
        class_rows = self._build_class_matrix_rows(class_name, rows, score_headers)
        if not class_rows:
            QMessageBox.information(self, "Xuất điểm lớp", "Lớp đã chọn chưa có dữ liệu học sinh để xuất.")
            return
        score_headers = self._trim_subject_headers_without_scores(class_rows, score_headers)
        wb = Workbook()
        ws = wb.active
        ws.title = self._safe_sheet_name(class_name, fallback="class_scores")
        self._write_subject_score_matrix_sheet(ws, base_headers, score_headers, class_rows)
        wb.save(path)
        QMessageBox.information(self, "Xuất điểm lớp", f"Đã xuất dữ liệu:\n{path}")

    def _export_all_classes_subject_scores(self) -> None:
        subjects = self._iter_export_subjects()
        if not subjects:
            QMessageBox.information(self, "Xuất điểm các lớp", "Chưa có môn nào để xuất.")
            return
        class_options = self._class_options_for_export()
        if not class_options:
            QMessageBox.information(self, "Xuất điểm các lớp", "Chưa có dữ liệu lớp để xuất.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Xuất điểm các lớp", "all_class_subject_scores.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook

        base_headers, score_headers, rows = self._build_subject_score_matrix_rows(subjects)
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        exported_count = 0
        for class_name in class_options:
            ws = wb.create_sheet(self._safe_sheet_name(class_name, fallback="class_scores"))
            class_rows = self._build_class_matrix_rows(class_name, rows, score_headers)
            if not class_rows:
                wb.remove(ws)
                continue
            exported_count += 1
            class_score_headers = self._trim_subject_headers_without_scores(class_rows, score_headers)
            self._write_subject_score_matrix_sheet(ws, base_headers, class_score_headers, class_rows)
        if exported_count <= 0:
            QMessageBox.information(self, "Xuất điểm các lớp", "Chưa có dữ liệu học sinh để xuất.")
            return
        wb.save(path)
        QMessageBox.information(self, "Xuất điểm các lớp", f"Đã xuất dữ liệu:\n{path}")

    def _export_all_subject_scores(self) -> None:
        subjects = self._iter_export_subjects()
        if not subjects:
            QMessageBox.information(self, "Xuất điểm chi tiết các môn", "Chưa có môn nào để xuất.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Xuất điểm chi tiết các môn", "all_subject_scores.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "summary"
        ws_summary.append(["subject_key", "subject_label", "student_count", "avg_score", "min_score", "max_score"])
        for idx, (label, key) in enumerate(subjects):
            rows = self._ensure_export_score_rows_for_subject(key)
            sheet_name = self._safe_sheet_name(label, fallback=f"subject_{idx + 1}")
            ws = wb.create_sheet(sheet_name)
            headers = ["student_id", "name", "exam_code", "correct", "wrong", "blank", "score", "status", "note"]
            ws.append(headers)
            scores: list[float] = []
            for row in rows:
                ws.append([
                    str(row.get("student_id", "") or ""),
                    str(row.get("name", "") or ""),
                    str(row.get("exam_code", "") or ""),
                    row.get("correct", 0),
                    row.get("wrong", 0),
                    row.get("blank", 0),
                    row.get("score", 0),
                    str(row.get("status", "") or ""),
                    str(row.get("note", "") or ""),
                ])
                score_value = self._score_value_for_statistics(row)
                if score_value is not None:
                    scores.append(score_value)
            if scores:
                ws_summary.append([key, label, len(scores), round(sum(scores) / len(scores), 4), min(scores), max(scores)])
            else:
                ws_summary.append([key, label, 0, 0, 0, 0])
        wb.save(path)
        QMessageBox.information(self, "Xuất điểm chi tiết các môn", f"Đã xuất dữ liệu:\n{path}")

    @staticmethod
    def _safe_file_component(value: str, fallback: str = "unknown") -> str:
        text = str(value or "").strip()
        if not text:
            return fallback
        normalized = unicodedata.normalize("NFKD", text)
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r"[\\/:*?\"<>|]+", "_", normalized)
        normalized = re.sub(r"\s+", "_", normalized).strip("._ ")
        return normalized or fallback

    def _export_return_by_class(self) -> None:
        if not self.session:
            QMessageBox.information(self, "Trả bài theo lớp", "Chưa có kỳ thi hiện tại.")
            return
        class_options = self._class_options_for_export()
        if not class_options:
            QMessageBox.information(self, "Trả bài theo lớp", "Chưa có dữ liệu lớp.")
            return
        picked_class, ok = QInputDialog.getItem(self, "Trả bài theo lớp", "Chọn lớp cần trả bài:", class_options, 0, False)
        if not ok:
            return
        class_name = str(picked_class or "").strip()
        if not class_name:
            return
        output_root = QFileDialog.getExistingDirectory(self, "Chọn thư mục lưu trả bài theo lớp")
        if not output_root:
            return

        student_meta = self._student_meta_by_sid()
        target_class_fold = class_name.casefold()
        class_dir = Path(output_root) / self._safe_file_component(class_name, fallback="class")
        class_dir.mkdir(parents=True, exist_ok=True)

        copied_count = 0
        missing_image_count = 0
        skipped_count = 0
        for subject_label, subject_key in self._iter_export_subjects():
            rows = self._scan_rows_for_subject(subject_key)
            if not rows:
                continue
            subject_dir = class_dir / self._safe_file_component(subject_label, fallback=self._safe_sheet_name(subject_key, fallback="subject"))
            subject_dir.mkdir(parents=True, exist_ok=True)
            used_names: set[str] = set()
            for row in rows:
                image_path = str(getattr(row, "image_path", "") or "").strip()
                sid = str(getattr(row, "student_id", "") or "").strip()
                meta = student_meta.get(sid, {})
                row_class = str(getattr(row, "class_name", "") or meta.get("class_name", "") or "").strip()
                if row_class.casefold() != target_class_fold:
                    continue
                if not image_path:
                    skipped_count += 1
                    continue
                source_path = Path(image_path)
                if not source_path.exists() or not source_path.is_file():
                    missing_image_count += 1
                    continue
                student_name = str(getattr(row, "full_name", "") or meta.get("name", "") or "").strip()
                base_name = f"{self._safe_file_component(sid, fallback='SBD')}_{self._safe_file_component(student_name, fallback='ho_ten')}"
                final_name = base_name
                ext = source_path.suffix
                dup_idx = 2
                while final_name in used_names or (subject_dir / f"{final_name}{ext}").exists():
                    final_name = f"{base_name}_{dup_idx}"
                    dup_idx += 1
                used_names.add(final_name)
                shutil.copy2(source_path, subject_dir / f"{final_name}{ext}")
                copied_count += 1

        QMessageBox.information(
            self,
            "Trả bài theo lớp",
            f"Đã xử lý xong.\n- Lớp: {class_name}\n- File đã copy: {copied_count}\n- File thiếu/không tồn tại: {missing_image_count}\n- Bản ghi bỏ qua (không có đường dẫn ảnh): {skipped_count}\n- Thư mục đích: {class_dir}",
        )

    def _export_recheck_package(self, group_by: str = "subject") -> None:
        from gui.export_reports_dialog import ExportReportsDialog

        dlg = ExportReportsDialog(self)
        dlg._package_recheck(group_by)

    def _export_subject_api_payload(self, subject_key: str) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        rows = self._scan_rows_for_subject(subject)
        if not rows:
            cfg = self._subject_config_by_subject_key(subject) or {}
            if self._subject_uses_direct_score_import(cfg):
                for payload in self._build_direct_score_payloads_for_subject(subject):
                    sid = str(payload.get("student_id", "") or "").strip()
                    rows.append(OMRResult(
                        image_path="",
                        student_id=sid,
                        exam_code=str(payload.get("exam_code", "") or ""),
                        answer_string="",
                    ))
                    if rows:
                        row_obj = rows[-1]
                        setattr(row_obj, "full_name", str(payload.get("name", "") or ""))
                        setattr(row_obj, "class_name", str(payload.get("class_name", "") or ""))
                        setattr(row_obj, "exam_room", str(payload.get("exam_room", "") or ""))
        if not rows:
            QMessageBox.information(self, "Xuất API bài làm", "Môn này chưa có dữ liệu bài làm.")
            return
        cfg = self._subject_config_by_subject_key(subject) or {}
        suggested = self._safe_sheet_name(str(cfg.get("name", "") or "subject_api"), fallback="subject_api")
        path, _ = QFileDialog.getSaveFileName(self, "Xuất API bài làm", f"{suggested}_api.csv", "CSV (*.csv)")
        if not path:
            return
        headers = ["STT", "student_id", "name", "class_name", "exam_room", "subject_key", "exam_code", "answer_string"]
        student_meta = self._student_meta_by_sid()
        sid_display_by_normalized: dict[str, str] = {}
        for st in (self.session.students if self.session else []):
            sid_origin = str(getattr(st, "student_id", "") or "").strip()
            sid_norm = self._normalized_student_id_for_match(sid_origin)
            if sid_origin and sid_norm and sid_norm not in sid_display_by_normalized:
                sid_display_by_normalized[sid_norm] = sid_origin

        def _sorted_question_numbers(payload: object) -> list[int]:
            out: list[int] = []
            if isinstance(payload, dict):
                for key in payload.keys():
                    key_text = str(key or "").strip()
                    if key_text.lstrip("-").isdigit():
                        out.append(int(key_text))
            return sorted(set(out))

        def _ordered_question_numbers(valid_map: object, invalid_map: object, fallback_map: object) -> list[int]:
            primary_nums: set[int] = set()
            for src in [valid_map or {}, invalid_map or {}]:
                if not isinstance(src, dict):
                    continue
                for key in src.keys():
                    key_text = str(key or "").strip()
                    if key_text.lstrip("-").isdigit():
                        primary_nums.add(int(key_text))
            if primary_nums:
                return sorted(primary_nums)
            return _sorted_question_numbers(fallback_map)

        def _answer_string_for_api(result) -> str:
            manual_text = str(getattr(result, "manual_content_override", "") or "").strip()
            if manual_text:
                return manual_text.replace(",", ";")

            answer_key = self._subject_answer_key_for_result(result, subject)
            mcq_map = dict(getattr(result, "mcq_answers", {}) or {})
            tf_map = dict(getattr(result, "true_false_answers", {}) or {})
            numeric_map = dict(getattr(result, "numeric_answers", {}) or {})

            invalid_rows = getattr(answer_key, "invalid_answer_rows", {}) or {} if answer_key is not None else {}
            mcq_qs = _ordered_question_numbers(
                getattr(answer_key, "answers", {}) or {} if answer_key is not None else {},
                (invalid_rows.get("MCQ", {}) or {}),
                mcq_map,
            )
            tf_qs = _ordered_question_numbers(
                getattr(answer_key, "true_false_answers", {}) or {} if answer_key is not None else {},
                (invalid_rows.get("TF", {}) or {}),
                tf_map,
            )
            numeric_qs = _ordered_question_numbers(
                getattr(answer_key, "numeric_answers", {}) or {} if answer_key is not None else {},
                (invalid_rows.get("NUMERIC", {}) or {}),
                numeric_map,
            )

            parts: list[str] = []
            for q_no in mcq_qs:
                value = str(mcq_map.get(q_no, "") or "").strip().upper()[:1]
                parts.append(value or "_")
            for q_no in tf_qs:
                flags = dict(tf_map.get(q_no, {}) or {})
                token = "".join(
                    "Đ" if key in flags and bool(flags.get(key)) else ("S" if key in flags else "_")
                    for key in ["a", "b", "c", "d"]
                )
                parts.append(token or "____")
            for q_no in numeric_qs:
                value = str(numeric_map.get(q_no, "") or "").strip().replace(" ", "").lstrip("+").replace(".", ",")
                parts.append(value or "_")
            if parts:
                return ";".join(parts)

            fallback = str(getattr(result, "answer_string", "") or "").strip()
            return fallback.replace(",", ";") if fallback else ""

        with Path(path).open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for idx, result in enumerate(rows, start=1):
                sid_raw = str(getattr(result, "student_id", "") or "").strip()
                sid_display = sid_display_by_normalized.get(self._normalized_student_id_for_match(sid_raw), sid_raw)
                meta = student_meta.get(sid_display, student_meta.get(sid_raw, {}))
                sid_export = f'="{sid_display}"' if sid_display and sid_display.isdigit() else sid_display
                exam_room_text = str(self._subject_room_for_student_id(sid_display or sid_raw, cfg) or "").strip()
                if not exam_room_text:
                    exam_room_text = str(getattr(result, "exam_room", "") or meta.get("exam_room", ""))
                writer.writerow({
                    "STT": idx,
                    "student_id": sid_export,
                    "name": str(getattr(result, "full_name", "") or meta.get("name", "")),
                    "class_name": str(getattr(result, "class_name", "") or meta.get("class_name", "")),
                    "exam_room": exam_room_text,
                    "subject_key": subject,
                    "exam_code": str(getattr(result, "exam_code", "") or ""),
                    "answer_string": _answer_string_for_api(result),
                })
        QMessageBox.information(self, "Xuất API bài làm", f"Đã xuất API bài làm:\n{path}")
