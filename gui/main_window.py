from __future__ import annotations

import copy
import csv
import gc
import json
import re
import shutil
import sys
import unicodedata
from collections import deque
from datetime import date, datetime
from pathlib import Path
import time
import uuid

sys.dont_write_bytecode = True

from PySide6.QtCore import Qt, QEvent, QTimer
from PySide6.QtGui import QAction, QColor, QImage, QKeySequence, QPixmap, QTransform, QPainter, QPen
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QCheckBox,
    QCompleter,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QDoubleSpinBox,
    QFileDialog,
    QFrame,
    QFormLayout,
    QGridLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMessageBox,
    QProgressBar,
    QProgressDialog,
    QPushButton,
    QRadioButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QInputDialog,
    QToolBar,
    QStyle,
    QGroupBox,
    QStackedWidget,
    QScrollArea,
)

from core.answer_key_importer import ImportedAnswerKey, ImportedAnswerKeyPackage, import_answer_key
from core.omr_engine import OMRProcessor, OMRResult, RecognitionContext
from core.scoring_engine import ScoreResult, ScoringEngine
from editor.template_editor import TemplateEditorWindow
from gui.import_answer_key_dialog import ImportAnswerKeyDialog
from gui.export_reports_dialog import ExportReportsDialog
from gui.main_window_scoring import open_scoring_review_editor_dialog
from gui.main_window_recheck import open_recheck_dialog
from models.answer_key import AnswerKeyRepository, SubjectKey
from models.database import OMRDatabase, bootstrap_application_db
from models.exam_session import ExamSession, Student
from models.template import Template, ZoneType
from models.template_repository import TemplateRepository


class PreviewImageWidget(QLabel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAlignment(Qt.AlignCenter)
        self.setMinimumHeight(260)
        self._overlay_markers: list[dict[str, float]] = []

    def clear_markers(self) -> None:
        self._overlay_markers = []
        self.update()

    def set_overlay_markers(self, markers: list[dict[str, float]]) -> None:
        self._overlay_markers = [dict(m) for m in markers]
        self.update()

    def paintEvent(self, event):  # type: ignore[override]
        super().paintEvent(event)
        if not self._overlay_markers:
            return
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, True)

        if self._overlay_markers:
            painter.setPen(QPen(QColor(40, 130, 255), 2))
            for m in self._overlay_markers:
                x = float(m.get("x", 0.0))
                y = float(m.get("y", 0.0))
                r = 5
                painter.drawLine(int(x - r), int(y - r), int(x + r), int(y + r))
                painter.drawLine(int(x - r), int(y + r), int(x + r), int(y - r))


class SubjectConfigDialog(QDialog):
    @staticmethod
    def default_section_scores() -> dict:
        return {
            "MCQ": {"total_points": 3.0, "distribution": "auto_by_question_count"},
            "TF": {
                "total_points": 2.0,
                "rule_per_question": {"1": 0.1, "2": 0.25, "3": 0.5, "4": 1.0},
            },
            "NUMERIC": {"total_points": 2.0, "distribution": "auto_by_question_count"},
        }

    @staticmethod
    def default_question_scores() -> dict:
        return {
            "MCQ": {"per_question": 0.25},
            "TF": {"1": 0.1, "2": 0.25, "3": 0.5, "4": 1.0},
            "NUMERIC": {"per_question": 1.0},
        }

    @staticmethod
    def _to_float(text: str, fallback: float = 0.0) -> float:
        try:
            return float((text or "").strip().replace(",", "."))
        except Exception:
            return fallback

    @staticmethod
    def _template_question_counts(template_path: str) -> dict[str, int]:
        counts = {"MCQ": 0, "TF": 0, "NUMERIC": 0}
        if not template_path:
            return counts
        path = Path(template_path)
        if not path.exists():
            return counts
        try:
            tpl = Template.load_json(path)
        except Exception:
            return counts
        for z in tpl.zones:
            if not z.grid:
                continue
            if z.zone_type.value == "MCQ_BLOCK":
                c = int(z.grid.question_count or z.grid.rows or 0)
                counts["MCQ"] += max(0, c)
            elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                # TF grid usually has rows = questions * statements_per_question.
                # `grid.question_count` can be stale/legacy in some templates, so derive from rows first.
                spq = max(1, int(z.metadata.get("statements_per_question", 4) or 4))
                from_rows = int((z.grid.rows or 0) // spq)
                from_meta = int(z.metadata.get("questions_per_block", 0) or 0)
                from_grid = int(z.grid.question_count or 0)
                candidates = [x for x in [from_rows, from_meta, from_grid] if x > 0]
                c = min(candidates) if candidates else 0
                counts["TF"] += max(0, c)
            elif z.zone_type.value == "NUMERIC_BLOCK":
                # Numeric grid usually has cols = questions * digits_per_answer.
                dpa = max(1, int(z.metadata.get("digits_per_answer", 3) or 3))
                from_cols = int((z.grid.cols or 0) // dpa)
                from_meta = int(z.metadata.get("questions_per_block", z.metadata.get("total_questions", 0)) or 0)
                from_grid = int(z.grid.question_count or 0)
                candidates = [x for x in [from_cols, from_meta, from_grid] if x > 0]
                c = min(candidates) if candidates else 0
                counts["NUMERIC"] += max(0, c)
        return counts

    @staticmethod
    def _answer_key_question_counts(answer_key_data: dict) -> dict[str, int]:
        counts = {"MCQ": 0, "TF": 0, "NUMERIC": 0}
        if not isinstance(answer_key_data, dict) or not answer_key_data:
            return counts
        # Total score is defined per one exam code; use the first code as representative.
        first_code = sorted(answer_key_data.keys())[0]
        key_data = answer_key_data.get(first_code, {}) or {}
        for sec, bucket_name in [("MCQ", "mcq_answers"), ("TF", "true_false_answers"), ("NUMERIC", "numeric_answers")]:
            valid_qs = {
                int(q) for q in (key_data.get(bucket_name, {}) or {}).keys()
                if str(q).strip().lstrip("-").isdigit()
            }
            full_qs = {
                int(q) for q in ((key_data.get("full_credit_questions", {}) or {}).get(sec, []) or [])
                if str(q).strip().lstrip("-").isdigit()
            }
            invalid_qs = {
                int(q) for q in ((key_data.get("invalid_answer_rows", {}) or {}).get(sec, {}) or {}).keys()
                if str(q).strip().lstrip("-").isdigit()
            }
            counts[sec] = len(valid_qs | full_qs | invalid_qs)
        return counts

    @staticmethod
    def _template_part_count(template_path: str, fallback: int = 3) -> int:
        counts = SubjectConfigDialog._template_question_counts(template_path)
        parts = sum(1 for k in counts if counts[k] > 0)
        return parts if parts > 0 else fallback

    def __init__(
        self,
        data: dict | None = None,
        subject_options: list[str] | None = None,
        block_options: list[str] | None = None,
        paper_part_count: int = 3,
        common_template_path: str = "",
        template_repo: TemplateRepository | None = None,
        parent=None,
    ):
        super().__init__(parent)
        self.setWindowTitle("Cấu hình môn học")
        self.resize(980, 760)
        self.setMinimumSize(880, 680)
        self.setWindowState(self.windowState() | Qt.WindowMaximized)
        data = data or {}
        subject_options = subject_options or []
        block_options = block_options or ["10", "11", "12"]

        self.common_template_path = common_template_path
        self.template_repo = template_repo or TemplateRepository()
        self.paper_part_count_default = paper_part_count
        self.answer_key_data: dict = dict(data.get("imported_answer_keys", {}))
        if not self.answer_key_data:
            db = getattr(parent, "database", None)
            subject_key_seed = str(data.get("answer_key_key", "") or "").strip()
            if not subject_key_seed:
                seed_name = str(data.get("name", "") or "").strip()
                seed_block = str(data.get("block", "") or "").strip()
                subject_key_seed = f"{seed_name}_{seed_block}" if seed_name and seed_block else ""
            if db is not None and subject_key_seed:
                try:
                    scoped_seed = subject_key_seed
                    if parent is not None and hasattr(parent, "_answer_key_scope_key"):
                        try:
                            scoped_seed = parent._answer_key_scope_key(subject_key_seed, str(data.get("block", "") or ""))
                        except Exception:
                            scoped_seed = subject_key_seed
                    fetched = db.fetch_answer_keys_for_subject(scoped_seed)
                    if not fetched and scoped_seed != subject_key_seed:
                        fetched = db.fetch_answer_keys_for_subject(subject_key_seed)
                    if fetched:
                        self.answer_key_data = fetched
                except Exception:
                    pass

        lay = QVBoxLayout(self)
        form = QFormLayout()
        form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

        self.subject_name = QComboBox(); self.subject_name.setEditable(True); self.subject_name.addItems(subject_options)
        if str(data.get("name", "")).strip():
            self.subject_name.setCurrentText(str(data.get("name", "")).strip())

        self.block_name = QComboBox(); self.block_name.setEditable(True); self.block_name.addItems(block_options)
        self.block_name.setCurrentText(str(data.get("block", block_options[0] if block_options else "10")))
        self.is_essay_subject = QCheckBox("Môn tự luận")
        self.is_essay_subject.setChecked(bool(data.get("is_essay_subject", False)))
        self.auto_recognize = QCheckBox("Tự động nhận dạng")
        self.auto_recognize.setChecked(bool(data.get("auto_recognize", False)))

        self.template_path = QLineEdit(str(data.get("template_path", "")))
        self.scan_folder = QLineEdit(str(data.get("scan_folder", "")))
        self.answer_key = QLineEdit(str(data.get("answer_key_path", "")))
        self.answer_key_key = QLineEdit(str(data.get("answer_key_key", ""))); self.answer_key_key.setReadOnly(True)
        self._fit_line_edit_to_text(self.template_path, min_chars=42, max_chars=78, padding=30)
        self._fit_line_edit_to_text(self.scan_folder, min_chars=42, max_chars=78, padding=30)
        self._fit_line_edit_to_text(self.answer_key, min_chars=42, max_chars=78, padding=30)
        self._fit_line_edit_to_text(self.answer_key_key, min_chars=20, max_chars=36, padding=28)
        self.template_path.setClearButtonEnabled(True)
        self.scan_folder.setClearButtonEnabled(True)
        self.answer_key.setClearButtonEnabled(True)
        self.template_path.textChanged.connect(lambda _text: self._fit_line_edit_to_text(self.template_path, min_chars=42, max_chars=78, padding=30))
        self.scan_folder.textChanged.connect(lambda _text: self._fit_line_edit_to_text(self.scan_folder, min_chars=42, max_chars=78, padding=30))
        self.answer_key.textChanged.connect(lambda _text: self._fit_line_edit_to_text(self.answer_key, min_chars=42, max_chars=78, padding=30))
        self.answer_key_key.textChanged.connect(lambda _text: self._fit_line_edit_to_text(self.answer_key_key, min_chars=20, max_chars=36, padding=28))
        self._exam_room_mapping_cache: dict[str, list[str]] = {}
        self.exam_room_name = QComboBox(); self.exam_room_name.setEditable(True)
        seed_room = str(data.get("exam_room_name", "") or "").strip()
        if seed_room:
            self.exam_room_name.addItem(seed_room)
            self.exam_room_name.setCurrentText(seed_room)
        self.exam_room_mapping_selector = QComboBox()
        self.exam_room_mapping_selector.setMinimumWidth(260)
        self.exam_room_mapping_selector.currentIndexChanged.connect(self._on_exam_room_mapping_selected)
        self.btn_import_exam_room_mapping = QPushButton("Import...")
        self.btn_import_exam_room_mapping.clicked.connect(self._import_exam_room_mapping_from_file)
        self.btn_delete_exam_room_mapping = QPushButton("Xóa phòng")
        self.btn_delete_exam_room_mapping.clicked.connect(self._remove_selected_exam_room_mapping)
        self.exam_room_mapping_hint = QLabel("Chưa nạp mapping SBD/phòng.")
        self.exam_room_mapping_hint.setWordWrap(False)
        room_map_wrap = QWidget()
        room_map_lay = QHBoxLayout(room_map_wrap)
        room_map_lay.setContentsMargins(0, 0, 0, 0)
        room_map_lay.setSpacing(8)
        room_map_lay.addWidget(self.exam_room_mapping_selector, 0)
        room_map_lay.addWidget(self.btn_import_exam_room_mapping, 0)
        room_map_lay.addWidget(self.btn_delete_exam_room_mapping, 0)
        room_map_lay.addWidget(self.exam_room_mapping_hint, 1)
        self._restore_exam_room_mapping_from_data(data, seed_room)
        self.answer_codes = QLineEdit(", ".join(sorted((data.get("imported_answer_keys") or {}).keys()))); self.answer_codes.setReadOnly(True)
        self._fit_line_edit_to_text(self.answer_codes, min_chars=24, max_chars=56, padding=28)
        self.answer_codes.textChanged.connect(lambda _text: self._fit_line_edit_to_text(self.answer_codes, min_chars=24, max_chars=56, padding=28))
        self.answer_summary = QTextEdit()
        self.answer_summary.setReadOnly(True)
        self.answer_summary.setFixedSize(640, 170)
        self.answer_summary.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.answer_summary.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        self.direct_score_import_data: dict = copy.deepcopy(data.get("direct_score_import", {}) or {})
        self.direct_score_import_path = QLineEdit(str(self.direct_score_import_data.get("path", "") or ""))
        self.direct_score_import_path.setReadOnly(True)
        self._fit_line_edit_to_text(self.direct_score_import_path, min_chars=42, max_chars=78, padding=30)
        self.direct_score_import_path.textChanged.connect(lambda _text: self._fit_line_edit_to_text(self.direct_score_import_path, min_chars=42, max_chars=78, padding=30))
        self.btn_import_direct_score = QPushButton("Import file điểm...")
        self.btn_import_direct_score.clicked.connect(self._import_direct_score_file)
        self.direct_score_import_summary = QTextEdit()
        self.direct_score_import_summary.setReadOnly(True)
        self.direct_score_import_summary.setFixedSize(640, 160)
        self.direct_score_import_summary.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.direct_score_import_summary.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._refresh_direct_score_import_summary()

        self.paper_part_label = QLabel(str(paper_part_count))

        self.score_mode = QComboBox(); self.score_mode.addItems(["Điểm theo phần", "Điểm theo câu"])
        self.score_mode.setCurrentText(str(data.get("score_mode", "Điểm theo phần")))

        sec = data.get("section_scores", self.default_section_scores())
        self.sec_mcq_total = QLineEdit(str((sec.get("MCQ") or {}).get("total_points", 3.0)))
        self.sec_tf_total = QLineEdit(str((sec.get("TF") or {}).get("total_points", 2.0)))
        self.sec_tf_1 = QLineEdit(str(((sec.get("TF") or {}).get("rule_per_question") or {}).get("1", 0.1)))
        self.sec_tf_2 = QLineEdit(str(((sec.get("TF") or {}).get("rule_per_question") or {}).get("2", 0.25)))
        self.sec_tf_3 = QLineEdit(str(((sec.get("TF") or {}).get("rule_per_question") or {}).get("3", 0.5)))
        self.sec_tf_4 = QLineEdit(str(((sec.get("TF") or {}).get("rule_per_question") or {}).get("4", 1.0)))
        self.sec_numeric_total = QLineEdit(str((sec.get("NUMERIC") or {}).get("total_points", 2.0)))

        qsc = data.get("question_scores", self.default_question_scores())
        self.q_mcq = QLineEdit(str((qsc.get("MCQ") or {}).get("per_question", 0.25)))
        self.q_tf_1 = QLineEdit(str((qsc.get("TF") or {}).get("1", 0.1)))
        self.q_tf_2 = QLineEdit(str((qsc.get("TF") or {}).get("2", 0.25)))
        self.q_tf_3 = QLineEdit(str((qsc.get("TF") or {}).get("3", 0.5)))
        self.q_tf_4 = QLineEdit(str((qsc.get("TF") or {}).get("4", 1.0)))
        self.q_numeric = QLineEdit(str((qsc.get("NUMERIC") or {}).get("per_question", 1.0)))

        self.total_score = QLineEdit(); self.total_score.setReadOnly(True)

        self._score_line_edits = [
            self.sec_mcq_total, self.sec_tf_total, self.sec_tf_1, self.sec_tf_2, self.sec_tf_3, self.sec_tf_4, self.sec_numeric_total,
            self.q_mcq, self.q_tf_1, self.q_tf_2, self.q_tf_3, self.q_tf_4, self.q_numeric, self.total_score,
        ]
        for _score_edit in self._score_line_edits:
            _score_edit.setAlignment(Qt.AlignRight)
            self._fit_line_edit_to_text(_score_edit)
            _score_edit.textChanged.connect(lambda _text, widget=_score_edit: self._fit_line_edit_to_text(widget))

        row_tpl = QHBoxLayout(); row_tpl.setContentsMargins(0, 0, 0, 0); row_tpl.addWidget(self.template_path); b_tpl = QPushButton("..."); row_tpl.addWidget(b_tpl); b_tpl_repo = QPushButton("Kho mẫu..."); row_tpl.addWidget(b_tpl_repo); row_tpl.addStretch(1)
        b_tpl.clicked.connect(self._browse_template)
        b_tpl_repo.clicked.connect(self._pick_template_from_repo)
        row_scan = QHBoxLayout(); row_scan.setContentsMargins(0, 0, 0, 0); row_scan.addWidget(self.scan_folder); b_scan = QPushButton("..."); row_scan.addWidget(b_scan); row_scan.addStretch(1)
        b_scan.clicked.connect(self._browse_scan_folder)
        row_key = QHBoxLayout(); row_key.setContentsMargins(0, 0, 0, 0); row_key.addWidget(self.answer_key); b_key = QPushButton("..."); row_key.addWidget(b_key)
        b_key_view = QPushButton("Xem/Sửa đáp án...")
        row_key.addWidget(b_key_view)
        row_key.addStretch(1)
        b_key.clicked.connect(self._browse_answer_key)
        b_key_view.clicked.connect(self._edit_current_answer_keys)

        subject_row_wrap = QWidget()
        subject_row_lay = QHBoxLayout(subject_row_wrap)
        subject_row_lay.setContentsMargins(0, 0, 0, 0)
        subject_row_lay.setSpacing(10)
        subject_row_lay.addWidget(self.subject_name, 1)
        subject_row_lay.addWidget(self.is_essay_subject, 0, Qt.AlignRight)

        top_form = QFormLayout()
        top_form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        top_form.addRow("Tên môn", subject_row_wrap)
        top_form.addRow("Khối", self.block_name)
        lay.addLayout(top_form)

        self.standard_config_widget = QWidget()
        standard_lay = QVBoxLayout(self.standard_config_widget)
        standard_lay.setContentsMargins(0, 0, 0, 0)
        standard_lay.setSpacing(10)
        standard_form = QFormLayout()
        standard_form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        standard_form.addRow("Giấy thi riêng (tùy chọn)", row_tpl)
        standard_form.addRow("Thư mục bài thi môn", row_scan)
        standard_form.addRow("", self.auto_recognize)
        standard_form.addRow("Đáp án môn", row_key)
        standard_form.addRow("Mã đáp án môn_khối", self.answer_key_key)
        standard_form.addRow("Danh sách SBD phòng thi", room_map_wrap)
        standard_form.addRow("Các mã đề của môn", self.answer_codes)
        standard_form.addRow("Tóm tắt đáp án", self.answer_summary)
        standard_form.addRow("Số phần giấy thi", self.paper_part_label)
        standard_form.addRow("Cách nhập điểm", self.score_mode)

        self.section_group = QGroupBox("Điểm theo phần")
        sec_form = QFormLayout(self.section_group)
        sec_form.setFieldGrowthPolicy(QFormLayout.FieldsStayAtSizeHint)
        sec_form.addRow("MCQ tổng điểm", self.sec_mcq_total)
        sec_form.addRow("TF tổng điểm", self.sec_tf_total)
        sec_form.addRow("TF đúng 1 ý", self.sec_tf_1)
        sec_form.addRow("TF đúng 2 ý", self.sec_tf_2)
        sec_form.addRow("TF đúng 3 ý", self.sec_tf_3)
        sec_form.addRow("TF đúng 4 ý", self.sec_tf_4)
        sec_form.addRow("NUMERIC tổng điểm", self.sec_numeric_total)

        self.question_group = QGroupBox("Điểm theo câu")
        q_form = QFormLayout(self.question_group)
        q_form.setFieldGrowthPolicy(QFormLayout.FieldsStayAtSizeHint)
        q_form.addRow("MCQ điểm/câu", self.q_mcq)
        q_form.addRow("TF đúng 1 ý", self.q_tf_1)
        q_form.addRow("TF đúng 2 ý", self.q_tf_2)
        q_form.addRow("TF đúng 3 ý", self.q_tf_3)
        q_form.addRow("TF đúng 4 ý", self.q_tf_4)
        q_form.addRow("NUMERIC điểm/câu", self.q_numeric)

        standard_form.addRow("Tổng điểm bài thi", self.total_score)
        standard_lay.addLayout(standard_form)
        standard_lay.addWidget(self.section_group)
        standard_lay.addWidget(self.question_group)
        lay.addWidget(self.standard_config_widget)

        self.essay_import_group = QGroupBox("Import điểm trực tiếp theo SBD")
        essay_form = QFormLayout(self.essay_import_group)
        essay_form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        direct_file_row = QHBoxLayout()
        direct_file_row.setContentsMargins(0, 0, 0, 0)
        direct_file_row.addWidget(self.direct_score_import_path)
        direct_file_row.addWidget(self.btn_import_direct_score)
        direct_file_row.addStretch(1)
        essay_form.addRow("File điểm", direct_file_row)
        direct_note = QLabel("Bắt buộc chọn cột SBD và cột Điểm. Cột Phòng thi là tùy chọn; nếu không có, hệ thống sẽ dùng phòng thi chung của kỳ thi.")
        direct_note.setWordWrap(True)
        essay_form.addRow("Quy tắc import", direct_note)
        essay_form.addRow("Tóm tắt import", self.direct_score_import_summary)
        lay.addWidget(self.essay_import_group)

        self.subject_name.currentTextChanged.connect(self._update_answer_key_key)
        self.block_name.currentTextChanged.connect(self._update_answer_key_key)
        self.is_essay_subject.toggled.connect(self._refresh_subject_mode_ui)
        self.score_mode.currentTextChanged.connect(self._refresh_score_mode_ui)
        self.template_path.textChanged.connect(self._update_paper_parts)

        for w in [self.sec_mcq_total, self.sec_tf_total, self.sec_tf_1, self.sec_tf_2, self.sec_tf_3, self.sec_tf_4, self.sec_numeric_total,
                  self.q_mcq, self.q_tf_1, self.q_tf_2, self.q_tf_3, self.q_tf_4, self.q_numeric]:
            w.textChanged.connect(self._update_total_score)

        self._update_answer_key_key()
        self._refresh_answer_key_summary()
        self._update_paper_parts()
        self._refresh_score_mode_ui()
        self._refresh_subject_mode_ui()

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        lay.addWidget(bb)

    @staticmethod
    def _normalize_exam_room_sid_values(raw_values: object) -> list[str]:
        tokens: list[str] = []
        if isinstance(raw_values, str):
            tokens = re.split(r"[,\n;]+", raw_values)
        elif isinstance(raw_values, (list, tuple, set)):
            tokens = [str(x) for x in raw_values]
        else:
            return []
        clean = [str(x or "").strip() for x in tokens if str(x or "").strip()]
        return sorted(set(clean))

    def _restore_exam_room_mapping_from_data(self, data: dict, seed_room: str = "") -> None:
        rebuilt: dict[str, list[str]] = {}
        mapping_by_room_seed = data.get("exam_room_sbd_mapping_by_room", {})
        if isinstance(mapping_by_room_seed, dict) and mapping_by_room_seed:
            for room_key, values in mapping_by_room_seed.items():
                room_text = str(room_key or "").strip()
                if not room_text:
                    continue
                normalized_values = self._normalize_exam_room_sid_values(values)
                if normalized_values:
                    rebuilt[room_text] = normalized_values

        if not rebuilt:
            mapping_seed = str(data.get("exam_room_sbd_mapping", "") or "").strip()
            normalized_values = self._normalize_exam_room_sid_values(mapping_seed)
            if normalized_values:
                room_key = str(seed_room or "").strip() or "[Không rõ phòng]"
                rebuilt[room_key] = normalized_values

        self._exam_room_mapping_cache = dict(rebuilt)
        self._refresh_exam_room_mapping_selector()

        preferred_room = str(seed_room or "").strip()
        if preferred_room and preferred_room in self._exam_room_mapping_cache:
            self.exam_room_name.setCurrentText(preferred_room)
            idx = self.exam_room_mapping_selector.findData(preferred_room)
            if idx >= 0:
                self.exam_room_mapping_selector.setCurrentIndex(idx)
        elif self._exam_room_mapping_cache:
            if not self.exam_room_name.currentText().strip():
                first_room = sorted(self._exam_room_mapping_cache.keys())[0]
                self.exam_room_name.setCurrentText(first_room)
            current_room = self.exam_room_name.currentText().strip()
            idx = self.exam_room_mapping_selector.findData(current_room)
            if idx >= 0:
                self.exam_room_mapping_selector.setCurrentIndex(idx)
        if not self._exam_room_mapping_cache:
            self.exam_room_mapping_hint.setText("Chưa nạp mapping SBD/phòng.")

    def _update_paper_parts(self) -> None:
        tpl = self.template_path.text().strip() or self.common_template_path
        part_count = self._template_part_count(tpl, self.paper_part_count_default)
        self.paper_part_label.setText(str(part_count))
        self._update_total_score()

    def _question_counts(self) -> dict[str, int]:
        key_counts = self._answer_key_question_counts(self.answer_key_data)
        if any(key_counts.values()):
            return key_counts
        tpl = self.template_path.text().strip() or self.common_template_path
        return self._template_question_counts(tpl)

    @staticmethod
    def _fit_line_edit_to_text(widget: QLineEdit, min_chars: int = 5, max_chars: int = 14, padding: int = 24) -> None:
        text = str(widget.text() or "").strip()
        metrics = widget.fontMetrics()
        min_width = metrics.horizontalAdvance("0" * max(1, min_chars)) + padding
        sample = text if text else ("0" * max(1, min_chars))
        width = metrics.horizontalAdvance(sample[:max_chars]) + padding
        width = max(min_width, width)
        width = min(width, metrics.horizontalAdvance("0" * max_chars) + padding)
        widget.setFixedWidth(width)
        widget.setToolTip(widget.text() or widget.placeholderText() or "")

    def _refresh_score_mode_ui(self) -> None:
        keep_size = self.size()
        section_mode = self.score_mode.currentText() == "Điểm theo phần"
        self.section_group.setVisible(section_mode)
        self.question_group.setVisible(not section_mode)
        self.layout().activate()
        if keep_size.isValid() and not self.isMaximized():
            self.resize(keep_size)
        self._update_total_score()

    def _update_total_score(self) -> None:
        if self.score_mode.currentText() == "Điểm theo phần":
            total = (
                self._to_float(self.sec_mcq_total.text())
                + self._to_float(self.sec_tf_total.text())
                + self._to_float(self.sec_numeric_total.text())
            )
        else:
            counts = self._question_counts()
            tf_max = max(
                self._to_float(self.q_tf_1.text()),
                self._to_float(self.q_tf_2.text()),
                self._to_float(self.q_tf_3.text()),
                self._to_float(self.q_tf_4.text()),
            )
            total = (
                self._to_float(self.q_mcq.text()) * counts.get("MCQ", 0)
                + tf_max * counts.get("TF", 0)
                + self._to_float(self.q_numeric.text()) * counts.get("NUMERIC", 0)
            )
        self.total_score.setText(f"{round(total, 4)}")

    def _update_answer_key_key(self) -> None:
        subject = self.subject_name.currentText().strip()
        block = self.block_name.currentText().strip()
        self.answer_key_key.setText(f"{subject}_{block}" if subject and block else "")

    def _browse_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Chọn giấy thi", "", "JSON (*.json)")
        if path:
            self.template_repo.register(path)
            self.template_path.setText(path)


    def _pick_template_from_repo(self) -> None:
        items = [f"{name} | {path}" for name, path in self.template_repo.list_templates()]
        if not items:
            QMessageBox.information(self, "Kho mẫu giấy thi", "Kho mẫu đang trống. Hãy thêm mẫu bằng nút ...")
            return
        chosen, ok = QInputDialog.getItem(self, "Kho mẫu giấy thi", "Chọn mẫu:", items, 0, False)
        if ok and chosen:
            self.template_path.setText(chosen.split(" | ", 1)[1])

    def _browse_scan_folder(self) -> None:
        path = QFileDialog.getExistingDirectory(self, "Chọn thư mục bài thi môn")
        if path:
            self.scan_folder.setText(path)

    def _browse_answer_key(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Chọn đáp án môn", "", "Answer files (*.json *.xlsx *.csv)")
        if not path:
            return
        try:
            imported_package = import_answer_key(path)
        except Exception as exc:
            message = (
                f"Không thể import đáp án:\n{exc}\n\n"
                "Bạn có muốn tiếp tục import các câu hợp lệ không?\n"
                "- Yes: Vẫn import và CHO ĐIỂM TỐI ĐA cho câu đáp án không đúng chuẩn.\n"
                "- No: Vẫn import nhưng BỎ QUA câu đáp án không đúng chuẩn (không chấm câu đó).\n"
                "- Cancel: Hủy import."
            )
            choose = QMessageBox.question(
                self,
                "Import đáp án",
                message,
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.Yes,
            )
            if choose == QMessageBox.Cancel:
                return
            try:
                imported_package = import_answer_key(
                    path,
                    strict=False,
                    award_full_credit_for_invalid=(choose == QMessageBox.Yes),
                )
            except Exception as inner_exc:
                QMessageBox.warning(self, "Import đáp án", f"Không thể import đáp án:\n{inner_exc}")
                return

        if imported_package.warnings:
            QMessageBox.information(
                self,
                "Import đáp án",
                "Import hoàn tất với cảnh báo:\n- " + "\n- ".join(imported_package.warnings[:20])
                + ("\n..." if len(imported_package.warnings) > 20 else ""),
            )

        dlg = ImportAnswerKeyDialog(imported_package, self)
        if dlg.exec() != QDialog.Accepted:
            return
        edited_package = dlg.result_answer_key()
        if not edited_package.exam_keys:
            QMessageBox.warning(self, "Import đáp án", "Không có mã đề nào trong file đáp án.")
            return

        # One subject-block can have multiple exam codes.
        self.answer_key_data = {}
        for code, key in edited_package.exam_keys.items():
            self.answer_key_data[code] = {
                "mcq_answers": key.mcq_answers,
                "true_false_answers": key.true_false_answers,
                "numeric_answers": key.numeric_answers,
                "full_credit_questions": key.full_credit_questions,
                "invalid_answer_rows": key.invalid_answer_rows,
            }
        self.answer_codes.setText(", ".join(sorted(self.answer_key_data.keys())))
        self.answer_key.setText(path)
        self._refresh_answer_key_summary()
        self._update_total_score()
        QMessageBox.information(self, "Import đáp án", "Đã gắn toàn bộ mã đề của file đáp án cho môn đang cấu hình.")

    @staticmethod
    def _build_imported_package_from_answer_data(answer_key_data: dict) -> ImportedAnswerKeyPackage:
        package = ImportedAnswerKeyPackage()
        for exam_code, payload in sorted((answer_key_data or {}).items()):
            if not isinstance(payload, dict):
                continue
            key = ImportedAnswerKey()
            key.mcq_answers = {
                int(k): str(v)
                for k, v in (payload.get("mcq_answers", {}) or {}).items()
                if str(k).strip().lstrip("-").isdigit()
            }
            key.true_false_answers = {
                int(k): {str(sub): bool(flag) for sub, flag in (flags or {}).items()}
                for k, flags in (payload.get("true_false_answers", {}) or {}).items()
                if str(k).strip().lstrip("-").isdigit()
            }
            key.numeric_answers = {
                int(k): str(v)
                for k, v in (payload.get("numeric_answers", {}) or {}).items()
                if str(k).strip().lstrip("-").isdigit()
            }
            key.full_credit_questions = {
                str(sec): [int(x) for x in (vals or []) if str(x).strip().lstrip("-").isdigit()]
                for sec, vals in (payload.get("full_credit_questions", {}) or {}).items()
            }
            key.invalid_answer_rows = {
                str(sec): {int(k): str(v) for k, v in (vals or {}).items()}
                for sec, vals in (payload.get("invalid_answer_rows", {}) or {}).items()
            }
            package.exam_keys[str(exam_code)] = key
        return package

    @staticmethod
    def _answer_payload_from_package(package: ImportedAnswerKeyPackage) -> dict[str, dict]:
        payload: dict[str, dict] = {}
        for code, key in (package.exam_keys or {}).items():
            payload[str(code)] = {
                "mcq_answers": {int(k): str(v) for k, v in (key.mcq_answers or {}).items()},
                "true_false_answers": {
                    int(k): {str(sub): bool(flag) for sub, flag in (flags or {}).items()}
                    for k, flags in (key.true_false_answers or {}).items()
                },
                "numeric_answers": {int(k): str(v) for k, v in (key.numeric_answers or {}).items()},
                "full_credit_questions": {
                    str(sec): [int(x) for x in (vals or []) if str(x).strip().lstrip("-").isdigit()]
                    for sec, vals in (key.full_credit_questions or {}).items()
                },
                "invalid_answer_rows": {
                    str(sec): {int(k): str(v) for k, v in (vals or {}).items()}
                    for sec, vals in (key.invalid_answer_rows or {}).items()
                },
            }
        return payload

    @staticmethod
    def _describe_answer_key_data(answer_key_data: dict) -> str:
        if not isinstance(answer_key_data, dict) or not answer_key_data:
            return "Chưa có đáp án cho môn này."
        lines: list[str] = []
        for exam_code, payload in sorted(answer_key_data.items()):
            if not isinstance(payload, dict):
                continue
            lines.append(f"Mã đề {exam_code}:")
            mcq = ", ".join(
                f"C{int(q)}:{str(a)}" for q, a in sorted((payload.get('mcq_answers', {}) or {}).items(), key=lambda item: int(item[0]))
            ) or "-"
            tf = ", ".join(
                f"C{int(q)}:{''.join('Đ' if bool((flags or {}).get(ch)) else 'S' for ch in ['a','b','c','d'])}"
                for q, flags in sorted((payload.get("true_false_answers", {}) or {}).items(), key=lambda item: int(item[0]))
            ) or "-"
            numeric = ", ".join(
                f"C{int(q)}:{str(a)}" for q, a in sorted((payload.get('numeric_answers', {}) or {}).items(), key=lambda item: int(item[0]))
            ) or "-"
            invalid_descriptions: list[str] = []
            for sec, invalid_rows in sorted((payload.get("invalid_answer_rows", {}) or {}).items()):
                if not invalid_rows:
                    continue
                mode = "cho điểm tối đa/giữ mô tả nhập sai"
                vals = ", ".join(f"C{int(q)}:{str(v)}" for q, v in sorted(invalid_rows.items(), key=lambda item: int(item[0])))
                invalid_descriptions.append(f"{sec} [{mode}] {vals}")
            lines.append(f"  - MCQ: {mcq}")
            lines.append(f"  - TF: {tf}")
            lines.append(f"  - NUMERIC: {numeric}")
            if invalid_descriptions:
                lines.append("  - Dòng nhập sai vẫn giữ để chấm:")
                lines.extend(f"    * {item}" for item in invalid_descriptions)
        return "\n".join(lines)

    def _refresh_answer_key_summary(self) -> None:
        self.answer_codes.setText(", ".join(sorted(self.answer_key_data.keys())))
        self.answer_summary.setPlainText(self._describe_answer_key_data(self.answer_key_data))

    def _refresh_subject_mode_ui(self) -> None:
        is_essay = self.is_essay_subject.isChecked()
        self.standard_config_widget.setVisible(not is_essay)
        self.essay_import_group.setVisible(is_essay)

    def _refresh_direct_score_import_summary(self) -> None:
        payload = self.direct_score_import_data if isinstance(self.direct_score_import_data, dict) else {}
        rows = payload.get("rows", []) or []
        path_text = str(payload.get("path", "") or "")
        mapping = payload.get("mapping", {}) or {}
        self.direct_score_import_path.setText(path_text)
        if not rows:
            self.direct_score_import_summary.setPlainText(
                "Chưa nạp file điểm trực tiếp.\n\n"
                "Yêu cầu tối thiểu:\n"
                "- Có cột SBD\n"
                "- Có cột Điểm\n"
                "- Cột Phòng thi là tùy chọn; nếu không có sẽ dùng phòng thi chung của kỳ thi."
            )
            return
        preview_lines = []
        for rec in rows[:8]:
            sid = str((rec or {}).get("sid", "") or "").strip()
            score = str((rec or {}).get("score", "") or "").strip()
            room = str((rec or {}).get("room", "") or "").strip() or "[dùng phòng chung của kỳ thi]"
            preview_lines.append(f"- SBD {sid}: điểm {score} | phòng {room}")
        if len(rows) > 8:
            preview_lines.append(f"- ... còn {len(rows) - 8} dòng")
        summary = [
            f"Số dòng hợp lệ: {len(rows)}",
            f"Cột SBD: {mapping.get('student_id_column', '-')}",
            f"Cột Điểm: {mapping.get('score_column', '-')}",
            f"Cột Phòng thi: {mapping.get('room_column', '[không dùng]')}",
            "",
            "Xem trước:",
            *preview_lines,
        ]
        self.direct_score_import_summary.setPlainText("\n".join(summary))

    @staticmethod
    def _find_header_index(keys: set[str], combo: QComboBox) -> int:
        normalized_keys = {str(k or "").strip().lower().replace(" ", "").replace("_", "") for k in keys}
        for i in range(combo.count()):
            text = str(combo.itemText(i) or "").strip().lower().replace(" ", "").replace("_", "")
            if text in normalized_keys:
                return i
        return 0

    def _import_direct_score_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import điểm trực tiếp theo SBD",
            "",
            "Data files (*.xlsx *.csv *.txt *.tsv);;All files (*.*)",
        )
        if not path:
            return
        try:
            headers, raw_rows = self._load_exam_room_mapping_rows(Path(path))
        except Exception as exc:
            QMessageBox.warning(self, "Import điểm trực tiếp", f"Không đọc được file:\n{exc}")
            return
        if not headers or not raw_rows:
            QMessageBox.warning(self, "Import điểm trực tiếp", "Không tìm thấy dữ liệu hợp lệ trong file điểm.")
            return
        pick = QDialog(self)
        pick.setWindowTitle("Chọn cột import điểm trực tiếp")
        pick_l = QFormLayout(pick)
        sid_col = QComboBox(); sid_col.addItems(headers)
        score_col = QComboBox(); score_col.addItems(headers)
        room_col = QComboBox(); room_col.addItem("[Không dùng]", ""); room_col.addItems(headers)
        sid_col.setCurrentIndex(self._find_header_index({"sbd", "studentid", "student_id", "sobaodanh"}, sid_col))
        score_col.setCurrentIndex(self._find_header_index({"diem", "score", "mark", "point"}, score_col))
        room_col.setCurrentIndex(self._find_header_index({"phongthi", "examroom", "exam_room", "room"}, room_col))
        pick_l.addRow("Cột SBD *", sid_col)
        pick_l.addRow("Cột Điểm *", score_col)
        pick_l.addRow("Cột Phòng thi", room_col)
        pick_note = QLabel("Nếu không chọn cột Phòng thi, hệ thống sẽ dùng phòng thi chung của kỳ thi khi import.")
        pick_note.setWordWrap(True)
        pick_l.addRow("", pick_note)
        pick_btn = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        pick_btn.accepted.connect(pick.accept)
        pick_btn.rejected.connect(pick.reject)
        pick_l.addRow(pick_btn)
        if pick.exec() != QDialog.Accepted:
            return
        sid_key = sid_col.currentText().strip()
        score_key = score_col.currentText().strip()
        room_key = str(room_col.currentData() or room_col.currentText() or "").strip()
        rows: list[dict[str, str]] = []
        skipped = 0
        for rec in raw_rows:
            sid = str(rec.get(sid_key, "") or "").strip()
            score_text = str(rec.get(score_key, "") or "").strip()
            room = str(rec.get(room_key, "") or "").strip() if room_key and room_key != "[Không dùng]" else ""
            if not sid or not score_text:
                skipped += 1
                continue
            try:
                score_value = str(float(score_text.replace(",", "."))).rstrip("0").rstrip(".")
            except Exception:
                skipped += 1
                continue
            rows.append({"sid": sid, "score": score_value, "room": room})
        if not rows:
            QMessageBox.warning(self, "Import điểm trực tiếp", "Không có dòng hợp lệ sau khi chọn cột SBD và Điểm.")
            return
        self.direct_score_import_data = {
            "path": path,
            "mapping": {
                "student_id_column": sid_key,
                "score_column": score_key,
                "room_column": room_key if room_key and room_key != "[Không dùng]" else "",
                "fallback_room_policy": "use_exam_common_room_if_missing",
            },
            "rows": rows,
            "skipped_rows": skipped,
        }
        self._refresh_direct_score_import_summary()
        QMessageBox.information(self, "Import điểm trực tiếp", f"Đã nạp {len(rows)} dòng hợp lệ cho môn tự luận.")

    def _edit_current_answer_keys(self) -> None:
        if not self.answer_key_data:
            QMessageBox.information(self, "Đáp án môn", "Môn này chưa có đáp án. Hãy import file hoặc thêm đáp án trước.")
            return
        dlg = ImportAnswerKeyDialog(self._build_imported_package_from_answer_data(self.answer_key_data), self)
        if dlg.exec() != QDialog.Accepted:
            return
        edited_package = dlg.result_answer_key()
        self.answer_key_data = self._answer_payload_from_package(edited_package)
        self._refresh_answer_key_summary()
        self._update_total_score()
        QMessageBox.information(self, "Đáp án môn", "Đã cập nhật đáp án hiện tại. Bạn có thể tiếp tục sửa hoặc thay đáp án khác.")

    def _import_exam_room_mapping_from_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import mapping SBD và phòng thi",
            "",
            "Data files (*.xlsx *.csv *.txt *.tsv);;All files (*.*)",
        )
        if not path:
            return
        try:
            headers, raw_rows = self._load_exam_room_mapping_rows(Path(path))
        except Exception as exc:
            QMessageBox.warning(self, "Import mapping phòng thi", f"Không đọc được file:\n{exc}")
            return
        if not raw_rows:
            QMessageBox.warning(self, "Import mapping phòng thi", "Không tìm thấy dữ liệu hợp lệ (cần cột SBD).")
            return
        pick = QDialog(self)
        pick.setWindowTitle("Chọn cột mapping SBD/phòng")
        pick_l = QFormLayout(pick)
        sid_col = QComboBox(); sid_col.addItems(headers)
        room_col = QComboBox(); room_col.addItem("[Không dùng]", ""); room_col.addItems(headers)
        def _find_idx(keys: set[str], combo: QComboBox) -> int:
            for i in range(combo.count()):
                text = str(combo.itemText(i) or "").strip().lower().replace(" ", "").replace("_", "")
                if text in keys:
                    return i
            return 0
        sid_col.setCurrentIndex(_find_idx({"sbd", "studentid", "student_id", "sobaodanh"}, sid_col))
        room_col.setCurrentIndex(_find_idx({"phongthi", "examroom", "exam_room", "room"}, room_col))
        pick_l.addRow("Cột SBD", sid_col)
        pick_l.addRow("Cột phòng thi", room_col)
        pick_btn = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        pick_btn.accepted.connect(pick.accept)
        pick_btn.rejected.connect(pick.reject)
        pick_l.addRow(pick_btn)
        if pick.exec() != QDialog.Accepted:
            return
        sid_key = sid_col.currentText().strip()
        room_key = str(room_col.currentData() or room_col.currentText() or "").strip()
        rows: list[dict[str, str]] = []
        for rec in raw_rows:
            sid = str(rec.get(sid_key, "")).strip()
            room = str(rec.get(room_key, "")).strip() if room_key and room_key != "[Không dùng]" else ""
            if sid:
                rows.append({"sid": sid, "room": room})
        if not rows:
            QMessageBox.warning(self, "Import mapping phòng thi", "Không có dữ liệu hợp lệ sau khi chọn cột.")
            return
        grouped: dict[str, set[str]] = {}
        for r in rows:
            sid = str(r.get("sid", "")).strip()
            room = str(r.get("room", "")).strip() or "[Không rõ phòng]"
            if not sid:
                continue
            grouped.setdefault(room, set()).add(sid)
        self._exam_room_mapping_cache = {k: sorted(v) for k, v in grouped.items()}
        self._refresh_exam_room_mapping_selector()
        QMessageBox.information(self, "Import mapping phòng thi", f"Đã nạp mapping cho {len(self._exam_room_mapping_cache)} phòng thi.")

    def _refresh_exam_room_mapping_selector(self) -> None:
        self.exam_room_mapping_selector.clear()
        if not self._exam_room_mapping_cache:
            self.exam_room_mapping_selector.addItem("[Chưa có danh sách]", "")
            self.exam_room_mapping_hint.setText("Chưa nạp mapping SBD/phòng.")
            self.btn_delete_exam_room_mapping.setEnabled(False)
            return
        for room in sorted(self._exam_room_mapping_cache.keys()):
            cnt = len(self._exam_room_mapping_cache.get(room, []))
            self.exam_room_mapping_selector.addItem(f"{room} ({cnt} SBD)", room)
        current_room = self.exam_room_name.currentText().strip()
        if current_room:
            idx = self.exam_room_mapping_selector.findData(current_room)
            if idx >= 0:
                self.exam_room_mapping_selector.setCurrentIndex(idx)
        self.btn_delete_exam_room_mapping.setEnabled(self.exam_room_mapping_selector.count() > 0)
        self._on_exam_room_mapping_selected(self.exam_room_mapping_selector.currentIndex())

    def _on_exam_room_mapping_selected(self, _index: int) -> None:
        room = str(self.exam_room_mapping_selector.currentData() or "").strip()
        if room:
            self.exam_room_name.setCurrentText(room)
            count = len(self._exam_room_mapping_cache.get(room, []))
            self.exam_room_mapping_hint.setText(f"{room}: {count} SBD")
            self.btn_delete_exam_room_mapping.setEnabled(True)
        else:
            self.exam_room_mapping_hint.setText("Chưa nạp mapping SBD/phòng.")
            self.btn_delete_exam_room_mapping.setEnabled(False)

    def _remove_selected_exam_room_mapping(self) -> None:
        room = str(self.exam_room_mapping_selector.currentData() or "").strip()
        if not room:
            return
        self._exam_room_mapping_cache.pop(room, None)
        if self.exam_room_name.currentText().strip() == room:
            self.exam_room_name.setCurrentText("")
        if self._exam_room_mapping_cache and not self.exam_room_name.currentText().strip():
            self.exam_room_name.setCurrentText(sorted(self._exam_room_mapping_cache.keys())[0])
        self._refresh_exam_room_mapping_selector()

    @staticmethod
    def _load_exam_room_mapping_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
        ext = path.suffix.lower()
        rows: list[dict[str, str]] = []
        headers: list[str] = []
        if ext in {".csv", ".txt", ".tsv"}:
            raw = path.read_text(encoding="utf-8-sig", errors="ignore")
            dialect = csv.Sniffer().sniff(raw[:2048]) if raw.strip() else csv.excel
            reader = csv.DictReader(raw.splitlines(), dialect=dialect)
            headers = [str(h or "") for h in (reader.fieldnames or [])]
            for row in reader:
                rec = {str(k): str(v or "") for k, v in (row or {}).items()}
                if rec:
                    rows.append(rec)
            return headers, rows
        if ext == ".xlsx":
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.values)
            if not values:
                return headers, rows
            headers = [str(x or "") for x in values[0]]
            for data in values[1:]:
                rec: dict[str, str] = {}
                for idx, key in enumerate(headers):
                    rec[str(key)] = str(data[idx] if idx < len(data) and data[idx] is not None else "")
                if rec:
                    rows.append(rec)
            return headers, rows
        raise RuntimeError("Chỉ hỗ trợ .xlsx/.csv/.txt/.tsv")

    def payload(self) -> dict:
        def f(v: str, label: str) -> float:
            try:
                return float(v.strip().replace(",", "."))
            except Exception as exc:
                raise ImportError(f"Giá trị điểm '{label}' không hợp lệ: {v}") from exc

        section_scores = {
            "MCQ": {"total_points": f(self.sec_mcq_total.text(), "MCQ tổng điểm"), "distribution": "auto_by_question_count"},
            "TF": {
                "total_points": f(self.sec_tf_total.text(), "TF tổng điểm"),
                "rule_per_question": {
                    "1": f(self.sec_tf_1.text(), "TF đúng 1 ý"),
                    "2": f(self.sec_tf_2.text(), "TF đúng 2 ý"),
                    "3": f(self.sec_tf_3.text(), "TF đúng 3 ý"),
                    "4": f(self.sec_tf_4.text(), "TF đúng 4 ý"),
                },
            },
            "NUMERIC": {"total_points": f(self.sec_numeric_total.text(), "NUMERIC tổng điểm"), "distribution": "auto_by_question_count"},
        }
        question_scores = {
            "MCQ": {"per_question": f(self.q_mcq.text(), "MCQ điểm/câu")},
            "TF": {
                "1": f(self.q_tf_1.text(), "TF đúng 1 ý"),
                "2": f(self.q_tf_2.text(), "TF đúng 2 ý"),
                "3": f(self.q_tf_3.text(), "TF đúng 3 ý"),
                "4": f(self.q_tf_4.text(), "TF đúng 4 ý"),
            },
            "NUMERIC": {"per_question": f(self.q_numeric.text(), "NUMERIC điểm/câu")},
        }

        return {
            "name": self.subject_name.currentText().strip(),
            "block": self.block_name.currentText().strip(),
            "is_essay_subject": bool(self.is_essay_subject.isChecked()),
            "template_path": self.template_path.text().strip(),
            "scan_folder": self.scan_folder.text().strip(),
            "auto_recognize": bool(self.auto_recognize.isChecked()),
            "answer_key_path": self.answer_key.text().strip(),
            "answer_key_key": self.answer_key_key.text().strip(),
            "exam_room_name": str(self.exam_room_mapping_selector.currentData() or self.exam_room_name.currentText() or "").strip(),
            "exam_room_sbd_mapping": ",".join(self._exam_room_mapping_cache.get(str(self.exam_room_mapping_selector.currentData() or self.exam_room_name.currentText() or "").strip(), [])),
            "exam_room_sbd_mapping_by_room": {str(k): list(v) for k, v in (self._exam_room_mapping_cache or {}).items()},
            "imported_answer_keys": self.answer_key_data,
            "direct_score_import": copy.deepcopy(self.direct_score_import_data),
            "score_mode": self.score_mode.currentText(),
            "section_scores": section_scores,
            "question_scores": question_scores,
            "question_counts": self._question_counts(),
            "total_exam_points": self._to_float(self.total_score.text()),
            "paper_part_count": int(self.paper_part_label.text() or self.paper_part_count_default),
        }


class NewExamDialog(QDialog):
    def __init__(
        self,
        subject_options: list[str],
        block_options: list[str],
        data: dict | None = None,
        parent=None,
        on_batch_scan_subject=None,
        on_save_exam=None,
        stay_open_on_save: bool = False,
        template_repo: TemplateRepository | None = None,
    ):
        super().__init__(parent)
        data = data or {}
        self.setWindowTitle("Sửa kỳ thi" if data else "Tạo kỳ thi mới")
        self.resize(860, 640)
        self.setWindowState(self.windowState() | Qt.WindowMaximized)
        self.subject_configs: list[dict] = list(data.get("subject_configs", []))
        self.student_list_path_value = str(data.get("student_list_path", "") or "")
        self.student_rows: list[dict] = list(data.get("students", [])) if isinstance(data.get("students", []), list) else []
        self.on_batch_scan_subject = on_batch_scan_subject
        self.on_save_exam = on_save_exam
        self.stay_open_on_save = bool(stay_open_on_save)
        self.subject_options = subject_options
        self.block_options = block_options
        self.template_repo = template_repo or TemplateRepository()

        lay = QVBoxLayout(self)
        form = QFormLayout()

        self.exam_name = QLineEdit(str(data.get("exam_name", "")))
        self.common_template = QLineEdit(str(data.get("common_template", "")))
        self.scan_root = QLineEdit(str(data.get("scan_root", "")))
        self.student_list_path = QLineEdit(self.student_list_path_value)
        self.student_list_path.setReadOnly(True)
        self.student_count_label = QLabel(f"{len(self.student_rows)} học sinh")
        self.scan_mode = QComboBox(); self.scan_mode.addItems(["Ảnh trong thư mục gốc", "Ảnh theo phòng thi (thư mục con)"])
        self.scan_mode.setCurrentText(str(data.get("scan_mode", "Ảnh trong thư mục gốc")))
        self.paper_part_count = QComboBox(); self.paper_part_count.addItems(["1", "2", "3", "4", "5"]); self.paper_part_count.setCurrentText(str(data.get("paper_part_count", "3")))

        row_tpl = QHBoxLayout(); row_tpl.addWidget(self.common_template); btn_tpl = QPushButton("..."); row_tpl.addWidget(btn_tpl); btn_tpl_repo = QPushButton("Kho mẫu..."); row_tpl.addWidget(btn_tpl_repo)
        btn_tpl.clicked.connect(self._browse_common_template)
        btn_tpl_repo.clicked.connect(self._pick_common_template_from_repo)
        row_scan = QHBoxLayout(); row_scan.addWidget(self.scan_root); btn_scan = QPushButton("..."); row_scan.addWidget(btn_scan)
        btn_scan.clicked.connect(self._browse_scan_root)
        row_students = QHBoxLayout(); row_students.addWidget(self.student_list_path)
        btn_students = QPushButton("Import Excel...")
        self.btn_view_students = QPushButton("Xem danh sách")
        row_students.addWidget(btn_students)
        row_students.addWidget(self.btn_view_students)
        row_students.addWidget(self.student_count_label)
        btn_students.clicked.connect(self._import_student_list)
        self.btn_view_students.clicked.connect(self._open_student_list_preview)
        self.btn_view_students.setEnabled(bool(self.student_rows))

        form.addRow("Tên kỳ thi", self.exam_name)
        form.addRow("Giấy thi dùng chung", row_tpl)
        form.addRow("Thư mục gốc bài thi", row_scan)
        form.addRow("Danh sách học sinh", row_students)
        form.addRow("Cơ chế thư mục bài thi", self.scan_mode)
        form.addRow("Số phần trên giấy thi", self.paper_part_count)
        lay.addLayout(form)

        lay.addWidget(QLabel("Các môn trong kỳ thi"))
        self.subject_table = QTableWidget(0, 11)
        self.subject_table.setHorizontalHeaderLabels(["STT", "Môn", "Khối", "Key", "Mã đề", "Chế độ điểm", "Tổng điểm", "Template", "Cơ chế", "Trạng thái", "Thao tác"])
        self.subject_table.verticalHeader().setVisible(False)
        self.subject_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.subject_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.subject_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.subject_table.setShowGrid(True)
        self.subject_table.setGridStyle(Qt.SolidLine)
        self.subject_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.subject_table.customContextMenuRequested.connect(self._open_subject_row_context_menu)
        # double click navigation: open subject editor directly from row.
        self.subject_table.cellDoubleClicked.connect(self._handle_subject_table_double_click)
        hdr = self.subject_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(6, QHeaderView.Stretch)
        hdr.setSectionResizeMode(7, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(8, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(9, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(10, QHeaderView.ResizeToContents)
        lay.addWidget(self.subject_table)

        self._refresh_subject_list()

    def _answer_key_scope_key(self, subject_key: str, block: str = "") -> str:
        base = str(subject_key or "").strip()
        if not base:
            return ""
        if "::" in base:
            return base
        session_id = str(getattr(self.parent(), "current_session_id", "") or "").strip()
        exam_name = str(self.exam_name.text() if hasattr(self, "exam_name") else "").strip().lower()
        block_text = str(block or "").strip()
        if not block_text and "_" in base:
            block_text = str(base.rsplit("_", 1)[-1]).strip()
        scope_prefix = f"{session_id}::{exam_name}" if session_id and exam_name else session_id
        if scope_prefix and block_text:
            return f"{scope_prefix}::{base}::{block_text}"
        if scope_prefix:
            return f"{scope_prefix}::{base}"
        return base

    @staticmethod
    def _normalized_student_id_for_match(student_id: str) -> str:
        sid = str(student_id or "").strip()
        if not sid:
            return ""
        compact = sid.replace(" ", "")
        if compact.endswith(".0"):
            prefix = compact[:-2]
            if prefix.isdigit():
                compact = prefix
        if compact.isdigit():
            compact = compact.lstrip("0") or "0"
        return compact.upper()

    @staticmethod
    def _normalized_room_for_match(room_text: str) -> str:
        room = str(room_text or "").strip().casefold()
        if room.isdigit():
            room = room.lstrip("0") or "0"
        return room

    def _browse_common_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Chọn giấy thi dùng chung", "", "JSON (*.json)")
        if path:
            self.template_repo.register(path)
            self.common_template.setText(path)


    def _pick_common_template_from_repo(self) -> None:
        items = [f"{name} | {path}" for name, path in self.template_repo.list_templates()]
        if not items:
            QMessageBox.information(self, "Kho mẫu giấy thi", "Kho mẫu đang trống. Hãy thêm mẫu bằng nút ...")
            return
        chosen, ok = QInputDialog.getItem(self, "Kho mẫu giấy thi", "Chọn mẫu dùng chung:", items, 0, False)
        if ok and chosen:
            self.common_template.setText(chosen.split(" | ", 1)[1])

    def _browse_scan_root(self) -> None:
        path = QFileDialog.getExistingDirectory(self, "Chọn thư mục gốc bài thi")
        if path:
            self.scan_root.setText(path)

    def _import_student_list(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Import danh sách học sinh", "", "Excel/CSV (*.xlsx *.xls *.csv)")
        if not path:
            return
        try:
            import pandas as pd
            if Path(path).suffix.lower() in {".xlsx", ".xls"}:
                df = pd.read_excel(path, dtype=str)
            else:
                df = pd.read_csv(path, dtype=str, keep_default_na=False)
        except Exception as exc:
            QMessageBox.warning(self, "Danh sách học sinh", f"Không đọc được file học sinh:\n{exc}")
            return
        if df.empty:
            QMessageBox.warning(self, "Danh sách học sinh", "File học sinh rỗng.")
            return

        columns = [str(c) for c in df.columns]
        dlg = QDialog(self)
        dlg.setWindowTitle("Mapping cột danh sách học sinh")
        lay = QVBoxLayout(dlg)
        frm = QFormLayout()
        c_sid = QComboBox(); c_sid.addItems(columns)
        c_name = QComboBox(); c_name.addItems(columns)
        c_birth = QComboBox(); c_birth.addItems(["[Không dùng]"] + columns)
        c_class = QComboBox(); c_class.addItems(["[Không dùng]"] + columns)
        c_room = QComboBox(); c_room.addItems(["[Không dùng]"] + columns)
        # best-effort default picks
        lower_cols = {x.lower(): x for x in columns}
        for key, cb in [
            ("studentid", c_sid), ("sobaodanh", c_sid), ("student_id", c_sid),
            ("name", c_name), ("hoten", c_name), ("họ tên", c_name),
        ]:
            if key in lower_cols:
                cb.setCurrentText(lower_cols[key])
        frm.addRow("Số báo danh (bắt buộc)", c_sid)
        frm.addRow("Họ tên (bắt buộc)", c_name)
        frm.addRow("Ngày sinh", c_birth)
        frm.addRow("Lớp", c_class)
        frm.addRow("Phòng thi", c_room)
        lay.addLayout(frm)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)
        lay.addWidget(bb)
        if dlg.exec() != QDialog.Accepted:
            return

        sid_col = c_sid.currentText().strip()
        name_col = c_name.currentText().strip()
        if not sid_col or not name_col:
            QMessageBox.warning(self, "Danh sách học sinh", "Bắt buộc map cột Số báo danh và Họ tên.")
            return

        def _col_value(row_obj, col_name: str) -> str:
            if not col_name or col_name == "[Không dùng]":
                return ""
            v = row_obj.get(col_name, "")
            return "" if v is None else str(v).strip()

        out: list[dict] = []
        for _, row_obj in df.iterrows():
            sid = _col_value(row_obj, sid_col)
            name = _col_value(row_obj, name_col)
            if not sid or not name:
                continue
            out.append(
                {
                    "student_id": sid,
                    "name": name,
                    "birth_date": _col_value(row_obj, c_birth.currentText()),
                    "class_name": _col_value(row_obj, c_class.currentText()),
                    "exam_room": _col_value(row_obj, c_room.currentText()),
                }
            )

        if not out:
            QMessageBox.warning(self, "Danh sách học sinh", "Không có dòng hợp lệ (thiếu Số báo danh/Họ tên).")
            return

        action_text = "thay thế"
        if self.student_rows:
            msg = QMessageBox(self)
            msg.setWindowTitle("Danh sách học sinh")
            msg.setText(
                f"Đã import được {len(out)} học sinh từ file mới.\n"
                "Bạn muốn thêm vào danh sách hiện tại hay thay thế toàn bộ?"
            )
            btn_append = msg.addButton("Thêm vào", QMessageBox.AcceptRole)
            btn_replace = msg.addButton("Thay thế", QMessageBox.DestructiveRole)
            msg.addButton(QMessageBox.Cancel)
            msg.exec()
            clicked = msg.clickedButton()
            if clicked == btn_append:
                action_text = "thêm vào"
                merged: dict[str, dict] = {}
                for row in self.student_rows:
                    sid_key = self._normalized_student_id_for_match(str(row.get("student_id", "") or ""))
                    merged[sid_key or str(row.get("student_id", "") or "")] = row
                for row in out:
                    sid_key = self._normalized_student_id_for_match(str(row.get("student_id", "") or ""))
                    merged[sid_key or str(row.get("student_id", "") or "")] = row
                self.student_rows = list(merged.values())
            elif clicked == btn_replace:
                self.student_rows = out
            else:
                return
        else:
            self.student_rows = out
        duplicate_sids = self._duplicate_student_ids(self.student_rows)
        self.student_list_path_value = path
        self.student_list_path.setText(path)
        self.student_count_label.setText(f"{len(self.student_rows)} học sinh")
        self.btn_view_students.setEnabled(bool(self.student_rows))
        QMessageBox.information(
            self,
            "Danh sách học sinh",
            f"Đã {action_text} danh sách học sinh. Tổng hiện tại: {len(self.student_rows)} học sinh.",
        )
        if duplicate_sids:
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Danh sách học sinh")
            msg.setText(
                f"Phát hiện {len(duplicate_sids)} SBD bị trùng trong danh sách vừa import.\n"
                "Bạn có muốn mở màn hình \"Xem danh sách\" để kiểm tra và xoá bản ghi không?"
            )
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg.setDefaultButton(QMessageBox.Yes)
            if msg.exec() == QMessageBox.Yes:
                self._open_student_list_preview()

    @staticmethod
    def _duplicate_student_ids(rows: list[dict]) -> set[str]:
        counts: dict[str, int] = {}
        for row in rows or []:
            sid = str((row or {}).get("student_id", "") or "").strip()
            if sid:
                counts[sid] = counts.get(sid, 0) + 1
        return {sid for sid, count in counts.items() if count > 1}

    def _open_student_list_preview(self) -> None:
        if not self.student_rows:
            QMessageBox.information(self, "Danh sách học sinh", "Chưa có dữ liệu học sinh để xem.")
            return
        dlg = StudentListPreviewDialog(self.student_rows, self)
        if dlg.exec() == QDialog.Accepted:
            self.student_rows = dlg.rows()
            self.student_count_label.setText(f"{len(self.student_rows)} học sinh")
            self.btn_view_students.setEnabled(bool(self.student_rows))


    def _refresh_subject_list(self) -> None:
        self.subject_table.setRowCount(len(self.subject_configs))
        style = self.style()
        for row_idx, cfg in enumerate(self.subject_configs):
            tpl = cfg.get("template_path") or "[dùng mẫu chung]"
            key = cfg.get("answer_key_key", "")
            mode = cfg.get("score_mode", "Điểm theo phần")
            total = cfg.get("total_exam_points", "-")
            codes = ",".join(sorted((cfg.get("imported_answer_keys") or {}).keys()))
            self.subject_table.setItem(row_idx, 0, QTableWidgetItem(str(row_idx + 1)))
            self.subject_table.setItem(row_idx, 1, QTableWidgetItem(str(cfg.get("name", "") or "-")))
            self.subject_table.setItem(row_idx, 2, QTableWidgetItem(str(cfg.get("block", "") or "-")))
            self.subject_table.setItem(row_idx, 3, QTableWidgetItem(str(key or "-")))
            self.subject_table.setItem(row_idx, 4, QTableWidgetItem(codes or "-"))
            self.subject_table.setItem(row_idx, 5, QTableWidgetItem(str(mode or "-")))
            self.subject_table.setItem(row_idx, 6, QTableWidgetItem(str(total or "-")))
            self.subject_table.setItem(row_idx, 7, QTableWidgetItem(str(tpl or "-")))
            auto_mode = "Nhận dạng tự động" if bool(cfg.get("auto_recognize", False)) else "Thủ công"
            self.subject_table.setItem(row_idx, 8, QTableWidgetItem(auto_mode))
            self.subject_table.setItem(row_idx, 9, QTableWidgetItem(self._subject_status_text(cfg)))

            btn_batch_scan = QPushButton("Nhận dạng")
            btn_batch_scan.setIcon(style.standardIcon(QStyle.SP_MediaPlay))
            btn_batch_scan.setToolTip("Batch Scan theo môn")
            btn_batch_scan.setEnabled(callable(self.on_batch_scan_subject))
            btn_batch_scan.clicked.connect(lambda _=False, i=row_idx: self._trigger_subject_batch_scan(i))
            wrap = QWidget()
            wrap_l = QHBoxLayout(wrap)
            wrap_l.setContentsMargins(0, 0, 0, 0)
            wrap_l.addWidget(btn_batch_scan)
            self.subject_table.setCellWidget(row_idx, 10, wrap)
        self.subject_table.resizeRowsToContents()

    def _subject_status_text(self, cfg: dict) -> str:
        if not isinstance(cfg, dict):
            return "-"
        count = int(cfg.get("batch_result_count", 0) or 0)
        if count > 0 or bool(cfg.get("batch_saved")):
            return "Đã nhận dạng"
        for key in ("batch_saved_rows", "batch_saved_preview", "batch_saved_results"):
            payload = cfg.get(key, [])
            if isinstance(payload, list) and payload:
                return "Đã nhận dạng"
        parent = self.parent()
        if parent is not None and hasattr(parent, "_is_subject_marked_batched"):
            try:
                if bool(parent._is_subject_marked_batched(cfg)):
                    return "Đã nhận dạng"
            except Exception:
                pass
        return "-"

    def _current_subject_index(self) -> int:
        idx = self.subject_table.currentRow()
        return idx if 0 <= idx < len(self.subject_configs) else -1

    def _handle_subject_table_double_click(self, row: int, col: int) -> None:
        if row < 0 or row >= len(self.subject_configs):
            return
        self.subject_table.selectRow(row)
        self._edit_subject()

    def _open_subject_row_context_menu(self, pos) -> None:
        row = self.subject_table.rowAt(pos.y())
        if row < 0 or row >= len(self.subject_configs):
            return
        self.subject_table.selectRow(row)
        menu = QMenu(self)
        act_scan = menu.addAction("Nhận dạng")
        act_edit = menu.addAction("Sửa cấu hình môn")
        act_delete = menu.addAction("Xoá")
        chosen = menu.exec(self.subject_table.viewport().mapToGlobal(pos))
        if chosen == act_scan:
            self._trigger_subject_batch_scan(row)
            return
        if chosen == act_edit:
            self._edit_subject()
            return
        if chosen == act_delete:
            self._delete_subject()

    def _trigger_subject_batch_scan(self, idx: int) -> None:
        if idx < 0 or idx >= len(self.subject_configs):
            return
        if not callable(self.on_batch_scan_subject):
            QMessageBox.information(self, "Batch Scan", "Vui lòng lưu kỳ thi trước khi chạy Batch Scan theo từng môn.")
            return
        cfg = dict(self.subject_configs[idx])
        cfg["template_path"] = self._normalize_template_path(str(cfg.get("template_path", "")))
        cfg["scan_folder"] = str(cfg.get("scan_folder", "") or self.scan_root.text().strip())
        proceed = self.on_batch_scan_subject(
            {
                "exam_name": self.exam_name.text().strip(),
                "common_template": self.common_template.text().strip(),
                "scan_root": self.scan_root.text().strip(),
                "scan_mode": self.scan_mode.currentText(),
                "paper_part_count": int(self.paper_part_count.currentText()),
                "subject_configs": list(self.subject_configs),
                "selected_subject_index": idx,
                "subject_config": cfg,
            }
        )
        if proceed is False:
            return
        if proceed is True:
            return
        self.reject()

    @staticmethod
    def _normalize_template_path(path_text: str) -> str:
        t = str(path_text or "").strip()
        if not t:
            return ""
        if t.lower() in {"[dùng mẫu chung]", "[dung mau chung]", "none", "null", "-"}:
            return ""
        return t

    @staticmethod
    def _subject_imported_answer_keys(subject_cfg: dict) -> dict:
        if not isinstance(subject_cfg, dict):
            return {}
        raw = subject_cfg.get("imported_answer_keys", {})
        return dict(raw) if isinstance(raw, dict) else {}

    @staticmethod
    def _subject_identity_changed(old_cfg: dict, new_cfg: dict) -> bool:
        def _norm(v: object) -> str:
            return str(v or "").strip().lower()
        watched = ["template_path", "scan_folder", "name", "block", "answer_key_key"]
        return any(_norm(old_cfg.get(k)) != _norm(new_cfg.get(k)) for k in watched)

    @staticmethod
    def _copy_noncritical_subject_updates(old_cfg: dict, new_cfg: dict) -> dict:
        merged = dict(old_cfg)
        merged.update(new_cfg)
        # Preserve batch/scoring artifacts when identity is unchanged.
        for k in [
            "batch_saved",
            "batch_saved_at",
            "batch_result_count",
            "batch_saved_rows",
            "batch_saved_preview",
            "batch_saved_results",
        ]:
            if k in old_cfg:
                merged[k] = old_cfg.get(k)
        return merged

    def _confirm_subject_identity_change(self, old_cfg: dict, new_cfg: dict) -> str:
        has_existing_batch = bool(old_cfg.get("batch_saved")) or bool(old_cfg.get("batch_result_count"))
        if not has_existing_batch:
            return "apply"

        msg = QMessageBox(self)
        msg.setWindowTitle("Môn đã có dữ liệu nhận dạng")
        msg.setIcon(QMessageBox.Warning)
        msg.setText(
            "Bạn đã thay đổi thông tin ảnh hưởng đến nhận dạng/chấm điểm (mẫu giấy, thư mục quét, tên môn, khối hoặc mã đáp án môn_khối)."
        )
        msg.setInformativeText(
            "Giữ dữ liệu Batch cũ có thể làm sai kết quả.\n"
            "Yes: Xóa trạng thái/dữ liệu nhận dạng của môn này và áp dụng thay đổi.\n"
            "No: Vẫn giữ dữ liệu nhận dạng cũ và áp dụng thay đổi.\n"
            "Cancel: Hủy chỉnh sửa."
        )
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
        msg.setDefaultButton(QMessageBox.Yes)
        ch = msg.exec()
        if ch == QMessageBox.Cancel:
            return "cancel"
        if ch == QMessageBox.Yes:
            return "reset"
        return "apply"

    def _add_subject(self) -> None:
        dlg = SubjectConfigDialog(
            subject_options=self.subject_options,
            block_options=self.block_options,
            paper_part_count=int(self.paper_part_count.currentText()),
            common_template_path=self.common_template.text().strip(),
            template_repo=self.template_repo,
            parent=self,
        )
        if dlg.exec() != QDialog.Accepted:
            return
        payload = dlg.payload()
        self.subject_configs.append(payload)
        try:
            scoped_key = self._answer_key_scope_key(str(payload.get("answer_key_key", "") or ""), str(payload.get("block", "") or ""))
            self.database.replace_answer_keys_for_subject(scoped_key, payload.get("imported_answer_keys", {}) or {})
            self.database.log_change(
                "answer_keys",
                scoped_key,
                "imported_answer_keys",
                "",
                payload.get("imported_answer_keys", {}) or {},
                "add_subject",
            )
        except Exception:
            pass
        self._refresh_subject_list()

    def _edit_subject(self) -> None:
        idx = self._current_subject_index()
        if idx < 0:
            return
        dlg = SubjectConfigDialog(
            self.subject_configs[idx],
            subject_options=self.subject_options,
            block_options=self.block_options,
            paper_part_count=int(self.paper_part_count.currentText()),
            common_template_path=self.common_template.text().strip(),
            template_repo=self.template_repo,
            parent=self,
        )
        if dlg.exec() != QDialog.Accepted:
            return
        old_cfg = dict(self.subject_configs[idx])
        edited = dlg.payload()
        try:
            old_subject_key = str(old_cfg.get("answer_key_key", "") or "")
            new_subject_key = str(edited.get("answer_key_key", "") or "")
            old_scoped_key = self._answer_key_scope_key(old_subject_key, str(old_cfg.get("block", "") or ""))
            new_scoped_key = self._answer_key_scope_key(new_subject_key, str(edited.get("block", "") or ""))
            if old_scoped_key and old_scoped_key != new_scoped_key:
                self.database.replace_answer_keys_for_subject(old_scoped_key, {})
            self.database.replace_answer_keys_for_subject(new_scoped_key, edited.get("imported_answer_keys", {}) or {})
            if old_cfg.get("imported_answer_keys", {}) != edited.get("imported_answer_keys", {}):
                self.database.log_change(
                    "answer_keys",
                    new_scoped_key,
                    "imported_answer_keys",
                    old_cfg.get("imported_answer_keys", {}) or {},
                    edited.get("imported_answer_keys", {}) or {},
                    "edit_subject",
                )
        except Exception:
            pass

        if not self._subject_identity_changed(old_cfg, edited):
            self.subject_configs[idx] = self._copy_noncritical_subject_updates(old_cfg, edited)
            self._refresh_subject_list()
            return

        decision = self._confirm_subject_identity_change(old_cfg, edited)
        if decision == "cancel":
            return
        if decision == "reset":
            updated = dict(edited)
            updated["batch_saved"] = False
            updated["batch_saved_at"] = "-"
            updated["batch_result_count"] = 0
            updated["batch_saved_rows"] = []
            updated["batch_saved_preview"] = []
            updated["batch_saved_results"] = []
            self.subject_configs[idx] = updated
        else:
            self.subject_configs[idx] = dict(old_cfg) | dict(edited)
        self._refresh_subject_list()

    def _delete_subject(self) -> None:
        idx = self._current_subject_index()
        if idx < 0:
            return
        name = str((self.subject_configs[idx] or {}).get("name", "") or "").strip() or "môn đã chọn"
        if QMessageBox.question(
            self,
            "Xoá môn",
            f"Bạn có chắc muốn xoá {name} khỏi kỳ thi?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        del self.subject_configs[idx]
        self._refresh_subject_list()

    def _validate_and_accept(self) -> None:
        if not self.exam_name.text().strip():
            QMessageBox.warning(self, "Thiếu dữ liệu", "Vui lòng nhập tên kỳ thi.")
            return
        if not self.subject_configs:
            QMessageBox.warning(self, "Thiếu dữ liệu", "Vui lòng thêm ít nhất 1 môn học.")
            return
        for cfg in self.subject_configs:
            if not cfg.get("name"):
                QMessageBox.warning(self, "Thiếu dữ liệu", "Mỗi môn phải có tên.")
                return
            if not cfg.get("block"):
                QMessageBox.warning(self, "Thiếu dữ liệu", "Mỗi môn phải có khối.")
                return
        if self.stay_open_on_save:
            if callable(self.on_save_exam):
                ok = self.on_save_exam()
                if ok is False:
                    return
            return
        self.accept()

    def payload(self) -> dict:
        return {
            "exam_name": self.exam_name.text().strip(),
            "common_template": self.common_template.text().strip(),
            "scan_root": self.scan_root.text().strip(),
            "student_list_path": self.student_list_path_value,
            "students": self.student_rows,
            "scan_mode": self.scan_mode.currentText(),
            "paper_part_count": int(self.paper_part_count.currentText()),
            "subject_configs": self.subject_configs,
        }

class StudentListPreviewDialog(QDialog):
    def __init__(self, rows: list[dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Xem danh sách học sinh")
        self.resize(980, 640)
        self._rows: list[dict] = [dict(r or {}) for r in (rows or [])]

        lay = QVBoxLayout(self)
        note = QLabel("Các dòng trùng SBD được tô đỏ. Có thể chọn nhiều dòng và bấm Xoá.")
        note.setWordWrap(True)
        lay.addWidget(note)

        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["STT", "SBD", "Họ tên", "Ngày sinh", "Lớp", "Phòng thi"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.Stretch)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        lay.addWidget(self.table)

        action_row = QHBoxLayout()
        self.delete_btn = QPushButton("Xoá bản ghi đã chọn")
        self.delete_btn.clicked.connect(self._delete_selected_rows)
        action_row.addWidget(self.delete_btn)
        action_row.addStretch(1)
        lay.addLayout(action_row)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        bb.button(QDialogButtonBox.Save).setText("Lưu danh sách")
        bb.button(QDialogButtonBox.Cancel).setText("Đóng")
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        lay.addWidget(bb)

        self._reload_table()

    def rows(self) -> list[dict]:
        return [dict(r or {}) for r in self._rows]

    def _reload_table(self) -> None:
        duplicate_sids = NewExamDialog._duplicate_student_ids(self._rows)
        self.table.setRowCount(len(self._rows))
        for idx, row in enumerate(self._rows):
            sid = str((row or {}).get("student_id", "") or "").strip()
            values = [
                str(idx + 1),
                sid,
                str((row or {}).get("name", "") or ""),
                str((row or {}).get("birth_date", "") or ""),
                str((row or {}).get("class_name", "") or ""),
                str((row or {}).get("exam_room", "") or ""),
            ]
            for col, text in enumerate(values):
                item = QTableWidgetItem(text)
                if sid and sid in duplicate_sids:
                    item.setBackground(QColor(255, 220, 220))
                    item.setForeground(QColor(180, 0, 0))
                self.table.setItem(idx, col, item)
        self.table.resizeRowsToContents()

    def _delete_selected_rows(self) -> None:
        selected_indexes = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not selected_indexes:
            QMessageBox.information(self, "Xem danh sách học sinh", "Vui lòng chọn bản ghi cần xoá.")
            return
        target_rows = sorted({int(index.row()) for index in selected_indexes}, reverse=True)
        for row_idx in target_rows:
            if 0 <= row_idx < len(self._rows):
                self._rows.pop(row_idx)
        self._reload_table()


class MainWindow(QMainWindow):
    SCAN_COL_STT = 0
    SCAN_COL_STUDENT_ID = 1
    SCAN_COL_EXAM_ROOM = 2
    SCAN_COL_EXAM_CODE = 3
    SCAN_COL_FULL_NAME = 4
    SCAN_COL_BIRTH_DATE = 5
    SCAN_COL_CONTENT = 6
    SCAN_COL_STATUS = 7
    SCAN_COL_ACTIONS = 8

    def __init__(self):
        super().__init__()
        self.setWindowTitle("OMR Exam Grading System")
        self.resize(1200, 800)
        self.setWindowState(self.windowState() | Qt.WindowMaximized)

        self.session: ExamSession | None = None
        self.template: Template | None = None
        self.answer_keys: AnswerKeyRepository | None = None
        self.scan_results = []
        self.scan_results_by_subject: dict[str, list] = {}
        self.batch_working_state_by_subject: dict[str, dict] = {}
        self.scan_files: list[Path] = []
        self.scan_blank_questions: dict[int, list[int]] = {}
        self.scan_blank_summary: dict[int, dict[str, list[int]]] = {}
        self.scan_manual_adjustments: dict[str, list[str]] = {}
        self.scan_edit_history: dict[str, list[str]] = {}
        self.scan_last_adjustment: dict[str, str] = {}
        self.score_rows = []
        self.scoring_results_by_subject: dict[str, dict[str, dict]] = {}
        self.scoring_phases: list[dict] = []
        self._scoring_dirty_subjects: set[str] = set()
        self.imported_exam_codes: list[str] = []
        self.active_batch_subject_key: str | None = None
        self.subject_catalog: list[str] = ["Toán", "Ngữ văn", "Tiếng Anh", "Vật lý", "Hóa học", "Sinh học"]
        self.block_catalog: list[str] = ["10", "11", "12"]
        self.subjects: list[str] = list(self.subject_catalog)
        self.grades: list[str] = list(self.block_catalog)
        self.subject_management_mode = "subjects"
        self.subject_edit_index: int | None = None
        self.batch_editor_return_payload: dict | None = None
        self.batch_editor_return_session_id: str | None = None
        self._current_batch_data_source: str = "empty"
        self._route_history: list[dict] = []
        self._current_route_name: str = "exam_list"
        self._current_route_context: dict = {}
        self._suspend_route_history_push: bool = False

        self.omr_processor = OMRProcessor()
        self.scoring_engine = ScoringEngine()
        self.current_session_path: Path | None = None
        self.current_session_id: str | None = None
        self.session_dirty = False
        self._session_saved_signature = ""

        self.database = OMRDatabase.default()
        self.session_registry: list[dict[str, str | bool]] = self._load_session_registry()
        self.template_repo = self._load_template_repository()
        self.template_editor_embedded: TemplateEditorWindow | None = None
        self.template_editor_mode = "library"

        self.stack = QStackedWidget()
        self.stack.addWidget(self._build_exam_list_page())
        self.stack.addWidget(self._build_workspace_page())
        self.stack.addWidget(self._build_subject_management_page())
        self.stack.addWidget(self._build_template_management_page())
        self.template_editor_page = QWidget()
        self.template_editor_layout = QVBoxLayout(self.template_editor_page)
        self.template_editor_layout.setContentsMargins(0, 0, 0, 0)
        self.stack.addWidget(self.template_editor_page)
        self.exam_editor_page = QWidget()
        self.exam_editor_layout = QVBoxLayout(self.exam_editor_page)
        self.exam_editor_layout.setContentsMargins(0, 0, 0, 0)
        self.stack.addWidget(self.exam_editor_page)
        self.embedded_exam_dialog: NewExamDialog | None = None
        self.embedded_exam_session_id: str | None = None
        self.embedded_exam_session: ExamSession | None = None
        self.embedded_exam_original_payload: dict | None = None
        self.embedded_exam_is_new: bool = False
        self.preview_zoom_factor = 0.3
        self.preview_source_pixmap = QPixmap()
        self.preview_rotation_by_index: dict[int, int] = {}
        self.scan_forced_status_by_index: dict[str, str] = {}
        self.deleted_scan_images_by_subject: dict[str, set[str]] = {}
        self._student_option_cache_session_id: str = ""
        self._student_option_labels_cache: list[str] = []
        self._student_option_sid_map: dict[str, str] = {}
        self._student_option_profile_map: dict[str, dict[str, str]] = {}
        self._student_option_sid_set: set[str] = set()
        self._student_option_cache_signature: str = ""
        self._scan_grid_loading = False
        self._switching_batch_subject = False
        self._template_cache_by_path: dict[str, Template] = {}
        self._answer_keys_ready_subjects: set[str] = set()
        self._batch_scan_running = False
        self._batch_cancel_requested = False
        self._batch_loaded_runtime_key: str = ""
        self._batch_loaded_subject_signature: str = ""
        self._auto_recognition_busy = False
        self._auto_recognition_queue: deque[str] = deque()
        self._auto_recognition_enqueued: set[str] = set()
        self._auto_recognition_last_seen: dict[str, tuple[int, float, int]] = {}
        self._auto_recognition_pause_requested = False
        self._auto_recognition_active_subject: str = ""
        self.preview_drag_active = False
        self.preview_last_pos = None
        self.setCentralWidget(self.stack)

        self._build_menu()
        self.stack.currentChanged.connect(self._handle_stack_changed)
        db_subjects = self.database.fetch_catalog("subjects")
        db_blocks = self.database.fetch_catalog("blocks")
        if db_subjects:
            self.subject_catalog = db_subjects
            self.subjects = list(db_subjects)
        else:
            self.database.replace_catalog("subjects", self.subject_catalog)
        if db_blocks:
            self.block_catalog = db_blocks
            self.grades = list(db_blocks)
        else:
            self.database.replace_catalog("blocks", self.block_catalog)
        self._refresh_exam_list()
        self._refresh_batch_subject_controls()
        self._handle_stack_changed(self.stack.currentIndex())
        self.stack.setCurrentIndex(0)
        self._setup_auto_recognition_timer()

    def _setup_auto_recognition_timer(self) -> None:
        self.auto_recognition_discovery_timer = QTimer(self)
        self.auto_recognition_discovery_timer.setInterval(4000)
        self.auto_recognition_discovery_timer.timeout.connect(self._poll_auto_recognition)
        self.auto_recognition_discovery_timer.start()

        self.auto_recognition_worker_timer = QTimer(self)
        self.auto_recognition_worker_timer.setInterval(600)
        self.auto_recognition_worker_timer.timeout.connect(self._process_auto_recognition_queue)
        self.auto_recognition_worker_timer.start()

    @staticmethod
    def _is_subfolder_scan_mode(mode_text: str) -> bool:
        mode = str(mode_text or "").strip().lower()
        if not mode:
            return False
        return any(token in mode for token in ["thư mục con", "folder con", "sub", "phòng thi", "room"])

    def _scan_folder_signature(self, cfg: dict | None) -> tuple[int, float, int]:
        if not isinstance(cfg, dict):
            return (0, 0.0, 0)
        scan_folder = str(cfg.get("scan_folder", "") or ((self.session.config or {}).get("scan_root", "") if self.session else "") or "").strip()
        if not scan_folder or scan_folder == "-":
            return (0, 0.0, 0)
        scan_dir = Path(scan_folder)
        if not scan_dir.exists() or not scan_dir.is_dir():
            return (0, 0.0, 0)
        mode = str(cfg.get("scan_mode", "") or ((self.session.config or {}).get("scan_mode", "") if self.session else "") or "")
        use_subfolders = self._is_subfolder_scan_mode(mode)
        image_exts = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}
        count = 0
        latest_mtime = 0.0
        fingerprint = 0
        iterator = scan_dir.rglob("*") if use_subfolders else scan_dir.iterdir()
        for p in iterator:
            if not p.is_file() or p.suffix.lower() not in image_exts:
                continue
            count += 1
            try:
                stat = p.stat()
                latest_mtime = max(latest_mtime, float(stat.st_mtime))
                # Include path + size + mtime into a rolling fingerprint so
                # replacing a whole folder with same file count still triggers.
                item_sig = hash((str(p.relative_to(scan_dir)).lower(), int(stat.st_size), int(stat.st_mtime_ns)))
                fingerprint ^= int(item_sig)
            except Exception:
                continue
        return (count, latest_mtime, fingerprint)

    @staticmethod
    def _scan_signature_has_files(signature: tuple[int, float, int]) -> bool:
        return int((signature or (0, 0.0, 0))[0] or 0) > 0

    def _enqueue_auto_recognition_subject(self, subject_key: str) -> None:
        key = str(subject_key or "").strip()
        if not key or key in self._auto_recognition_enqueued:
            return
        self._auto_recognition_queue.append(key)
        self._auto_recognition_enqueued.add(key)
        self._update_auto_recognition_progress()

    def _pop_auto_recognition_subject(self) -> str:
        if not self._auto_recognition_queue:
            return ""
        key = self._auto_recognition_queue.popleft()
        self._auto_recognition_enqueued.discard(key)
        self._update_auto_recognition_progress()
        return key

    def _poll_auto_recognition(self) -> None:
        if not self.session or self._auto_recognition_pause_requested:
            return
        for cfg in self._effective_subject_configs_for_batch():
            if not isinstance(cfg, dict):
                continue
            if not bool(cfg.get("auto_recognize", False)):
                continue
            if self._subject_uses_direct_score_import(cfg):
                continue
            subject_key = self._subject_instance_key_from_cfg(cfg)
            if not subject_key:
                continue
            new_signature = self._scan_folder_signature(cfg)
            old_signature = self._auto_recognition_last_seen.get(subject_key)
            self._auto_recognition_last_seen[subject_key] = new_signature
            if old_signature is None:
                if self._scan_signature_has_files(new_signature):
                    self._enqueue_auto_recognition_subject(subject_key)
                continue
            if new_signature == old_signature:
                continue
            self._enqueue_auto_recognition_subject(subject_key)

    def _process_auto_recognition_queue(self) -> None:
        if self._auto_recognition_pause_requested or self._auto_recognition_busy or self._batch_scan_running:
            return
        if not self._auto_recognition_queue:
            self._auto_recognition_active_subject = ""
            self._update_auto_recognition_progress()
            return
        subject_key = self._pop_auto_recognition_subject()
        if not subject_key:
            return
        self._auto_recognition_active_subject = subject_key
        self._auto_recognition_busy = True
        try:
            self._run_auto_recognition_for_subject(subject_key)
        finally:
            self._auto_recognition_busy = False
            self._auto_recognition_active_subject = ""
            gc.collect()
            self._update_auto_recognition_progress()

    def _run_auto_recognition_for_subject(self, subject_key: str) -> None:
        if not hasattr(self, "batch_subject_combo") or self.batch_subject_combo.count() <= 1:
            return
        previous_subject_index = self.batch_subject_combo.currentIndex()
        previous_scope_index = self.batch_file_scope_combo.currentIndex() if hasattr(self, "batch_file_scope_combo") else -1

        target_index = -1
        for i in range(1, self.batch_subject_combo.count()):
            if str(self.batch_subject_combo.itemData(i) or "").strip() == str(subject_key or "").strip():
                target_index = i
                break
        if target_index <= 0:
            return

        try:
            self.batch_subject_combo.setCurrentIndex(target_index)
            cfg = self._selected_batch_subject_config()
            if not isinstance(cfg, dict):
                return

            scan_paths = self._configured_scan_file_paths(cfg)
            if not scan_paths:
                return
            recognized = self._recognized_image_paths_for_subject(self._subject_key_from_cfg(cfg))
            if not any(str(path).strip() not in recognized for path in scan_paths):
                return

            if hasattr(self, "batch_file_scope_combo"):
                for i in range(self.batch_file_scope_combo.count()):
                    if str(self.batch_file_scope_combo.itemData(i) or "") == "new_only":
                        self.batch_file_scope_combo.setCurrentIndex(i)
                        break
            self._update_auto_recognition_progress()
            self.run_batch_scan(auto_triggered=True)
        finally:
            if previous_scope_index >= 0 and hasattr(self, "batch_file_scope_combo"):
                self.batch_file_scope_combo.setCurrentIndex(previous_scope_index)
            if previous_subject_index >= 0 and previous_subject_index < self.batch_subject_combo.count():
                self.batch_subject_combo.setCurrentIndex(previous_subject_index)

    def _update_auto_recognition_progress(self) -> None:
        active_key = str(self._auto_recognition_active_subject or "").strip()
        queue_count = len(self._auto_recognition_queue)
        has_work = bool(active_key or queue_count > 0)
        if not has_work:
            if hasattr(self, "auto_recognition_progress_dialog") and self.auto_recognition_progress_dialog is not None:
                self.auto_recognition_progress_dialog.hide()
            return

        if not hasattr(self, "auto_recognition_progress_dialog") or self.auto_recognition_progress_dialog is None:
            dlg = QDialog(self)
            dlg.setWindowTitle("Tiến trình nhận dạng tự động")
            dlg.setModal(False)
            lay = QVBoxLayout(dlg)
            self.auto_recognition_progress_label = QLabel("Đang khởi tạo cơ chế nhận dạng tự động...")
            self.auto_recognition_progress_label.setWordWrap(True)
            self.auto_recognition_progress_bar = QProgressBar()
            self.auto_recognition_progress_bar.setRange(0, 0)
            self.auto_recognition_progress_bar.setTextVisible(False)
            btn_pause = QPushButton("Tạm dừng")
            btn_pause.clicked.connect(self._toggle_auto_recognition_pause)
            lay.addWidget(self.auto_recognition_progress_label)
            lay.addWidget(self.auto_recognition_progress_bar)
            lay.addWidget(btn_pause)
            self.auto_recognition_progress_dialog = dlg
        else:
            dlg = self.auto_recognition_progress_dialog

        active_label = active_key or "-"
        paused = " (Tạm dừng)" if self._auto_recognition_pause_requested else ""
        text = (
            f"Cơ chế: Nhận dạng tự động{paused}\n"
            f"Đang xử lý: {active_label}\n"
            f"Hàng đợi còn: {queue_count} môn"
        )
        self.auto_recognition_progress_label.setText(text)
        if not dlg.isVisible():
            dlg.show()
            dlg.raise_()

    def _toggle_auto_recognition_pause(self) -> None:
        self._auto_recognition_pause_requested = not self._auto_recognition_pause_requested
        self._update_auto_recognition_progress()

    def _build_template_management_page(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        split = QSplitter()
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(0)
        self.template_library_table = QTableWidget(0, 2)
        self.template_library_table.setHorizontalHeaderLabels(["STT", "Tên mẫu"])
        self.template_library_table.verticalHeader().setVisible(False)
        self.template_library_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.template_library_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.template_library_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.template_library_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.template_library_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.template_library_table.itemSelectionChanged.connect(self._handle_template_library_selection)
        # double click navigation: open selected template directly.
        self.template_library_table.cellDoubleClicked.connect(self._handle_template_library_double_click)
        left_layout.addWidget(self.template_library_table)
        split.addWidget(left)

        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(0)
        self.template_preview_title = QLabel("Chưa chọn mẫu giấy thi")
        self.template_preview_title.setWordWrap(True)
        self.template_preview_title.setContentsMargins(0, 0, 0, 4)
        self.template_preview_image = QLabel("Chọn mẫu giấy thi ở danh sách bên trái")
        self.template_preview_image.setAlignment(Qt.AlignCenter)
        self.template_preview_image.setMinimumHeight(420)
        self.template_preview_image.setStyleSheet("border: 1px solid #cfcfcf; background: #fafafa;")
        right_layout.addWidget(self.template_preview_title)
        right_layout.addWidget(self.template_preview_image, 1)
        split.addWidget(right)
        split.setStretchFactor(0, 1)
        split.setStretchFactor(1, 2)
        layout.addWidget(split, 1)
        return w

    def _refresh_template_library(self) -> None:
        rows = self.template_repo.list_templates()
        self.template_library_table.setRowCount(len(rows))
        selected_row = 0 if rows else -1
        for idx, (name, path) in enumerate(rows):
            num_item = QTableWidgetItem(str(idx + 1))
            name_item = QTableWidgetItem(name)
            name_item.setData(Qt.UserRole, path)
            self.template_library_table.setItem(idx, 0, num_item)
            self.template_library_table.setItem(idx, 1, name_item)
        if rows:
            self.template_library_table.selectRow(selected_row)
            self._update_template_preview_by_row(selected_row)
        else:
            self.template_preview_title.setText("Kho mẫu giấy thi đang trống")
            self.template_preview_image.setPixmap(QPixmap())
            self.template_preview_image.setText("Chưa có mẫu giấy thi trong kho")

    def _handle_template_library_selection(self) -> None:
        row = self.template_library_table.currentRow()
        self._update_template_preview_by_row(row)

    def _update_template_preview_by_row(self, row: int) -> None:
        if row < 0:
            self.template_preview_title.setText("Chưa chọn mẫu giấy thi")
            self.template_preview_image.setPixmap(QPixmap())
            self.template_preview_image.setText("Chọn mẫu giấy thi ở danh sách bên trái")
            return
        item = self.template_library_table.item(row, 1)
        template_path = str(item.data(Qt.UserRole) if item else "")
        if not template_path:
            return
        try:
            tpl = Template.load_json(template_path)
            img_path = Path(tpl.image_path)
            if not img_path.is_absolute():
                img_path = (Path(template_path).parent / img_path).resolve()
            pix = QPixmap(str(img_path))
            self.template_preview_title.setText(f"{row + 1}. {tpl.name}")
            if pix.isNull():
                self.template_preview_image.setPixmap(QPixmap())
                self.template_preview_image.setText("Không thể tải ảnh mẫu giấy thi")
                return
            scaled = pix.scaled(self.template_preview_image.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.template_preview_image.setPixmap(scaled)
            self.template_preview_image.setText("")
        except Exception:
            self.template_preview_title.setText(f"{row + 1}. Không thể đọc mẫu giấy thi")
            self.template_preview_image.setPixmap(QPixmap())
            self.template_preview_image.setText("Không thể đọc dữ liệu mẫu giấy thi")

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        if hasattr(self, "scan_list"):
            idx = self.scan_list.currentRow()
            if 0 <= idx < len(self.scan_results):
                self._update_scan_preview(idx)
            elif idx >= 0:
                self._update_scan_preview_from_saved_row(idx)
        elif hasattr(self, "scan_image_scroll") and not self.preview_source_pixmap.isNull():
            self._render_preview_pixmap()

    def _confirm(self, title: str, message: str) -> bool:
        return (
            QMessageBox.question(
                self,
                title,
                message,
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            == QMessageBox.Yes
        )

    def _build_session_persistence_payload(self) -> dict:
        if not self.session:
            return {}
        current_payload = self.session.to_dict()
        current_cfg = dict(current_payload.get("config", {}) or {})
        # Scoring rows/phases không còn mirror vào session config.
        current_cfg.pop("scoring_phases", None)
        current_cfg.pop("scoring_results", None)
        current_payload["config"] = current_cfg
        return current_payload
    def _session_payload_signature(self, payload: dict) -> str:
        return json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)

    def _remember_current_session_snapshot(self) -> None:
        if not self.session:
            self._session_saved_signature = ""
            return
        try:
            self._session_saved_signature = self._session_payload_signature(self._build_session_persistence_payload())
        except Exception:
            self._session_saved_signature = ""

    def _has_pending_unsaved_work(self) -> bool:
        if hasattr(self, "btn_save_batch_subject") and self.btn_save_batch_subject.isEnabled():
            return True
        if self.stack.currentIndex() == 5 and self._embedded_exam_has_real_changes():
            return True
        if bool(getattr(self, "session_dirty", False)):
            return True
        return self._session_has_real_changes()

    def _embedded_exam_has_real_changes(self) -> bool:
        if not self.embedded_exam_dialog:
            return False
        try:
            current_payload = self.embedded_exam_dialog.payload()
        except Exception:
            return False
        return self._payload_changed(current_payload, self.embedded_exam_original_payload)

    def _session_has_real_changes(self) -> bool:
        if not self.session:
            return False

        if not self.current_session_id:
            return bool(getattr(self, "session_dirty", False))

        current_payload = self._build_session_persistence_payload()
        current_signature = self._session_payload_signature(current_payload)
        saved_signature = str(getattr(self, "_session_saved_signature", "") or "")
        if saved_signature:
            return current_signature != saved_signature

        saved_payload = self.database.fetch_exam_session(self.current_session_id)
        if not isinstance(saved_payload, dict):
            return bool(getattr(self, "session_dirty", False))

        return current_signature != self._session_payload_signature(saved_payload)

    def _confirm_before_switching_work(self, target_text: str) -> bool:
        if not self._has_pending_unsaved_work():
            return True
        return self._handle_pending_changes_before_switch(target_text)

    def _has_active_workflows(self) -> bool:
        return bool(
            self._batch_scan_running
            or self._auto_recognition_busy
            or self._auto_recognition_queue
            or self._auto_recognition_active_subject
        )

    def _interrupt_active_workflows(self) -> None:
        self._batch_cancel_requested = True
        self._auto_recognition_pause_requested = True
        self._auto_recognition_queue.clear()
        self._auto_recognition_enqueued.clear()
        self._auto_recognition_active_subject = ""
        self._update_auto_recognition_progress()

    def _confirm_interrupt_active_workflows(self, destination_text: str) -> bool:
        if not self._has_active_workflows():
            return True
        answer = QMessageBox.question(
            self,
            "Xác nhận chuyển màn hình",
            (
                f"Chuyển sang \"{destination_text}\" có thể ảnh hưởng đến luồng đang chạy.\n"
                "Bạn có đồng ý ngắt tất cả tiến trình nhận dạng đang chạy không?"
            ),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return False
        self._interrupt_active_workflows()
        return True

    def _prompt_save_changes_word_style(self, title: str, message: str) -> str:
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setInformativeText("Bạn muốn lưu thay đổi trước khi tiếp tục không?")
        msg.setStandardButtons(QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
        msg.setDefaultButton(QMessageBox.Save)
        choice = msg.exec()
        if choice == QMessageBox.Save:
            return "save"
        if choice == QMessageBox.Discard:
            return "discard"
        return "cancel"

    def _save_current_work(self) -> bool:
        if self._has_batch_unsaved_changes():
            return self._save_batch_for_selected_subject(
                show_success_message=False,
                reload_after_save=False,
                refresh_exam_list=False,
            )
        if self._has_scoring_unsaved_changes():
            subject_key = self._resolve_preferred_scoring_subject()
            if subject_key:
                rows = self._collect_scoring_preview_rows(subject_key)
                cfg = self._subject_config_by_subject_key(subject_key) or {}
                mode = str((cfg or {}).get("scoring_last_mode", "Tính lại toàn bộ") or "Tính lại toàn bộ")
                note = str((cfg or {}).get("scoring_phase_last_note", "") or "")
                self._persist_scoring_results_for_subject(subject_key, rows, mode, note, mark_saved=True)
                self._persist_runtime_session_state_quietly()
                return True
            return False
        return True

    def _handle_pending_changes_before_switch(self, target_text: str) -> bool:
        if not self._has_pending_unsaved_work():
            return True
        choice = self._prompt_save_changes_word_style(
            "Dữ liệu chưa lưu",
            f"Bạn đang có dữ liệu chưa lưu. Trước khi chuyển sang {target_text}, bạn có muốn lưu không?",
        )
        if choice == "cancel":
            return False
        if choice == "discard":
            return True
        return self._save_current_work()

    def _session_storage_dir(self) -> Path:
        d = Path.home() / ".omr_exam" / "sessions"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _generate_session_id(self, seed: str = "") -> str:
        raw = f"{seed}-{datetime.now().isoformat()}"
        return str(abs(hash(raw)))

    def _session_path_from_id(self, session_id: str) -> Path:
        return self._session_storage_dir() / f"{session_id}.json"

    def _load_session_registry(self) -> list[dict[str, str | bool]]:
        try:
            rows = self.database.list_exam_sessions()
            return [dict(x) for x in rows]
        except Exception:
            return []

    def _save_session_registry(self) -> None:
        self.session_registry = self._load_session_registry()

    def _upsert_session_registry(self, session_id: str, name: str | None = None) -> None:
        payload = self.database.fetch_exam_session(session_id) or {}
        exam_name = str(name or payload.get("exam_name") or "Kỳ thi")
        if payload:
            self.database.save_exam_session(session_id, exam_name, payload)
        self.session_registry = self._load_session_registry()

    def _load_template_repository(self) -> TemplateRepository:
        payload = self.database.get_app_state("template_repository", {})
        if isinstance(payload, dict):
            try:
                return TemplateRepository.from_dict(payload)
            except Exception:
                pass
        return TemplateRepository()

    def _save_template_repository(self) -> None:
        try:
            self.database.set_app_state("template_repository", self.template_repo.to_dict())
        except Exception:
            pass

    def _session_name_exists(self, exam_name: str, exclude_session_id: str = "") -> bool:
        name_norm = str(exam_name or "").strip().casefold()
        if not name_norm:
            return False
        for row in self.session_registry:
            sid = str(row.get("session_id", "") or "")
            if exclude_session_id and sid == exclude_session_id:
                continue
            if str(row.get("name", "") or "").strip().casefold() == name_norm:
                return True
        return False

    def _build_exam_list_page(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.addWidget(QLabel("Danh sách các kỳ thi"))

        self.exam_list_table = QTableWidget(0, 9)
        self.exam_list_table.setHorizontalHeaderLabels(["STT", "Tên kỳ thi", "Số môn", "Thư mục quét", "Môn học", "Trạng thái", "Xem", "Xoá", "Mặc định"])
        self.exam_list_table.verticalHeader().setVisible(False)
        self.exam_list_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.exam_list_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.exam_list_table.setSelectionMode(QAbstractItemView.SingleSelection)
        hdr = self.exam_list_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.Stretch)
        hdr.setSectionResizeMode(4, QHeaderView.Stretch)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(7, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(8, QHeaderView.ResizeToContents)
        # double click navigation: open selected exam directly.
        self.exam_list_table.cellDoubleClicked.connect(self._handle_exam_list_double_click)
        self.exam_list_table.itemSelectionChanged.connect(lambda: self._handle_stack_changed(self.stack.currentIndex()))
        layout.addWidget(self.exam_list_table)
        return w

    def _build_subject_management_page(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.addWidget(QLabel("Quản lý môn học và khối"))

        tables_row = QHBoxLayout()

        self.subjects_table = QTableWidget(0, 1)
        self.subjects_table.setHorizontalHeaderLabels(["Subject Name"])
        self.subjects_table.verticalHeader().setVisible(False)
        self.subjects_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.subjects_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.subjects_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.subjects_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.subjects_table.itemSelectionChanged.connect(lambda: self._handle_subject_management_selection("subjects"))

        self.grades_table = QTableWidget(0, 1)
        self.grades_table.setHorizontalHeaderLabels(["Grade"])
        self.grades_table.verticalHeader().setVisible(False)
        self.grades_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.grades_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.grades_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.grades_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.grades_table.itemSelectionChanged.connect(lambda: self._handle_subject_management_selection("grades"))

        tables_row.addWidget(self.subjects_table)
        tables_row.addWidget(self.grades_table)
        layout.addLayout(tables_row)

        form = QFormLayout()
        self.subject_management_label = QLabel("Subject Name")
        self.subject_management_editor = QLineEdit()
        self.subject_management_editor.setPlaceholderText("Nhập giá trị đang chọn")
        form.addRow(self.subject_management_label, self.subject_management_editor)
        layout.addLayout(form)

        btn_row = QHBoxLayout()
        btn_add = QPushButton("Add")
        btn_add.clicked.connect(self._subject_management_add)
        btn_edit = QPushButton("Edit")
        btn_edit.clicked.connect(self._subject_management_edit)
        btn_delete = QPushButton("Delete")
        btn_delete.clicked.connect(self._subject_management_delete)
        btn_save = QPushButton("Save")
        btn_save.clicked.connect(self._save_subject_management)
        btn_back = QPushButton("Đóng")
        btn_back.clicked.connect(lambda: self._navigate_to("exam_list", context={}, push_current=False, require_confirm=True, reason="close_subject_management"))
        for btn in [btn_add, btn_edit, btn_delete, btn_save, btn_back]:
            btn_row.addWidget(btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)
        self._refresh_subject_management_tables()
        return w

    def _subject_management_values(self, mode: str) -> list[str]:
        return self.subjects if mode == "subjects" else self.grades

    def _subject_management_table(self, mode: str) -> QTableWidget:
        return self.subjects_table if mode == "subjects" else self.grades_table

    def _set_subject_management_mode(self, mode: str) -> None:
        self.subject_management_mode = mode
        self.subject_management_label.setText("Subject Name" if mode == "subjects" else "Grade")

    def _refresh_subject_management_tables(self) -> None:
        self.subjects = list(self.subject_catalog)
        self.grades = list(self.block_catalog)
        for mode, values in (("subjects", self.subjects), ("grades", self.grades)):
            table = self._subject_management_table(mode)
            table.blockSignals(True)
            table.setRowCount(len(values))
            for row, value in enumerate(values):
                table.setItem(row, 0, QTableWidgetItem(value))
            table.clearSelection()
            table.blockSignals(False)
        self._subject_management_add()

    def _handle_subject_management_selection(self, mode: str) -> None:
        table = self._subject_management_table(mode)
        row = table.currentRow()
        if row < 0:
            return
        other_mode = "grades" if mode == "subjects" else "subjects"
        other_table = self._subject_management_table(other_mode)
        other_table.blockSignals(True)
        other_table.clearSelection()
        other_table.blockSignals(False)
        values = self._subject_management_values(mode)
        self._set_subject_management_mode(mode)
        self.subject_edit_index = row
        self.subject_management_editor.setText(values[row] if row < len(values) else "")

    def _subject_management_add(self) -> None:
        self.subject_edit_index = None
        self.subject_management_editor.clear()
        self._set_subject_management_mode(self.subject_management_mode or "subjects")
        self.subject_management_editor.setFocus()

    def _subject_management_edit(self) -> None:
        table = self._subject_management_table(self.subject_management_mode)
        row = table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Quản lý môn học", "Vui lòng chọn một dòng để chỉnh sửa.")
            return
        values = self._subject_management_values(self.subject_management_mode)
        self.subject_edit_index = row
        self.subject_management_editor.setText(values[row] if row < len(values) else "")
        self.subject_management_editor.setFocus()

    def _subject_management_delete(self) -> None:
        table = self._subject_management_table(self.subject_management_mode)
        row = table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Quản lý môn học", "Vui lòng chọn một dòng để xoá.")
            return
        values = self._subject_management_values(self.subject_management_mode)
        if row >= len(values):
            return
        del values[row]
        self._apply_subject_management_values()
        self._refresh_subject_management_tables()
        self._set_subject_management_mode(self.subject_management_mode)

    def _apply_subject_management_values(self) -> None:
        old_subjects = list(self.subject_catalog)
        old_blocks = list(self.block_catalog)
        self.subject_catalog = list(self.subjects)
        self.block_catalog = list(self.grades)
        self.database.replace_catalog("subjects", self.subject_catalog)
        self.database.replace_catalog("blocks", self.block_catalog)
        if old_subjects != self.subject_catalog:
            self.database.log_change("catalog", "subjects", "subject_catalog", old_subjects, self.subject_catalog, "subject_management")
        if old_blocks != self.block_catalog:
            self.database.log_change("catalog", "blocks", "block_catalog", old_blocks, self.block_catalog, "subject_management")

    def _sync_subject_configs_with_catalog(self) -> bool:
        if not self.session:
            return False
        cfg = dict(self.session.config or {})
        subject_configs = cfg.get("subject_configs", []) if isinstance(cfg.get("subject_configs", []), list) else []
        allowed = {x.casefold() for x in self.subject_catalog}
        filtered = [x for x in subject_configs if str(x.get("name", "")).strip().casefold() in allowed]
        changed = len(filtered) != len(subject_configs)
        if changed:
            cfg["subject_configs"] = filtered
            self.session.config = cfg
            self.session.subjects = [
                f"{str(x.get('name', '')).strip()}_{str(x.get('block', '')).strip()}"
                for x in filtered
                if str(x.get("name", "")).strip()
            ]
            self.session_dirty = True
            self._refresh_batch_subject_controls()
            self._refresh_session_info()
        return changed

    def _save_subject_management(self) -> None:
        value = self.subject_management_editor.text().strip()
        values = self._subject_management_values(self.subject_management_mode)
        label = "môn học" if self.subject_management_mode == "subjects" else "khối"
        if value:
            if self.subject_edit_index is None:
                values.append(value)
            else:
                values[self.subject_edit_index] = value
        if not self.subjects:
            QMessageBox.warning(self, "Quản lý môn học", "Danh sách môn học không được để trống.")
            return
        if not self.grades:
            QMessageBox.warning(self, "Quản lý khối", "Danh sách khối không được để trống.")
            return

        normalized = [item.strip() for item in values if item.strip()]
        if len(normalized) != len(set(x.casefold() for x in normalized)):
            QMessageBox.warning(self, "Quản lý môn học", f"Danh sách {label} không được trùng lặp.")
            return
        if self.subject_management_mode == "subjects":
            self.subjects = normalized
        else:
            self.grades = normalized
        self._apply_subject_management_values()
        self.session_dirty = True

        if self.session:
            cfg = dict(self.session.config or {})
            cfg["subject_catalog"] = list(self.subject_catalog)
            cfg["block_catalog"] = list(self.block_catalog)
            self.session.config = cfg
            removed = self._sync_subject_configs_with_catalog()
            if removed:
                QMessageBox.information(self, "Quản lý môn học", "Đã cập nhật danh sách môn/khối và đồng bộ các môn trong kỳ thi hiện tại.")
            else:
                QMessageBox.information(self, "Quản lý môn học", "Đã cập nhật danh sách môn và khối.")
        else:
            QMessageBox.information(self, "Quản lý môn học", "Đã cập nhật danh sách môn và khối.")
        self._refresh_subject_management_tables()
        self._set_subject_management_mode(self.subject_management_mode)

    def _build_workspace_page(self) -> QWidget:
        central = QWidget()
        root_layout = QVBoxLayout(central)

        # Keep only Batch Scan UI visible in workspace.
        group_scan = QGroupBox("Batch Scan")
        l2 = QVBoxLayout(group_scan); l2.addWidget(self._build_scan_tab())
        root_layout.addWidget(group_scan)

        # Initialize hidden widgets still used by existing logic.
        self._hidden_session_tab = self._build_session_tab()
        self._hidden_correction_tab = self._build_correction_tab()
        return central

    def _session_id_for_row(self, row_idx: int) -> str | None:
        item = self.exam_list_table.item(row_idx, 1)
        if not item:
            return None
        sid = item.data(Qt.UserRole)
        return str(sid) if sid else None

    def _handle_exam_list_double_click(self, row: int, col: int) -> None:
        # double click navigation: route row double click to exam editor action.
        sid = self._session_id_for_row(row)
        if not sid:
            return
        self.exam_list_table.selectRow(row)
        self._edit_registry_session_by_id(sid)

    def _handle_template_library_double_click(self, row: int, col: int) -> None:
        # double click navigation: route row double click to template edit action.
        if row < 0:
            return
        self.template_library_table.selectRow(row)
        self._edit_selected_template()

    def _handle_scan_list_double_click(self, row: int, col: int) -> None:
        # double click navigation: route row double click to scan edit action.
        if row < 0 or row >= self.scan_list.rowCount():
            return
        self.scan_list.selectRow(row)
        self._open_edit_selected_scan()

    def _make_row_icon_button(self, icon, tooltip: str, cb):
        btn = QPushButton()
        btn.setIcon(icon)
        btn.setToolTip(tooltip)
        btn.setFlat(True)
        btn.clicked.connect(cb)
        return btn

    def _begin_scan_grid_update(self) -> None:
        if not hasattr(self, "scan_list"):
            return
        self._scan_grid_loading = True
        self.scan_list.setUpdatesEnabled(False)
        self.scan_list.blockSignals(True)
        try:
            self.scan_list.setSortingEnabled(False)
        except Exception:
            pass

    def _end_scan_grid_update(self) -> None:
        if not hasattr(self, "scan_list"):
            self._scan_grid_loading = False
            return
        try:
            self.scan_list.setSortingEnabled(False)
        except Exception:
            pass
        self.scan_list.blockSignals(False)
        self.scan_list.setUpdatesEnabled(True)
        self._scan_grid_loading = False

    def _open_wait_progress(self, label_text: str, title: str = "Đang xử lý...") -> QProgressDialog:
        dlg = QProgressDialog(label_text, "", 0, 0, self)
        dlg.setWindowTitle(title)
        dlg.setCancelButton(None)
        dlg.setMinimumDuration(0)
        dlg.setWindowModality(Qt.WindowModal)
        dlg.setValue(0)
        dlg.show()
        QApplication.processEvents()
        return dlg

    @staticmethod
    def _close_wait_progress(dlg: QProgressDialog | None) -> None:
        if dlg is None:
            return
        try:
            dlg.close()
            dlg.deleteLater()
        except Exception:
            pass

    @staticmethod
    def _format_eta_text(seconds: float) -> str:
        sec = max(0, int(round(float(seconds or 0.0))))
        if sec <= 0:
            return "~0s"
        m, s = divmod(sec, 60)
        h, m = divmod(m, 60)
        if h > 0:
            return f"~{h}h {m:02d}m {s:02d}s"
        if m > 0:
            return f"~{m}m {s:02d}s"
        return f"~{s}s"

    def _open_batch_progress_screen(self, total_items: int, title: str = "Đang nhận dạng Batch Scan") -> QDialog:
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.setModal(True)
        lay = QVBoxLayout(dlg)
        lbl_title = QLabel("Đang nhận dạng bài thi...")
        lbl_total = QLabel(f"Tổng số bài cần nhận dạng: {max(0, int(total_items or 0))}")
        lbl_current = QLabel("Đang nhận dạng bài thứ: 0/0")
        lbl_eta = QLabel("Thời gian còn lại ước tính: -")
        prog = QProgressBar(dlg)
        prog.setMinimum(0)
        prog.setMaximum(max(1, int(total_items or 0)))
        prog.setValue(0)
        prog.setFormat("%v/%m bài")
        lay.addWidget(lbl_title)
        lay.addWidget(lbl_total)
        lay.addWidget(lbl_current)
        lay.addWidget(lbl_eta)
        lay.addWidget(prog)
        setattr(dlg, "_batch_lbl_total", lbl_total)
        setattr(dlg, "_batch_lbl_current", lbl_current)
        setattr(dlg, "_batch_lbl_eta", lbl_eta)
        setattr(dlg, "_batch_progress", prog)
        dlg.resize(520, 180)
        dlg.show()
        QApplication.processEvents()
        return dlg

    def _update_batch_progress_screen(self, dlg: QDialog | None, current: int, total: int, image_path: str, started_at: float) -> None:
        if dlg is None:
            return
        current_safe = max(0, int(current or 0))
        total_safe = max(1, int(total or 1))
        lbl_total = getattr(dlg, "_batch_lbl_total", None)
        lbl_current = getattr(dlg, "_batch_lbl_current", None)
        lbl_eta = getattr(dlg, "_batch_lbl_eta", None)
        prog = getattr(dlg, "_batch_progress", None)
        if isinstance(lbl_total, QLabel):
            lbl_total.setText(f"Tổng số bài cần nhận dạng: {total_safe}")
        if isinstance(lbl_current, QLabel):
            lbl_current.setText(f"Đang nhận dạng bài thứ: {min(current_safe, total_safe)}/{total_safe} - {Path(str(image_path or '')).name or '-'}")
        elapsed = max(0.0, float(time.perf_counter() - float(started_at or 0.0)))
        remain_items = max(0, total_safe - current_safe)
        eta_sec = (elapsed / current_safe) * remain_items if current_safe > 0 else 0.0
        if isinstance(lbl_eta, QLabel):
            lbl_eta.setText(f"Thời gian còn lại ước tính: {self._format_eta_text(eta_sec)}")
        if isinstance(prog, QProgressBar):
            prog.setMaximum(total_safe)
            prog.setValue(min(current_safe, total_safe))
        QApplication.processEvents()

    @staticmethod
    def _close_batch_progress_screen(dlg: QDialog | None) -> None:
        if dlg is None:
            return
        try:
            dlg.close()
            dlg.deleteLater()
        except Exception:
            pass

    def _set_scan_action_widget(self, row: int) -> None:
        if row < 0 or row >= self.scan_list.rowCount():
            return
        style = self.style()
        holder = QWidget()
        lay = QHBoxLayout(holder)
        lay.setContentsMargins(2, 0, 2, 0)
        lay.setSpacing(2)
        btn_edit = self._make_row_icon_button(style.standardIcon(QStyle.SP_FileDialogDetailedView), "Sửa bài thi", lambda _=False, r=row: self._edit_scan_row_by_index(r))
        btn_delete = self._make_row_icon_button(style.standardIcon(QStyle.SP_TrashIcon), "Xoá bài thi", lambda _=False, r=row: self._delete_scan_row_by_index(r))
        lay.addWidget(btn_edit)
        lay.addWidget(btn_delete)
        self.scan_list.setCellWidget(row, self.SCAN_COL_ACTIONS, holder)

    def _ensure_scan_action_widget(self, row: int) -> None:
        if row < 0 or row >= self.scan_list.rowCount():
            return
        if self.scan_list.cellWidget(row, self.SCAN_COL_ACTIONS) is not None:
            return
        self._set_scan_action_widget(row)

    def _open_scan_row_context_menu(self, pos) -> None:
        if not hasattr(self, "scan_list"):
            return
        row = self.scan_list.rowAt(pos.y())
        if row < 0 or row >= self.scan_list.rowCount():
            return
        self.scan_list.selectRow(row)
        menu = QMenu(self)
        act_edit = menu.addAction("Sửa bài thi")
        act_delete = menu.addAction("Xoá bài thi")
        act_rerecognize = menu.addAction("Nhận dạng lại bài thi")
        chosen = menu.exec(self.scan_list.viewport().mapToGlobal(pos))
        if chosen == act_edit:
            self._edit_scan_row_by_index(row)
            return
        if chosen == act_delete:
            self._delete_scan_row_by_index(row, require_confirm=True)
            return
        if chosen == act_rerecognize:
            self.scan_list.selectRow(row)
            self._on_scan_selected()
            self._rerecognize_selected_scan()

    def _edit_scan_row_by_index(self, row: int) -> None:
        if row < 0 or row >= self.scan_list.rowCount():
            return
        self.scan_list.selectRow(row)
        self._on_scan_selected()
        self._open_edit_selected_scan()

    def _delete_scan_row_by_index(self, row: int, require_confirm: bool = False) -> None:
        self._ensure_correction_state()
        if row < 0 or row >= self.scan_list.rowCount():
            return
        self.scan_list.selectRow(row)
        if self.correction_save_timer.isActive():
            self.correction_save_timer.stop()
            self._flush_pending_correction_updates()

        idx = self.scan_list.currentRow()
        sid_item = self.scan_list.item(idx, self.SCAN_COL_STUDENT_ID) if 0 <= idx < self.scan_list.rowCount() else None
        image_path = str(sid_item.data(Qt.UserRole) if sid_item else "").strip()
        sid_for_score = str(sid_item.text() if sid_item else "").strip()
        if not image_path:
            result = self.scan_results[idx] if 0 <= idx < len(self.scan_results) else self._build_result_from_saved_table_row(idx)
            if result is None:
                return
            image_path = str(getattr(result, "image_path", "") or "").strip()
            sid_for_score = str(getattr(result, "student_id", "") or sid_for_score).strip()
        if not image_path:
            return
        confirm_message = f"Bạn có chắc muốn xoá bản ghi này?\n\nẢnh: {Path(image_path).name}"
        if require_confirm:
            confirm_message = (
                "Bạn có chắc muốn xoá bài thi đã chọn?\n\n"
                f"Ảnh: {Path(image_path).name}"
            )
        confirm = QMessageBox.question(
            self,
            "Xoá bài thi",
            confirm_message,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        wait_dlg = self._open_wait_progress("Đang xoá bản ghi và cập nhật danh sách...")
        subject_key = self._current_batch_subject_key()
        self._mark_deleted_scan_image(subject_key, image_path)
        if sid_for_score and sid_for_score != "-":
            self._invalidate_scoring_for_student_ids([sid_for_score], subject_key=subject_key, reason="delete_scan_row")
        scoped_subject = self._batch_result_subject_key(subject_key)
        candidate_subject_keys: list[str] = [scoped_subject]
        sid = str(self.current_session_id or "").strip()
        legacy_scoped = f"{sid}::{subject_key}" if sid and subject_key else ""
        if legacy_scoped and legacy_scoped not in candidate_subject_keys:
            candidate_subject_keys.append(legacy_scoped)
        if subject_key and subject_key not in candidate_subject_keys:
            candidate_subject_keys.append(subject_key)
        try:
            QApplication.processEvents()
            # Keep direct delete by current subject key for compatibility with legacy flows/tests.
            self.database.delete_scan_result(subject_key, image_path)
            for key in candidate_subject_keys:
                self.database.delete_scan_result(key, image_path)
            self.scan_results = [x for x in self._refresh_scan_results_from_db(subject_key) if str(getattr(x, "image_path", "") or "") != ""]
            self.scan_results_by_subject[scoped_subject] = list(self.scan_results)
            self._populate_scan_grid_from_results(self.scan_results)
            self._rebuild_error_list()
            self._refresh_all_statuses()
            self._update_batch_scan_bottom_status_text()
            if self.scan_list.rowCount() > 0:
                target_row = min(row, self.scan_list.rowCount() - 1)
                self.scan_list.selectRow(target_row)
                self._on_scan_selected()
            else:
                self.scan_result_preview.setRowCount(0)
                self.result_preview.clear()
                self.manual_edit.clear()
                self.scan_image_preview.setText("Chọn bài thi ở danh sách bên trái")
                self.scan_image_preview.clear_markers()
        finally:
            self._close_wait_progress(wait_dlg)

    def _refresh_exam_list(self) -> None:
        self.exam_list_table.setRowCount(len(self.session_registry))
        style = self.style()
        for idx, row in enumerate(self.session_registry):
            sid = str(row.get("session_id", ""))
            name = str(row.get("name") or f"Kỳ thi {idx+1}")
            subject_text = "-"
            subject_count = "0"
            scan_root = "-"
            payload = self.database.fetch_exam_session(sid) if sid else None
            if sid and payload:
                try:
                    ses = ExamSession.from_dict(payload)
                    cfg = ses.config or {}
                    subject_cfgs = cfg.get("subject_configs", []) if isinstance(cfg.get("subject_configs", []), list) else []
                    subject_count = str(len(subject_cfgs))
                    subject_text = ", ".join(f"{x.get('name','?')}-{x.get('block','?')}" for x in subject_cfgs[:4])
                    if len(subject_cfgs) > 4:
                        subject_text += f" ...(+{len(subject_cfgs)-4})"
                    scan_root = str(cfg.get("scan_root", "") or "-")
                except Exception:
                    pass
            status = "Mặc định" if bool(row.get("default")) else "Thường"
            if sid and not payload:
                status = "Không tìm thấy"

            self.exam_list_table.setItem(idx, 0, QTableWidgetItem(str(idx + 1)))
            name_item = QTableWidgetItem(name)
            name_item.setData(Qt.UserRole, sid)
            self.exam_list_table.setItem(idx, 1, name_item)
            self.exam_list_table.setItem(idx, 2, QTableWidgetItem(subject_count))
            self.exam_list_table.setItem(idx, 3, QTableWidgetItem(scan_root))
            self.exam_list_table.setItem(idx, 4, QTableWidgetItem(subject_text or "-"))
            self.exam_list_table.setItem(idx, 5, QTableWidgetItem(status))

            b_edit = self._make_row_icon_button(style.standardIcon(QStyle.SP_DialogOpenButton), "Xem kỳ thi", lambda _=False, s=sid: self._edit_registry_session_by_id(s))
            b_del = self._make_row_icon_button(style.standardIcon(QStyle.SP_TrashIcon), "Xoá kỳ thi", lambda _=False, s=sid: self._delete_registry_session_by_id(s))
            b_def = self._make_row_icon_button(style.standardIcon(QStyle.SP_DialogApplyButton), "Đặt mặc định", lambda _=False, s=sid: self._set_default_registry_session_by_id(s))
            edit_wrap = QWidget(); e_l = QHBoxLayout(edit_wrap); e_l.setContentsMargins(0, 0, 0, 0); e_l.addWidget(b_edit)
            del_wrap = QWidget(); d_l = QHBoxLayout(del_wrap); d_l.setContentsMargins(0, 0, 0, 0); d_l.addWidget(b_del)
            def_wrap = QWidget(); f_l = QHBoxLayout(def_wrap); f_l.setContentsMargins(0, 0, 0, 0); f_l.addWidget(b_def)
            self.exam_list_table.setCellWidget(idx, 6, edit_wrap)
            self.exam_list_table.setCellWidget(idx, 7, del_wrap)
            self.exam_list_table.setCellWidget(idx, 8, def_wrap)

        self.exam_list_table.resizeRowsToContents()

    def _selected_registry_path(self) -> Path | None:
        row = self.exam_list_table.currentRow()
        if row < 0:
            return None
        sid = self._session_id_for_row(row)
        if not sid:
            return None
        return self._session_path_from_id(sid)

    def _open_selected_registry_session(self) -> None:
        path = self._selected_registry_path()
        if not path:
            QMessageBox.warning(self, "Mở kỳ thi", "Chọn kỳ thi trong danh sách trước.")
            return
        if not self._confirm("Mở kỳ thi", "Bạn có chắc muốn mở kỳ thi này?"):
            return
        self._open_session_path(path)

    def _edit_selected_registry_session(self) -> None:
        row = self.exam_list_table.currentRow()
        sid = self._session_id_for_row(row) if row >= 0 else None
        if not sid:
            QMessageBox.warning(self, "Sửa kỳ thi", "Chọn kỳ thi trong danh sách trước.")
            return
        self._edit_registry_session_by_id(sid)

    def _edit_registry_session_by_id(self, session_id: str) -> None:
        payload = self.database.fetch_exam_session(session_id)
        if not payload:
            QMessageBox.warning(self, "Sửa kỳ thi", "Không tìm thấy kỳ thi trong kho lưu trữ hệ thống.")
            return False        
        try:
            session = ExamSession.from_dict(payload)
            cfg = session.config or {}
            payload = {
                "exam_name": session.exam_name,
                "common_template": session.template_path,
                "scan_root": cfg.get("scan_root", ""),
                "student_list_path": cfg.get("student_list_path", ""),
                "students": [
                    {
                        "student_id": s.student_id,
                        "name": s.name,
                        "birth_date": str((s.extra or {}).get("birth_date", "") or ""),
                        "class_name": str((s.extra or {}).get("class_name", "") or ""),
                        "exam_room": str((s.extra or {}).get("exam_room", "") or ""),
                    }
                    for s in (session.students or [])
                ],
                "scan_mode": cfg.get("scan_mode", "Ảnh trong thư mục gốc"),
                "paper_part_count": cfg.get("paper_part_count", 3),
                "subject_configs": cfg.get("subject_configs", []),
            }
            self._open_embedded_exam_editor(session_id, session, payload)
        except Exception as exc:
            QMessageBox.warning(self, "Sửa kỳ thi", f"Không thể sửa kỳ thi\n{exc}")

    def _open_embedded_exam_editor(self, session_id: str, session: ExamSession, payload: dict, *, is_new: bool = False) -> None:
        while self.exam_editor_layout.count():
            item = self.exam_editor_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        if session_id:
            self.current_session_id = session_id
            self.current_session_path = self._session_path_from_id(session_id)
        if session is not None:
            self.session = session
            if not is_new:
                self.session_dirty = False
            self._refresh_session_info()
            self._refresh_batch_subject_controls()
            self._refresh_scoring_phase_table()
            self._refresh_ribbon_action_states()

        self.embedded_exam_session_id = session_id
        self.embedded_exam_session = session
        self.embedded_exam_original_payload = dict(payload)
        self.embedded_exam_is_new = bool(is_new)

        dlg = NewExamDialog(
            self.subject_catalog,
            self.block_catalog,
            data=payload,
            parent=self,
            on_batch_scan_subject=lambda x: self._handle_batch_request_from_editor(x),
            on_save_exam=self._save_embedded_exam_editor,
            stay_open_on_save=True,
            template_repo=self.template_repo,
        )
        dlg.setWindowFlags(Qt.Widget)
        dlg.rejected.connect(self._close_embedded_exam_editor)
        self.embedded_exam_dialog = dlg
        self.exam_editor_layout.addWidget(dlg)
        self._navigate_to("exam_editor", context={"session_id": session_id}, push_current=True, require_confirm=False, reason="open_exam_editor")

    def _save_embedded_exam_editor(self) -> bool:
        if not self.embedded_exam_dialog or not self.embedded_exam_session_id:
            return False
        edited = self.embedded_exam_dialog.payload()
        self._register_templates_from_payload(edited)
        session_id = self.embedded_exam_session_id
        saved_payload = self.database.fetch_exam_session(session_id)
        if not saved_payload and not self.embedded_exam_is_new:
            QMessageBox.warning(self, "Sửa kỳ thi", "Không tìm thấy kỳ thi trong kho lưu trữ hệ thống.")
            return False
        try:
            if saved_payload:
                session = ExamSession.from_dict(saved_payload)
            else:
                session = ExamSession(
                    exam_name=str(edited.get("exam_name", "") or "").strip() or "Kỳ thi",
                    exam_date=str(date.today()),
                    subjects=[],
                    template_path=str(edited.get("common_template", "") or ""),
                    answer_key_path="",
                    students=[],
                    config={},
                )
            session.exam_name = edited.get("exam_name", session.exam_name)
            session.template_path = edited.get("common_template", session.template_path)
            session.subjects = [
                f"{str(x.get('name', '')).strip()}_{str(x.get('block', '')).strip()}"
                for x in edited.get("subject_configs", [])
                if str(x.get("name", "")).strip()
            ] or session.subjects
            incoming_students = edited.get("students", []) if isinstance(edited.get("students", []), list) else []
            session.students = self._students_from_editor_rows(incoming_students)
            session.config = {
                **(session.config or {}),
                "scan_mode": edited.get("scan_mode", "Ảnh trong thư mục gốc"),
                "scan_root": edited.get("scan_root", ""),
                "student_list_path": edited.get("student_list_path", ""),
                "paper_part_count": edited.get("paper_part_count", 3),
                "subject_configs": edited.get("subject_configs", []),
                "subject_catalog": self.subject_catalog,
                "block_catalog": self.block_catalog,
            }
            self.database.save_exam_session(session_id, session.exam_name, session.to_dict())
            self.embedded_exam_original_payload = edited
            self.embedded_exam_is_new = False
            self.session = session
            self.current_session_id = session_id
            self.current_session_path = self._session_path_from_id(session_id)
            self.session_dirty = False
            self._upsert_session_registry(session_id, session.exam_name)
            self._save_session_registry()
            self._refresh_exam_list()
            self._refresh_session_info()
            self._refresh_batch_subject_controls()
            self._refresh_ribbon_action_states()
            if hasattr(self, "scoring_subject_combo"):
                preferred_scoring = self._resolve_preferred_scoring_subject()
                self._populate_scoring_subjects(preferred_scoring)
                self._refresh_scoring_phase_table()
            QMessageBox.information(self, "Xem kỳ thi", "Đã lưu thông số kỳ thi.")
            if self.embedded_exam_dialog:
                self.embedded_exam_dialog.show()
            return True
        except Exception as exc:
            QMessageBox.warning(self, "Sửa kỳ thi", f"Không thể sửa kỳ thi\n{exc}")
            return False

    @staticmethod
    def _students_from_editor_rows(rows: list[dict] | None) -> list[Student]:
        """
        Convert student rows from exam-editor payload into canonical Student objects
        without dropping valid SID-only records.
        """
        normalized: dict[str, Student] = {}
        ordered_keys: list[str] = []
        for raw in rows or []:
            if not isinstance(raw, dict):
                continue
            sid = str(raw.get("student_id", "") or "").strip()
            if not sid:
                continue
            key = sid.casefold()
            name = str(raw.get("name", "") or "").strip()
            birth_date = str(raw.get("birth_date", "") or "")
            class_name = str(raw.get("class_name", "") or "")
            exam_room = str(raw.get("exam_room", "") or "")
            existing = normalized.get(key)
            if existing is None:
                normalized[key] = Student(
                    student_id=sid,
                    name=name,
                    extra={
                        "birth_date": birth_date,
                        "class_name": class_name,
                        "exam_room": exam_room,
                    },
                )
                ordered_keys.append(key)
                continue
            # Merge duplicates by SID: prefer non-empty values from the newest row.
            if name:
                existing.name = name
            merged_extra = dict(getattr(existing, "extra", {}) or {})
            if birth_date:
                merged_extra["birth_date"] = birth_date
            if class_name:
                merged_extra["class_name"] = class_name
            if exam_room:
                merged_extra["exam_room"] = exam_room
            existing.extra = merged_extra
        return [normalized[k] for k in ordered_keys if k in normalized]

    def _close_embedded_exam_editor(self) -> None:
        self.embedded_exam_dialog = None
        self.embedded_exam_session = None
        self.embedded_exam_session_id = None
        self.embedded_exam_original_payload = None
        self.embedded_exam_is_new = False
        self._navigate_to("exam_list", context={}, push_current=False, require_confirm=False, reason="close_exam_editor")

    @staticmethod
    def _payload_changed(a: dict | None, b: dict | None) -> bool:
        return json.dumps(a or {}, ensure_ascii=False, sort_keys=True) != json.dumps(b or {}, ensure_ascii=False, sort_keys=True)

    def _handle_batch_request_from_editor(self, batch_payload: dict) -> bool:
        if not self.embedded_exam_dialog or not self.embedded_exam_session or not self.embedded_exam_session_id:
            return False

        current_payload = self.embedded_exam_dialog.payload()
        if self._payload_changed(current_payload, self.embedded_exam_original_payload):
            msg = QMessageBox(self)
            msg.setWindowTitle("Xác nhận")
            msg.setText("Dữ liệu đã thay đổi. Bạn muốn lưu trước khi chuyển sang nhận dạng?")
            msg.setStandardButtons(QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
            msg.setDefaultButton(QMessageBox.Save)
            ch = msg.exec()
            if ch == QMessageBox.Cancel:
                return False
            if ch == QMessageBox.Save and not self._save_embedded_exam_editor():
                return False

        session_id = self.embedded_exam_session_id
        base_session = self.embedded_exam_session
        self._close_embedded_exam_editor()
        if not session_id or not base_session:
            return False
        self._open_batch_scan_from_exam_editor(session_id, base_session, batch_payload)
        return True


    def _open_batch_scan_from_exam_editor(self, session_id: str, base_session: ExamSession, payload: dict) -> None:
        subject_cfg = dict(payload.get("subject_config") or {})
        if not subject_cfg:
            QMessageBox.warning(self, "Batch Scan", "Không tìm thấy cấu hình môn để nhận dạng.")
            return

        exam_name = str(payload.get("exam_name") or base_session.exam_name or "Kỳ thi")
        common_template = str(payload.get("common_template") or base_session.template_path or "")
        all_subjects = payload.get("subject_configs")
        if not isinstance(all_subjects, list) or not all_subjects:
            all_subjects = list((base_session.config or {}).get("subject_configs", []))
        self.batch_editor_return_payload = {
            "exam_name": exam_name,
            "common_template": common_template,
            "scan_root": str(payload.get("scan_root") or (base_session.config or {}).get("scan_root", "") or ""),
            "student_list_path": str(payload.get("student_list_path") or (base_session.config or {}).get("student_list_path", "") or ""),
            "students": list(payload.get("students", [])) if isinstance(payload.get("students", []), list) else [
                {
                    "student_id": s.student_id,
                    "name": s.name,
                    "birth_date": str((s.extra or {}).get("birth_date", "") or ""),
                    "class_name": str((s.extra or {}).get("class_name", "") or ""),
                    "exam_room": str((s.extra or {}).get("exam_room", "") or ""),
                }
                for s in (base_session.students or [])
            ],
            "scan_mode": str(payload.get("scan_mode") or (base_session.config or {}).get("scan_mode", "Ảnh trong thư mục gốc")),
            "paper_part_count": int(payload.get("paper_part_count") or (base_session.config or {}).get("paper_part_count", 3) or 3),
            "subject_configs": all_subjects,
        }
        self.batch_editor_return_session_id = session_id
        selected_subject_index = int(payload.get("selected_subject_index", 0) or 0)
        scan_root = str(payload.get("scan_root") or (base_session.config or {}).get("scan_root", "") or "")
        scan_mode = str(payload.get("scan_mode") or (base_session.config or {}).get("scan_mode", "Ảnh trong thư mục gốc"))
        paper_part_count = int(payload.get("paper_part_count") or (base_session.config or {}).get("paper_part_count", 3) or 3)

        self.session = ExamSession(
            exam_name=exam_name,
            exam_date=str(date.today()),
            subjects=[f"{subject_cfg.get('name', '')}_{subject_cfg.get('block', '')}"],
            template_path=common_template,
            answer_key_path=str(base_session.answer_key_path or ""),
            students=list(base_session.students or []),
            config={
                "scan_mode": scan_mode,
                "scan_root": scan_root,
                "student_list_path": str(payload.get("student_list_path") or (base_session.config or {}).get("student_list_path", "") or ""),
                "paper_part_count": paper_part_count,
                "subject_configs": all_subjects,
                "subject_catalog": self.subject_catalog,
                "block_catalog": self.block_catalog,
            },
        )

        self.current_session_id = session_id
        self.current_session_path = self._session_path_from_id(session_id)
        self.session_dirty = True
        self._refresh_session_info()
        self._refresh_batch_subject_controls()
        self.batch_subject_combo.setCurrentIndex(max(1, selected_subject_index + 1))
        selected_instance_key = self._subject_instance_key_from_cfg(subject_cfg)
        self._navigate_to(
            "workspace_batch_scan",
            context={
                "session_id": session_id,
                "origin": "exam_editor",
                "selected_subject_instance_key": selected_instance_key,
            },
            push_current=True,
            require_confirm=False,
            reason="open_batch_from_exam_editor",
        )

    def _delete_selected_registry_session(self) -> None:
        row = self.exam_list_table.currentRow()
        sid = self._session_id_for_row(row) if row >= 0 else None
        if not sid:
            QMessageBox.warning(self, "Xoá kỳ thi", "Chọn kỳ thi trong danh sách trước.")
            return
        self._delete_registry_session_by_id(sid)

    def _delete_registry_session_by_id(self, session_id: str) -> None:
        if not self._confirm("Xoá kỳ thi", "Bạn có chắc muốn xoá kỳ thi khỏi danh sách?"):
            return
        self.database.delete_exam_session(session_id)
        self.session_registry = self._load_session_registry()
        self._save_session_registry()
        self._refresh_exam_list()

    def _set_default_selected_registry_session(self) -> None:
        row = self.exam_list_table.currentRow()
        sid = self._session_id_for_row(row) if row >= 0 else None
        if not sid:
            QMessageBox.warning(self, "Đặt mặc định", "Chọn kỳ thi trong danh sách trước.")
            return
        self._set_default_registry_session_by_id(sid)

    def _set_default_registry_session_by_id(self, session_id: str) -> None:
        if not self._confirm("Đặt mặc định", "Đặt kỳ thi này làm mặc định?"):
            return
        self.database.set_app_state("default_session_id", str(session_id))
        self.session_registry = self._load_session_registry()
        self._save_session_registry()
        self._refresh_exam_list()

    def _open_session_path(self, path: Path) -> None:
        try:
            self._release_batch_runtime_state()
            payload = self.database.fetch_exam_session(path.stem)
            if not payload:
                raise FileNotFoundError(f"Không tìm thấy session '{path.stem}' trong SQLite.")
            self.session = ExamSession.from_dict(payload)
            self.current_session_path = path
            self.current_session_id = path.stem
            if self.session.template_path:
                t = Path(self.session.template_path)
                if t.exists():
                    self.template = Template.load_json(t)
            cfg = dict(self.session.config or {})
            cfg.pop("scoring_phases", None)
            cfg.pop("scoring_results", None)
            self.session.config = cfg
            self.scoring_phases = []
            self.scoring_results_by_subject = {}
            self._scoring_dirty_subjects = set()
            # DB-only mode: không phục hồi Batch Scan từ subject config.
            # scan_results_by_subject chỉ được nạp lại khi người dùng mở môn, bằng truy vấn DB.
            self.scan_results_by_subject = {}
            self.batch_working_state_by_subject = {}
            self.subject_catalog = list(cfg.get("subject_catalog", self.subject_catalog)) or self.subject_catalog
            self.block_catalog = list(cfg.get("block_catalog", self.block_catalog)) or self.block_catalog
            if self.session.answer_key_path:
                p = Path(self.session.answer_key_path)
                if p.exists() and p.suffix.lower() == ".json":
                    self.answer_keys = AnswerKeyRepository.load_json(p)
                    self.imported_exam_codes = sorted({k.split("::", 1)[1] for k in self.answer_keys.keys.keys() if "::" in k})
            self._upsert_session_registry(path.stem, self.session.exam_name if self.session else None)
            self._save_session_registry()
            self._refresh_exam_list()
            self.session_dirty = False
            self._remember_current_session_snapshot()
            self.batch_editor_return_payload = None
            self.batch_editor_return_session_id = None
            self._refresh_session_info()
            self._refresh_batch_subject_controls()
            self._refresh_scoring_phase_table()
            cfg = self.session.config or {}
            editor_payload = {
                "exam_name": self.session.exam_name,
                "common_template": self.session.template_path,
                "scan_root": cfg.get("scan_root", ""),
                "student_list_path": cfg.get("student_list_path", ""),
                "students": [
                    {
                        "student_id": s.student_id,
                        "name": s.name,
                        "birth_date": str((s.extra or {}).get("birth_date", "") or ""),
                        "class_name": str((s.extra or {}).get("class_name", "") or ""),
                        "exam_room": str((s.extra or {}).get("exam_room", "") or ""),
                    }
                    for s in (self.session.students or [])
                ],
                "scan_mode": cfg.get("scan_mode", "Ảnh trong thư mục gốc"),
                "paper_part_count": cfg.get("paper_part_count", 3),
                "subject_configs": cfg.get("subject_configs", []),
            }
            self._open_embedded_exam_editor(path.stem, self.session, editor_payload)
            self._refresh_ribbon_action_states()
            QMessageBox.information(self, "Open session", "Đã mở kỳ thi thành công.")
        except Exception as exc:
            QMessageBox.warning(self, "Open session", f"Không thể mở kỳ thi:\n{exc}")

    def _build_menu(self) -> None:
        file_menu = self.menuBar().addMenu("File")
        self.act_new_session = file_menu.addAction("Tạo kỳ thi mới")
        self.act_new_session.setShortcut(QKeySequence("Ctrl+N"))
        self.act_new_session.triggered.connect(self.action_create_session)

        self.act_open_from_list = file_menu.addAction("Mở từ danh sách")
        self.act_open_from_list.setShortcut(QKeySequence("Ctrl+O"))
        self.act_open_from_list.triggered.connect(self.action_open_session)

        self.act_save_session = file_menu.addAction("Lưu kỳ thi")
        self.act_save_session.setShortcut(QKeySequence("Ctrl+S"))
        self.act_save_session.triggered.connect(self.action_save_session)

        self.act_save_as_subject = file_menu.addAction("Lưu dưới tên khác")
        self.act_save_as_subject.triggered.connect(self.action_save_session_as)

        self.act_close_current_session = file_menu.addAction("Đóng kỳ thi hiện tại")
        self.act_close_current_session.triggered.connect(self.action_close_current_session)

        file_menu.addSeparator()
        self.act_manage_template = file_menu.addAction("Quản lý mẫu giấy thi")
        self.act_manage_template.triggered.connect(self.action_manage_template)
        self.act_close_template_module = file_menu.addAction("Đóng quản lý mẫu giấy thi")
        self.act_close_template_module.triggered.connect(self._close_template_module)

        act_manage_subject = file_menu.addAction("Quản lý môn học")
        act_manage_subject.triggered.connect(self.action_manage_subjects)

        file_menu.addSeparator()
        act_exit = file_menu.addAction("Thoát")
        act_exit.triggered.connect(self.action_exit)

        exam_menu = self.menuBar().addMenu("Exam")
        exam_menu.addAction("Load Template JSON", self.action_load_template)
        exam_menu.addAction("Load Answer Keys JSON", self.action_load_answer_keys)
        exam_menu.addAction("Import Answer Key", self.action_import_answer_key)
        exam_menu.addAction("Export Answer Key Sample", self.action_export_answer_key_sample)
        exam_menu.addAction("Batch Scan Images", self.action_run_batch_scan)
        exam_menu.addAction("Sửa bài thi được chọn", self.action_edit_selected_scan)
        exam_menu.addAction("Load Selected Scan Result", self.action_load_selected_scan_result)
        exam_menu.addAction("Apply Manual Correction", self.action_apply_manual_correction)

        scoring_menu = self.menuBar().addMenu("Scoring")
        scoring_menu.addAction("Calculate & Preview Scores", self.action_calculate_scores)
        scoring_menu.addAction("Export Results", self.action_export_results)

        self.export_menu = self.menuBar().addMenu("Export")
        self.act_export_subject_scores = self.export_menu.addAction("Xuất điểm môn...")
        self.act_export_subject_scores.triggered.connect(self.action_export_subject_scores)
        self.act_export_subject_score_matrix = self.export_menu.addAction("Xuất điểm các môn...")
        self.act_export_subject_score_matrix.triggered.connect(self.action_export_subject_score_matrix)
        self.act_export_class_subject_scores = self.export_menu.addAction("Xuất điểm lớp...")
        self.act_export_class_subject_scores.triggered.connect(self.action_export_class_subject_scores)
        self.act_export_all_classes_subject_scores = self.export_menu.addAction("Xuất điểm các lớp...")
        self.act_export_all_classes_subject_scores.triggered.connect(self.action_export_all_classes_subject_scores)
        self.act_export_all_scores = self.export_menu.addAction("Xuất điểm chi tiết các môn...")
        self.act_export_all_scores.triggered.connect(self.action_export_all_subject_scores)
        self.act_export_return_by_class = self.export_menu.addAction("Trả bài theo lớp...")
        self.act_export_return_by_class.triggered.connect(self.action_export_return_by_class)
        self.export_menu.addSeparator()
        self.act_export_subject_api = self.export_menu.addAction("Xuất API bài làm theo môn (;)")
        self.act_export_subject_api.triggered.connect(self.action_export_subject_api_payload)
        self.export_menu.addSeparator()
        self.act_export_reports_center = self.export_menu.addAction("Báo cáo thống kê...")
        self.act_export_reports_center.triggered.connect(self.action_open_export_reports_center)
        self.act_export_range_report = self.export_menu.addAction("Báo cáo thống kê khoảng điểm...")
        self.act_export_range_report.triggered.connect(self.action_export_score_range_report)
        self.act_export_class_report = self.export_menu.addAction("Báo cáo thống kê theo lớp...")
        self.act_export_class_report.triggered.connect(self.action_export_class_report)
        self.act_export_management_report = self.export_menu.addAction("Báo cáo tổng hợp quản lý...")
        self.act_export_management_report.triggered.connect(self.action_export_management_report)

        self.template_module_menu = self.menuBar().addMenu("Template Editor")
        self.template_module_menu.addAction("Tạo mới", self._create_new_template)
        self.template_module_menu.addAction("Sửa", self._edit_selected_template)
        self.template_module_menu.addAction("Xoá", self._delete_selected_template)
        self.template_module_menu.addSeparator()
        self.template_module_menu.addAction("Save", self._save_current_template)
        self.template_module_menu.addAction("Save As", self._save_current_template_as)
        self.template_module_menu.addSeparator()
        self.template_module_menu.addAction("Close", self._close_template_module)

        toolbar = QToolBar("Ribbon")
        toolbar.setMovable(False)
        toolbar.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        self.addToolBar(toolbar)
        self.main_ribbon = toolbar

        style = self.style()
        # Session actions
        self.ribbon_new_exam_action = toolbar.addAction(style.standardIcon(QStyle.SP_FileIcon), "Tạo mới", self.action_create_session)
        self.ribbon_view_exam_action = toolbar.addAction(style.standardIcon(QStyle.SP_FileDialogListView), "Danh sách kỳ thi", self.action_open_session)
        # self.ribbon_delete_exam_action = toolbar.addAction(style.standardIcon(QStyle.SP_TrashIcon), "Xoá", self._delete_selected_registry_session)
        toolbar.addSeparator()
        # Workflow actions
        self.ribbon_subject_list_action = toolbar.addAction(style.standardIcon(QStyle.SP_DirIcon), "Danh sách môn thi", self.action_open_current_exam_subjects)
        self.ribbon_batch_scan_action = toolbar.addAction(style.standardIcon(QStyle.SP_ComputerIcon), "Xử lý ảnh", self.action_run_batch_scan)
        self.ribbon_scoring_action = toolbar.addAction(style.standardIcon(QStyle.SP_CommandLink), "Tính điểm", self.action_calculate_scores)
        self.ribbon_recheck_action = toolbar.addAction(style.standardIcon(QStyle.SP_BrowserReload), "Phúc tra", self.action_open_recheck)
        self.ribbon_export_action = QAction(style.standardIcon(QStyle.SP_DriveNetIcon), "Export", self)
        self.ribbon_export_action.triggered.connect(self.action_open_export_reports_center)
        self.ribbon_export_action.setMenu(self.export_menu)
        toolbar.addAction(self.ribbon_export_action)
        toolbar.addSeparator()
        self.ribbon_batch_execute_action = toolbar.addAction(style.standardIcon(QStyle.SP_MediaPlay), "Nhận dạng", self.action_execute_batch_scan)
        self.ribbon_batch_save_action = toolbar.addAction(style.standardIcon(QStyle.SP_DialogSaveButton), "Lưu", self._save_batch_for_selected_subject)
        self.ribbon_batch_close_action = toolbar.addAction(style.standardIcon(QStyle.SP_DialogCloseButton), "Đóng", self._close_batch_scan_view)
        toolbar.addSeparator()
        self.ribbon_exam_editor_add_subject_action = toolbar.addAction(style.standardIcon(QStyle.SP_FileDialogNewFolder), "Thêm môn", self._exam_editor_add_subject)
        self.ribbon_exam_editor_edit_subject_action = toolbar.addAction(style.standardIcon(QStyle.SP_FileDialogDetailedView), "Sửa môn", self._exam_editor_edit_subject)
        self.ribbon_exam_editor_delete_subject_action = toolbar.addAction(style.standardIcon(QStyle.SP_TrashIcon), "Xoá môn", self._exam_editor_delete_subject)
        self.ribbon_exam_editor_save_action = toolbar.addAction(style.standardIcon(QStyle.SP_DialogSaveButton), "Save", self._exam_editor_save)
        self.ribbon_exam_editor_close_action = toolbar.addAction(style.standardIcon(QStyle.SP_DialogCloseButton), "Đóng", self._exam_editor_close)
        toolbar.addSeparator()
        self.ribbon_add_subject_action = toolbar.addAction(style.standardIcon(QStyle.SP_FileDialogNewFolder), "Add Subject", self._subject_management_add)
        self.ribbon_edit_subject_action = toolbar.addAction(style.standardIcon(QStyle.SP_FileDialogDetailedView), "Edit", self._subject_management_edit)
        self.ribbon_delete_subject_action = toolbar.addAction(style.standardIcon(QStyle.SP_TrashIcon), "Delete Subject", self._subject_management_delete)
        self.ribbon_save_subject_action = toolbar.addAction(style.standardIcon(QStyle.SP_DialogSaveButton), "Save", self._save_subject_management)
        toolbar.addSeparator()
        self.ribbon_new_template_action = QAction(style.standardIcon(QStyle.SP_FileIcon), "Tạo mới", self)
        self.ribbon_new_template_action.triggered.connect(self._create_new_template)
        toolbar.addAction(self.ribbon_new_template_action)
        self.ribbon_edit_template_action = QAction(style.standardIcon(QStyle.SP_DialogOpenButton), "Sửa", self)
        self.ribbon_edit_template_action.triggered.connect(self._edit_selected_template)
        toolbar.addAction(self.ribbon_edit_template_action)
        self.ribbon_delete_template_action = QAction(style.standardIcon(QStyle.SP_TrashIcon), "Xoá", self)
        self.ribbon_delete_template_action.triggered.connect(self._delete_selected_template)
        toolbar.addAction(self.ribbon_delete_template_action)
        self.ribbon_save_template_action = QAction(style.standardIcon(QStyle.SP_DialogSaveButton), "Save", self)
        self.ribbon_save_template_action.triggered.connect(self._save_current_template)
        toolbar.addAction(self.ribbon_save_template_action)
        self.ribbon_save_template_as_action = QAction(style.standardIcon(QStyle.SP_DriveFDIcon), "Save As", self)
        self.ribbon_save_template_as_action.triggered.connect(self._save_current_template_as)
        toolbar.addAction(self.ribbon_save_template_as_action)
        self.ribbon_close_template_action = QAction(style.standardIcon(QStyle.SP_DialogCloseButton), "Close", self)
        self.ribbon_close_template_action.triggered.connect(self._close_template_module)
        toolbar.addAction(self.ribbon_close_template_action)

    def open_session(self) -> None:
        if self.session and self.session_dirty:
            if not self._confirm("Dữ liệu chưa lưu", "Kỳ thi hiện tại có thay đổi chưa lưu. Vẫn mở kỳ thi khác?"):
                return
        row = self.exam_list_table.currentRow()
        if row >= 0:
            path = self._selected_registry_path()
            if path:
                self._open_session_path(path)
                return
        default_rows = [x for x in self.session_registry if bool(x.get("default"))]
        if default_rows:
            sid = str(default_rows[0].get("session_id", ""))
            if sid:
                self._open_session_path(self._session_path_from_id(sid))
                return
        QMessageBox.information(self, "Mở kỳ thi", "Vui lòng chọn kỳ thi trong danh sách để mở.")

    def save_session(self) -> None:
        if not self.session:
            self.create_session()
        if not self.current_session_id:
            self.current_session_id = self._generate_session_id(self.session.exam_name if self.session else "exam")
            self.current_session_path = self._session_path_from_id(self.current_session_id)
        try:
            if self.session:
                payload = self._build_session_persistence_payload()
                self.session.config = dict(payload.get("config", {}) or {})
                self.database.save_exam_session(self.current_session_id, self.session.exam_name, payload)
            self._upsert_session_registry(self.current_session_id, self.session.exam_name if self.session else None)
            self._save_session_registry()
            self._refresh_exam_list()
            self.session_dirty = False
            self._remember_current_session_snapshot()
            QMessageBox.information(self, "Save session", "Đã lưu kỳ thi vào kho hệ thống.")
        except Exception as exc:
            QMessageBox.warning(self, "Save session", f"Không thể lưu kỳ thi:\n{exc}")

    def _persist_session_quietly(self) -> bool:
        if not self.session:
            return False
        if not self.current_session_id:
            self.current_session_id = self._generate_session_id(self.session.exam_name if self.session else "exam")
            self.current_session_path = self._session_path_from_id(self.current_session_id)
        try:
            payload = self._build_session_persistence_payload()
            self.session.config = dict(payload.get("config", {}) or {})
            self.database.save_exam_session(self.current_session_id, self.session.exam_name, payload)
            self._upsert_session_registry(self.current_session_id, self.session.exam_name if self.session else None)
            self._save_session_registry()
            self._refresh_exam_list()
            self.session_dirty = False
            self._remember_current_session_snapshot()
            return True
        except Exception:
            self.session_dirty = True
            return False

    def save_session_as(self) -> None:
        if self.stack.currentIndex() != 0:
            QMessageBox.information(self, "Lưu dưới tên khác", "Chức năng này chỉ sử dụng trong màn hình danh sách kỳ thi.")
            return
        row = self.exam_list_table.currentRow() if hasattr(self, "exam_list_table") else -1
        source_session_id = self._session_id_for_row(row) if row >= 0 else ""
        if not source_session_id:
            QMessageBox.information(self, "Lưu dưới tên khác", "Vui lòng chọn kỳ thi nguồn trong danh sách.")
            return
        source_payload = self.database.fetch_exam_session(source_session_id)
        if not isinstance(source_payload, dict) or not source_payload:
            QMessageBox.warning(self, "Lưu dưới tên khác", "Không đọc được dữ liệu kỳ thi nguồn.")
            return

        # 1) Chỉ nhập và kiểm tra tên mới trong while; không làm gì khác.
        current_name = str(source_payload.get("exam_name", "") or "").strip() or "Kỳ thi"
        while True:
            new_name, ok = QInputDialog.getText(self, "Lưu dưới tên khác", "Nhập tên kỳ thi mới:", text=current_name)
            if not ok:
                return
            new_name = str(new_name or "").strip()
            if not new_name:
                QMessageBox.warning(self, "Lưu dưới tên khác", "Tên kỳ thi không được để trống.")
                continue
            if self._session_name_exists(new_name):
                QMessageBox.warning(self, "Lưu dưới tên khác", "Tên kỳ thi đã tồn tại. Vui lòng chọn tên khác.")
                continue
            break

        # 2) Chỉ sau khi tên hợp lệ mới thực hiện copy kỳ thi.
        try:
            new_session_id = self._generate_session_id(new_name)
            source_exam_name = str(source_payload.get("exam_name", "") or "").strip().lower()
            target_exam_name = str(new_name or "").strip().lower()
            source_prefix = f"{source_session_id}::{source_exam_name}" if source_session_id and source_exam_name else source_session_id
            target_prefix = f"{new_session_id}::{target_exam_name}" if new_session_id and target_exam_name else new_session_id

            payload = copy.deepcopy(source_payload)
            payload["exam_name"] = new_name
            cfg_root = payload.get("config", {}) if isinstance(payload.get("config", {}), dict) else {}
            payload["config"] = cfg_root
            subject_cfgs = list(cfg_root.get("subject_configs", []) if isinstance(cfg_root.get("subject_configs", []), list) else [])

            subject_key_map: dict[str, str] = {}
            for subject_cfg in subject_cfgs:
                if not isinstance(subject_cfg, dict):
                    continue
                old_key = str(subject_cfg.get("subject_instance_key", "") or "").strip()
                logical = str(subject_cfg.get("logical_subject_key", "") or "").strip() or str(self._logical_subject_key_from_cfg(subject_cfg) or "General")
                new_uid = str(uuid.uuid4())
                new_key = f"{target_prefix}::{logical}::{new_uid}" if target_prefix else f"{logical}::{new_uid}"
                subject_cfg["subject_uid"] = new_uid
                subject_cfg["logical_subject_key"] = logical
                subject_cfg["subject_instance_key"] = new_key
                if old_key:
                    subject_key_map[old_key] = new_key

            self.database.save_exam_session(new_session_id, new_name, payload)

            for old_subject_key, new_subject_key in subject_key_map.items():
                source_scan_key = f"{source_prefix}::{old_subject_key}" if source_prefix else old_subject_key
                target_scan_key = f"{target_prefix}::{new_subject_key}" if target_prefix else new_subject_key
                source_rows = list(self.database.fetch_scan_results_for_subject(source_scan_key) or [])
                self.database.replace_scan_results_for_subject(target_scan_key, source_rows)

                source_keys = self.database.fetch_answer_keys_for_subject(old_subject_key)
                self.database.replace_answer_keys_for_subject(new_subject_key, source_keys)

                source_scores = list(self.database.fetch_scores_for_subject(old_subject_key) or [])
                self.database.conn.execute("DELETE FROM scores WHERE subject_key = ?", (new_subject_key,))
                for score_row in source_scores:
                    self.database.upsert_score_row(
                        new_subject_key,
                        str((score_row or {}).get("student_id", "") or ""),
                        str((score_row or {}).get("exam_code", "") or ""),
                        dict(score_row or {}),
                    )

            source_histories = list(self.database.fetch_recheck_history(source_session_id) or [])
            for item in source_histories:
                old_hist_subject = str(item.get("subject_key", "") or "")
                new_hist_subject = subject_key_map.get(old_hist_subject, old_hist_subject)
                self.database.add_recheck_history(
                    session_id=new_session_id,
                    exam_name=new_name,
                    subject_key=new_hist_subject,
                    student_code=str(item.get("student_code", "") or ""),
                    exam_code=str(item.get("exam_code", "") or ""),
                    change_text=str(item.get("change_text", "") or ""),
                    old_score=float(item.get("old_score", 0.0) or 0.0),
                    new_score=float(item.get("new_score", 0.0) or 0.0),
                    payload=dict(item.get("payload", {}) or {}),
                )

            self._upsert_session_registry(new_session_id, new_name)
            self._save_session_registry()
            self._refresh_exam_list()
            QMessageBox.information(self, "Lưu dưới tên khác", f"Đã sao chép kỳ thi thành '{new_name}'.")
        except Exception as exc:
            QMessageBox.warning(self, "Lưu dưới tên khác", f"Không thể sao chép kỳ thi:\n{exc}")

    def close_current_session(self) -> None:
        if self.session_dirty:
            if not self._confirm("Dữ liệu chưa lưu", "Kỳ thi hiện tại có thay đổi chưa lưu. Vẫn đóng?"):
                return
        self._release_batch_runtime_state()
        self._release_preview_resources()
        self._release_template_cache()
        self._release_editor_resources()
        self.session = None
        self.template = None
        self.answer_keys = None
        self.scan_results = []
        self.scan_results_by_subject = {}
        self.batch_working_state_by_subject = {}
        self.scoring_phases = []
        self.scoring_results_by_subject = {}
        self._scoring_dirty_subjects = set()
        self.scan_results_by_subject = {}
        self.current_session_path = None
        self.current_session_id = None
        self.session_dirty = False
        self._session_saved_signature = ""
        self.session_info.clear()
        self.exam_code_preview.setText("Mã đề trên phiếu trả lời mẫu: -")
        self.scan_list.setRowCount(0)
        self.score_preview_table.setRowCount(0)
        self.error_list.clear()
        self.result_preview.clear()
        self.scan_result_preview.setRowCount(0)
        self.manual_edit.clear()
        self._refresh_batch_subject_controls()
        self.stack.setCurrentIndex(0)

    def _release_batch_runtime_state(self) -> None:
        # resource cleanup / release: aggressively clear heavy per-subject runtime maps.
        self.scan_results = []
        self.scan_results_by_subject = {}
        self.batch_working_state_by_subject = {}
        self.scoring_results_by_subject = {}
        self.scoring_phases = []
        self._scoring_dirty_subjects = set()
        self.scan_files = []
        self.score_rows = []
        self.imported_exam_codes = []
        self.preview_rotation_by_index = {}
        self.scan_forced_status_by_index = {}
        self.deleted_scan_images_by_subject = {}
        self.scan_blank_questions = {}
        self.scan_blank_summary = {}
        self.scan_manual_adjustments = {}
        self.scan_edit_history = {}
        self.scan_last_adjustment = {}
        self.batch_editor_return_payload = None
        self.batch_editor_return_session_id = None
        self.active_batch_subject_key = None
        self._batch_loaded_runtime_key = ""
        self._batch_loaded_subject_signature = ""
        self._current_batch_data_source = "empty"
        self.batch_status_filter_mode = "all"
        self.scoring_status_filter_mode = "all"

    def _release_preview_resources(self) -> None:
        if hasattr(self, "scan_image_preview"):
            self.scan_image_preview.clear()
            if hasattr(self.scan_image_preview, "clear_markers"):
                self.scan_image_preview.clear_markers()
        if hasattr(self, "result_preview"):
            self.result_preview.clear()
        if hasattr(self, "preview_source_pixmap"):
            self.preview_source_pixmap = None

    def _release_template_cache(self) -> None:
        if hasattr(self, "_template_cache_by_path"):
            self._template_cache_by_path = {}

    def _release_editor_resources(self) -> None:
        for attr in ["template_editor_embedded", "embedded_exam_dialog"]:
            obj = getattr(self, attr, None)
            if obj is not None:
                try:
                    obj.deleteLater()
                except Exception:
                    pass
                setattr(self, attr, None)
        for attr in ["embedded_exam_session", "embedded_exam_original_payload"]:
            if hasattr(self, attr):
                setattr(self, attr, None)

    def manage_subjects(self) -> None:
        if not self._confirm_interrupt_active_workflows("Danh sách môn thi"):
            return
        self._refresh_subject_management_tables()
        self._set_subject_management_mode("subjects")
        self._navigate_to("subject_management", context={"session_id": self.current_session_id}, push_current=True, require_confirm=False, reason="manage_subjects")

    def action_open_current_exam_subjects(self) -> None:
        if not self._confirm_interrupt_active_workflows("Danh sách môn thi"):
            return
        if self._open_current_session_in_exam_editor():
            return
        QMessageBox.information(
            self,
            "Danh sách môn thi",
            "Chưa có kỳ thi hiện tại. Vui lòng mở hoặc tạo kỳ thi trước.",
        )

    def action_create_session(self) -> None:
        if not self._confirm_before_switching_work("kỳ thi mới"):
            return
        if not self._confirm("Tạo kỳ thi mới", "Bạn có chắc muốn tạo kỳ thi mới?"):
            return
        session_id = self._generate_session_id("new_exam")
        payload = {
            "exam_name": "",
            "common_template": "",
            "scan_root": "",
            "student_list_path": "",
            "students": [],
            "scan_mode": "Ảnh trong thư mục gốc",
            "paper_part_count": 3,
            "subject_configs": [],
        }
        draft = ExamSession(
            exam_name="Kỳ thi mới",
            exam_date=str(date.today()),
            subjects=[],
            template_path="",
            answer_key_path="",
            students=[],
            config={
                "scan_mode": "Ảnh trong thư mục gốc",
                "scan_root": "",
                "student_list_path": "",
                "paper_part_count": 3,
                "subject_configs": [],
                "subject_catalog": self.subject_catalog,
                "block_catalog": self.block_catalog,
            },
        )
        self._open_embedded_exam_editor(session_id, draft, payload, is_new=True)

    def action_open_session(self) -> None:
        if not self._confirm_interrupt_active_workflows("Danh sách kỳ thi"):
            return
        self._navigate_to("exam_list", context={}, push_current=True, require_confirm=False, reason="open_exam_list")

    def action_save_session(self) -> None:
        if self._confirm("Lưu kỳ thi", "Bạn có chắc muốn lưu kỳ thi?"):
            if self.stack.currentIndex() == 5 and self.embedded_exam_dialog:
                self._save_embedded_exam_editor()
                return
            if self.stack.currentIndex() == 4 and self.template_editor_embedded:
                self._save_current_template()
                return
            self.save_session()

    def action_save_session_as(self) -> None:
        if self._confirm("Lưu dưới tên khác", "Bạn có chắc muốn sao chép toàn bộ kỳ thi sang kỳ thi mới?"):
            self.save_session_as()

    def action_close_current_session(self) -> None:
        if not self._confirm_before_switching_work("đóng kỳ thi hiện tại"):
            return
        if self._confirm("Đóng kỳ thi", "Bạn có chắc muốn đóng kỳ thi hiện tại?"):
            self.close_current_session()

    def action_manage_template(self) -> None:
        self.open_template_editor()

    def _rebuild_template_module_menu(self, *, library_mode: bool, editor_mode: bool) -> None:
        if not hasattr(self, "template_module_menu"):
            return
        self.template_module_menu.clear()
        if editor_mode and self.template_editor_embedded:
            editor = self.template_editor_embedded
            self.template_module_menu.addAction(editor.act_load_blank)
            self.template_module_menu.addAction(editor.act_open_template)
            self.template_module_menu.addSeparator()
            self.template_module_menu.addAction(editor.act_save)
            self.template_module_menu.addAction(editor.act_save_as)
            self.template_module_menu.addSeparator()
            self.template_module_menu.addAction(editor.act_preview)
            self.template_module_menu.addAction(editor.act_test_recognition)
            self.template_module_menu.addAction(editor.act_template_qc)
            self.template_module_menu.addAction(editor.act_snap_grid)
            self.template_module_menu.addSeparator()
            self.template_module_menu.addAction(editor.act_copy)
            self.template_module_menu.addAction(editor.act_paste)
            self.template_module_menu.addAction(editor.act_duplicate)
            self.template_module_menu.addAction(editor.act_delete)
            self.template_module_menu.addAction(editor.act_delete_anchor)
            self.template_module_menu.addSeparator()
            self.template_module_menu.addAction(editor.act_zoom_in)
            self.template_module_menu.addAction(editor.act_zoom_out)
            self.template_module_menu.addSeparator()
            self.template_module_menu.addAction(self.ribbon_close_template_action)
            return
        if library_mode:
            self.template_module_menu.addAction(self.ribbon_new_template_action)
            self.template_module_menu.addAction(self.ribbon_edit_template_action)
            self.template_module_menu.addAction(self.ribbon_delete_template_action)
            self.template_module_menu.addSeparator()
            self.template_module_menu.addAction(self.ribbon_close_template_action)

    def _handle_stack_changed(self, index: int) -> None:
        route_name = self._stack_index_to_route_name(index)
        self._current_route_name = route_name
        subject_management_visible = index == 2
        template_library_visible = index == 3
        template_editor_visible = index == 4
        exam_editor_visible = index == 5
        batch_scan_visible = route_name == "workspace_batch_scan"
        template_visible = template_library_visible or template_editor_visible
        for action in [
            getattr(self, "ribbon_new_exam_action", None),
            getattr(self, "ribbon_view_exam_action", None),
            getattr(self, "ribbon_subject_list_action", None),
            getattr(self, "ribbon_delete_exam_action", None),
            getattr(self, "ribbon_batch_scan_action", None),
            getattr(self, "ribbon_scoring_action", None),
            getattr(self, "ribbon_recheck_action", None),
            getattr(self, "ribbon_export_action", None),
        ]:
            if action is not None:
                action.setVisible(not subject_management_visible and not template_visible)
        for action in [
            getattr(self, "ribbon_exam_editor_add_subject_action", None),
            getattr(self, "ribbon_exam_editor_edit_subject_action", None),
            getattr(self, "ribbon_exam_editor_delete_subject_action", None),
            getattr(self, "ribbon_exam_editor_save_action", None),
            getattr(self, "ribbon_exam_editor_close_action", None),
        ]:
            if action is not None:
                action.setVisible(exam_editor_visible)
        for action in [
            getattr(self, "ribbon_batch_execute_action", None),
            getattr(self, "ribbon_batch_save_action", None),
            getattr(self, "ribbon_batch_close_action", None),
        ]:
            if action is not None:
                action.setVisible(batch_scan_visible)
        for action in [
            getattr(self, "ribbon_add_subject_action", None),
            getattr(self, "ribbon_edit_subject_action", None),
            getattr(self, "ribbon_delete_subject_action", None),
            getattr(self, "ribbon_save_subject_action", None),
        ]:
            if action is not None:
                action.setVisible(subject_management_visible)
        template_library_actions = [
            getattr(self, "ribbon_new_template_action", None),
            getattr(self, "ribbon_edit_template_action", None),
            getattr(self, "ribbon_delete_template_action", None),
        ]
        template_editor_actions = [
            getattr(self, "ribbon_save_template_action", None),
            getattr(self, "ribbon_save_template_as_action", None),
        ]
        for action in template_library_actions:
            if action is not None:
                action.setVisible(template_library_visible)
        for action in template_editor_actions:
            if action is not None:
                action.setVisible(False)
        if getattr(self, "ribbon_close_template_action", None) is not None:
            self.ribbon_close_template_action.setVisible(template_library_visible)

        if hasattr(self, "main_ribbon"):
            self.main_ribbon.setVisible(index != 4)
        if hasattr(self, "template_module_menu"):
            self.template_module_menu.menuAction().setVisible(template_visible)
            self._rebuild_template_module_menu(library_mode=template_library_visible, editor_mode=template_editor_visible)
        if hasattr(self, "act_close_template_module"):
            self.act_close_template_module.setVisible(template_visible)
        if hasattr(self, "act_save_as_subject"):
            row = self.exam_list_table.currentRow() if hasattr(self, "exam_list_table") else -1
            sid = self._session_id_for_row(row) if row >= 0 else ""
            can_save_as = (index == 0) and bool(str(sid or "").strip())
            self.act_save_as_subject.setVisible(True)
            self.act_save_as_subject.setEnabled(can_save_as)
        self._refresh_ribbon_action_states()

    def _refresh_ribbon_action_states(self) -> None:
        has_session = self._has_session_context_for_export()
        has_subject_cfg = bool(self._effective_subject_configs_for_batch())
        if getattr(self, "ribbon_new_exam_action", None) is not None:
            self.ribbon_new_exam_action.setEnabled(True)
        if getattr(self, "ribbon_view_exam_action", None) is not None:
            self.ribbon_view_exam_action.setEnabled(True)
        if getattr(self, "ribbon_subject_list_action", None) is not None:
            self.ribbon_subject_list_action.setEnabled(bool(self.current_session_id))
        if getattr(self, "ribbon_batch_scan_action", None) is not None:
            self.ribbon_batch_scan_action.setEnabled(has_session and has_subject_cfg)
        if getattr(self, "ribbon_scoring_action", None) is not None:
            self.ribbon_scoring_action.setEnabled(has_session)
        if getattr(self, "ribbon_recheck_action", None) is not None:
            self.ribbon_recheck_action.setEnabled(has_session)
        has_embedded_exam_editor = bool(self.embedded_exam_dialog is not None)
        for attr_name in [
            "ribbon_exam_editor_add_subject_action",
            "ribbon_exam_editor_edit_subject_action",
            "ribbon_exam_editor_delete_subject_action",
            "ribbon_exam_editor_save_action",
            "ribbon_exam_editor_close_action",
        ]:
            action = getattr(self, attr_name, None)
            if action is not None:
                action.setEnabled(has_embedded_exam_editor)
        if getattr(self, "ribbon_batch_execute_action", None) is not None:
            self.ribbon_batch_execute_action.setEnabled(bool(has_session))
        if getattr(self, "ribbon_batch_save_action", None) is not None:
            save_enabled = bool(hasattr(self, "btn_save_batch_subject") and self.btn_save_batch_subject.isEnabled())
            self.ribbon_batch_save_action.setEnabled(save_enabled)
        if getattr(self, "ribbon_batch_close_action", None) is not None:
            self.ribbon_batch_close_action.setEnabled(True)
        has_export_data = self._has_exportable_data()
        if getattr(self, "ribbon_export_action", None) is not None:
            self.ribbon_export_action.setEnabled(has_session)
        self._refresh_export_action_states(has_session=has_session, has_export_data=has_export_data)

    def _exam_editor_add_subject(self) -> None:
        if self.embedded_exam_dialog is not None:
            self.embedded_exam_dialog._add_subject()
            self._refresh_ribbon_action_states()

    def _exam_editor_edit_subject(self) -> None:
        if self.embedded_exam_dialog is not None:
            self.embedded_exam_dialog._edit_subject()
            self._refresh_ribbon_action_states()

    def _exam_editor_delete_subject(self) -> None:
        if self.embedded_exam_dialog is not None:
            self.embedded_exam_dialog._delete_subject()
            self._refresh_ribbon_action_states()

    def _exam_editor_save(self) -> None:
        if self.embedded_exam_dialog is not None:
            self.embedded_exam_dialog._validate_and_accept()
            self._refresh_ribbon_action_states()

    def _exam_editor_close(self) -> None:
        if self.embedded_exam_dialog is not None:
            self.embedded_exam_dialog.reject()
            self._refresh_ribbon_action_states()

    def _has_session_context_for_export(self) -> bool:
        if bool(str(getattr(self, "current_session_id", "") or "").strip()):
            return True
        if getattr(self, "session", None) is not None:
            sid = str(getattr(self.session, "session_id", "") or "").strip()
            if sid:
                return True
        if hasattr(self, "exam_list_table"):
            row = self.exam_list_table.currentRow()
            if row >= 0 and bool(str(self._session_id_for_row(row) or "").strip()):
                return True
        return False

    def _ensure_current_session_loaded(self) -> bool:
        if self.session is not None:
            return True
        sid = str(getattr(self, "current_session_id", "") or "").strip()
        if not sid:
            return False
        payload = self.database.fetch_exam_session(sid) or {}
        if not payload:
            return False
        try:
            self.session = ExamSession.from_dict(payload)
            self.current_session_path = self._session_path_from_id(sid)
        except Exception:
            return False
        return self.session is not None

    def _refresh_export_action_states(self, *, has_session: bool | None = None, has_export_data: bool | None = None) -> None:
        if has_session is None:
            has_session = self._has_session_context_for_export()
        if has_export_data is None:
            has_export_data = self._has_exportable_data()
        for attr_name in [
            "act_export_subject_scores",
            "act_export_subject_score_matrix",
            "act_export_class_subject_scores",
            "act_export_all_classes_subject_scores",
            "act_export_all_scores",
            "act_export_return_by_class",
            "act_export_subject_api",
            "act_export_reports_center",
            "act_export_range_report",
            "act_export_class_report",
            "act_export_management_report",
        ]:
            action = getattr(self, attr_name, None)
            if action is not None:
                action.setEnabled(bool(has_session))

    def _iter_export_subjects(self) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        for cfg in self._subject_configs_for_scoring():
            key = str(self._subject_key_from_cfg(cfg) or "").strip()
            if not key:
                continue
            label = self._display_subject_label(cfg)
            pairs.append((label, key))
        if not pairs:
            for key in sorted(str(k) for k in (self.scoring_results_by_subject or {}).keys() if str(k).strip()):
                pairs.append((key, key))
        return pairs

    def _has_exportable_data(self) -> bool:
        for _label, key in self._iter_export_subjects():
            if self._load_scoring_results_for_subject_from_storage(key, force_db=True):
                return True
            if self._refresh_scan_results_from_db(key):
                return True
            cfg = self._subject_config_by_subject_key(key) or {}
            if self._subject_uses_direct_score_import(cfg) and self._direct_score_import_rows_for_subject(cfg):
                return True
        return False

    def _pick_subject_for_export(self, title: str, prompt: str) -> str:
        subjects = self._iter_export_subjects()
        if not subjects:
            return ""
        if len(subjects) == 1:
            return str(subjects[0][1])
        labels = [x[0] for x in subjects]
        choice, ok = QInputDialog.getItem(self, title, prompt, labels, 0, False)
        if not ok:
            return ""
        picked = str(choice or "").strip()
        for label, key in subjects:
            if label == picked:
                return key
        return ""

    @staticmethod
    def _safe_sheet_name(sheet_name: str, fallback: str = "Sheet") -> str:
        cleaned = re.sub(r"[\\/*?:\\[\\]]+", "_", str(sheet_name or "").strip())
        cleaned = cleaned[:31].strip()
        return cleaned or fallback

    @staticmethod
    def _parse_score_ranges(range_text: str) -> list[tuple[float, float]]:
        ranges: list[tuple[float, float]] = []
        for chunk in str(range_text or "").split(","):
            text = chunk.strip()
            if not text:
                continue
            if "-" not in text:
                raise ValueError(f"Khoảng không hợp lệ: '{text}'. Định dạng đúng: min-max.")
            left, right = text.split("-", 1)
            start = float(left.strip().replace(",", "."))
            end = float(right.strip().replace(",", "."))
            if end < start:
                start, end = end, start
            ranges.append((start, end))
        if not ranges:
            raise ValueError("Không có khoảng điểm hợp lệ.")
        return ranges

    @staticmethod
    def _format_birth_date_for_export(value: object) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        candidates = [raw]
        if "T" in raw:
            candidates.append(raw.split("T", 1)[0].strip())
        if " " in raw:
            candidates.append(raw.split(" ", 1)[0].strip())
        for text in candidates:
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y"):
                try:
                    return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
                except Exception:
                    continue
        return raw

    def _score_rows_for_subject(self, subject_key: str, *, force_db: bool = False) -> list[dict]:
        payload = self._load_scoring_results_for_subject_from_storage(subject_key, force_db=force_db)
        rows = [dict(v) for v in payload.values() if isinstance(v, dict)]
        rows.sort(key=lambda x: str(x.get("student_id", "") or ""))
        return rows

    def _align_export_rows_with_session_students(self, rows: list[dict], subject_key: str) -> list[dict]:
        subject = str(subject_key or "").strip()
        if not subject:
            return list(rows or [])

        normalized_rows: list[dict] = [dict(item) for item in (rows or []) if isinstance(item, dict)]
        if not self.session or not getattr(self.session, "students", None):
            return normalized_rows

        student_meta = self._student_meta_by_sid()
        row_by_sid: dict[str, dict] = {}
        for row in normalized_rows:
            sid = str(row.get("student_id", "") or "").strip()
            if sid and sid not in row_by_sid:
                row_by_sid[sid] = row

        ordered: list[dict] = []
        seen: set[str] = set()
        for st in (self.session.students or []):
            sid = str(getattr(st, "student_id", "") or "").strip()
            if not sid or sid in seen:
                continue
            seen.add(sid)
            existing = row_by_sid.get(sid, {})
            meta = student_meta.get(sid, {})
            base = {
                "student_id": sid,
                "name": str(getattr(st, "name", "") or meta.get("name", "") or ""),
                "subject": subject,
                "class_name": str(getattr(st, "class_name", "") or meta.get("class_name", "") or ""),
                "birth_date": str(getattr(st, "birth_date", "") or meta.get("birth_date", "") or ""),
                "exam_room": str(getattr(st, "exam_room", "") or meta.get("exam_room", "") or ""),
                "exam_code": "",
                "mcq_correct": "",
                "tf_correct": "",
                "numeric_correct": "",
                "tf_compare": "",
                "numeric_compare": "",
                "correct": "",
                "wrong": "",
                "blank": "",
                "score": "",
                "recheck_score": "",
                "baithiphuctra": "",
                "phase": "",
                "phase_timestamp": "",
                "phase_mode": "",
                "note": "",
                "status": "",
            }
            base.update(existing)
            ordered.append(base)

        for row in normalized_rows:
            sid = str(row.get("student_id", "") or "").strip()
            if not sid or sid in seen:
                continue
            ordered.append(row)
        return ordered

    def _ensure_export_score_rows_for_subject(self, subject_key: str) -> list[dict]:
        subject = str(subject_key or "").strip()
        if not subject:
            return []
        rows = self._score_rows_for_subject(subject, force_db=True)
        if rows:
            return rows
        cfg = self._subject_config_by_subject_key(subject) or {}
        if self._subject_uses_direct_score_import(cfg) and self._direct_score_import_rows_for_subject(cfg):
            self.calculate_scores(subject_key=subject, mode="Nhập điểm trực tiếp", note="auto_prepare_export_essay")
            rows = self._score_rows_for_subject(subject, force_db=True)
        return rows

    @staticmethod
    def _score_value_for_statistics(row: dict) -> float | None:
        if not isinstance(row, dict):
            return None
        status_text = str(row.get("status", "") or "").strip()
        status_fold = status_text.casefold()
        if status_fold.startswith("lỗi"):
            return None
        if status_fold in {"chưa chấm", "cần chấm lại", "chưa có điểm"}:
            return None
        score_raw = row.get("score", "")
        if score_raw in {"", None}:
            return None
        try:
            return float(score_raw)
        except Exception:
            return None

    def _scan_rows_for_subject(self, subject_key: str) -> list[OMRResult]:
        rows = list(self.scan_results_by_subject.get(self._batch_result_subject_key(subject_key), []) or [])
        if rows:
            return rows
        return self._refresh_scan_results_from_db(subject_key) or []

    def _scoring_source_student_ids(self, subject_key: str) -> tuple[set[str], int]:
        subject = str(subject_key or "").strip()
        if not subject:
            return set(), 0
        cfg = self._subject_config_by_subject_key(subject) or {}
        if self._subject_uses_direct_score_import(cfg):
            import_rows = self._direct_score_import_rows_for_subject(cfg)
            student_ids = {
                str((row or {}).get("student_id", "") or "").strip()
                for row in import_rows
                if str((row or {}).get("student_id", "") or "").strip()
            }
            return student_ids, len(import_rows)
        scan_rows = self._scan_rows_for_subject(subject)
        student_ids = {
            str(getattr(scan, "student_id", "") or "").strip()
            for scan in scan_rows
            if str(getattr(scan, "student_id", "") or "").strip()
        }
        return student_ids, len(scan_rows)

    def _subject_uses_direct_score_import(self, subject_or_cfg: str | dict | None) -> bool:
        cfg = subject_or_cfg if isinstance(subject_or_cfg, dict) else self._subject_config_by_subject_key(str(subject_or_cfg or "").strip())
        if not isinstance(cfg, dict):
            return False
        return bool(cfg.get("is_essay_subject"))

    def _direct_score_import_rows_for_subject(self, subject_or_cfg: str | dict | None) -> list[dict]:
        cfg = subject_or_cfg if isinstance(subject_or_cfg, dict) else self._subject_config_by_subject_key(str(subject_or_cfg or "").strip())
        if not isinstance(cfg, dict) or not self._subject_uses_direct_score_import(cfg):
            return []

        payload = cfg.get("direct_score_import", {}) or {}
        raw_rows = payload.get("rows", []) or []
        normalized: list[dict] = []
        for raw in raw_rows:
            if not isinstance(raw, dict):
                continue
            sid = str(raw.get("student_id", "") or raw.get("sid", "") or raw.get("sbd", "") or "").strip()
            score_text = str(raw.get("score", "") or raw.get("point", "") or raw.get("mark", "") or "").strip()
            if not sid or score_text == "":
                continue
            normalized.append({
                "student_id": sid,
                "score": score_text,
                "exam_room": str(raw.get("exam_room", "") or raw.get("room", "") or "").strip(),
                "manually_edited": bool(raw.get("manually_edited", False)),
            })
        normalized.sort(key=lambda row: (str(row.get("student_id", "") or ""), str(row.get("exam_room", "") or "")))
        return normalized

    def _build_direct_score_payloads_for_subject(self, subject_key: str) -> list[dict]:
        subject = str(subject_key or "").strip()
        if not subject:
            return []

        cfg = self._subject_config_by_subject_key(subject) or {}
        if not self._subject_uses_direct_score_import(cfg):
            return []

        default_room = str((cfg or {}).get("exam_room_name", "") or "").strip()
        payloads: list[dict] = []
        for row in self._direct_score_import_rows_for_subject(cfg):
            sid = str(row.get("student_id", "") or "").strip()
            score_text = str(row.get("score", "") or "").strip().replace(",", ".")
            if not sid or score_text == "":
                continue
            try:
                score_value = float(score_text)
            except Exception:
                continue

            profile = self._student_profile_by_id(sid)
            room_text = (
                str(row.get("exam_room", "") or "").strip()
                or self._subject_room_for_student_id(sid, cfg)
                or default_room
                or str(profile.get("exam_room", "") or "").strip()
            )

            is_manual_edit = bool(row.get("manually_edited", False))
            payloads.append({
                "student_id": sid,
                "name": str(profile.get("name", "") or "").strip(),
                "subject": subject,
                "exam_code": "",
                "mcq_correct": 0,
                "tf_correct": 0,
                "numeric_correct": 0,
                "tf_compare": "",
                "numeric_compare": "",
                "correct": 0,
                "wrong": 0,
                "blank": 0,
                "score": score_value,
                "class_name": str(profile.get("class_name", "") or "").strip(),
                "birth_date": str(profile.get("birth_date", "") or "").strip(),
                "exam_room": room_text,
                "status": "OK",
                "note": "Nhập điểm trực tiếp (đã sửa)" if is_manual_edit else "Nhập điểm trực tiếp",
                "manually_edited": is_manual_edit,
            })
        payloads.sort(key=lambda row: str(row.get("student_id", "") or ""))
        return payloads

    def _materialize_direct_score_rows_for_subject(self, subject_key: str, *, persist: bool = False) -> dict[str, dict]:
        subject = str(subject_key or "").strip()
        if not subject:
            return {}

        payloads = self._build_direct_score_payloads_for_subject(subject)
        packed: dict[str, dict] = {}
        for payload in payloads:
            sid_key = str(payload.get("student_id", "") or "").strip()
            if sid_key:
                packed[sid_key] = dict(payload)

        if packed:
            self.scoring_results_by_subject[subject] = dict(packed)
            if persist:
                self._persist_scoring_results_for_subject(
                    subject,
                    list(packed.values()),
                    "Nhập điểm trực tiếp",
                    "essay_direct_import",
                    mark_saved=True,
                )
        return packed

    def _student_meta_by_sid(self) -> dict[str, dict[str, str]]:
        lookup: dict[str, dict[str, str]] = {}
        for st in (self.session.students if self.session else []):
            sid = str(getattr(st, "student_id", "") or "").strip()
            if not sid:
                continue
            extra = getattr(st, "extra", {}) or {}
            lookup[sid] = {
                "name": str(getattr(st, "name", "") or "").strip(),
                "class_name": str((extra or {}).get("class_name", "") or "").strip(),
                "exam_room": str((extra or {}).get("exam_room", "") or "").strip(),
                "birth_date": str((extra or {}).get("birth_date", "") or "").strip(),
            }
        return lookup

    def _route_to_stack_index(self, route_name: str) -> int:
        mapping = {
            "exam_list": 0,
            "workspace_batch_scan": 1,
            "workspace_scoring": 1,
            "subject_management": 2,
            "template_library": 3,
            "template_editor": 4,
            "exam_editor": 5,
        }
        return int(mapping.get(str(route_name or "").strip(), 0))

    def _stack_index_to_route_name(self, index: int) -> str:
        if int(index) == 1 and hasattr(self, "scoring_panel") and self.scoring_panel.isVisible():
            return "workspace_scoring"
        mapping = {0: "exam_list", 1: "workspace_batch_scan", 2: "subject_management", 3: "template_library", 4: "template_editor", 5: "exam_editor"}
        return str(mapping.get(int(index), "exam_list"))

    def _push_route_history(self, name: str, context: dict | None = None) -> None:
        if self._suspend_route_history_push:
            return
        entry = {"name": str(name or "exam_list"), "context": dict(context or {})}
        if self._route_history and self._route_history[-1] == entry:
            return
        self._route_history.append(entry)

    def _navigate_to(
        self,
        route_name: str,
        *,
        context: dict | None = None,
        push_current: bool = True,
        require_confirm: bool = True,
        reason: str = "",
    ) -> bool:
        # route-based navigation + confirm before destructive/switch action.
        target = str(route_name or "exam_list").strip() or "exam_list"
        if require_confirm and target != self._current_route_name:
            if not self._confirm_before_switching_work(reason or target):
                return False
        if push_current:
            self._push_route_history(self._current_route_name, self._current_route_context)
        self._current_route_name = target
        self._current_route_context = dict(context or {})
        self.stack.setCurrentIndex(self._route_to_stack_index(target))
        if target == "workspace_batch_scan":
            self._show_batch_scan_panel()
        elif target == "workspace_scoring":
            self._show_scoring_panel()
        return True

    def _navigate_back(self, default_route: str = "exam_list") -> None:
        if self._route_history:
            target = self._route_history.pop()
        else:
            parent_map = {
                "subject_management": "exam_list",
                "template_library": "exam_list",
                "template_editor": "template_library",
                "exam_editor": "exam_list",
                "workspace_scoring": "workspace_batch_scan",
                "workspace_batch_scan": "exam_editor" if self.current_session_id else "exam_list",
            }
            fallback_name = parent_map.get(self._current_route_name, default_route)
            target = {"name": fallback_name, "context": {"session_id": self.current_session_id} if self.current_session_id else {}}
        self._navigate_to(
            str(target.get("name", default_route) or default_route),
            context=dict(target.get("context", {}) or {}),
            push_current=False,
            require_confirm=False,
            reason="back",
        )

    def action_manage_subjects(self) -> None:
        self.manage_subjects()

    def action_exit(self) -> None:
        if not self._confirm_before_switching_work("thoát ứng dụng"):
            return
        if self._confirm("Thoát", "Bạn có chắc muốn thoát ứng dụng?"):
            self.close()

    def action_load_template(self) -> None:
        if self._confirm("Load Template", "Bạn có chắc muốn tải Template JSON?"):
            self.load_template()

    def action_load_answer_keys(self) -> None:
        if self._confirm("Load Answer Keys", "Bạn có chắc muốn tải Answer Keys JSON?"):
            self.load_answer_keys()

    def action_import_answer_key(self) -> None:
        if self._confirm("Import Answer Key", "Bạn có chắc muốn import Answer Key?"):
            self.import_answer_key_file()

    def action_export_answer_key_sample(self) -> None:
        if self._confirm("Export Sample", "Bạn có chắc muốn export file mẫu?"):
            self.export_answer_key_sample()

    def _start_batch_scan_from_ui(self) -> None:
        if not self._ensure_current_session_loaded():
            QMessageBox.warning(self, "Batch Scan", "Chưa có kỳ thi hiện tại. Vui lòng mở hoặc tạo kỳ thi trước.")
            return
        cfgs = self._effective_subject_configs_for_batch()
        if not cfgs:
            QMessageBox.warning(self, "Batch Scan", "Kỳ thi hiện tại chưa có môn thi để nhận dạng.")
            return
        if self.stack.currentIndex() != 1:
            self._navigate_to(
                "workspace_batch_scan",
                context={"session_id": self.current_session_id, "origin": self._current_route_name},
                push_current=True,
                require_confirm=False,
                reason="open_batch_scan",
            )
        self._refresh_batch_subject_controls()
        self._show_batch_scan_panel()
        if hasattr(self, "batch_subject_combo") and self.batch_subject_combo.currentIndex() <= 0 and self.batch_subject_combo.count() > 1:
            self.batch_subject_combo.setCurrentIndex(1)

    def action_run_batch_scan(self) -> None:
        self._start_batch_scan_from_ui()
        if hasattr(self, "batch_subject_combo") and self.batch_subject_combo.currentIndex() > 0:
            has_unsaved = bool(hasattr(self, "btn_save_batch_subject") and self.btn_save_batch_subject.isEnabled())
            if not has_unsaved:
                self._on_batch_subject_changed(self.batch_subject_combo.currentIndex(), force_reload=True)
        self._refresh_ribbon_action_states()

    def action_execute_batch_scan(self) -> None:
        if self.stack.currentIndex() != 1:
            self._start_batch_scan_from_ui()
            return
        self.run_batch_scan()

    def action_edit_selected_scan(self) -> None:
        if self._confirm("Sửa bài thi", "Bạn có chắc muốn sửa bài thi được chọn?"):
            self._open_edit_selected_scan()

    def action_load_selected_scan_result(self) -> None:
        if self._confirm("Load Selected", "Bạn có chắc muốn load kết quả bài thi đang chọn?"):
            self._load_selected_result_for_correction()

    def action_apply_manual_correction(self) -> None:
        if self._confirm("Apply Correction", "Bạn có chắc muốn áp dụng manual correction?"):
            self.apply_manual_correction()

    def action_calculate_scores(self) -> None:
        # From Batch Scan -> Scoring: do not prompt save when no real batch edits.
        if self.stack.currentIndex() != 1 or bool(hasattr(self, "btn_save_batch_subject") and self.btn_save_batch_subject.isEnabled()):
            if not self._confirm_before_switching_work("màn hình Tính điểm"):
                return
        if self.stack.currentIndex() == 1 and hasattr(self, "batch_subject_combo") and self.batch_subject_combo.currentIndex() > 0:
            has_unsaved = bool(hasattr(self, "btn_save_batch_subject") and self.btn_save_batch_subject.isEnabled())
            if not has_unsaved:
                self._on_batch_subject_changed(self.batch_subject_combo.currentIndex(), force_reload=True)
        self._open_scoring_view()
        self._refresh_ribbon_action_states()

    def action_export_results(self) -> None:
        if self._confirm("Export Results", "Bạn có chắc muốn export kết quả?"):
            self.export_results()

    def action_export_subject_scores(self) -> None:
        subject_key = self._pick_subject_for_export("Xuất điểm môn", "Chọn môn cần xuất điểm:")
        if not subject_key:
            return
        self._export_subject_scores(subject_key)

    def action_export_subject_score_matrix(self) -> None:
        self._export_subject_score_matrix()

    def action_export_class_subject_scores(self) -> None:
        self._export_class_subject_scores()

    def action_export_all_classes_subject_scores(self) -> None:
        self._export_all_classes_subject_scores()

    def action_export_all_subject_scores(self) -> None:
        self._export_all_subject_scores()

    def action_export_return_by_class(self) -> None:
        self._export_return_by_class()

    def action_export_subject_api_payload(self) -> None:
        subject_key = self._pick_subject_for_export("Xuất API bài làm", "Chọn môn cần xuất API bài làm:")
        if not subject_key:
            return
        self._export_subject_api_payload(subject_key)

    def action_export_score_range_report(self) -> None:
        if not self._has_exportable_data():
            QMessageBox.information(self, "Báo cáo khoảng điểm", "Chưa có dữ liệu để xuất.")
            return
        default_ranges = "0-2,2-4,4-6,6-8,8-10"
        range_text, ok = QInputDialog.getText(
            self,
            "Báo cáo khoảng điểm",
            "Nhập các khoảng điểm (ngăn cách bằng dấu phẩy, ví dụ 0-2,2-4,4-6,6-8,8-10):",
            text=default_ranges,
        )
        if not ok:
            return
        try:
            ranges = self._parse_score_ranges(range_text)
        except Exception as exc:
            QMessageBox.warning(self, "Báo cáo khoảng điểm", str(exc))
            return
        self._export_score_range_report(ranges)

    def action_open_export_reports_center(self) -> None:
        dlg = ExportReportsDialog(self)
        dlg.exec()

    def action_export_class_report(self) -> None:
        self._export_class_report()

    def action_export_management_report(self) -> None:
        self._export_management_report()

    def import_answer_key_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Answer Key",
            "",
            "Answer key files (*.xlsx *.csv)",
        )
        if not file_path:
            return
        try:
            imported_package = import_answer_key(file_path)
        except Exception as exc:
            message = (
                f"Cannot import answer key:\n{exc}\n\n"
                "Continue importing only valid rows?\n"
                "- Yes: Continue and AWARD FULL SCORE for invalid-answer questions.\n"
                "- No: Continue but SKIP invalid-answer questions.\n"
                "- Cancel: Stop import."
            )
            choose = QMessageBox.question(
                self,
                "Import failed",
                message,
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.Yes,
            )
            if choose == QMessageBox.Cancel:
                return
            try:
                imported_package = import_answer_key(
                    file_path,
                    strict=False,
                    award_full_credit_for_invalid=(choose == QMessageBox.Yes),
                )
            except Exception as inner_exc:
                QMessageBox.warning(self, "Import failed", f"Cannot import answer key:\n{inner_exc}")
                return

        if imported_package.warnings:
            QMessageBox.information(
                self,
                "Import warnings",
                "Imported with warnings:\n- " + "\n- ".join(imported_package.warnings[:20])
                + ("\n..." if len(imported_package.warnings) > 20 else ""),
            )

        dlg = ImportAnswerKeyDialog(imported_package, self)
        if dlg.exec() != QDialog.Accepted:
            return
        edited_package = dlg.result_answer_key()

        if not self.session:
            self.create_session()
        subject = self.session.subjects[0] if self.session and self.session.subjects else "General"

        if self.answer_keys is None:
            self.answer_keys = AnswerKeyRepository()

        imported_count = 0
        self.imported_exam_codes = sorted(set(edited_package.exam_keys.keys()))
        for exam_code, edited in edited_package.exam_keys.items():
            code = exam_code.strip() or "DEFAULT"
            self.answer_keys.upsert(
                SubjectKey(
                    subject=subject,
                    exam_code=code,
                    answers=edited.mcq_answers,
                    true_false_answers=edited.true_false_answers,
                    numeric_answers=edited.numeric_answers,
                    full_credit_questions=edited.full_credit_questions,
                    invalid_answer_rows=edited.invalid_answer_rows,
                )
            )
            imported_count += 1

        if self.session:
            self.session.answer_key_path = file_path
        self.active_batch_subject_key = subject
        self.session_dirty = True
        self._refresh_session_info()
        self._refresh_batch_subject_controls()
        self._refresh_batch_subject_controls()
        self._retrim_batch_results_to_answer_key_scope()
        QMessageBox.information(self, "Import successful", f"Imported {imported_count} exam code(s) into current session.")

    def export_answer_key_sample(self) -> None:
        save_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Sample Answer Key",
            "answer_key_sample.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)",
        )
        if not save_path:
            return

        import pandas as pd

        data = {
            "Question": list(range(1, 19)) + [1, 2, 3, 4] + [1, 2, 3, 4, 5, 6],
            "0101": ["C", "C", "C", "B", "D", "D", "D", "B", "B", "B", "B", "B", "D", "C", "D", "B", "C", "C", "ĐSĐĐ", "ĐĐĐS", "ĐSĐS", "ĐDDS", "5", "69", "0,61", "58,3", "49,6", "2"],
            "0102": ["B", "C", "C", "C", "D", "D", "B", "D", "B", "B", "B", "C", "C", "B", "D", "C", "D", "B", "ĐĐSĐ", "ĐĐĐS", "ĐĐSĐ", "ĐSDS", "2", "69", "58,3", "5", "49,6", "0,61"],
            "0103": ["C", "C", "C", "B", "C", "B", "D", "D", "C", "B", "B", "C", "B", "D", "C", "D", "B", "B", "SĐĐĐ", "ĐĐĐS", "ĐĐSĐ", "SĐDS", "2", "0,61", "58,3", "49,6", "5", "69"],
            "0104": ["C", "C", "C", "B", "C", "B", "D", "D", "C", "B", "B", "C", "B", "D", "C", "D", "B", "B", "ĐĐĐS", "SĐSĐ", "ĐĐSĐ", "SĐDD", "5", "2", "49,6", "0,61", "58,3", "69"],
        }
        df = pd.DataFrame(data)
        path = Path(save_path)
        if path.suffix.lower() == ".csv" or "CSV" in selected_filter:
            if path.suffix.lower() != ".csv":
                path = path.with_suffix(".csv")
            df.to_csv(path, index=False)
        else:
            if path.suffix.lower() != ".xlsx":
                path = path.with_suffix(".xlsx")
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, startrow=1)
                ws = writer.sheets[next(iter(writer.sheets))]
                ws["A1"] = "Câu hỏi"
                ws["B1"] = "Mã đề thi"
                ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
                ws["B2"] = "0101"
                ws["C2"] = "0102"
                ws["D2"] = "0103"
                ws["E2"] = "0104"
        QMessageBox.information(self, "Sample exported", f"Saved sample answer key file to:\n{path}")

    def _build_session_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)

        self.session_info = QTextEdit()
        self.session_info.setReadOnly(True)
        self.exam_code_preview = QLabel("Mã đề trên phiếu trả lời mẫu: -")
        self.exam_code_preview.setWordWrap(True)

        layout.addWidget(self.session_info)
        layout.addWidget(self.exam_code_preview)
        return w

    def _build_scan_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)

        batch_group = QGroupBox("Nhận dạng theo môn đã cấu hình")
        batch_form = QFormLayout(batch_group)
        self.batch_subject_combo = QComboBox()
        self.batch_subject_combo.currentIndexChanged.connect(self._on_batch_subject_changed)
        self.batch_file_scope_combo = QComboBox()
        self.batch_file_scope_combo.addItem("Nhận dạng file mới", "new_only")
        self.batch_file_scope_combo.addItem("Nhận dạng toàn bộ", "all")
        self.batch_file_scope_combo.currentIndexChanged.connect(lambda _=0: self._update_batch_scan_scope_summary())
        self.batch_recognition_mode_combo = QComboBox()
        self.batch_recognition_mode_combo.addItem("Tự động (khuyến nghị)", "auto")
        self.batch_recognition_mode_combo.addItem("Mẫu cũ / Anchor chuẩn", "legacy")
        self.batch_recognition_mode_combo.addItem("Mẫu mới / Anchor sát biên", "border")
        self.batch_recognition_mode_combo.addItem("Anchor 1 phía (ruler theo dòng)", "one_side")
        self.batch_recognition_mode_combo.addItem("Lai (thử nhiều cơ chế)", "hybrid")
        self.batch_recognition_mode_combo.setCurrentIndex(0)
        self.batch_recognition_mode_combo.setVisible(False)
        self.batch_api_file_value = QLineEdit("-"); self.batch_api_file_value.setReadOnly(True)
        self.btn_pick_batch_api_file = QPushButton("Chọn file API")
        self.btn_pick_batch_api_file.clicked.connect(self._pick_batch_api_file)
        api_row = QHBoxLayout()
        api_row.addWidget(self.batch_api_file_value, 1)
        api_row.addWidget(self.btn_pick_batch_api_file)
        self.batch_template_value = QLineEdit("-"); self.batch_template_value.setReadOnly(True)
        self.batch_template_path_value = "-"
        self.batch_answer_codes_value = QLineEdit("-"); self.batch_answer_codes_value.setReadOnly(True)
        self.batch_student_id_value = QLineEdit("-"); self.batch_student_id_value.setReadOnly(True)
        self.batch_scan_folder_value = QLineEdit("-"); self.batch_scan_folder_value.setReadOnly(True)
        self.batch_scan_state_value = QLineEdit("-"); self.batch_scan_state_value.setReadOnly(True)
        self.batch_context_value = QLineEdit("-"); self.batch_context_value.setReadOnly(True)
        style = self.style()
        self.btn_batch_recognize = QPushButton("Nhận dạng")
        self.btn_batch_recognize.setIcon(style.standardIcon(QStyle.SP_MediaPlay))
        self.btn_batch_recognize.clicked.connect(self.action_execute_batch_scan)
        self.btn_save_batch_subject = QPushButton("Lưu")
        self.btn_save_batch_subject.setIcon(style.standardIcon(QStyle.SP_DialogSaveButton))
        self.btn_save_batch_subject.clicked.connect(self._save_batch_for_selected_subject)
        self.btn_save_batch_subject.setEnabled(False)
        self.btn_close_batch_view = QPushButton("Đóng")
        self.btn_close_batch_view.setIcon(style.standardIcon(QStyle.SP_DialogCloseButton))
        self.btn_close_batch_view.clicked.connect(self._close_batch_scan_view)
        for b in [self.btn_batch_recognize, self.btn_save_batch_subject, self.btn_close_batch_view]:
            b.setMaximumWidth(140)
            b.setVisible(False)

        batch_form.addRow("Môn", self.batch_subject_combo)
        batch_form.addRow("Phạm vi", self.batch_file_scope_combo)
        batch_form.addRow("API bài thi", api_row)
        batch_form.addRow("Mẫu giấy dùng", self.batch_template_value)
        batch_form.addRow("Mã đề", self.batch_answer_codes_value)
        batch_form.addRow("Thư mục quét", self.batch_scan_folder_value)

        self.filter_column = QComboBox()
        self.filter_column.addItems(["Tất cả", "STT", "STUDENT ID", "Phòng thi", "Mã đề", "Họ tên", "Ngày sinh", "Nội dung", "Status"])
        self.filter_column.currentTextChanged.connect(self._apply_scan_filter)
        self.search_value = QLineEdit()
        self.search_value.setPlaceholderText("Tìm trong cột đã chọn hoặc toàn bảng")
        self.search_value.textChanged.connect(self._apply_scan_filter)

        search_row = QHBoxLayout()
        search_row.addWidget(self.filter_column)
        search_row.addWidget(self.search_value)

        self.scan_list = QTableWidget(0, 9)
        self.scan_list.setHorizontalHeaderLabels(["STT", "STUDENT ID", "Phòng thi", "Mã đề", "Họ tên", "Ngày sinh", "Nội dung", "Status", "Chức năng"])
        self.scan_list.verticalHeader().setVisible(False)
        self.scan_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.scan_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.scan_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.scan_list.customContextMenuRequested.connect(self._open_scan_row_context_menu)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_STT, QHeaderView.ResizeToContents)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_STUDENT_ID, QHeaderView.ResizeToContents)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_EXAM_ROOM, QHeaderView.ResizeToContents)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_EXAM_CODE, QHeaderView.ResizeToContents)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_FULL_NAME, QHeaderView.ResizeToContents)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_BIRTH_DATE, QHeaderView.ResizeToContents)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_CONTENT, QHeaderView.Stretch)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_STATUS, QHeaderView.Stretch)
        self.scan_list.horizontalHeader().setSectionResizeMode(self.SCAN_COL_ACTIONS, QHeaderView.ResizeToContents)
        self.scan_list.horizontalHeader().sectionClicked.connect(self._on_scan_header_clicked)
        self.scan_list.itemSelectionChanged.connect(self._on_scan_selected)
        # double click navigation: open selected scan directly.
        self.scan_list.cellDoubleClicked.connect(self._handle_scan_list_double_click)
        self.scan_list.cellClicked.connect(self._on_scan_cell_clicked)
        self.progress = QProgressBar()
        self.progress.setVisible(False)

        self.scan_image_preview = PreviewImageWidget(); self.scan_image_preview.setText("Chọn bài thi ở danh sách bên trái")
        self.scan_image_scroll = QScrollArea()
        self.scan_image_scroll.setWidgetResizable(False)
        self.scan_image_scroll.setAlignment(Qt.AlignCenter)
        self.scan_image_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.scan_image_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.scan_image_scroll.setWidget(self.scan_image_preview)
        self.scan_image_scroll.viewport().installEventFilter(self)

        self.btn_zoom_out = QPushButton("-")
        self.btn_zoom_out.setMaximumWidth(36)
        self.btn_zoom_out.clicked.connect(self._zoom_preview_out)
        self.btn_zoom_reset = QPushButton("30%")
        self.btn_zoom_reset.setMaximumWidth(60)
        self.btn_zoom_reset.clicked.connect(self._zoom_preview_reset)
        self.btn_zoom_in = QPushButton("+")
        self.btn_zoom_in.setMaximumWidth(36)
        self.btn_zoom_in.clicked.connect(self._zoom_preview_in)
        zoom_row = QHBoxLayout()
        zoom_row.addWidget(self.btn_zoom_out)
        zoom_row.addWidget(self.btn_zoom_reset)
        zoom_row.addWidget(self.btn_zoom_in)
        self.btn_rotate_left = QPushButton("⟲ 90°")
        self.btn_rotate_left.setToolTip("Xoay trái ảnh đang chọn")
        self.btn_rotate_left.clicked.connect(lambda: self._rotate_selected_scan(-90))
        self.btn_rotate_right = QPushButton("⟳ 90°")
        self.btn_rotate_right.setToolTip("Xoay phải ảnh đang chọn")
        self.btn_rotate_right.clicked.connect(lambda: self._rotate_selected_scan(90))
        self.btn_rerecognize_selected = QPushButton("Nhận dạng lại ảnh chọn")
        self.btn_rerecognize_selected.clicked.connect(self._rerecognize_selected_scan)
        zoom_row.addWidget(self.btn_rotate_left)
        zoom_row.addWidget(self.btn_rotate_right)
        zoom_row.addWidget(self.btn_rerecognize_selected)
        zoom_row.addStretch()

        self.scan_result_preview = QTableWidget(0, 2)
        self.scan_result_preview.setHorizontalHeaderLabels(["Mục nhận dạng", "Kết quả"])
        self.scan_result_preview.verticalHeader().setVisible(False)
        self.scan_result_preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.scan_result_preview.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.scan_result_preview.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        left = QWidget()
        left_l = QVBoxLayout(left)
        left_l.addWidget(batch_group)
        left_l.addLayout(search_row)
        left_l.addWidget(self.scan_list)

        right = QWidget()
        right_l = QVBoxLayout(right)
        right_l.setContentsMargins(0, 0, 0, 0)
        right_l.addLayout(zoom_row)
        right_l.addWidget(self.scan_image_scroll, 7)
        right_l.addWidget(self.scan_result_preview, 3)

        self.scan_lr_split = QSplitter(Qt.Horizontal)
        self.scan_lr_split.addWidget(left)
        self.scan_lr_split.addWidget(right)
        self.scan_lr_split.setStretchFactor(0, 68)
        self.scan_lr_split.setStretchFactor(1, 32)
        self.scan_lr_split.setSizes([680, 320])
        self.batch_scan_status_bottom = QLabel("Trạng thái file: - | Lọc: 0/0")
        self.batch_scan_status_bottom.setWordWrap(False)
        self.batch_scan_status_bottom.setFixedHeight(22)
        self.batch_scan_status_bottom.setStyleSheet("QLabel { padding: 2px 8px; color: #444; }")
        self.batch_status_filter_mode = "all"
        self.batch_scan_status_bottom.setOpenExternalLinks(False)
        self.batch_scan_status_bottom.linkActivated.connect(self._handle_batch_status_filter_link)

        # Create scoring widgets with explicit parent to avoid lifecycle issues
        # on some PySide6 builds (preventing "Internal C++ object ... already deleted").
        self.scoring_panel = QWidget(w)

        self.score_preview_table = QTableWidget(0, 13, self.scoring_panel)
        self.score_preview_table.setHorizontalHeaderLabels([
            "Student ID", "Name", "Lớp", "Ngày sinh", "Exam Code", "MCQ đúng", "TF đúng", "NUM đúng", "Correct", "Wrong", "Blank", "Score", "Trạng thái"
        ])
        self.score_preview_table.verticalHeader().setVisible(False)
        self.score_preview_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.score_preview_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.score_preview_table.horizontalHeader().setSectionResizeMode(12, QHeaderView.Stretch)
        self.score_preview_table.cellDoubleClicked.connect(self._open_scoring_review_editor_from_table)

        self.scoring_subject_combo = QComboBox()
        self.scoring_subject_combo.currentIndexChanged.connect(self._handle_scoring_subject_changed)
        self.scoring_mode_combo = QComboBox()
        self.scoring_mode_combo.addItems(["Tính lại toàn bộ", "Chỉ tính bài chưa có điểm"])
        self.scoring_phase_note = QLineEdit()
        self.scoring_phase_note.setPlaceholderText("Ghi chú pha chấm điểm (tuỳ chọn)")
        self.scoring_phase_note.setVisible(False)
        self.scoring_filter_column = QComboBox()
        self.scoring_filter_column.addItems(["Tất cả", "Student ID", "Name", "Lớp", "Ngày sinh", "Exam Code", "Trạng thái"])
        self.scoring_filter_column.currentTextChanged.connect(self._apply_scoring_filter)
        self.scoring_search_value = QLineEdit()
        self.scoring_search_value.setPlaceholderText("Tìm kiếm trên lưới Tính điểm")
        self.scoring_search_value.textChanged.connect(self._apply_scoring_filter)
        self.btn_scoring_run = QPushButton("Chấm điểm")
        self.btn_scoring_run.clicked.connect(self._run_scoring_from_panel)
        self.btn_scoring_save = QPushButton("Lưu điểm")
        self.btn_scoring_save.clicked.connect(self._save_current_work)
        self.btn_scoring_save.setEnabled(False)
        self.btn_scoring_back = QPushButton("Quay lại Batch Scan")
        self.btn_scoring_back.clicked.connect(self._back_to_batch_scan)
        scoring_top = QHBoxLayout()
        scoring_top.addWidget(QLabel("Môn"))
        scoring_top.addWidget(self.scoring_subject_combo, 2)
        scoring_top.addWidget(QLabel("Cơ chế"))
        scoring_top.addWidget(self.scoring_mode_combo, 2)
        scoring_top.addWidget(self.scoring_filter_column, 2)
        scoring_top.addWidget(self.scoring_search_value, 3)
        scoring_top.addWidget(self.btn_scoring_run)
        scoring_top.addWidget(self.btn_scoring_save)
        scoring_top.addWidget(self.btn_scoring_back)

        self.scoring_phase_table = QTableWidget(0, 5, self.scoring_panel)
        self.scoring_phase_table.setHorizontalHeaderLabels(["Thời gian", "Môn", "Cơ chế", "Số bài", "Ghi chú"])
        self.scoring_phase_table.verticalHeader().setVisible(False)
        self.scoring_phase_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.scoring_phase_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.scoring_phase_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.scoring_phase_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.scoring_phase_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.scoring_phase_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)

        scoring_panel_layout = QVBoxLayout(self.scoring_panel)
        scoring_panel_layout.setContentsMargins(0, 0, 0, 0)
        scoring_panel_layout.addLayout(scoring_top)
        scoring_panel_layout.addWidget(self.score_preview_table, 7)
        self.scoring_status_bar = QLabel("Thống kê chấm điểm: Thành công 0 | Lỗi 0 | Đã sửa 0")
        self.scoring_status_bar.setStyleSheet("QLabel { padding: 4px 8px; background: #f4f6f8; border: 1px solid #d0d7de; }")
        self.scoring_status_filter_mode = "all"
        self.scoring_status_bar.setTextFormat(Qt.RichText)
        self.scoring_status_bar.setOpenExternalLinks(False)
        self.scoring_status_bar.linkActivated.connect(self._handle_scoring_status_filter_link)
        scoring_panel_layout.addWidget(self.scoring_status_bar)

        layout.addWidget(self.progress)
        layout.addWidget(self.scan_lr_split)
        layout.addWidget(self.batch_scan_status_bottom)
        layout.addWidget(self.scoring_panel)
        self.scoring_panel.setVisible(False)
        return w

    def _close_batch_scan_view(self) -> None:
        if self._has_batch_unsaved_changes():
            choice = self._prompt_save_changes_word_style(
                "Batch Scan chưa lưu",
                "Bạn có thay đổi chưa lưu cho môn hiện tại. Bạn muốn lưu trước khi đóng không?",
            )
            if choice == "cancel":
                return
            if choice == "save" and not self._save_batch_for_selected_subject(
                show_success_message=False,
                reload_after_save=False,
                refresh_exam_list=False,
            ):
                return
        # On close, always route back to the current exam subject list view.
        self._return_to_current_exam_from_batch_scan()
    def _clear_batch_display_caches(self) -> None:
        for result in list(getattr(self, "scan_results", []) or []):
            for attr in ["cached_status", "cached_content", "cached_recognized_short", "cached_blank_summary", "cached_forced_status"]:
                try:
                    setattr(result, attr, "" if attr != "cached_blank_summary" else {})
                except Exception:
                    pass

    def _has_batch_unsaved_changes(self) -> bool:
        return bool(hasattr(self, "btn_save_batch_subject") and self.btn_save_batch_subject.isEnabled())

    def _has_scoring_unsaved_changes(self) -> bool:
        if hasattr(self, "btn_scoring_save") and self.btn_scoring_save.isEnabled():
            return True

        subject_key = ""
        try:
            subject_key = str(self._resolve_preferred_scoring_subject() or getattr(self, "_current_scoring_subject", "") or "").strip()
        except Exception:
            subject_key = str(getattr(self, "_current_scoring_subject", "") or "").strip()

        dirty_subjects = {
            str(s or "").strip()
            for s in (getattr(self, "_scoring_dirty_subjects", set()) or set())
            if str(s or "").strip()
        }
        if subject_key and subject_key in dirty_subjects:
            return True

        if subject_key:
            state = self._subject_scoring_state(subject_key) if hasattr(self, "_subject_scoring_state") else {}
            saved_flag = bool((state or {}).get("scoring_saved", False)) and subject_key not in dirty_subjects
            result_count = int((state or {}).get("scoring_result_count", 0) or 0)
            if result_count > 0 and not saved_flag:
                return True

        return bool(dirty_subjects)

    def _open_current_session_in_exam_editor(self) -> bool:
        # return to current exam on batch close: open current runtime session, avoid stale batch return payload.
        if not self.current_session_id:
            return False
        if self.embedded_exam_dialog and str(self.embedded_exam_session_id or "").strip() == str(self.current_session_id or "").strip():
            self._navigate_to(
                "exam_editor",
                context={"session_id": self.current_session_id},
                push_current=False,
                require_confirm=False,
                reason="return_to_existing_exam_editor",
            )
            return True
        session = self.session if self.session is not None else None
        if session is None:
            payload = self.database.fetch_exam_session(self.current_session_id) or {}
            if not payload:
                return False
            try:
                session = ExamSession.from_dict(payload)
            except Exception:
                return False
        cfg = session.config or {}
        editor_payload = {
            "exam_name": session.exam_name,
            "common_template": session.template_path,
            "scan_root": cfg.get("scan_root", ""),
            "student_list_path": cfg.get("student_list_path", ""),
            "students": [
                {
                    "student_id": s.student_id,
                    "name": s.name,
                    "birth_date": str((s.extra or {}).get("birth_date", "") or ""),
                    "class_name": str((s.extra or {}).get("class_name", "") or ""),
                    "exam_room": str((s.extra or {}).get("exam_room", "") or ""),
                }
                for s in (session.students or [])
            ],
            "scan_mode": cfg.get("scan_mode", "Ảnh trong thư mục gốc"),
            "paper_part_count": cfg.get("paper_part_count", 3),
            "subject_configs": cfg.get("subject_configs", []),
        }
        self._open_embedded_exam_editor(self.current_session_id, session, editor_payload)
        return True

    def _return_to_current_exam_from_batch_scan(self) -> None:
        ctx = dict(self._current_route_context or {})
        origin = str(ctx.get("origin", "") or "")
        target = "exam_list"
        if self.embedded_exam_dialog and self.current_session_id:
            target = "exam_editor"
        elif self.current_session_id:
            target = "exam_editor"
        elif origin == "workspace_scoring":
            # keep deterministic fallback when context was opened from scoring but exam context is gone.
            target = "exam_list"
        if target == "exam_editor":
            if self.embedded_exam_dialog and str(self.embedded_exam_session_id or "").strip() == str(self.current_session_id or "").strip():
                self._navigate_to(
                    "exam_editor",
                    context={"session_id": self.current_session_id},
                    push_current=False,
                    require_confirm=False,
                    reason="return_existing_exam_editor",
                )
                return
            if self._open_current_session_in_exam_editor():
                return
        self._navigate_back(default_route="exam_list")

    def _strip_transient_scan_artifacts(self, result: OMRResult) -> OMRResult:
        for attr_name in [
            "aligned_image",
            "aligned_binary",
            "alignment_debug",
            "detected_anchors",
            "detected_digit_anchors",
            "bubble_states_by_zone",
            "digit_zone_debug",
        ]:
            if hasattr(result, attr_name):
                try:
                    delattr(result, attr_name)
                except Exception:
                    setattr(result, attr_name, None)
        return result

    def _lightweight_result_copy(self, result: OMRResult) -> OMRResult:
        return self._deserialize_omr_result(self._serialize_omr_result(result))

    def _collect_current_subject_results_for_save(self, subject_key: str) -> list[OMRResult]:
        # scan_list columns: 0 stt, 1 sid, 2 room, 3 exam_code, 4 full_name, 5 birth_date, 6 content, 7 status, 8 actions
        key = str(subject_key or "").strip()
        base_results = list(self.scan_results_by_subject.get(self._batch_result_subject_key(key)) or self.scan_results or [])
        row_count = self.scan_list.rowCount() if hasattr(self, "scan_list") else 0
        if row_count <= 0:
            return base_results

        by_image: dict[str, OMRResult] = {}
        for res in base_results:
            img = str(getattr(res, "image_path", "") or "").strip()
            if img and img not in by_image:
                by_image[img] = res

        out: list[OMRResult] = []
        for r in range(row_count):
            sid_item = self.scan_list.item(r, self.SCAN_COL_STUDENT_ID)
            image_path = str(sid_item.data(Qt.UserRole) if sid_item else "")
            exam_code = str(sid_item.data(Qt.UserRole + 1) if sid_item else "")
            sid_text = str(sid_item.text() if sid_item else "-")
            base = by_image.get(image_path) if image_path else None
            if base is not None:
                res = self._lightweight_result_copy(base)
            else:
                res = OMRResult(image_path=image_path)
            res.student_id = "" if sid_text == "-" else sid_text
            res.exam_code = exam_code
            if hasattr(self, "scan_list"):
                res.full_name = str(self.scan_list.item(r, self.SCAN_COL_FULL_NAME).text() if self.scan_list.item(r, self.SCAN_COL_FULL_NAME) else "")
                res.birth_date = str(self.scan_list.item(r, self.SCAN_COL_BIRTH_DATE).text() if self.scan_list.item(r, self.SCAN_COL_BIRTH_DATE) else "")
            res.sync_legacy_aliases()
            out.append(res)

        return out


    def _refresh_scan_results_from_db(self, subject_key: str) -> list[OMRResult]:
        subject = str(subject_key or "").strip()
        if not subject:
            return []
        scoped_subject = self._batch_result_subject_key(subject)
        has_session_scope = bool(str(self.current_session_id or "").strip())
        rows = self.database.fetch_scan_results_for_subject(scoped_subject)
        if not rows:
            sid = str(self.current_session_id or "").strip()
            legacy_scoped = f"{sid}::{subject}" if sid else ""
            if legacy_scoped and legacy_scoped != scoped_subject:
                rows = self.database.fetch_scan_results_for_subject(legacy_scoped)
        # Never fallback to unscoped key when current session exists to avoid cross-exam result bleed.
        if not rows and (not has_session_scope) and scoped_subject != subject:
            # Legacy fallback (older sessions saved by raw subject key).
            rows = self.database.fetch_scan_results_for_subject(subject)
        refreshed: list[OMRResult] = []
        for item in rows:
            try:
                refreshed.append(self._deserialize_omr_result(item))
            except Exception:
                continue
        self.scan_results_by_subject[scoped_subject] = list(refreshed)
        return refreshed

    def _refresh_dashboard_summary_from_db(self, subject_key: str) -> None:
        if not hasattr(self, "dashboard_summary_label"):
            return
        summary = self.database.dashboard_summary(subject_key)
        avg_score = float(summary.get("average_score", 0.0) or 0.0)
        distribution = summary.get("distribution", []) if isinstance(summary.get("distribution", []), list) else []
        top_students = summary.get("top_students", []) if isinstance(summary.get("top_students", []), list) else []
        dist_text = ", ".join(f"{item.get('bucket', 0)}: {item.get('count', 0)}" for item in distribution[:8]) or "-"
        top_text = ", ".join(f"{item.get('student_code', '-')}: {item.get('score', 0)}" for item in top_students[:5]) or "-"
        self.dashboard_summary_label.setText(
            f"Dashboard DB | Điểm TB: {avg_score:.2f} | Phổ điểm: {dist_text} | Top học sinh: {top_text}"
        )

    def _subject_scoring_state(self, subject_key: str) -> dict:
        cfg = self._subject_config_by_subject_key(subject_key) or {}
        return {
            "scoring_saved": bool(cfg.get("scoring_saved", False)),
            "scoring_saved_at": str(cfg.get("scoring_saved_at", "") or ""),
            "scoring_result_count": int(cfg.get("scoring_result_count", 0) or 0),
            "scoring_phase_last_note": str(cfg.get("scoring_phase_last_note", "") or ""),
            "scoring_last_mode": str(cfg.get("scoring_last_mode", "") or ""),
        }

    def _refresh_scoring_state_label(self, subject_key: str) -> None:
        if not hasattr(self, "scoring_state_label"):
            return
        state = self._subject_scoring_state(subject_key)
        saved_flag = bool(state.get("scoring_saved", False)) and subject_key not in self._scoring_dirty_subjects
        status_text = "Đã lưu" if saved_flag else "Chưa lưu"
        saved_at = str(state.get("scoring_saved_at", "") or "-")
        count = int(state.get("scoring_result_count", 0) or 0)
        mode = str(state.get("scoring_last_mode", "") or "-")
        note = str(state.get("scoring_phase_last_note", "") or "-")
        self.scoring_state_label.setText(
            f"Trạng thái: {status_text} | Lần lưu cuối: {saved_at} | Số bài đã chấm: {count} | Cơ chế: {mode} | Ghi chú: {note}"
        )
        if hasattr(self, "btn_scoring_save"):
            self.btn_scoring_save.setEnabled((subject_key in self._scoring_dirty_subjects) or (not saved_flag and count > 0))

    def _set_subject_scoring_metadata(self, subject_key: str, updates: dict) -> None:
        cfg = self._subject_config_by_subject_key(subject_key)
        if not isinstance(cfg, dict):
            return
        cfg.update(dict(updates or {}))
        if self.session and isinstance(self.session.config, dict):
            self.session_dirty = True

    def _mark_subject_scoring_saved(self, subject_key: str, count: int, mode: str, note: str) -> None:
        # persisted scoring state
        now_text = datetime.now().isoformat(timespec="seconds")
        self._set_subject_scoring_metadata(subject_key, {
            "scoring_saved": True,
            "scoring_saved_at": now_text,
            "scoring_result_count": int(count or 0),
            "scoring_phase_last_note": str(note or ""),
            "scoring_last_mode": str(mode or ""),
        })
        self._scoring_dirty_subjects.discard(subject_key)
        self._refresh_scoring_state_label(subject_key)

    def _clear_subject_scoring_saved_state(self, subject_key: str, count: int = 0, mode: str = "", note: str = "") -> None:
        self._set_subject_scoring_metadata(subject_key, {
            "scoring_saved": False,
            "scoring_saved_at": "",
            "scoring_result_count": int(count or 0),
            "scoring_phase_last_note": str(note or ""),
            "scoring_last_mode": str(mode or ""),
        })
        self._scoring_dirty_subjects.add(subject_key)
        self._refresh_scoring_state_label(subject_key)

    def _persist_scoring_results_for_subject(self, subject_key: str, rows: list[dict], mode: str, note: str, *, mark_saved: bool) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        packed: dict[str, dict] = {}
        for row in (rows or []):
            sid_key = str((row or {}).get("student_id", "") or "").strip()
            if sid_key:
                packed[sid_key] = dict(row or {})
        self.scoring_results_by_subject[subject] = packed
        try:
            self.database.conn.execute("DELETE FROM scores WHERE subject_key = ?", (subject,))
            for payload in packed.values():
                self.database.upsert_score_row(
                    subject,
                    str(payload.get("student_id", "") or ""),
                    str(payload.get("exam_code", "") or ""),
                    dict(payload),
                )
        except Exception:
            pass
        if mark_saved:
            self._mark_subject_scoring_saved(subject, len(packed), mode, note)

    def _persist_scoring_results(self, subject_key: str, phase: dict | None = None) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        payloads = list((self.scoring_results_by_subject.get(subject, {}) or {}).values())
        phase = dict(phase or {})
        mode = str(phase.get("mode", "Tính lại toàn bộ") or "Tính lại toàn bộ")
        note = str(phase.get("note", "") or "")
        self._persist_scoring_results_for_subject(subject, payloads, mode, note, mark_saved=True)

    def _load_scoring_results_for_subject_from_storage(self, subject_key: str, *, force_db: bool = False) -> dict[str, dict]:
        # DB là nguồn chuẩn cho scoring; chỉ dùng runtime cache đã có nếu cùng subject trong phiên hiện tại.
        subject = str(subject_key or "").strip()
        if not subject:
            return {}
        cached = self.scoring_results_by_subject.get(subject)
        if (not force_db) and isinstance(cached, dict) and cached:
            return dict(cached)
        db_rows = self.database.fetch_scores_for_subject(subject) or []
        loaded: dict[str, dict] = {}
        for payload in db_rows:
            if not isinstance(payload, dict):
                continue
            sid_key = str(payload.get("student_id", "") or "").strip()
            if sid_key:
                loaded[sid_key] = dict(payload)
        if not loaded and self._subject_uses_direct_score_import(subject):
            loaded = self._materialize_direct_score_rows_for_subject(subject, persist=True)
        if loaded:
            self.scoring_results_by_subject[subject] = dict(loaded)
        return loaded
    def _load_cached_scoring_results_for_subject(self, subject_key: str) -> None:
        subject = str(subject_key or "").strip()
        if not subject or not hasattr(self, "score_preview_table"):
            return
        stored_rows = self._load_scoring_results_for_subject_from_storage(subject)
        loaded_rows = []
        for _, payload in (stored_rows or {}).items():
            if isinstance(payload, dict):
                loaded_rows.append(dict(payload))
        source_student_ids, source_row_count = self._scoring_source_student_ids(subject)
        if source_row_count > 0:
            filtered_rows: list[dict] = []
            for row in loaded_rows:
                sid = str((row or {}).get("student_id", "") or "").strip()
                status_text = str((row or {}).get("status", "") or (row or {}).get("note", "") or "").strip()
                if sid and sid in source_student_ids:
                    filtered_rows.append(row)
                    continue
                if not sid and status_text.startswith("Lỗi"):
                    filtered_rows.append(row)
            loaded_rows = filtered_rows
        loaded_rows.sort(key=lambda row: (
            0 if str((row or {}).get("status", "") or "").startswith("Lỗi") else 1,
            str((row or {}).get("student_id", "") or ""),
        ))

        success_count = 0
        error_count = 0
        edited_count = 0
        pending_count = 0
        self.score_preview_table.setSortingEnabled(False)
        self.score_preview_table.setUpdatesEnabled(False)
        try:
            self.score_preview_table.setRowCount(0)
            for row in loaded_rows:
                row = self._hydrate_scoring_row_with_student_profile(subject, row)
                r = self.score_preview_table.rowCount()
                self.score_preview_table.insertRow(r)
                status_text = str((row or {}).get("status", "") or (row or {}).get("note", "") or "OK")
                score_raw = (row or {}).get("score", "")
                score_display = str(score_raw) if score_raw not in {"", None} else ""
                if status_text == "OK" and score_display == "":
                    status_text = "Chưa có điểm"
                if status_text.startswith("Lỗi"):
                    error_count += 1
                else:
                    success_count += 1
                if status_text == "Đã sửa":
                    edited_count += 1
                if status_text in {"Chưa chấm", "Cần chấm lại", "Chưa có điểm"}:
                    pending_count += 1
                values = [
                    str((row or {}).get("student_id", "") or "-"),
                    str((row or {}).get("name", "") or "-"),
                    str((row or {}).get("class_name", "") or "-"),
                    str((row or {}).get("birth_date", "") or "-"),
                    str((row or {}).get("exam_code", "") or "-"),
                    str((row or {}).get("mcq_correct", 0)),
                    str(self._tf_statement_correct_count(str((row or {}).get("tf_compare", "") or ""))),
                    str((row or {}).get("numeric_correct", 0)),
                    str((row or {}).get("correct", 0)),
                    str((row or {}).get("wrong", 0)),
                    str((row or {}).get("blank", 0)),
                    str((row or {}).get("recheck_score", "") if (row or {}).get("recheck_score", "") not in {"", None} else score_display),
                    status_text,
                ]
                for col, value in enumerate(values):
                    item = QTableWidgetItem(value)
                    if col == 12 and status_text != "OK":
                        item.setForeground(QColor("red"))
                        category = "success"
                        if status_text.startswith("Lỗi"):
                            category = "error"
                        elif status_text == "Đã sửa":
                            category = "edited"
                        elif status_text in {"Chưa chấm", "Cần chấm lại", "Chưa có điểm"}:
                            category = "pending"
                        item.setData(Qt.UserRole, category)
                    self.score_preview_table.setItem(r, col, item)
                if status_text.startswith("Lỗi"):
                    for col in range(self.score_preview_table.columnCount()):
                        item = self.score_preview_table.item(r, col)
                        if item:
                            item.setBackground(QColor(255, 225, 225))
        finally:
            self.score_preview_table.setUpdatesEnabled(True)
            self.score_preview_table.viewport().update()
        self._current_scoring_subject = subject
        self._refresh_scoring_state_label(subject)
        self._update_scoring_status_bar(success_count, error_count, edited_count, pending_count)
        self._apply_scoring_filter()

    def _hydrate_scoring_row_with_student_profile(self, subject_key: str, row: dict) -> dict:
        payload = dict(row or {})
        sid = str(payload.get("student_id", "") or "").strip()
        if not sid:
            return payload
        cfg = self._subject_config_by_subject_key(subject_key) or {}
        profile = self._student_profile_by_id(sid)
        room_text = (
            self._subject_room_for_student_id(sid, cfg)
            or str(profile.get("exam_room", "") or "").strip()
            or str(payload.get("exam_room", "") or "").strip()
        )
        payload["name"] = str(profile.get("name", "") or payload.get("name", "") or "").strip()
        payload["class_name"] = str(profile.get("class_name", "") or payload.get("class_name", "") or "").strip()
        payload["birth_date"] = str(profile.get("birth_date", "") or payload.get("birth_date", "") or "").strip()
        payload["exam_room"] = room_text
        return payload

    @staticmethod
    def _scoring_filter_column_from_combo_index(combo_index: int) -> int | None:
        mapping = {
            0: None,  # Tất cả
            1: 0,     # Student ID
            2: 1,     # Name
            3: 2,     # Lớp
            4: 3,     # Ngày sinh
            5: 4,     # Exam Code
            6: 12,    # Trạng thái
        }
        return mapping.get(int(combo_index), None)

    def _apply_scoring_filter(self) -> None:
        if not hasattr(self, "score_preview_table"):
            return

        def _normalize(text: str) -> str:
            return " ".join(str(text or "").strip().lower().split())

        value = _normalize(self.scoring_search_value.text() if hasattr(self, "scoring_search_value") else "")
        col = self._scoring_filter_column_from_combo_index(self.scoring_filter_column.currentIndex()) if hasattr(self, "scoring_filter_column") else None
        for i in range(self.score_preview_table.rowCount()):
            status_item = self.score_preview_table.item(i, 12)
            status_category = str(status_item.data(Qt.UserRole) if status_item else "success")
            status_ok = self.scoring_status_filter_mode in {"all", "", status_category}
            if not value:
                self.score_preview_table.setRowHidden(i, not status_ok)
                continue
            if col is None:
                row_text = " | ".join(_normalize(self.score_preview_table.item(i, j).text() if self.score_preview_table.item(i, j) else "") for j in range(self.score_preview_table.columnCount()))
            else:
                item = self.score_preview_table.item(i, col)
                row_text = _normalize(item.text() if item else "")
            self.score_preview_table.setRowHidden(i, (value not in row_text) or (not status_ok))

    def _handle_scoring_status_filter_link(self, link: str) -> None:
        self.scoring_status_filter_mode = str(link or "all").strip() or "all"
        self._apply_scoring_filter()

    def _update_scoring_status_bar(self, success_count: int, error_count: int, edited_count: int, pending_count: int = 0) -> None:
        if not hasattr(self, "scoring_status_bar"):
            return
        self.scoring_status_bar.setText(
            f"Thống kê chấm điểm: "
            f"<a href='all'><b>Tất cả</b></a> | "
            f"<a href='success' style='color:#0b7a0b;font-weight:600'>Thành công {int(success_count)}</a> | "
            f"<a href='error' style='color:#c62828;font-weight:600'>Lỗi {int(error_count)}</a> | "
            f"<a href='edited' style='color:#1565c0;font-weight:600'>Đã sửa {int(edited_count)}</a> | "
            f"<a href='pending' style='color:#6d4c41;font-weight:600'>Chưa chấm/Cần chấm lại {int(pending_count)}</a>"
        )

    @staticmethod
    def _tf_statement_correct_count(tf_compare_text: str) -> int:
        total = 0
        for token in [x.strip() for x in str(tf_compare_text or "").split(";") if x.strip()]:
            _, _, pair = token.partition(":")
            key_txt, _, marked_txt = pair.partition("|")
            key_norm = "".join(ch for ch in str(key_txt or "").upper() if ch in {"T", "F", "Đ", "D", "S"})
            mark_norm = "".join(ch for ch in str(marked_txt or "").upper() if ch in {"T", "F", "Đ", "D", "S"})
            limit = min(len(key_norm), len(mark_norm))
            for i in range(limit):
                if key_norm[i] == mark_norm[i]:
                    total += 1
        return total

    def _find_scoring_scan_result(self, subject_key: str, student_id: str, exam_code: str = "") -> OMRResult | None:
        subject = str(subject_key or "").strip()
        sid = str(student_id or "").strip()
        code = str(exam_code or "").strip()
        if not subject or not sid:
            return None
        sid_norm = self._normalized_student_id_for_match(sid)
        rows = self.database.fetch_scan_results_for_subject(self._batch_result_subject_key(subject)) or []
        best_match: OMRResult | None = None
        for payload in rows:
            res = self._deserialize_omr_result(payload)
            res_sid = str(getattr(res, "student_id", "") or "").strip()
            if res_sid != sid and self._normalized_student_id_for_match(res_sid) != sid_norm:
                continue
            if code and str(getattr(res, "exam_code", "") or "").strip() == code:
                return res
            if best_match is None:
                best_match = res
        return best_match

    def _open_scoring_review_editor_from_table(self, row: int, _col: int) -> None:
        if row < 0 or row >= self.score_preview_table.rowCount():
            return
        sid = str(self.score_preview_table.item(row, 0).text() if self.score_preview_table.item(row, 0) else "").strip()
        exam_code = str(self.score_preview_table.item(row, 4).text() if self.score_preview_table.item(row, 4) else "").strip()
        subject = str(self.scoring_subject_combo.currentData() or self.scoring_subject_combo.currentText() or "").strip() if hasattr(self, "scoring_subject_combo") else ""
        if not sid or not subject:
            return
        if self._subject_uses_direct_score_import(subject):
            self._open_essay_score_editor_from_table(row, subject)
            return
        result = self._find_scoring_scan_result(subject, sid, exam_code)
        if result is None:
            QMessageBox.warning(self, "Tính điểm", "Không tìm thấy bài scan gốc để mở màn hình sửa.")
            return
        self._open_scoring_review_editor(subject, result)

    def _update_direct_score_import_row(self, subject_key: str, student_id: str, score_text: str, exam_room: str = "") -> bool:
        subject = str(subject_key or "").strip()
        sid = str(student_id or "").strip()
        if not subject or not sid:
            return False
        cfg = self._subject_config_by_subject_key(subject)
        if not isinstance(cfg, dict) or not self._subject_uses_direct_score_import(cfg):
            return False
        direct_payload = cfg.setdefault("direct_score_import", {})
        rows = direct_payload.setdefault("rows", [])
        if not isinstance(rows, list):
            rows = []
            direct_payload["rows"] = rows

        target_room = str(exam_room or "").strip()
        matched = False
        for row in rows:
            if not isinstance(row, dict):
                continue
            row_sid = str(row.get("student_id", "") or "").strip()
            row_room = str(row.get("exam_room", "") or row.get("room", "") or "").strip()
            if row_sid != sid:
                continue
            if target_room and row_room and row_room != target_room:
                continue
            row["score"] = str(score_text or "").strip()
            row["manually_edited"] = True
            if target_room:
                row["exam_room"] = target_room
            matched = True
            break

        if not matched:
            rows.append({
                "student_id": sid,
                "score": str(score_text or "").strip(),
                "exam_room": target_room,
                "manually_edited": True,
            })

        cfg["direct_score_import"] = direct_payload
        subject_cfgs = list(self.subject_configs or [])
        idx = self._find_subject_config_index_for_batch_save(cfg, subject_cfgs)
        if idx >= 0:
            subject_cfgs[idx] = cfg
            self.subject_configs = subject_cfgs
        if self.session:
            self.session.config = {**(self.session.config or {}), "subject_configs": list(self.subject_configs or [])}
        self._persist_current_session_subject_configs(list(self.subject_configs or []))
        self._persist_runtime_session_state_quietly()
        return True

    def _open_essay_score_editor_from_table(self, row: int, subject_key: str) -> None:
        sid = str(self.score_preview_table.item(row, 0).text() if self.score_preview_table.item(row, 0) else "").strip()
        name = str(self.score_preview_table.item(row, 1).text() if self.score_preview_table.item(row, 1) else "").strip()
        exam_room = str(self.score_preview_table.item(row, 3).text() if self.score_preview_table.item(row, 3) else "").strip()
        current_score = str(self.score_preview_table.item(row, 6).text() if self.score_preview_table.item(row, 6) else "").strip()
        if not sid:
            return
        new_score, ok = QInputDialog.getText(
            self,
            "Sửa điểm tự luận",
            f"SBD: {sid}\nHọ tên: {name or '-'}\nPhòng thi: {exam_room or '-'}\n\nĐiểm:",
            text=current_score,
        )
        if not ok:
            return
        score_text = str(new_score or "").strip().replace(",", ".")
        if score_text == "":
            QMessageBox.warning(self, "Sửa điểm tự luận", "Điểm không được để trống.")
            return
        try:
            float(score_text)
        except Exception:
            QMessageBox.warning(self, "Sửa điểm tự luận", "Điểm không hợp lệ.")
            return
        if not self._update_direct_score_import_row(subject_key, sid, score_text, exam_room):
            QMessageBox.warning(self, "Sửa điểm tự luận", "Không cập nhật được dữ liệu điểm trực tiếp của môn.")
            return
        self.calculate_scores(subject_key, mode="Tính lại toàn bộ", note="essay_manual_edit")
        self._ensure_scoring_preview_current(subject_key, reason="essay_manual_edit", force=True)
        QMessageBox.information(self, "Sửa điểm tự luận", "Đã cập nhật điểm tự luận và đồng bộ lại bảng tính điểm.")

    def _open_scoring_review_editor(self, subject_key: str, result: OMRResult) -> None:
        return open_scoring_review_editor_dialog(self, subject_key, result)

    def _handle_scoring_subject_changed(self, _index: int) -> None:
        subject_key = str(self.scoring_subject_combo.currentData() or "").strip() if hasattr(self, "scoring_subject_combo") else ""
        if subject_key:
            self._ensure_scoring_preview_current(subject_key, reason="auto_refresh_subject_change", force=False)

    def _sync_current_batch_subject_snapshot(self, persist_to_db: bool = True) -> tuple[str, list[OMRResult]]:
        # Helper này chỉ còn phục vụ thao tác explicit save của Batch Scan.
        # Khi persist_to_db=False, trả snapshot runtime để UI tham chiếu, không được ghi DB
        # và không được cập nhật nguồn dữ liệu lõi.
        subject_cfg = self._selected_batch_subject_config()
        if not subject_cfg or not hasattr(self, "scan_list"):
            return "", []
        subject_key = self._subject_key_from_cfg(subject_cfg)
        if not subject_key:
            return "", []
        subject_db_key = self._batch_result_subject_key(subject_key)

        if self.scan_list.rowCount() <= 0:
            if persist_to_db:
                self.database.replace_scan_results_for_subject(subject_db_key, [])
                self.scan_results = []
                self.scan_results_by_subject[subject_db_key] = []
            return subject_key, []

        current_results = self._current_scan_results_snapshot()
        for result in current_results:
            result.answer_string = self._normalize_non_api_answer_string(result, subject_key)
            self._debug_scan_result_state("sync_snapshot", result)

        if not persist_to_db:
            return subject_key, list(current_results)

        self.scan_results = list(current_results)
        self.scan_results_by_subject[subject_db_key] = list(current_results)
        self.database.replace_scan_results_for_subject(
            subject_db_key,
            [self._serialize_omr_result(result) for result in current_results],
        )
        return subject_key, current_results
    def _show_batch_scan_panel(self) -> None:
        if hasattr(self, "scan_lr_split"):
            self.scan_lr_split.setVisible(True)
        if hasattr(self, "batch_scan_status_bottom"):
            self.batch_scan_status_bottom.setVisible(True)
        if hasattr(self, "scoring_panel"):
            self.scoring_panel.setVisible(False)
        self._current_route_name = "workspace_batch_scan"
        for action_name in ["ribbon_batch_execute_action", "ribbon_batch_save_action", "ribbon_batch_close_action"]:
            action = getattr(self, action_name, None)
            if action is not None:
                action.setVisible(True)
        self._refresh_ribbon_action_states()
        if hasattr(self, "batch_subject_combo") and self.batch_subject_combo.count() > 0:
            cfg = self._selected_batch_subject_config()
            if cfg:
                runtime_key = self._batch_runtime_key(cfg)
                need_reload = bool(self.scan_list.rowCount() <= 0 or runtime_key != self._batch_loaded_runtime_key)
                if need_reload:
                    self._load_batch_subject_state(cfg, source_hint="show_batch_panel", force_reload=False)
                else:
                    self._update_batch_scan_scope_summary()

    def _show_scoring_panel(self) -> None:
        if hasattr(self, "scan_lr_split"):
            self.scan_lr_split.setVisible(False)
        if hasattr(self, "batch_scan_status_bottom"):
            self.batch_scan_status_bottom.setVisible(False)
        if hasattr(self, "scoring_panel"):
            self.scoring_panel.setVisible(True)
        self._current_route_name = "workspace_scoring"
        for action_name in ["ribbon_batch_execute_action", "ribbon_batch_save_action", "ribbon_batch_close_action"]:
            action = getattr(self, action_name, None)
            if action is not None:
                action.setVisible(False)
        self._refresh_ribbon_action_states()
        selected_subject = str(self.scoring_subject_combo.currentData() or "").strip() if hasattr(self, "scoring_subject_combo") else ""
        if selected_subject:
            self._ensure_scoring_preview_current(selected_subject, reason="auto_refresh_open_scoring", force=False)

    def _back_to_batch_scan(self) -> None:
        # workspace sub-route
        if not self._navigate_to(
            "workspace_batch_scan",
            context={"session_id": self.current_session_id, "origin": "workspace_scoring"},
            push_current=False,
            require_confirm=True,
            reason="back_to_batch",
        ):
            return

    def _subject_configs_for_scoring(self) -> list[dict]:
        return self._effective_subject_configs_for_batch()

    def _subject_config_by_subject_key(self, subject_key: str) -> dict | None:
        key_norm = str(subject_key or "").strip()
        if not key_norm:
            return None
        exact_match: dict | None = None
        logical_matches: list[dict] = []
        for cfg in self._subject_configs_for_scoring():
            canonical = self._subject_instance_key_from_cfg(cfg)
            logical = self._logical_subject_key_from_cfg(cfg)
            if key_norm == canonical:
                exact_match = cfg
                break
            if key_norm == logical:
                logical_matches.append(cfg)
        if exact_match is not None:
            return exact_match
        # legacy fallback / migration: only allow logical-key lookup when unambiguous.
        if len(logical_matches) == 1:
            return logical_matches[0]
        return None

    @staticmethod
    def _serialize_omr_result(result: OMRResult) -> dict:
        return {
            "image_path": str(getattr(result, "image_path", "") or ""),
            "student_id": str(getattr(result, "student_id", "") or ""),
            "exam_code": str(getattr(result, "exam_code", "") or ""),
            "mcq_answers": {int(k): str(v) for k, v in (getattr(result, "mcq_answers", {}) or {}).items()},
            "true_false_answers": {int(k): dict(v) for k, v in (getattr(result, "true_false_answers", {}) or {}).items()},
            "numeric_answers": {int(k): str(v) for k, v in (getattr(result, "numeric_answers", {}) or {}).items()},
            "confidence_scores": {str(k): float(v) for k, v in (getattr(result, "confidence_scores", {}) or {}).items()},
            "recognition_errors": [str(x) for x in (getattr(result, "recognition_errors", []) or [])],
            "processing_time_sec": float(getattr(result, "processing_time_sec", 0.0) or 0.0),
            "debug_image_path": str(getattr(result, "debug_image_path", "") or ""),
            "full_name": str(getattr(result, "full_name", "") or ""),
            "birth_date": str(getattr(result, "birth_date", "") or ""),
            "exam_room": str(getattr(result, "exam_room", "") or ""),
            "class_name": str(getattr(result, "class_name", "") or ""),
            "cached_status": str(getattr(result, "cached_status", "") or ""),
            "cached_content": str(getattr(result, "cached_content", "") or ""),
            "cached_recognized_short": str(getattr(result, "cached_recognized_short", "") or ""),
            "manual_content_override": str(getattr(result, "manual_content_override", "") or ""),
            "cached_forced_status": str(getattr(result, "cached_forced_status", "") or ""),
            "manually_edited": bool(getattr(result, "manually_edited", False)),
            "edit_history": [str(x) for x in (getattr(result, "edit_history", []) or []) if str(x or "").strip()],
            "last_adjustment": str(getattr(result, "last_adjustment", "") or ""),
            "manual_adjustments": [str(x) for x in (getattr(result, "manual_adjustments", []) or []) if str(x or "").strip()],
            "cached_blank_summary": dict(getattr(result, "cached_blank_summary", {}) or {}),
            "recognized_template_path": str(getattr(result, "recognized_template_path", "") or ""),
            "recognized_alignment_profile": str(getattr(result, "recognized_alignment_profile", "") or ""),
            "recognized_fill_threshold": float(getattr(result, "recognized_fill_threshold", 0.45) or 0.45),
            "recognized_empty_threshold": float(getattr(result, "recognized_empty_threshold", 0.20) or 0.20),
            "recognized_certainty_margin": float(getattr(result, "recognized_certainty_margin", 0.08) or 0.08),
            "answer_string_api_mode": bool(getattr(result, "answer_string_api_mode", False)),
            "answer_string": str(getattr(result, "answer_string", "") or ""),
        }

    @staticmethod
    def _deserialize_omr_result(payload: dict) -> OMRResult:
        result = OMRResult(
            image_path=str(payload.get("image_path", "") or ""),
            student_id=str(payload.get("student_id", "") or ""),
            exam_code=str(payload.get("exam_code", "") or ""),
            mcq_answers={int(k): str(v) for k, v in (payload.get("mcq_answers", {}) or {}).items()},
            true_false_answers={int(k): dict(v) for k, v in (payload.get("true_false_answers", {}) or {}).items()},
            numeric_answers={int(k): str(v) for k, v in (payload.get("numeric_answers", {}) or {}).items()},
            confidence_scores={str(k): float(v) for k, v in (payload.get("confidence_scores", {}) or {}).items()},
            recognition_errors=[str(x) for x in (payload.get("recognition_errors", []) or [])],
            processing_time_sec=float(payload.get("processing_time_sec", 0.0) or 0.0),
            debug_image_path=str(payload.get("debug_image_path", "") or ""),
        )
        setattr(result, "full_name", str(payload.get("full_name", "") or ""))
        setattr(result, "birth_date", str(payload.get("birth_date", "") or ""))
        setattr(result, "exam_room", str(payload.get("exam_room", "") or ""))
        setattr(result, "class_name", str(payload.get("class_name", "") or ""))
        setattr(result, "cached_status", str(payload.get("cached_status", "") or ""))
        setattr(result, "cached_content", str(payload.get("cached_content", "") or ""))
        setattr(result, "cached_recognized_short", str(payload.get("cached_recognized_short", "") or ""))
        setattr(result, "manual_content_override", str(payload.get("manual_content_override", "") or ""))
        setattr(result, "cached_forced_status", str(payload.get("cached_forced_status", "") or ""))
        setattr(result, "manually_edited", bool(payload.get("manually_edited", False)))
        setattr(result, "edit_history", [str(x) for x in (payload.get("edit_history", []) or []) if str(x or "").strip()])
        setattr(result, "last_adjustment", str(payload.get("last_adjustment", "") or ""))
        setattr(result, "manual_adjustments", [str(x) for x in (payload.get("manual_adjustments", []) or []) if str(x or "").strip()])
        if (not bool(getattr(result, "manually_edited", False))) and bool(getattr(result, "edit_history", [])):
            setattr(result, "manually_edited", True)
            if not str(getattr(result, "cached_forced_status", "") or "").strip():
                setattr(result, "cached_forced_status", "Đã sửa")
        setattr(result, "cached_blank_summary", dict(payload.get("cached_blank_summary", {}) or {}))
        setattr(result, "recognized_template_path", str(payload.get("recognized_template_path", "") or ""))
        setattr(result, "recognized_alignment_profile", str(payload.get("recognized_alignment_profile", "") or ""))
        setattr(result, "recognized_fill_threshold", float(payload.get("recognized_fill_threshold", 0.45) or 0.45))
        setattr(result, "recognized_empty_threshold", float(payload.get("recognized_empty_threshold", 0.20) or 0.20))
        setattr(result, "recognized_certainty_margin", float(payload.get("recognized_certainty_margin", 0.08) or 0.08))
        setattr(result, "answer_string_api_mode", bool(payload.get("answer_string_api_mode", False)))
        setattr(result, "answer_string", str(payload.get("answer_string", "") or ""))
        result.sync_legacy_aliases()
        return result

    def _is_subject_marked_batched(self, cfg: dict) -> bool:
        key = self._subject_key_from_cfg(cfg)
        if not key:
            return False
        if self._subject_uses_direct_score_import(cfg):
            return bool(self._direct_score_import_rows_for_subject(cfg))
        scoped_key = self._batch_result_subject_key(key)
        if self.scan_results_by_subject.get(scoped_key):
            return True
        try:
            db_rows = self.database.fetch_scan_results_for_subject(scoped_key) or []
            if isinstance(db_rows, list) and db_rows:
                return True
        except Exception:
            pass
        return bool(cfg.get("batch_saved"))
    def _cached_subject_scans_from_config(self, subject_key: str) -> list[OMRResult]:
        # Legacy helper giữ lại để tương thích, nhưng luôn hydrate từ DB.
        return list(self._refresh_scan_results_from_db(subject_key) or [])
    def _ensure_answer_keys_for_subject(self, subject_key: str) -> bool:
        if self.answer_keys and any(str(k).startswith(f"{subject_key}::") for k in self.answer_keys.keys.keys()):
            return True
        cfg = self._subject_config_by_subject_key(subject_key)
        if not cfg:
            return self.answer_keys is not None

        imported_keys = cfg.get("imported_answer_keys", {}) or {}
        if isinstance(imported_keys, dict) and imported_keys:
            repo = self.answer_keys or AnswerKeyRepository()
            for exam_code, kd in imported_keys.items():
                repo.upsert(
                    SubjectKey(
                        subject=subject_key,
                        exam_code=str(exam_code),
                        answers={int(k): str(v) for k, v in (kd.get("mcq_answers", {}) or {}).items()},
                        true_false_answers={int(k): dict(v) for k, v in (kd.get("true_false_answers", {}) or {}).items()},
                        numeric_answers={int(k): str(v) for k, v in (kd.get("numeric_answers", {}) or {}).items()},
                        full_credit_questions={
                            str(sec): [int(x) for x in (vals or []) if str(x).strip().lstrip("-").isdigit()]
                            for sec, vals in (kd.get("full_credit_questions", {}) or {}).items()
                        },
                        invalid_answer_rows={
                            str(sec): {int(k): str(v) for k, v in (vals or {}).items()}
                            for sec, vals in (kd.get("invalid_answer_rows", {}) or {}).items()
                        },
                    )
                )
            self.answer_keys = repo
            self.imported_exam_codes = sorted(str(k) for k in imported_keys.keys())
            self.active_batch_subject_key = subject_key
            return True

        answer_key_path = str(cfg.get("answer_key_path", "") or (self.session.answer_key_path if self.session else "") or "")
        if answer_key_path:
            pth = Path(answer_key_path)
            if pth.exists() and pth.suffix.lower() == ".json":
                try:
                    loaded = AnswerKeyRepository.load_json(pth)
                    repo = self.answer_keys or AnswerKeyRepository()
                    for key_obj in loaded.keys.values():
                        repo.upsert(key_obj)
                    self.answer_keys = repo
                    all_codes: set[str] = set()
                    for key_name in self.answer_keys.keys.keys():
                        parts = str(key_name).split("::", 1)
                        if len(parts) == 2 and parts[0] == subject_key:
                            all_codes.add(parts[1])
                    if all_codes:
                        self.imported_exam_codes = sorted(all_codes)
                    self.active_batch_subject_key = subject_key
                    return True
                except Exception:
                    pass

        return self.answer_keys is not None

    @staticmethod
    def _logical_subject_key_from_cfg(cfg: dict) -> str:
        key = str(cfg.get("answer_key_key", "") or "").strip()
        if key:
            return key
        name = str(cfg.get("name", "") or "").strip()
        block = str(cfg.get("block", "") or "").strip()
        return f"{name}_{block}" if name and block else (name or "General")

    def _ensure_subject_instance_key(self, cfg: dict, index: int | None = None) -> str:
        # canonical subject_instance_key: scoped runtime/storage key, stable across edits of display fields.
        if not isinstance(cfg, dict):
            return ""
        existing = str(cfg.get("subject_instance_key", "") or "").strip()
        scope_prefix = self._session_scope_prefix() or "session"
        # hard isolation guard: if a key belongs to another session scope, remap it to current session scope.
        if existing and scope_prefix and not existing.startswith(f"{scope_prefix}::"):
            legacy_keys = list(cfg.get("legacy_subject_instance_keys", [])) if isinstance(cfg.get("legacy_subject_instance_keys", []), list) else []
            if existing not in legacy_keys:
                legacy_keys.append(existing)
            cfg["legacy_subject_instance_keys"] = legacy_keys[-10:]
            existing = ""
        if existing:
            return existing
        logical = self._logical_subject_key_from_cfg(cfg) or "General"
        uid = str(cfg.get("subject_uid", "") or "").strip()
        if not uid:
            uid = str(uuid.uuid4())
            cfg["subject_uid"] = uid
        suffix = f"{int(index)}" if isinstance(index, int) and index >= 0 else uid
        value = f"{scope_prefix}::{logical}::{suffix}"
        cfg["logical_subject_key"] = logical
        cfg["subject_instance_key"] = value
        return value

    def _subject_instance_key_from_cfg(self, cfg: dict) -> str:
        if not isinstance(cfg, dict):
            return ""
        return self._ensure_subject_instance_key(cfg)

    def _current_subject_instance_key(self) -> str:
        cfg = self._selected_batch_subject_config()
        if not isinstance(cfg, dict):
            return ""
        return self._subject_instance_key_from_cfg(cfg)

    def _current_batch_runtime_key(self) -> str:
        return self._batch_runtime_key(self._current_subject_instance_key())

    def _subject_key_from_cfg(self, cfg: dict) -> str:
        # backward compatibility shim: storage/runtime key is now subject_instance_key.
        return self._subject_instance_key_from_cfg(cfg)

    def _batch_result_subject_key(self, subject_key: str) -> str:
        base = str(subject_key or "").strip()
        scope_prefix = self._session_scope_prefix()
        if scope_prefix and base:
            return f"{scope_prefix}::{base}"
        return base

    def _session_scope_prefix(self) -> str:
        sid = str(self.current_session_id or "").strip()
        exam_name = str((self.session.exam_name if self.session else "") or "").strip().lower()
        if sid and exam_name:
            return f"{sid}::{exam_name}"
        return sid

    def _answer_key_subject_key(self, subject_key: str, subject_cfg: dict | None = None) -> str:
        base = str(subject_key or "").strip()
        if not base:
            return ""
        if "::" in base:
            return base
        block = ""
        if isinstance(subject_cfg, dict):
            block = str(subject_cfg.get("block", "") or "").strip()
        if not block and "_" in base:
            block = str(base.rsplit("_", 1)[-1]).strip()
        scope_prefix = self._session_scope_prefix()
        if scope_prefix and block:
            return f"{scope_prefix}::{base}::{block}"
        if scope_prefix:
            return f"{scope_prefix}::{base}"
        return base

    def _fetch_answer_keys_for_subject_scoped(self, subject_key: str, subject_cfg: dict | None = None) -> dict[str, dict[str, Any]]:
        candidates: list[str] = []

        def _push(value: str) -> None:
            key = str(value or "").strip()
            if key and key not in candidates:
                candidates.append(key)

        _push(subject_key)
        if isinstance(subject_cfg, dict):
            _push(str(subject_cfg.get("answer_key_key", "") or ""))
            _push(self._logical_subject_key_from_cfg(subject_cfg))
            _push(self._subject_instance_key_from_cfg(subject_cfg))

        has_session_scope = bool(str(self.current_session_id or "").strip())
        block = str((subject_cfg or {}).get("block", "") or "").strip() if isinstance(subject_cfg, dict) else ""
        sid = str(self.current_session_id or "").strip()

        for candidate in candidates:
            scoped = self._answer_key_subject_key(candidate, subject_cfg)
            rows = self.database.fetch_answer_keys_for_subject(scoped) if scoped else {}
            if not rows and sid:
                legacy_scoped = f"{sid}::{candidate}::{block}" if block else f"{sid}::{candidate}"
                rows = self.database.fetch_answer_keys_for_subject(legacy_scoped)
            if not rows and (not has_session_scope) and scoped and scoped != candidate:
                rows = self.database.fetch_answer_keys_for_subject(candidate)
            if rows:
                return rows or {}

        return {}

    def _display_subject_label(self, cfg: dict | None) -> str:
        if not isinstance(cfg, dict):
            return "-"
        exam_name = str((self.session.exam_name if self.session else "") or "").strip() or "Kỳ thi hiện tại"
        subject_name = str(cfg.get("name", "") or "").strip() or str(self._logical_subject_key_from_cfg(cfg) or "-")
        block = str(cfg.get("block", "") or "").strip() or "-"
        return f"{exam_name} | {subject_name} | Khối {block}"

    def _batch_cache_subject_key(self, cfg: dict | None, include_session: bool = True) -> str:
        if not isinstance(cfg, dict):
            return ""
        canonical = str(self._subject_instance_key_from_cfg(cfg) or "").strip().lower()
        if canonical:
            return canonical
        name = str(cfg.get("name", "") or "").strip().lower()
        block = str(cfg.get("block", "") or "").strip().lower()
        answer_key = str(cfg.get("answer_key_key", "") or "").strip().lower()
        base = f"{name}::{block}::{answer_key}"
        if include_session:
            scope_prefix = self._session_scope_prefix().strip().lower()
            if scope_prefix:
                return f"{scope_prefix}::{base}"
        return base

    def _resolve_preferred_scoring_subject(self) -> str:
        if self.stack.currentIndex() == 1:
            cfg = self._selected_batch_subject_config()
            if cfg:
                return self._subject_key_from_cfg(cfg)
        cfgs = self._subject_configs_for_scoring()
        if cfgs:
            return self._subject_key_from_cfg(cfgs[0])
        if self.session and self.session.subjects:
            return str(self.session.subjects[0])
        return "General"

    def _eligible_scoring_subject_keys(self) -> list[str]:
        out: list[str] = []
        for cfg in self._subject_configs_for_scoring():
            key = self._subject_key_from_cfg(cfg)
            if self._is_subject_marked_batched(cfg):
                out.append(key)
        return out

    def _populate_scoring_subjects(self, preferred_key: str = "") -> None:
        if not hasattr(self, "scoring_subject_combo"):
            return
        self.scoring_subject_combo.blockSignals(True)
        self.scoring_subject_combo.clear()
        eligible = set(self._eligible_scoring_subject_keys())
        for cfg in self._subject_configs_for_scoring():
            key = self._subject_key_from_cfg(cfg)
            if eligible and key not in eligible:
                continue
            label = self._display_subject_label(cfg)
            self.scoring_subject_combo.addItem(label, key)
        if self.scoring_subject_combo.count() == 0:
            fallback = self._resolve_preferred_scoring_subject()
            self.scoring_subject_combo.addItem(fallback, fallback)
        pick = preferred_key or self._resolve_preferred_scoring_subject()
        for i in range(self.scoring_subject_combo.count()):
            if str(self.scoring_subject_combo.itemData(i)) == pick:
                self.scoring_subject_combo.setCurrentIndex(i)
                break
        self.scoring_subject_combo.blockSignals(False)

    def _open_scoring_view(self) -> None:
        if not self._ensure_current_session_loaded():
            QMessageBox.warning(self, "Tính điểm", "Chưa có kỳ thi hiện tại. Vui lòng mở hoặc tạo kỳ thi trước.")
            return

        # Scoring là DB-first: nếu Batch Scan hiện tại còn thay đổi chưa lưu thì hỏi người dùng
        # lưu hay bỏ qua. Không dựng snapshot runtime để làm nguồn dữ liệu chấm.
        if self._has_batch_unsaved_changes():
            choice = self._prompt_save_changes_word_style(
                "Batch Scan chưa lưu",
                "Bạn có thay đổi chưa lưu cho môn hiện tại. Muốn lưu trước khi sang phần chấm điểm không?",
            )
            if choice == "cancel":
                return
            if choice == "save" and not self._save_batch_for_selected_subject():
                return

        if not self._eligible_scoring_subject_keys():
            QMessageBox.warning(self, "Tính điểm", "Cần có ít nhất 1 môn đã Batch Scan hoặc đã import điểm trực tiếp trước khi tính điểm.")
            return

        self._navigate_to(
            "workspace_scoring",
            context={"session_id": self.current_session_id},
            push_current=True,
            require_confirm=False,
            reason="open_scoring",
        )
        selected_subject = self._resolve_preferred_scoring_subject()
        self._populate_scoring_subjects(selected_subject)
        self._refresh_scoring_phase_table()
        self._refresh_dashboard_summary_from_db(selected_subject)
        self._show_scoring_panel()
    def _refresh_scoring_phase_table(self) -> None:
        if not hasattr(self, "score_phase_table"):
            return
        subjects = self._subject_configs_for_scoring()
        self.score_phase_table.setUpdatesEnabled(False)
        try:
            self.score_phase_table.setRowCount(0)
            for cfg in subjects:
                subject = self._subject_key_from_cfg(cfg)
                if not subject:
                    continue
                saved_at = str((cfg or {}).get("scoring_saved_at", "") or "").strip()
                mode = str((cfg or {}).get("scoring_last_mode", "") or "").strip()
                note = str((cfg or {}).get("scoring_phase_last_note", "") or "").strip()
                count = int((cfg or {}).get("scoring_result_count", 0) or 0)

                if (not saved_at and count <= 0) or (not mode and not note):
                    stored_rows = self._load_scoring_results_for_subject_from_storage(subject)
                    latest = None
                    latest_ts = ""
                    for payload in (stored_rows or {}).values():
                        if not isinstance(payload, dict):
                            continue
                        ts = str(payload.get("phase_timestamp", "") or "")
                        if ts >= latest_ts:
                            latest_ts = ts
                            latest = payload
                    if latest is not None:
                        saved_at = saved_at or latest_ts
                        mode = mode or str((latest or {}).get("phase_mode", "") or "")
                        note = note or str((latest or {}).get("note", "") or (latest or {}).get("status", "") or "")
                        count = count or len(stored_rows or {})

                r = self.score_phase_table.rowCount()
                self.score_phase_table.insertRow(r)
                self.score_phase_table.setItem(r, 0, QTableWidgetItem(subject or "-"))
                self.score_phase_table.setItem(r, 1, QTableWidgetItem(saved_at or "-"))
                self.score_phase_table.setItem(r, 2, QTableWidgetItem(mode or "-"))
                self.score_phase_table.setItem(r, 3, QTableWidgetItem(str(count)))
                self.score_phase_table.setItem(r, 4, QTableWidgetItem(note or "-"))
        finally:
            self.score_phase_table.setUpdatesEnabled(True)
            self.score_phase_table.viewport().update()

    def _run_scoring_from_panel(self) -> None:
        subject_key = str(self.scoring_subject_combo.currentData() or "").strip() if hasattr(self, "scoring_subject_combo") else ""
        mode = self.scoring_mode_combo.currentText() if hasattr(self, "scoring_mode_combo") else "Tính lại toàn bộ"
        note = self.scoring_phase_note.text().strip() if hasattr(self, "scoring_phase_note") else ""
        self.calculate_scores(subject_key=subject_key or self._resolve_preferred_scoring_subject(), mode=mode, note=note)
        self._refresh_scoring_state_label(subject_key or self._resolve_preferred_scoring_subject())



    def _resolve_current_session_subject_configs_for_update(self) -> tuple[list[dict], str]:
        # batch save source of truth: prefer runtime session currently opened in app.
        if self.session and isinstance(self.session.config, dict):
            cfgs = self.session.config.get("subject_configs", [])
            if isinstance(cfgs, list):
                return cfgs, "runtime"
        if self.current_session_id:
            payload = self.database.fetch_exam_session(self.current_session_id) or {}
            if isinstance(payload, dict) and payload:
                try:
                    db_session = ExamSession.from_dict(payload)
                    if db_session and isinstance(db_session.config, dict):
                        cfgs = db_session.config.get("subject_configs", [])
                        if isinstance(cfgs, list):
                            self.session = db_session
                            self.current_session_path = self._session_path_from_id(self.current_session_id)
                            return cfgs, "database"
                except Exception:
                    pass
        session_path = self.current_session_path
        if not session_path and self.current_session_id:
            candidate = self._session_path_from_id(self.current_session_id)
            if candidate.exists():
                session_path = candidate
        if session_path and session_path.exists():
            try:
                fallback = ExamSession.load_json(session_path)
                if isinstance(fallback.config, dict):
                    cfgs = fallback.config.get("subject_configs", [])
                    if isinstance(cfgs, list):
                        if not self.session:
                            self.session = fallback
                        return cfgs, "file_fallback"
            except Exception:
                pass
        return [], "empty"

    def _find_subject_config_index_for_batch_save(self, subject_cfg: dict, subject_cfgs: list[dict]) -> int:
        # canonical subject match: subject_instance_key > subject_uid > runtime_key > legacy logical (unambiguous).
        selected_instance = str(self._subject_instance_key_from_cfg(subject_cfg) or "").strip()
        selected_uid = str(subject_cfg.get("subject_uid", "") or "").strip()
        selected_runtime = str(self._batch_runtime_key(selected_instance) or "").strip()
        selected_logical = str(self._logical_subject_key_from_cfg(subject_cfg) or "").strip().lower()
        if selected_instance:
            for idx, item in enumerate(subject_cfgs):
                if str(self._subject_instance_key_from_cfg(item) or "").strip() == selected_instance:
                    return idx
        if selected_uid:
            for idx, item in enumerate(subject_cfgs):
                if str(item.get("subject_uid", "") or "").strip() == selected_uid:
                    return idx
        if selected_runtime:
            for idx, item in enumerate(subject_cfgs):
                item_instance = str(self._subject_instance_key_from_cfg(item) or "").strip()
                if item_instance and str(self._batch_runtime_key(item_instance) or "").strip() == selected_runtime:
                    return idx
        if selected_logical:
            matches = [
                idx for idx, item in enumerate(subject_cfgs)
                if str(self._logical_subject_key_from_cfg(item) or "").strip().lower() == selected_logical
            ]
            if len(matches) == 1:
                return matches[0]
        return -1

    def _persist_current_session_subject_configs(self, subject_cfgs: list[dict]) -> bool:
        if not self.current_session_id:
            return False
        if self.session is None:
            self.session = ExamSession(exam_name="Kỳ thi", exam_date=str(date.today()))
        self.session.config = {**(self.session.config or {}), "subject_configs": list(subject_cfgs)}
        try:
            self.database.save_exam_session(self.current_session_id, self.session.exam_name, self.session.to_dict())
            self.current_session_path = self._session_path_from_id(self.current_session_id)
        except Exception:
            return False
        if self.current_session_path:
            try:
                self.session.save_json(self.current_session_path)
            except Exception:
                pass
        return True

    def _persist_runtime_session_state_quietly(self) -> bool:
        if not self.session or not self.current_session_id:
            return False
        try:
            cfg = dict(self.session.config or {})
            # Không mirror scoring payload vào config nữa.
            cfg.pop("scoring_phases", None)
            cfg.pop("scoring_results", None)
            self.session.config = cfg
            self.database.save_exam_session(self.current_session_id, self.session.exam_name, self.session.to_dict())
            self.current_session_path = self._session_path_from_id(self.current_session_id)
            if self.current_session_path and self.session:
                try:
                    self.session.save_json(self.current_session_path)
                except Exception:
                    pass
            return True
        except Exception:
            return False
    def _save_batch_for_selected_subject(
        self,
        *,
        show_success_message: bool = True,
        reload_after_save: bool = False,
        refresh_exam_list: bool = False,
    ) -> bool:
        subject_cfg = self._selected_batch_subject_config()
        if not subject_cfg:
            if show_success_message:
                QMessageBox.warning(self, "Lưu Batch", "Vui lòng chọn môn trước khi lưu Batch.")
            return False
        row_count = self.scan_list.rowCount() if hasattr(self, "scan_list") else 0
        if row_count <= 0:
            if show_success_message:
                QMessageBox.warning(self, "Lưu Batch", "Chưa có dữ liệu Batch Scan để lưu.")
            return False

        saved_rows = self._serialize_scan_grid_rows_for_save()
        saved_results = [self._serialize_omr_result(result) for result in (self.scan_results or [])]
        subject_key = self._subject_key_from_cfg(subject_cfg)
        saved_at = datetime.now().isoformat(timespec="seconds")

        subject_cfg["batch_saved_rows"] = saved_rows
        subject_cfg["batch_saved"] = True
        subject_cfg["batch_saved_at"] = saved_at
        subject_cfg["batch_result_count"] = len(saved_results)

        try:
            self.sync_current_batch_subject_snapshot(persist_to_db=True)
        except Exception:
            pass

        if subject_key:
            try:
                self.database.replace_scan_results_for_subject(subject_key, saved_results, note="save_batch_subject")
                self._mark_subject_batch_saved(subject_key, len(saved_results), saved_at)
            except Exception as exc:
                QMessageBox.warning(self, "Lưu Batch", f"Không thể lưu dữ liệu Batch Scan vào CSDL: {exc}")
                return False
            self._batch_loaded_runtime_key = self._batch_runtime_key(subject_key)
            self._current_batch_data_source = "database"
            self._mark_scoring_stale_after_batch_save(subject_key)

        self.btn_save_batch_subject.setEnabled(False)
        self._persist_current_session_subject_configs()
        if reload_after_save:
            self._load_batch_subject_state(subject_cfg, source_hint="after_save", force_reload=False)
        else:
            self._update_batch_scan_scope_summary()

        current_route = str(getattr(self, "_current_route", "") or "").strip()
        if refresh_exam_list or current_route in {"exam_list", "dashboard"}:
            self._refresh_exam_list()

        if show_success_message:
            QMessageBox.information(self, "Lưu Batch", f"Đã lưu {len(saved_results)} bài của môn '{subject_key or 'hiện tại'}'.")
        return True

    def _mark_scoring_stale_after_batch_save(self, subject_key: str, note: str = "Batch Scan đã thay đổi") -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        cfg = self._subject_config_by_subject_key(subject) or {}
        existing_count = int((cfg or {}).get("scoring_result_count", 0) or 0)
        if existing_count <= 0:
            existing_count = len(self.scoring_results_by_subject.get(subject, {}) or {})
        self._clear_subject_scoring_saved_state(
            subject,
            count=existing_count,
            mode=str((cfg or {}).get("scoring_last_mode", "") or ""),
            note=note,
        )

    def _ensure_scoring_preview_current(self, subject_key: str, reason: str = "", force: bool = False) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        self._current_scoring_subject = subject
        self._apply_subject_section_visibility(subject)
        cached_rows = self._load_scoring_results_for_subject_from_storage(subject)
        is_dirty = subject in self._scoring_dirty_subjects
        if force or is_dirty or not cached_rows:
            self.calculate_scores(subject_key=subject, mode="Tính lại toàn bộ", note=reason)
            return
        self._load_cached_scoring_results_for_subject(subject)

    def _build_correction_tab(self) -> QWidget:
        w = QWidget()
        splitter = QSplitter(Qt.Horizontal)
        self.exam_code_correction_combo = QComboBox()
        self.exam_code_correction_combo.currentIndexChanged.connect(self._handle_exam_code_correction_changed)
        self.student_correction_combo = QComboBox()
        self.student_correction_combo.setEditable(True)
        self.student_correction_combo.setInsertPolicy(QComboBox.NoInsert)
        self.student_correction_combo.currentIndexChanged.connect(self._handle_student_correction_changed)
        self.answer_editor_scroll = QScrollArea()
        self.answer_editor_scroll.setWidgetResizable(True)
        self.answer_editor_container = QWidget()
        self.answer_editor_layout = QVBoxLayout(self.answer_editor_container)
        self.answer_editor_scroll.setWidget(self.answer_editor_container)
        left = QWidget()
        left_layout = QVBoxLayout(left)
        self.error_list = QListWidget()
        self.preview_label = QLabel("Image preview / bubble overlay placeholder")
        self.preview_label.setAlignment(Qt.AlignCenter)
        self.manual_edit = QTextEdit()
        self.manual_edit.setPlaceholderText("Manual corrections JSON (e.g. {'student_id': '1001', 'answers': {'1':'A'}})")
        self.result_preview = QTextEdit()
        self.result_preview.setReadOnly(True)
        self.result_preview.setPlaceholderText("Recognized result for selected scan")

        btn_load_selected = QPushButton("Load Selected Scan Result")
        btn_load_selected.clicked.connect(self._load_selected_result_for_correction)
        btn_apply_correction = QPushButton("Apply Manual Correction")
        btn_apply_correction.clicked.connect(self.apply_manual_correction)

        left_layout.addWidget(QLabel("Detected Errors"))
        left_layout.addWidget(self.error_list)
        left_layout.addWidget(btn_load_selected)
        left_layout.addWidget(btn_apply_correction)

        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(0)
        right_layout.addWidget(self.preview_label)
        correction_form = QFormLayout()
        correction_form.addRow("Mã đề (DB)", self.exam_code_correction_combo)
        correction_form.addRow("Học sinh (DB)", self.student_correction_combo)
        right_layout.addLayout(correction_form)
        right_layout.addWidget(QLabel("Chỉnh sửa đáp án trực quan"))
        right_layout.addWidget(self.answer_editor_scroll, 4)
        right_layout.addWidget(self.result_preview)
        right_layout.addWidget(QLabel("Manual Edit"))
        right_layout.addWidget(self.manual_edit)

        splitter.addWidget(left)
        splitter.addWidget(right)

        layout = QHBoxLayout(w)
        layout.addWidget(splitter)
        return w

    def _correction_selected_result(self) -> tuple[int, OMRResult] | tuple[None, None]:
        idx = self.scan_list.currentRow() if hasattr(self, "scan_list") else -1
        if idx < 0 or idx >= len(self.scan_results):
            return None, None
        return idx, self.scan_results[idx]

    def _load_exam_code_correction_options(self, subject_key: str, current_code: str) -> None:
        self.exam_code_correction_combo.blockSignals(True)
        self.exam_code_correction_combo.clear()
        codes = set(self._fetch_answer_keys_for_subject_scoped(subject_key).keys())
        codes.update(str(x).strip() for x in (self.imported_exam_codes or []) if str(x).strip())
        if current_code:
            codes.add(str(current_code).strip())
        for code in sorted(codes):
            self.exam_code_correction_combo.addItem(code, code)
        if self.exam_code_correction_combo.count() == 0:
            self.exam_code_correction_combo.addItem(current_code or "-", current_code or "")
        match_index = max(0, self.exam_code_correction_combo.findData(current_code))
        self.exam_code_correction_combo.setCurrentIndex(match_index)
        self.exam_code_correction_combo.blockSignals(False)

    def _load_student_correction_options(self, current_student_id: str) -> None:
        cache_session_id = str(getattr(self, "_student_option_cache_session_id", "") or "")
        current_session_id = str(self.current_session_id or "")
        session_signature = ""
        if self.session and str(getattr(self.session, "session_id", "") or "") == current_session_id:
            sig_parts: list[str] = []
            for st in (self.session.students or []):
                sid = str(getattr(st, "student_id", "") or "").strip()
                if not sid:
                    continue
                extra = getattr(st, "extra", {}) if isinstance(getattr(st, "extra", {}), dict) else {}
                sig_parts.append(f"{sid}|{str(getattr(st, 'name', '') or '').strip()}|{str(extra.get('class_name', '') or '').strip()}")
            session_signature = "||".join(sorted(sig_parts))
        if cache_session_id != current_session_id:
            self._student_option_cache_session_id = current_session_id
            self._student_option_labels_cache = []
            self._student_option_sid_map = {}
            self._student_option_profile_map = {}
            self._student_option_sid_set = set()
            self._student_option_cache_signature = ""
        if session_signature and session_signature != str(getattr(self, "_student_option_cache_signature", "") or ""):
            self._student_option_labels_cache = []
            self._student_option_sid_map = {}
            self._student_option_profile_map = {}
            self._student_option_sid_set = set()
            self._student_option_cache_signature = session_signature

        if not isinstance(getattr(self, "_student_option_labels_cache", None), list) or not self._student_option_labels_cache:
            students: list[tuple[str, str, str]] = []
            if self.session and str(getattr(self.session, "session_id", "") or "") == current_session_id:
                for st in (self.session.students or []):
                    sid = str(getattr(st, "student_id", "") or "").strip()
                    if not sid:
                        continue
                    extra = getattr(st, "extra", {}) if isinstance(getattr(st, "extra", {}), dict) else {}
                    students.append((sid, str(getattr(st, "name", "") or "").strip(), str(extra.get("class_name", "") or "").strip()))
            elif current_session_id:
                payload = self.database.fetch_exam_session(current_session_id) or {}
                session_students = payload.get("students", []) if isinstance(payload.get("students", []), list) else []
                for item in session_students:
                    if not isinstance(item, dict):
                        continue
                    extra = item.get("extra", {}) if isinstance(item.get("extra", {}), dict) else {}
                    sid = str(item.get("student_id", "") or "").strip()
                    if not sid:
                        continue
                    students.append((sid, str(item.get("name", "") or "").strip(), str(extra.get("class_name", "") or "").strip()))
            labels: list[str] = []
            sid_map: dict[str, str] = {}
            profile_map: dict[str, dict[str, str]] = {}
            sid_set: set[str] = set()
            for sid, name, class_name in students:
                if sid in sid_set:
                    continue
                sid_set.add(sid)
                label = f"[{sid}] - {name or '-'} - {class_name or '-'}"
                labels.append(label)
                sid_map[label] = sid
                profile_map[sid] = {"name": name, "class_name": class_name}
            self._student_option_labels_cache = labels
            self._student_option_sid_map = sid_map
            self._student_option_profile_map = profile_map
            self._student_option_sid_set = sid_set
            if session_signature:
                self._student_option_cache_signature = session_signature

        self.student_correction_combo.blockSignals(True)
        self.student_correction_combo.clear()
        for label in list(self._student_option_labels_cache):
            sid = str(self._student_option_sid_map.get(label, "") or "")
            if sid:
                self.student_correction_combo.addItem(label, sid)
        if current_student_id and current_student_id not in set(self._student_option_sid_set):
            label = f"[{current_student_id}] - - -"
            self.student_correction_combo.addItem(label, current_student_id)
        idx = self.student_correction_combo.findData(current_student_id)
        self.student_correction_combo.setCurrentIndex(max(0, idx))
        completer = QCompleter(list(self._student_option_labels_cache), self.student_correction_combo)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchContains)
        self.student_correction_combo.setCompleter(completer)
        self.student_correction_combo.blockSignals(False)

    def _build_visual_answer_editor(self, result: OMRResult) -> None:
        while self.answer_editor_layout.count():
            item = self.answer_editor_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        result = self._scoped_result_copy(result)
        expected = self._expected_questions_by_section(result)

        if expected.get("MCQ", []):
            mcq_box = QGroupBox("MCQ")
            mcq_layout = QVBoxLayout(mcq_box)
            for q_no in expected.get("MCQ", []):
                row = QHBoxLayout()
                row.addWidget(QLabel(f"Câu {q_no}"))
                group = QButtonGroup(mcq_box)
                current_value = str((result.mcq_answers or {}).get(q_no, "") or "")[:1]
                for choice in ["A", "B", "C", "D", "E"]:
                    radio = QRadioButton(choice)
                    if current_value == choice:
                        radio.setChecked(True)
                    radio.toggled.connect(lambda checked, q=q_no, v=choice: checked and self._handle_mcq_visual_change(q, v))
                    group.addButton(radio)
                    row.addWidget(radio)
                clear_btn = QPushButton("Clear")
                clear_btn.clicked.connect(lambda _=False, q=q_no: self._handle_mcq_visual_change(q, ""))
                row.addWidget(clear_btn)
                row.addStretch()
                mcq_layout.addLayout(row)
            self.answer_editor_layout.addWidget(mcq_box)

        if expected.get("TF", []):
            tf_box = QGroupBox("True/False")
            tf_layout = QVBoxLayout(tf_box)
            for q_no in expected.get("TF", []):
                row = QHBoxLayout()
                row.addWidget(QLabel(f"Câu {q_no}"))
                flags = (result.true_false_answers or {}).get(q_no, {}) or {}
                for key in ["a", "b", "c", "d"]:
                    cb = QCheckBox(key.upper())
                    cb.setChecked(bool(flags.get(key)))
                    cb.toggled.connect(lambda checked, q=q_no, k=key: self._handle_tf_visual_change(q, k, checked))
                    row.addWidget(cb)
                row.addStretch()
                tf_layout.addLayout(row)
            self.answer_editor_layout.addWidget(tf_box)

        if expected.get("NUMERIC", []):
            num_box = QGroupBox("Numeric")
            num_layout = QVBoxLayout(num_box)
            for q_no in expected.get("NUMERIC", []):
                row = QHBoxLayout()
                row.addWidget(QLabel(f"Câu {q_no}"))
                edit = QLineEdit(str((result.numeric_answers or {}).get(q_no, "") or ""))
                edit.editingFinished.connect(lambda q=q_no, w=edit: self._handle_numeric_visual_change(q, w.text()))
                row.addWidget(edit)
                num_layout.addLayout(row)
            self.answer_editor_layout.addWidget(num_box)
        self.answer_editor_layout.addStretch()

    def _ensure_correction_state(self) -> None:
        if not hasattr(self, "correction_ui_loading"):
            self.correction_ui_loading = False
        if not hasattr(self, "correction_pending_payload") or not isinstance(getattr(self, "correction_pending_payload", None), dict):
            self.correction_pending_payload = {}
        if not hasattr(self, "correction_save_timer") or not isinstance(getattr(self, "correction_save_timer", None), QTimer):
            self.correction_save_timer = QTimer(self)
            self.correction_save_timer.setSingleShot(True)
            self.correction_save_timer.timeout.connect(self._flush_pending_correction_updates)

    def _schedule_correction_update(self, field_name: str, old_value: object, new_value: object, apply_fn) -> None:
        self._ensure_correction_state()
        if self.correction_ui_loading or old_value == new_value:
            return
        apply_fn()
        self._refresh_all_statuses()
        self.correction_pending_payload[field_name] = {
            "old": old_value,
            "new": new_value,
        }
        self.correction_save_timer.start(150)

    def _flush_pending_correction_updates(self) -> None:
        self._ensure_correction_state()
        idx, result = self._correction_selected_result()
        if idx is None or result is None or not self.correction_pending_payload:
            return
        changes = [f"{field}: '{payload['old']}' -> '{payload['new']}'" for field, payload in self.correction_pending_payload.items()]
        self._refresh_student_profile_for_result(result, idx)
        scoped = self._scoped_result_copy(result)
        self.scan_blank_summary[idx] = self._compute_blank_questions(scoped)
        expected = self._expected_questions_by_section(scoped)
        self.scan_list.setItem(
            idx,
            self.SCAN_COL_CONTENT,
            QTableWidgetItem(self._build_recognition_content_text(result, self.scan_blank_summary[idx], expected)),
        )
        sid_item = self.scan_list.item(idx, self.SCAN_COL_STUDENT_ID)
        if sid_item:
            sid_item.setText((result.student_id or "").strip() or "-")
            sid_item.setData(Qt.UserRole + 1, result.exam_code or "")
            sid_item.setData(Qt.UserRole + 2, self._short_recognition_text_for_result(scoped))
        self._record_adjustment(idx, changes, "visual_correction")
        self._persist_single_scan_result_to_db(result, note="visual_correction")
        image_key = str(getattr(result, "image_path", "") or idx)
        for field_name, payload in self.correction_pending_payload.items():
            self.database.log_change("scan_results", image_key, field_name, payload["old"], payload["new"], "visual_correction")
        self.correction_pending_payload = {}
        self._refresh_all_statuses()
        self._update_scan_preview(idx)
        self.correction_ui_loading = True
        self._sync_correction_detail_panel(result, rebuild_editor=False)
        self.correction_ui_loading = False

    def _handle_exam_code_correction_changed(self, _index: int) -> None:
        idx, result = self._correction_selected_result()
        if idx is None or result is None:
            return
        new_code = str(self.exam_code_correction_combo.currentData() or "").strip()
        old_code = str(result.exam_code or "").strip()
        self._schedule_correction_update("exam_code", old_code, new_code, lambda: setattr(result, "exam_code", new_code))

    def _handle_student_correction_changed(self, _index: int) -> None:
        idx, result = self._correction_selected_result()
        if idx is None or result is None:
            return
        new_sid = str(self.student_correction_combo.currentData() or self.student_correction_combo.currentText() or "").strip()
        old_sid = str(result.student_id or "").strip()
        self._schedule_correction_update("student_id", old_sid, new_sid, lambda: setattr(result, "student_id", new_sid))

    def _handle_mcq_visual_change(self, question_no: int, answer_value: str) -> None:
        idx, result = self._correction_selected_result()
        if idx is None or result is None:
            return
        old_value = str((result.mcq_answers or {}).get(question_no, "") or "")[:1]

        def _apply() -> None:
            current = dict(result.mcq_answers or {})
            if answer_value:
                current[int(question_no)] = str(answer_value)[:1]
            else:
                current.pop(int(question_no), None)
            result.mcq_answers = current

        self._schedule_correction_update(f"mcq_answers[{question_no}]", old_value, str(answer_value)[:1], _apply)

    def _handle_tf_visual_change(self, question_no: int, key: str, checked: bool) -> None:
        idx, result = self._correction_selected_result()
        if idx is None or result is None:
            return
        old_flags = dict((result.true_false_answers or {}).get(question_no, {}) or {})
        new_flags = dict(old_flags)
        new_flags[key] = bool(checked)

        def _apply() -> None:
            current = dict(result.true_false_answers or {})
            current[int(question_no)] = dict(new_flags)
            result.true_false_answers = current

        self._schedule_correction_update(f"true_false_answers[{question_no}].{key}", old_flags.get(key), bool(checked), _apply)

    def _handle_numeric_visual_change(self, question_no: int, text: str) -> None:
        idx, result = self._correction_selected_result()
        if idx is None or result is None:
            return
        old_value = str((result.numeric_answers or {}).get(question_no, "") or "")
        new_value = str(text or "").strip()

        def _apply() -> None:
            current = dict(result.numeric_answers or {})
            if new_value:
                current[int(question_no)] = new_value
            else:
                current.pop(int(question_no), None)
            result.numeric_answers = current

        self._schedule_correction_update(f"numeric_answers[{question_no}]", old_value, new_value, _apply)


    def _selected_template_repository_entry(self) -> tuple[str, str] | None:
        row = self.template_library_table.currentRow() if hasattr(self, "template_library_table") else -1
        if row < 0:
            return None
        item = self.template_library_table.item(row, 1)
        if not item:
            return None
        return item.text(), str(item.data(Qt.UserRole) or "")

    def _register_single_template(self, template_path: str, display_name: str | None = None) -> None:
        self.template_repo.register(template_path, display_name=display_name)
        self._save_template_repository()
        self._refresh_template_library()

    def _open_embedded_template_editor(self, template_path: str = "") -> bool:
        if self.template_editor_embedded and not self._close_embedded_template_editor():
            return False
        while self.template_editor_layout.count():
            item = self.template_editor_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        editor = TemplateEditorWindow(self, on_template_saved=lambda path, name: self._handle_template_saved(path, name))
        editor.setWindowFlags(Qt.Widget)
        editor.menuBar().setVisible(False)
        close_action = editor.template_toolbar.addAction(self.style().standardIcon(QStyle.SP_DialogCloseButton), "Close", self._close_template_module)
        close_action.setToolTip("Close")
        self.template_editor_embedded = editor
        self.template_editor_layout.addWidget(editor)
        self.template_editor_mode = "editor"
        self.stack.setCurrentIndex(4)
        if template_path:
            return bool(editor.load_template_from_path(template_path))
        return True

    def _handle_template_saved(self, template_path: str, display_name: str) -> None:
        self._register_single_template(template_path, display_name=display_name)

    def _create_new_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Chọn ảnh mẫu giấy thi", "", "Images (*.png *.jpg *.jpeg *.tif *.tiff)")
        if not path:
            return
        if not self._open_embedded_template_editor():
            return
        if not self.template_editor_embedded or not self.template_editor_embedded.load_image_from_path(path):
            return
        self.stack.setCurrentIndex(4)

    def _edit_selected_template(self) -> None:
        selected = self._selected_template_repository_entry()
        if not selected:
            return
        _, template_path = selected
        if not template_path:
            return
        if not self._open_embedded_template_editor(template_path):
            return
        self.stack.setCurrentIndex(4)

    def _delete_selected_template(self) -> None:
        selected = self._selected_template_repository_entry()
        if not selected:
            return
        name, template_path = selected
        if not self._confirm("Xoá mẫu giấy thi", f"Bạn có chắc muốn xoá mẫu giấy thi '{name}' khỏi kho?"):
            return
        self.template_repo.templates.pop(name, None)
        self._save_template_repository()
        if template_path and Path(template_path).exists():
            try:
                Path(template_path).unlink()
            except Exception:
                pass
        self._refresh_template_library()

    def _save_current_template(self) -> bool:
        if not self.template_editor_embedded:
            return False
        return bool(self.template_editor_embedded.save_template())

    def _save_current_template_as(self) -> bool:
        if not self.template_editor_embedded:
            return False
        return bool(self.template_editor_embedded.save_template_as())

    def _close_embedded_template_editor(self) -> bool:
        if self.template_editor_embedded and not self.template_editor_embedded._confirm_close():
            return False
        while self.template_editor_layout.count():
            item = self.template_editor_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        self.template_editor_embedded = None
        self.template_editor_mode = "library"
        self._refresh_template_library()
        self._navigate_to("template_library", context={"session_id": self.current_session_id}, push_current=False, require_confirm=False, reason="close_template_editor")
        return True

    def _close_template_module(self) -> bool:
        if self.template_editor_embedded:
            return self._close_embedded_template_editor()
        self._navigate_to("exam_list", context={}, push_current=False, require_confirm=False, reason="close_template_library")
        return True

    def _save_template_repository(self) -> None:
        try:
            self.database.set_app_state("template_repository", self.template_repo.to_dict())
        except Exception:
            pass

    def _register_templates_from_payload(self, payload: dict) -> None:
        common_template = str(payload.get("common_template", "") or "").strip()
        if common_template:
            self.template_repo.register(common_template)
        for cfg in payload.get("subject_configs", []) if isinstance(payload.get("subject_configs", []), list) else []:
            tp = str((cfg or {}).get("template_path", "") or "").strip()
            if tp:
                self.template_repo.register(tp)
        self._save_template_repository()

    def create_session(self, payload: dict | None = None) -> None:
        payload = payload or {}
        self._register_templates_from_payload(payload)
        exam_name = str(payload.get("exam_name", "Untitled Exam"))
        common_template = str(payload.get("common_template", ""))
        subject_cfgs = payload.get("subject_configs", [])
        subjects = [
            f"{str(x.get('name', '')).strip()}_{str(x.get('block', '')).strip()}"
            for x in subject_cfgs
            if str(x.get("name", "")).strip()
        ]
        if not subjects:
            subjects = ["General"]

        self.session = ExamSession(
            exam_name=exam_name,
            exam_date=str(date.today()),
            subjects=subjects,
            template_path=common_template,
            answer_key_path="",
            students=self._students_from_editor_rows(
                payload.get("students", []) if isinstance(payload.get("students", []), list) else []
            ),
            config={
                "scan_mode": payload.get("scan_mode", "Ảnh trong thư mục gốc"),
                "scan_root": payload.get("scan_root", ""),
                "student_list_path": payload.get("student_list_path", ""),
                "paper_part_count": payload.get("paper_part_count", 3),
                "subject_configs": subject_cfgs,
                "subject_catalog": self.subject_catalog,
                "block_catalog": self.block_catalog,
            },
        )
        self.scoring_phases = []
        self.scoring_results_by_subject = {}
        self._scoring_dirty_subjects = set()
        self.batch_working_state_by_subject = {}
        if common_template and Path(common_template).exists():
            try:
                self.template = Template.load_json(common_template)
            except Exception:
                self.template = None
        self.current_session_path = None
        self.current_session_id = None
        self._session_saved_signature = ""
        self.session_dirty = True
        self._refresh_session_info()
        self._refresh_batch_subject_controls()

    def load_template(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, "Load Template", "", "JSON (*.json)")
        if not file_path:
            return
        self.template = Template.load_json(file_path)
        self.template_repo.register(file_path)
        self._save_template_repository()
        if self.session:
            self.session.template_path = file_path
        self.session_dirty = True
        self._refresh_session_info()
        self._refresh_batch_subject_controls()

    def load_answer_keys(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, "Load Answer Keys", "", "JSON (*.json)")
        if not file_path:
            return
        self.answer_keys = AnswerKeyRepository.load_json(file_path)
        self.imported_exam_codes = sorted({k.split("::", 1)[1] for k in self.answer_keys.keys.keys() if "::" in k})
        if self.session:
            self.session.answer_key_path = file_path
        self.session_dirty = True
        self._refresh_session_info()

    def open_template_editor(self) -> None:
        self._refresh_template_library()
        self.template_editor_mode = "library"
        self._navigate_to("template_library", context={"session_id": self.current_session_id}, push_current=True, require_confirm=False, reason="open_template_library")

    def _subject_configs_in_session(self) -> list[dict]:
        if not self.session:
            return []
        cfg = self.session.config or {}
        raw = cfg.get("subject_configs", [])
        items = raw if isinstance(raw, list) else []
        for idx, item in enumerate(items):
            if isinstance(item, dict):
                # legacy fallback / migration: promote old logical-only configs to canonical subject_instance_key.
                self._ensure_subject_instance_key(item, idx)
        return items

    def _student_profile_by_id(self, student_id: str) -> dict:
        sid = str(student_id or "").strip()
        if not sid or not self.session:
            return {}
        normalized_sid = self._normalized_student_id_for_match(sid)
        for s in (self.session.students or []):
            candidate = str(getattr(s, "student_id", "") or "").strip()
            if candidate == sid:
                return {
                    "name": str(getattr(s, "name", "") or ""),
                    "birth_date": self._format_birth_date_mmddyyyy(str((getattr(s, "extra", {}) or {}).get("birth_date", "") or "")),
                    "class_name": str((getattr(s, "extra", {}) or {}).get("class_name", "") or ""),
                    "exam_room": str((getattr(s, "extra", {}) or {}).get("exam_room", "") or ""),
                }
            if normalized_sid and self._normalized_student_id_for_match(candidate) == normalized_sid:
                return {
                    "name": str(getattr(s, "name", "") or ""),
                    "birth_date": self._format_birth_date_mmddyyyy(str((getattr(s, "extra", {}) or {}).get("birth_date", "") or "")),
                    "class_name": str((getattr(s, "extra", {}) or {}).get("class_name", "") or ""),
                    "exam_room": str((getattr(s, "extra", {}) or {}).get("exam_room", "") or ""),
                }
        return {}

    @staticmethod
    def _format_birth_date_mmddyyyy(value: str) -> str:
        text = str(value or "").strip()
        if not text or text == "-":
            return ""
        normalized = text.replace("-", "/").replace(".", "/")
        candidates = [
            "%m/%d/%Y", "%m/%d/%y",
            "%d/%m/%Y", "%d/%m/%y",
            "%Y/%m/%d", "%Y/%d/%m",
        ]
        for fmt in candidates:
            try:
                parsed = datetime.strptime(normalized, fmt)
                return parsed.strftime("%m/%d/%Y")
            except Exception:
                pass
        chunks = [part for part in normalized.split("/") if part]
        if len(chunks) == 3 and all(part.isdigit() for part in chunks):
            nums = [int(part) for part in chunks]
            if len(chunks[0]) == 4:
                year, month, day = nums
            elif int(chunks[0]) > 12 and int(chunks[1]) <= 12:
                day, month, year = nums
            else:
                month, day, year = nums
            if year < 100:
                year += 2000 if year < 70 else 1900
            try:
                return datetime(year, month, day).strftime("%m/%d/%Y")
            except Exception:
                return text
        return text

    @staticmethod
    def _birth_date_missing(birth_date_text: str) -> bool:
        birth = str(birth_date_text or "").strip()
        return birth in {"", "-"}

    def _refresh_student_profile_for_result(self, result, row_idx: int | None = None) -> None:
        sid = str(getattr(result, "student_id", "") or "").strip()
        profile = self._student_profile_by_id(sid)
        setattr(result, "full_name", str(profile.get("name", "") or ""))
        setattr(result, "birth_date", self._format_birth_date_mmddyyyy(str(profile.get("birth_date", "") or "")))
        setattr(result, "class_name", str(profile.get("class_name", "") or ""))
        exam_room = str(self._subject_room_for_student_id(sid) or "").strip()
        if not exam_room:
            exam_room = str(profile.get("exam_room", "") or "").strip()
        setattr(result, "exam_room", exam_room)
        if row_idx is not None and 0 <= row_idx < self.scan_list.rowCount():
            self.scan_list.setItem(row_idx, self.SCAN_COL_EXAM_ROOM, QTableWidgetItem(str(getattr(result, "exam_room", "") or "-")))
            self.scan_list.setItem(row_idx, self.SCAN_COL_FULL_NAME, QTableWidgetItem(str(getattr(result, "full_name", "") or "-")))
            self.scan_list.setItem(row_idx, self.SCAN_COL_BIRTH_DATE, QTableWidgetItem(str(getattr(result, "birth_date", "") or "-")))

    @staticmethod
    def _normalized_student_id_for_match(student_id: str) -> str:
        sid = str(student_id or "").strip()
        if not sid:
            return ""
        compact = sid.replace(" ", "")
        if compact.endswith(".0"):
            prefix = compact[:-2]
            if prefix.isdigit():
                compact = prefix
        if compact.isdigit():
            compact = compact.lstrip("0") or "0"
        return compact.upper()

    @staticmethod
    def _normalized_room_for_match(room_text: str) -> str:
        room = str(room_text or "").strip().casefold()
        if room.isdigit():
            room = room.lstrip("0") or "0"
        return room

    @staticmethod
    def _normalize_template_path(path_text: str) -> str:
        t = str(path_text or "").strip()
        if not t:
            return ""
        if t.lower() in {"[dùng mẫu chung]", "[dung mau chung]", "none", "null", "-"}:
            return ""
        return t

    @staticmethod
    def _subject_imported_answer_keys_for_main(subject_cfg: dict) -> dict:
        if not isinstance(subject_cfg, dict):
            return {}
        raw = subject_cfg.get("imported_answer_keys", {})
        return dict(raw) if isinstance(raw, dict) else {}

    def _effective_subject_configs_for_batch(self) -> list[dict]:
        cfgs = self._subject_configs_in_session()
        common_template = self._normalize_template_path(str(self.session.template_path if self.session else ""))
        if cfgs:
            # Default to exam common template, but keep subject-specific template if explicitly configured.
            out_cfgs: list[dict] = []
            for cfg in cfgs:
                if not isinstance(cfg, dict):
                    continue
                item = dict(cfg)
                subject_template = self._normalize_template_path(str(item.get("template_path", "") or ""))
                item["template_path"] = subject_template or common_template
                out_cfgs.append(item)
            return out_cfgs
        # Fallback for older sessions without subject_configs.
        if not self.session:
            return []
        scan_root = str((self.session.config or {}).get("scan_root", "") or "")
        out: list[dict] = []
        for raw in (self.session.subjects or ["General"]):
            subject = str(raw)
            name, block = (subject.split("_", 1) + [""])[:2] if "_" in subject else (subject, "")
            out.append({
                "name": name,
                "block": block,
                "template_path": common_template or str(self.session.template_path or ""),
                "scan_folder": scan_root,
                "answer_key_key": subject,
                "imported_answer_keys": {},
            })
        return out

    def _resolve_subject_config_for_batch(self) -> dict | None:
        cfg = self._selected_batch_subject_config()
        if cfg:
            return cfg
        cfgs = self._effective_subject_configs_for_batch()
        if not cfgs:
            return None
        # Do not show extra chooser dialog: Batch panel already has subject combo.
        # Fallback to first configured subject when none is currently selected.
        return cfgs[0]

    def _refresh_batch_subject_controls(self) -> None:
        if not hasattr(self, "batch_subject_combo"):
            return
        previous_active_key = str(getattr(self, "active_batch_subject_key", "") or "").strip()
        current_key = str(self.batch_subject_combo.currentData() or "").strip() if self.batch_subject_combo.count() > 0 else ""
        self.batch_subject_combo.blockSignals(True)
        self.batch_subject_combo.clear()
        self.batch_subject_combo.addItem("[Chọn môn]", "")
        target_index = 0
        for idx, cfg in enumerate(self._effective_subject_configs_for_batch()):
            if isinstance(cfg, dict):
                self._ensure_subject_instance_key(cfg, idx)
            label = self._display_subject_label(cfg)
            key = self._subject_instance_key_from_cfg(cfg)
            self.batch_subject_combo.addItem(label, key)
            if current_key and key == current_key:
                target_index = idx + 1
            elif not current_key and previous_active_key and key == previous_active_key:
                target_index = idx + 1
        self.batch_subject_combo.setCurrentIndex(target_index)
        self.batch_subject_combo.blockSignals(False)
        selected_key = str(self.batch_subject_combo.currentData() or "").strip() if self.batch_subject_combo.currentIndex() > 0 else ""
        selected_cfg = self._selected_batch_subject_config() if selected_key else None
        selected_sig = self._batch_subject_refresh_signature(selected_cfg) if isinstance(selected_cfg, dict) else ""
        has_signature_change = bool(selected_key) and bool(selected_sig) and selected_sig != str(getattr(self, "_batch_loaded_subject_signature", "") or "")
        should_load = bool(selected_key) and (
            selected_key != previous_active_key
            or self.scan_list.rowCount() <= 0
            or has_signature_change
        )
        if should_load:
            self._on_batch_subject_changed(self.batch_subject_combo.currentIndex(), force_reload=has_signature_change)
        self._handle_stack_changed(self.stack.currentIndex())

    def _selected_batch_subject_config(self) -> dict | None:
        if not hasattr(self, "batch_subject_combo"):
            return None
        idx = self.batch_subject_combo.currentIndex()
        if idx <= 0:
            return None
        key = str(self.batch_subject_combo.itemData(idx) or "").strip()
        if not key:
            return None
        return self._subject_config_by_subject_key(key)

    def _subject_section_question_counts(self, subject_key: str = "") -> dict[str, int]:
        counts = {"MCQ": 0, "TF": 0, "NUMERIC": 0}
        cfg = self._subject_config_by_subject_key(subject_key) if subject_key else (self._selected_batch_subject_config() or self._resolve_subject_config_for_batch())
        if not cfg:
            return counts
        raw_counts = cfg.get("question_counts", {}) if isinstance(cfg.get("question_counts", {}), dict) else {}
        for sec in counts:
            try:
                counts[sec] = max(0, int(raw_counts.get(sec, 0) or 0))
            except Exception:
                counts[sec] = 0
        if any(counts.values()):
            return counts
        template_path = self._normalize_template_path(str(cfg.get("template_path", "") or ""))
        if template_path:
            try:
                return SubjectConfigDialog._template_question_counts(template_path)
            except Exception:
                return counts
        return counts

    def _apply_subject_section_visibility(self, subject_key: str = "") -> None:
        if not hasattr(self, "score_preview_table"):
            return
        counts = self._subject_section_question_counts(subject_key)
        visibility = {
            5: counts.get("MCQ", 0) > 0,
            6: counts.get("TF", 0) > 0,
            7: counts.get("NUMERIC", 0) > 0,
        }
        for col, is_visible in visibility.items():
            self.score_preview_table.setColumnHidden(col, not is_visible)

    def _batch_context_session_path(self) -> Path | None:
        if self.current_session_id:
            return self._session_path_from_id(self.current_session_id)
        if self.current_session_path:
            return self.current_session_path
        if self.batch_editor_return_session_id:
            return self._session_path_from_id(self.batch_editor_return_session_id)
        return None

    def _recognized_image_paths_for_subject(self, subject_key: str) -> set[str]:
        key = str(subject_key or "").strip()
        scoped_key = self._batch_result_subject_key(key)
        rows = list(self.scan_results_by_subject.get(scoped_key) or [])
        if not rows and key:
            rows = list(self._refresh_scan_results_from_db(key) or [])
        recognized = {
            str(getattr(item, "image_path", "") or "").strip()
            for item in rows
            if str(getattr(item, "image_path", "") or "").strip()
        }
        recognized |= set(self.deleted_scan_images_by_subject.get(scoped_key, set()) or set())
        return recognized

    def _mark_deleted_scan_image(self, subject_key: str, image_path: str) -> None:
        subject = str(subject_key or "").strip()
        image_key = self._result_identity_key(image_path)
        if not subject or not image_key:
            return
        scoped_key = self._batch_result_subject_key(subject)
        deleted_set = self.deleted_scan_images_by_subject.setdefault(scoped_key, set())
        deleted_set.add(image_key)

    def _unmark_deleted_scan_images(self, subject_key: str, image_paths: list[str]) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        scoped_key = self._batch_result_subject_key(subject)
        deleted_set = self.deleted_scan_images_by_subject.get(scoped_key)
        if not deleted_set:
            return
        for path in image_paths or []:
            image_key = self._result_identity_key(path)
            if image_key:
                deleted_set.discard(image_key)
    def _configured_scan_file_paths(self, cfg: dict | None) -> list[str]:
        if not cfg:
            return []
        scan_folder = str(cfg.get("scan_folder", "") or ((self.session.config or {}).get("scan_root", "") if self.session else "") or "").strip()
        if not scan_folder or scan_folder == "-":
            return []
        scan_dir = Path(scan_folder)
        if not scan_dir.exists() or not scan_dir.is_dir():
            return []
        scan_mode = str(cfg.get("scan_mode", "") or (self.session.config or {}).get("scan_mode", "") if self.session else "")
        image_exts = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}
        if self._is_subfolder_scan_mode(scan_mode):
            return [str(p) for p in sorted(scan_dir.rglob("*")) if p.is_file() and p.suffix.lower() in image_exts]
        return [str(p) for p in sorted(scan_dir.iterdir()) if p.is_file() and p.suffix.lower() in image_exts]

    def _update_batch_scan_scope_summary(self) -> None:
        if not hasattr(self, "batch_scan_state_value"):
            return
        cfg = self._selected_batch_subject_config()
        if cfg:
            cfg = self._merge_saved_batch_snapshot(cfg)
        all_paths = self._configured_scan_file_paths(cfg)
        logical_key = self._logical_subject_key_from_cfg(cfg) if cfg else ""
        instance_key = self._subject_instance_key_from_cfg(cfg) if cfg else ""
        runtime_key = self._batch_runtime_key(instance_key) if instance_key else ""
        recognized = self._recognized_image_paths_for_subject(instance_key)
        recognized_count = sum(1 for path in all_paths if path in recognized)
        pending_count = max(0, len(all_paths) - recognized_count)
        mode = str(self.batch_file_scope_combo.currentData() or "new_only") if hasattr(self, "batch_file_scope_combo") else "new_only"
        mode_label = "File mới" if mode == "new_only" else "Toàn bộ"
        self.batch_scan_state_value.setText(f"{mode_label} | Đã nhận diện: {recognized_count} | Chưa nhận diện: {pending_count}")
        self._update_batch_scan_bottom_status_text()
        if hasattr(self, "batch_context_value"):
            self.batch_context_value.setText(
                f"{self._display_subject_label(cfg)} | Logical: {logical_key or '-'} | "
                f"SubjectInstance: {instance_key or '-'} | Runtime: {runtime_key or '-'} | "
                f"Nguồn: {self._current_batch_data_source}"
            )

    def _update_batch_scan_bottom_status_text(self) -> None:
        if not hasattr(self, "batch_scan_status_bottom"):
            return
        total_rows = self.scan_list.rowCount() if hasattr(self, "scan_list") else 0
        visible_rows = total_rows
        if hasattr(self, "scan_list"):
            visible_rows = sum(1 for r in range(total_rows) if not self.scan_list.isRowHidden(r))
        file_status = str(self.batch_scan_state_value.text() if hasattr(self, "batch_scan_state_value") else "-").strip() or "-"
        error_count = duplicate_count = wrong_code_count = edited_count = 0
        if hasattr(self, "scan_list"):
            for r in range(total_rows):
                status_item = self.scan_list.item(r, self.SCAN_COL_STATUS)
                status_txt = str(status_item.text() if status_item else "").strip()
                low = status_txt.lower()
                is_edited = "đã sửa" in low
                if is_edited:
                    edited_count += 1
                    continue
                if status_txt and status_txt != "OK":
                    error_count += 1
                if "trùng sbd" in low or "duplicate" in low:
                    duplicate_count += 1
                if "mã đề" in low and ("sai" in low or "không" in low or "?" in status_txt):
                    wrong_code_count += 1
        issue_total = error_count + edited_count
        bar_text = f"Trạng thái file: {file_status} | Lọc: {visible_rows}/{total_rows}"
        if issue_total > 0:
            bar_text += (
                " | "
                f"<a href='all'><b>Tất cả</b></a> | "
                f"<a href='error' style='color:#c62828;font-weight:600'>Lỗi nhận dạng: {error_count}</a> | "
                f"<a href='duplicate' style='color:#6a1b9a;font-weight:600'>Trùng SBD: {duplicate_count}</a> | "
                f"<a href='wrong_code' style='color:#ef6c00;font-weight:600'>Sai mã đề: {wrong_code_count}</a> | "
                f"<a href='edited' style='color:#1565c0;font-weight:600'>Đã sửa: {edited_count}</a>"
            )
            self.batch_scan_status_bottom.setTextFormat(Qt.RichText)
        else:
            self.batch_scan_status_bottom.setTextFormat(Qt.PlainText)
        self.batch_scan_status_bottom.setText(bar_text)
        self.batch_scan_status_bottom.setToolTip(bar_text)

    @staticmethod
    def _recommended_batch_timeout_sec(template: Template | None) -> float:
        if template is None:
            return 3.0
        zones = list(getattr(template, "zones", []) or [])
        answer_zone_count = sum(1 for z in zones if getattr(z, "zone_type", None) in {ZoneType.MCQ_BLOCK, ZoneType.TRUE_FALSE_BLOCK, ZoneType.NUMERIC_BLOCK})
        id_zone_count = sum(1 for z in zones if getattr(z, "zone_type", None) in {ZoneType.STUDENT_ID_BLOCK, ZoneType.EXAM_CODE_BLOCK})
        other_zone_count = max(0, len(zones) - answer_zone_count - id_zone_count)
        timeout = 2.5 + (0.25 * answer_zone_count) + (0.15 * id_zone_count) + (0.10 * other_zone_count)
        return float(min(6.0, max(2.5, timeout)))

    def _merge_saved_batch_snapshot(self, cfg: dict) -> dict:
        # DB là nguồn sự thật duy nhất cho Batch Scan.
        # Chỉ giữ metadata nhẹ trong subject config; bỏ qua snapshot/cache cũ.
        merged = dict(cfg or {})
        for key in ("batch_saved_rows", "batch_saved_preview", "batch_saved_results"):
            if key in merged:
                merged[key] = []
        return merged

    def _batch_subject_refresh_signature(self, cfg: dict | None) -> str:
        if not isinstance(cfg, dict):
            return ""
        subject_key = self._subject_key_from_cfg(cfg)
        imported_codes = sorted(str(k).strip() for k in (cfg.get("imported_answer_keys", {}) or {}).keys() if str(k).strip())
        fetched_codes = sorted(str(k).strip() for k in (self._fetch_answer_keys_for_subject_scoped(subject_key, cfg) or {}).keys() if str(k).strip())
        student_rows = []
        for st in (self.session.students or []) if self.session else []:
            extra = dict(getattr(st, "extra", {}) or {})
            student_rows.append(
                (
                    str(getattr(st, "student_id", "") or "").strip(),
                    str(getattr(st, "name", "") or "").strip(),
                    str(extra.get("birth_date", "") or "").strip(),
                    str(extra.get("class_name", "") or "").strip(),
                    str(extra.get("exam_room", "") or "").strip(),
                )
            )
        signature_payload = {
            "subject_key": str(subject_key or ""),
            "template_path": self._normalize_template_path(str(cfg.get("template_path", "") or "")),
            "scan_folder": str(cfg.get("scan_folder", "") or "").strip(),
            "answer_key_key": str(cfg.get("answer_key_key", "") or "").strip(),
            "imported_codes": imported_codes,
            "fetched_codes": fetched_codes,
            "students": student_rows,
        }
        return json.dumps(signature_payload, ensure_ascii=False, sort_keys=True)
    def _on_batch_subject_changed(self, _index: int, force_reload: bool = False) -> None:
        if self._switching_batch_subject:
            return
        self._switching_batch_subject = True
        try:
            previous_runtime_key = str(getattr(self, "active_batch_subject_key", "") or "").strip()
            cfg = self._selected_batch_subject_config()
            next_runtime_key = ""
            if cfg:
                cfg = self._merge_saved_batch_snapshot(cfg)
                next_runtime_key = self._batch_runtime_key(cfg)
            if previous_runtime_key and previous_runtime_key != next_runtime_key:
                pass
            if cfg:
                self.active_batch_subject_key = next_runtime_key
            else:
                self.active_batch_subject_key = None
            self._load_batch_subject_state(cfg, source_hint="subject_changed", force_reload=force_reload)
        finally:
            self._switching_batch_subject = False

    def _load_batch_subject_state(self, subject_cfg: dict | None, source_hint: str = "", force_reload: bool = False) -> bool:
        cfg = dict(subject_cfg or {}) if isinstance(subject_cfg, dict) else {}
        if cfg:
            cfg = self._merge_saved_batch_snapshot(cfg)

        if not cfg:
            self._reset_batch_subject_ui_state()
            wait_dlg = self._open_wait_progress("Đang tải dữ liệu môn, vui lòng chờ...", "Batch Scan")
            self.batch_template_value.setText("-")
            self.batch_template_path_value = "-"
            self.batch_answer_codes_value.setText("-")
            self.batch_student_id_value.setText("-")
            self.batch_scan_folder_value.setText("-")
            self.batch_scan_state_value.setText("-")
            if hasattr(self, "batch_context_value"):
                self.batch_context_value.setText("-")
            self._current_batch_data_source = "empty"
            self._batch_loaded_runtime_key = ""
            self._batch_loaded_subject_signature = ""
            self._close_wait_progress(wait_dlg)
            return False

        subject_key = self._subject_key_from_cfg(cfg)
        runtime_key = self._batch_runtime_key(subject_key)

        # Không clear/reload lại khi vẫn là đúng môn đang mở và không có yêu cầu ép reload.
        cfg_signature = self._batch_subject_refresh_signature(cfg)

        if (
            not force_reload
            and subject_key
            and runtime_key
            and str(getattr(self, "_batch_loaded_runtime_key", "") or "").strip() == runtime_key
            and str(getattr(self, "_batch_loaded_subject_signature", "") or "").strip() == cfg_signature
            and str(getattr(self, "_current_batch_data_source", "") or "").strip() == "database"
            and bool(self.scan_results or self.scan_list.rowCount() > 0)
        ):
            fetched_codes = sorted(str(k).strip() for k in (self._fetch_answer_keys_for_subject_scoped(subject_key, cfg) or {}).keys() if str(k).strip())
            imported_codes = sorted(str(k).strip() for k in (cfg.get("imported_answer_keys", {}) or {}).keys() if str(k).strip())
            codes = ", ".join(sorted(set(imported_codes) | set(fetched_codes))) or "-"
            self.batch_answer_codes_value.setText(codes)
            self._update_batch_scan_scope_summary()
            return True

        self._reset_batch_subject_ui_state()
        wait_dlg = self._open_wait_progress("Đang tải dữ liệu môn, vui lòng chờ...", "Batch Scan")

        if subject_key and subject_key not in self._answer_keys_ready_subjects:
            self._ensure_answer_keys_for_subject(subject_key)
            self._answer_keys_ready_subjects.add(subject_key)

        template_path = self._normalize_template_path(str(cfg.get("template_path", "") or "")) or self._normalize_template_path(str(self.session.template_path if self.session else "")) or "-"
        scan_folder = str(cfg.get("scan_folder", "") or ((self.session.config or {}).get("scan_root", "") if self.session else "") or "-")
        imported_codes = sorted(str(k).strip() for k in (cfg.get("imported_answer_keys", {}) or {}).keys() if str(k).strip())
        fetched_codes = sorted(str(k).strip() for k in (self._fetch_answer_keys_for_subject_scoped(subject_key, cfg) or {}).keys() if str(k).strip())
        codes = ", ".join(sorted(set(imported_codes) | set(fetched_codes))) or "-"
        self.batch_template_path_value = template_path
        template_display = Path(template_path).stem if template_path and template_path != "-" else "-"
        self.batch_template_value.setText(template_display)
        self.batch_answer_codes_value.setText(codes)
        self.batch_scan_folder_value.setText(scan_folder)

        tp = Path(template_path) if template_path and template_path != "-" else None
        template_cache_key = str(tp.resolve()) if tp and tp.exists() else ""
        tpl_for_view = self._template_cache_by_path.get(template_cache_key) if template_cache_key else None
        if tpl_for_view is None and tp and tp.exists():
            try:
                tpl_for_view = Template.load_json(tp)
                if template_cache_key:
                    self._template_cache_by_path[template_cache_key] = tpl_for_view
            except Exception:
                tpl_for_view = self.template
        if tpl_for_view is None:
            tpl_for_view = self.template
        has_sid = "Có" if (tpl_for_view and any(z.zone_type.value == "STUDENT_ID_BLOCK" for z in tpl_for_view.zones)) else "Không"
        self.batch_student_id_value.setText(has_sid)
        if tpl_for_view:
            self._apply_template_recognition_settings(tpl_for_view)
            self.template = tpl_for_view

        if self._subject_uses_direct_score_import(cfg):
            self.scan_results = []
            self.scan_files = []
            if hasattr(self, "scan_list"):
                self.scan_list.setRowCount(0)
            self.batch_scan_state_value.setText("Môn tự luận - dùng import điểm trực tiếp")
            self._current_batch_data_source = "essay_direct_import"
            self._batch_loaded_runtime_key = runtime_key
            self._batch_loaded_subject_signature = cfg_signature
            self._clear_batch_preview_panels()
            self._update_batch_scan_scope_summary()
            self._close_wait_progress(wait_dlg)
            return True

        loaded_results: list[OMRResult] = list(self._refresh_scan_results_from_db(subject_key) or [])
        scoped_subject = self._batch_result_subject_key(subject_key)
        if loaded_results:
            self.scan_results = list(loaded_results)
            self.scan_results_by_subject[scoped_subject] = list(self.scan_results)
            self._populate_scan_grid_from_results(self.scan_results, skip_expensive_checks=False)
            source = "database"
        else:
            self.scan_results = []
            self.scan_results_by_subject[scoped_subject] = []
            source = "empty"

        self._finalize_batch_scan_display(refresh_statuses=True)
        self.btn_save_batch_subject.setEnabled(False)
        self._current_batch_data_source = source
        self._batch_loaded_runtime_key = runtime_key if source != "empty" else ""
        self._batch_loaded_subject_signature = cfg_signature if source != "empty" else ""
        self._update_batch_scan_scope_summary()
        if self.scan_results:
            self._debug_scan_result_state("restore_subject_loaded_first_row", self.scan_results[0])
        self._close_wait_progress(wait_dlg)
        return source != "empty"
    @staticmethod
    def _normalized_mcq_answer_map(raw_map: object) -> dict[int, str]:
        normalized: dict[int, str] = {}
        if not isinstance(raw_map, dict):
            return normalized
        for q_raw, ans in raw_map.items():
            try:
                q_no = int(q_raw)
            except Exception:
                continue
            normalized[q_no] = str(ans or "").strip().upper()
        return normalized

    @staticmethod
    def _normalized_tf_answer_map(raw_map: object) -> dict[int, dict[str, bool]]:
        normalized: dict[int, dict[str, bool]] = {}
        if not isinstance(raw_map, dict):
            return normalized
        for q_raw, flags_raw in raw_map.items():
            try:
                q_no = int(q_raw)
            except Exception:
                continue
            flags_src = flags_raw if isinstance(flags_raw, dict) else {}
            flags: dict[str, bool] = {}
            for key_raw, flag in flags_src.items():
                key = str(key_raw or "").strip().lower()
                if key in {"a", "b", "c", "d"}:
                    flags[key] = bool(flag)
            if flags:
                normalized[q_no] = flags
        return normalized

    @staticmethod
    def _normalized_numeric_answer_map(raw_map: object) -> dict[int, str]:
        normalized: dict[int, str] = {}
        if not isinstance(raw_map, dict):
            return normalized
        for q_raw, value in raw_map.items():
            try:
                q_no = int(q_raw)
            except Exception:
                continue
            normalized[q_no] = str(value or "").strip()
        return normalized

    def _hydrate_result_from_saved_row_payload(self, result: OMRResult | None, row_payload: dict | None) -> OMRResult | None:
        if result is None and not isinstance(row_payload, dict):
            return result
        row = dict(row_payload or {}) if isinstance(row_payload, dict) else {}
        if result is None:
            result = OMRResult(
                image_path=str(row.get("image_path", "") or ""),
                student_id=str(row.get("student_id", "") or ""),
                exam_code=str(row.get("exam_code", "") or ""),
            )

        payload_sources: list[dict] = []
        serialized_payload = dict(row.get("serialized_result", {}) or {}) if isinstance(row.get("serialized_result", {}), dict) else {}
        if serialized_payload:
            payload_sources.append(serialized_payload)
        if row:
            payload_sources.append(row)

        current_mcq = self._normalized_mcq_answer_map(getattr(result, "mcq_answers", {}) or {})
        current_tf = self._normalized_tf_answer_map(getattr(result, "true_false_answers", {}) or {})
        current_numeric = self._normalized_numeric_answer_map(getattr(result, "numeric_answers", {}) or {})

        for payload in payload_sources:
            payload_mcq = self._normalized_mcq_answer_map(payload.get("mcq_answers", {}) or {})
            if payload_mcq and (not current_mcq or len(current_mcq) < len(payload_mcq)):
                current_mcq = dict(payload_mcq)

            payload_tf = self._normalized_tf_answer_map(payload.get("true_false_answers", {}) or {})
            if payload_tf:
                merged_tf = dict(current_tf)
                for q_no, flags in payload_tf.items():
                    existing = dict(merged_tf.get(q_no, {}) or {})
                    if len(existing) < len(flags):
                        merged_tf[q_no] = dict(flags)
                    else:
                        for key, value in flags.items():
                            if key not in existing:
                                existing[key] = bool(value)
                        merged_tf[q_no] = existing
                current_tf = merged_tf

            payload_numeric = self._normalized_numeric_answer_map(payload.get("numeric_answers", {}) or {})
            if payload_numeric:
                merged_numeric = dict(current_numeric)
                for q_no, value in payload_numeric.items():
                    current_value = str(merged_numeric.get(q_no, "") or "").strip()
                    payload_value = str(value or "").strip()
                    if (q_no not in merged_numeric) or (not current_value and payload_value):
                        merged_numeric[q_no] = payload_value
                current_numeric = merged_numeric

            for attr_name in [
                "image_path",
                "student_id",
                "exam_code",
                "full_name",
                "birth_date",
                "exam_room",
                "class_name",
                "answer_string",
                "manual_content_override",
                "cached_content",
                "cached_status",
                "cached_recognized_short",
                "recognized_template_path",
                "recognized_alignment_profile",
                "last_adjustment",
            ]:
                current_value = str(getattr(result, attr_name, "") or "").strip()
                payload_value = str(payload.get(attr_name, "") or "").strip()
                if payload_value and not current_value:
                    setattr(result, attr_name, payload_value)

            current_errors = list(getattr(result, "recognition_errors", []) or [])
            payload_errors = [str(x) for x in (payload.get("recognition_errors", []) or []) if str(x or "").strip()]
            if payload_errors and not current_errors:
                setattr(result, "recognition_errors", payload_errors)
            current_edit_history = [str(x) for x in (getattr(result, "edit_history", []) or []) if str(x or "").strip()]
            payload_edit_history = [str(x) for x in (payload.get("edit_history", []) or []) if str(x or "").strip()]
            if payload_edit_history and not current_edit_history:
                setattr(result, "edit_history", payload_edit_history)
            current_manual_adjustments = [str(x) for x in (getattr(result, "manual_adjustments", []) or []) if str(x or "").strip()]
            payload_manual_adjustments = [str(x) for x in (payload.get("manual_adjustments", []) or []) if str(x or "").strip()]
            if payload_manual_adjustments and not current_manual_adjustments:
                setattr(result, "manual_adjustments", payload_manual_adjustments)
            forced_status_value = str(payload.get("cached_forced_status", payload.get("forced_status", "")) or "").strip()
            current_forced_status = str(getattr(result, "cached_forced_status", "") or "").strip()
            if forced_status_value and not current_forced_status:
                setattr(result, "cached_forced_status", forced_status_value)
            if bool(payload.get("manually_edited", False)) or bool(payload_edit_history) or forced_status_value == "Đã sửa":
                setattr(result, "manually_edited", True)
                if not str(getattr(result, "cached_forced_status", "") or "").strip():
                    setattr(result, "cached_forced_status", "Đã sửa")
                if not str(getattr(result, "cached_status", "") or "").strip():
                    setattr(result, "cached_status", "Đã sửa")

            for attr_name, default_value in [
                ("recognized_fill_threshold", 0.45),
                ("recognized_empty_threshold", 0.20),
                ("recognized_certainty_margin", 0.08),
            ]:
                current_value = getattr(result, attr_name, None)
                payload_value = payload.get(attr_name, None)
                if payload_value is not None and (current_value is None or float(current_value or default_value) == float(default_value)):
                    try:
                        setattr(result, attr_name, float(payload_value))
                    except Exception:
                        pass

            cached_blank_summary = dict(payload.get("cached_blank_summary", {}) or {}) if isinstance(payload.get("cached_blank_summary", {}), dict) else {}
            if cached_blank_summary and not dict(getattr(result, "cached_blank_summary", {}) or {}):
                setattr(result, "cached_blank_summary", cached_blank_summary)

        if current_mcq:
            result.mcq_answers = current_mcq
        if current_tf:
            result.true_false_answers = current_tf
        if current_numeric:
            result.numeric_answers = current_numeric
        result.sync_legacy_aliases()
        return result

    def _cache_working_batch_state(self, subject_key: str) -> None:
        # DB-only mode: không giữ working cache theo môn.
        return
    def _restore_cached_working_batch_state(self, subject_key: str) -> bool:
        # DB-only mode: không phục hồi từ cache tạm.
        return False
    def _batch_runtime_key(self, subject_key_or_cfg) -> str:
        # canonical runtime key for batch UI caches/scans.
        if isinstance(subject_key_or_cfg, dict):
            subject_key = self._subject_instance_key_from_cfg(subject_key_or_cfg)
        else:
            subject_key = str(subject_key_or_cfg or "").strip()
        return self._normalize_subject_runtime_key(subject_key)

    def _normalize_subject_runtime_key(self, key: str) -> str:
        key_text = str(key or "").strip()
        if not key_text:
            return ""
        scope_prefix = self._session_scope_prefix()
        if scope_prefix and key_text.startswith(f"{scope_prefix}::"):
            return key_text
        return self._batch_result_subject_key(key_text)

    def _clear_batch_preview_panels(self) -> None:
        if hasattr(self, "scan_result_preview"):
            self.scan_result_preview.setRowCount(0)
        if hasattr(self, "result_preview"):
            self.result_preview.clear()
        if hasattr(self, "manual_edit"):
            self.manual_edit.clear()
        if hasattr(self, "scan_image_preview"):
            self.scan_image_preview.clear()
            self.scan_image_preview.setText("Chọn bài thi ở danh sách bên trái")
            if hasattr(self.scan_image_preview, "clear_markers"):
                self.scan_image_preview.clear_markers()

    def _reset_batch_subject_ui_state(self) -> None:
        # reset stale subject UI
        self.scan_results = []
        self.scan_files = []
        self.scan_blank_questions.clear()
        self.scan_blank_summary.clear()
        self.scan_manual_adjustments.clear()
        self.scan_edit_history.clear()
        self.scan_last_adjustment.clear()
        self.scan_forced_status_by_index.clear()
        self.preview_rotation_by_index.clear()
        self._batch_loaded_subject_signature = ""
        self.batch_status_filter_mode = "all"
        if hasattr(self, "scan_list"):
            self.scan_list.clearSelection()
            self.scan_list.setRowCount(0)
        self._clear_batch_preview_panels()
        if hasattr(self, "error_list"):
            self.error_list.clear()
        if hasattr(self, "progress"):
            self.progress.setValue(0)
        self.preview_source_pixmap = QPixmap()
        self.preview_zoom_factor = 0.3
        if hasattr(self, "btn_zoom_reset"):
            self.btn_zoom_reset.setText("30%")
        if hasattr(self, "btn_save_batch_subject"):
            self.btn_save_batch_subject.setEnabled(False)
        self._update_batch_scan_bottom_status_text()

    @staticmethod
    def _has_valid_identity(result) -> bool:
        sid = str(getattr(result, "student_id", "") or "").strip()
        code = str(getattr(result, "exam_code", "") or "").strip()
        has_id = bool(sid and "?" not in sid)
        has_code = bool(code and "?" not in code)
        return has_id or has_code

    @staticmethod
    def _result_has_meaningful_recognition(result) -> bool:
        has_identity = MainWindow._has_valid_identity(result)
        has_answers = bool((result.mcq_answers or {}) or (result.true_false_answers or {}) or (result.numeric_answers or {}))
        return has_answers or has_identity

    @staticmethod
    def _should_force_image_error_status(result) -> bool:
        issues = list(getattr(result, "issues", []) or [])
        if any(str(getattr(issue, "code", "") or "").upper() == "FILE" for issue in issues):
            return True
        return not MainWindow._result_has_meaningful_recognition(result)

    @staticmethod
    def _preferred_forced_status(result) -> str:
        issues = [str(getattr(issue, "code", "") or "").strip().upper() for issue in (getattr(result, "issues", []) or [])]
        if "FILE" in issues:
            return "Lỗi file ảnh"
        if "POOR_IDENTIFIER_ZONE" in issues:
            return "Không đủ chất lượng"
        if "POOR_IMAGE" in issues or "FAST_FAIL_POOR_SCAN" in issues:
            return "Ảnh xấu"
        if "IDENTIFIER_FAST_CAP" in issues or "STUDENT_ID_FAST_FAIL" in issues:
            return "Giới hạn SBD"
        if "SCANNER_LOCK_FAIL" in issues:
            return "Scanner lock fail"
        if "SAFE_FALLBACK_USED" in issues:
            return "Safe fallback used"
        if "IDENTIFIER_TIMEOUT" in issues:
            return "Timeout vùng SBD"
        if "TIMEOUT" in issues:
            return "Timeout nhận dạng"
        if not MainWindow._result_has_meaningful_recognition(result):
            return "Lỗi file ảnh"
        return ""

    @staticmethod
    def _result_is_poor_image(result) -> bool:
        issues = [str(getattr(issue, "code", "") or "").strip().upper() for issue in (getattr(result, "issues", []) or [])]
        if "POOR_IMAGE" in issues or "FAST_FAIL_POOR_SCAN" in issues:
            return True
        alignment_debug = dict(getattr(result, "alignment_debug", {}) or {})
        return bool(alignment_debug.get("poor_image", False))

    @staticmethod
    def _recognition_quality_score(result) -> int:
        sid = str(getattr(result, "student_id", "") or "").strip()
        code = str(getattr(result, "exam_code", "") or "").strip()
        has_id = 1 if sid and "?" not in sid else 0
        has_code = 1 if code and "?" not in code else 0
        answers_count = len(result.mcq_answers or {}) + len(result.true_false_answers or {}) + len(result.numeric_answers or {})
        penalty = len(getattr(result, "issues", []) or []) + len(getattr(result, "recognition_errors", []) or getattr(result, "errors", []) or [])
        return has_id * 3 + has_code * 3 + answers_count - penalty

    def _apply_template_recognition_settings(self, template: Template, *, sync_mode_selector: bool = True) -> None:
        if not template:
            return
        md = template.metadata if isinstance(template.metadata, dict) else {}

        mode = str(md.get("alignment_profile", "") or "").strip().lower()
        if sync_mode_selector and mode in {"auto", "legacy", "border", "hybrid", "one_side"}:
            setattr(self.omr_processor, "alignment_profile", mode)
            if hasattr(self, "batch_recognition_mode_combo"):
                for i in range(self.batch_recognition_mode_combo.count()):
                    if str(self.batch_recognition_mode_combo.itemData(i) or "") == mode:
                        self.batch_recognition_mode_combo.blockSignals(True)
                        self.batch_recognition_mode_combo.setCurrentIndex(i)
                        self.batch_recognition_mode_combo.blockSignals(False)
                        break

        # Optional recognition thresholds can be embedded in template metadata.
        for field, default in (("fill_threshold", 0.45), ("empty_threshold", 0.20), ("certainty_margin", 0.08)):
            raw = md.get(field, None)
            if raw is None:
                continue
            try:
                value = float(raw)
            except Exception:
                continue
            setattr(self.omr_processor, field, value if value >= 0 else default)

    def _allow_batch_auto_rotate_retry(self) -> bool:
        template_md = (self.template.metadata if self.template and isinstance(self.template.metadata, dict) else {})
        raw = template_md.get("batch_auto_rotate_retry", False)
        if isinstance(raw, bool):
            return raw
        text = str(raw or "").strip().lower()
        return text in {"1", "true", "yes", "on"}

    def _skip_retry_for_poor_images(self) -> bool:
        template_md = (self.template.metadata if self.template and isinstance(self.template.metadata, dict) else {})
        raw = template_md.get("batch_skip_retry_for_poor_images", True)
        if isinstance(raw, bool):
            return raw
        return str(raw or "").strip().lower() in {"1", "true", "yes", "on"}

    def _recognize_single_image(self, image_path: str, *, allow_retry: bool = False, context_tag: str = ""):
        cfg = self._selected_batch_subject_config() or self._resolve_subject_config_for_batch()
        template_path = self._resolve_template_path_for_subject(cfg)
        return self._recognize_image_with_exact_template(
            image_path,
            template_path,
            source_tag=context_tag or "single",
            allow_retry=allow_retry,
        )

    def _try_reprocess_result_rotated_180(self, result, template_path: str = "", source_tag: str = ""):
        image_path = str(getattr(result, "image_path", "") or "").strip()
        if not image_path or not Path(image_path).exists():
            return result, False
        pix = QPixmap(image_path)
        if pix.isNull():
            return result, False
        rotated = pix.transformed(QTransform().rotate(180.0), Qt.SmoothTransformation)
        temp_path = str(Path(image_path).with_name(f".{Path(image_path).stem}_tmp_auto180.png"))
        if not rotated.save(temp_path):
            return result, False
        try:
            chosen_template = template_path or self._resolve_template_path_for_subject(self._selected_batch_subject_config() or self._resolve_subject_config_for_batch())
            alt = self._recognize_image_with_exact_template(temp_path, chosen_template, source_tag=source_tag or "retry_180", allow_retry=False)
        finally:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        alt.image_path = image_path
        if self._recognition_quality_score(alt) > self._recognition_quality_score(result):
            return alt, True
        return result, False

    def _resolve_template_path_for_subject(self, subject_cfg: dict | None = None) -> str:
        cfg = subject_cfg if isinstance(subject_cfg, dict) else (self._selected_batch_subject_config() or self._resolve_subject_config_for_batch() or {})
        template_path = self._normalize_template_path(str((cfg or {}).get("template_path", "") or ""))
        if not template_path and self.session:
            template_path = self._normalize_template_path(str(self.session.template_path or ""))
        return str(template_path or "").strip()

    def _recognize_image_with_exact_template(self, image_path: str, template_path: str, source_tag: str = "", allow_retry: bool = False):
        path_text = str(template_path or "").strip()
        if not path_text:
            raise RuntimeError("Chưa cấu hình template cho môn hiện tại.")
        pth = Path(path_text)
        if not pth.exists():
            raise RuntimeError(f"Không tìm thấy template: {path_text}")
        loaded_template = Template.load_json(pth)
        self.template = loaded_template
        setattr(self, "_active_template_path", str(pth.resolve()))
        self._apply_template_recognition_settings(self.template, sync_mode_selector=False)
        # Keep batch recognition aligned with Template Editor's "Test Recognition" path:
        # run_recognition_test(..., fast_production_test=True) keeps fast production behavior
        # but also forces identifier recognition to reduce SID/ExamCode drift between screens.
        result = self.omr_processor.run_recognition_test(
            image_path,
            self.template,
            RecognitionContext(collect_diagnostics=False),
            fast_production_test=True,
            debug_deep=False,
        )
        result.sync_legacy_aliases()
        if allow_retry:
            retried, improved = self._try_reprocess_result_rotated_180(result, template_path=path_text, source_tag=f"{source_tag}_retry180")
            if improved:
                result = retried
        result.sync_legacy_aliases()
        subject_key = self._current_batch_subject_key() or self._resolve_preferred_scoring_subject()
        is_api_flow = "api" in str(source_tag or "").lower()
        setattr(result, "answer_string_api_mode", bool(is_api_flow))
        scoped = self._scoped_result_copy(result)
        result.answer_string = self._build_answer_string_for_result(scoped, subject_key)
        blank_map = self._compute_blank_questions(scoped)
        expected_by_section = self._expected_questions_by_section(scoped)
        cached_status = ", ".join(self._status_parts_for_result(result, 1)) or "OK"
        setattr(result, "cached_status", cached_status)
        setattr(result, "cached_content", self._build_recognition_content_text(scoped, blank_map, expected_by_section))
        setattr(result, "cached_recognized_short", self._short_recognition_text_for_result(scoped))
        if source_tag:
            setattr(result, "cached_forced_status", str(source_tag))
        setattr(result, "manual_content_override", "")
        md = self.template.metadata if isinstance(self.template.metadata, dict) else {}
        setattr(result, "recognized_template_path", path_text)
        setattr(result, "recognized_alignment_profile", str(getattr(self.omr_processor, "alignment_profile", md.get("alignment_profile", "")) or ""))
        setattr(result, "recognized_fill_threshold", float(getattr(self.omr_processor, "fill_threshold", md.get("fill_threshold", 0.45)) or 0.45))
        setattr(result, "recognized_empty_threshold", float(getattr(self.omr_processor, "empty_threshold", md.get("empty_threshold", 0.20)) or 0.20))
        setattr(result, "recognized_certainty_margin", float(getattr(self.omr_processor, "certainty_margin", md.get("certainty_margin", 0.08)) or 0.08))
        return result

    def run_batch_scan(self, auto_triggered: bool = False) -> None:
        if self._batch_scan_running:
            confirm_cancel = QMessageBox.question(
                self,
                "Huỷ Batch Scan",
                "Batch Scan đang chạy. Bạn có muốn huỷ không?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if confirm_cancel == QMessageBox.Yes:
                self._batch_cancel_requested = True
            return
        self._batch_cancel_requested = False

        subject_cfg = self._selected_batch_subject_config() or self._resolve_subject_config_for_batch()
        if subject_cfg:
            subject_cfg = self._merge_saved_batch_snapshot(subject_cfg)
        elif self.session:
            cfgs = self._effective_subject_configs_for_batch()
            if cfgs:
                subject_cfg = self._merge_saved_batch_snapshot(cfgs[0])
        if self.session and not subject_cfg:
            QMessageBox.warning(self, "Batch Scan", "Không có môn nào để nhận dạng trong kỳ thi hiện tại.")
            return

        if subject_cfg and self._subject_uses_direct_score_import(subject_cfg):
            QMessageBox.information(
                self,
                "Batch Scan",
                "Môn tự luận không đi qua nhận dạng OMR. Hãy import điểm trực tiếp theo SBD trong cấu hình môn, sau đó mở Tính điểm / Export.",
            )
            return

        if hasattr(self, "batch_recognition_mode_combo"):
            mode = str(self.batch_recognition_mode_combo.currentData() or "auto")
            setattr(self.omr_processor, "alignment_profile", mode)
        file_scope_mode = str(self.batch_file_scope_combo.currentData() or "new_only") if hasattr(self, "batch_file_scope_combo") else "new_only"
        subject_key_for_reset = self._subject_key_from_cfg(subject_cfg) if isinstance(subject_cfg, dict) else ""
        if file_scope_mode == "all" and subject_key_for_reset:
            has_scoring = bool(self.scoring_results_by_subject.get(subject_key_for_reset, {}))
            try:
                has_scoring = has_scoring or bool(self.database.fetch_scores_for_subject(subject_key_for_reset))
            except Exception:
                pass
            if has_scoring:
                confirm_reset = QMessageBox.question(
                    self,
                    "Nhận dạng toàn bộ",
                    "Môn này đã có điểm. Nhận dạng toàn bộ sẽ xoá toàn bộ điểm đã tính trước đó và chấm lại từ đầu.\n\nBạn có muốn tiếp tục không?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No,
                )
                if confirm_reset != QMessageBox.Yes:
                    return
                self._persist_scoring_results_for_subject(
                    subject_key_for_reset,
                    [],
                    mode="reset_by_batch_scan_all",
                    note="reset_before_batch_scan_all",
                    mark_saved=False,
                )
                self.scoring_phases = [p for p in (self.scoring_phases or []) if str((p or {}).get("subject", "") or "").strip() != subject_key_for_reset]
                if self.session:
                    self.session_dirty = True
                    self._persist_session_quietly()
                if hasattr(self, "score_preview_table"):
                    self.score_preview_table.setRowCount(0)
                self._update_scoring_status_bar(0, 0, 0, 0)
        api_file = str(getattr(self, "batch_api_file_value", QLineEdit("-")).text() if hasattr(self, "batch_api_file_value") else "").strip()
        if api_file and api_file != "-":
            self._run_batch_scan_from_api_file(subject_cfg or {}, file_scope_mode, api_file)
            return

        # Resolve template, scan folder and answer keys from selected subject config in session.
        subject_template_path = ""
        exam_template_path = self._normalize_template_path(str(self.session.template_path if self.session else ""))
        scan_folder = ""
        answer_key_key = None
        if subject_cfg:
            subject_template_path = self._normalize_template_path(str(subject_cfg.get("template_path", "") or ""))
            scan_folder = str(subject_cfg.get("scan_folder", "") or "")
            answer_key_key = str(subject_cfg.get("answer_key_key", "") or "")
            if not scan_folder and self.session:
                scan_folder = str((self.session.config or {}).get("scan_root", "") or "")

            imported_answer_keys_map = subject_cfg.get("imported_answer_keys", {}) if isinstance(subject_cfg.get("imported_answer_keys", {}), dict) else {}
            if answer_key_key and imported_answer_keys_map:
                repo = AnswerKeyRepository()
                for exam_code, kd in imported_answer_keys_map.items():
                    repo.upsert(SubjectKey(
                        subject=answer_key_key,
                        exam_code=str(exam_code),
                        answers={int(k): str(v) for k, v in (kd.get("mcq_answers", {}) or {}).items()},
                        true_false_answers={int(k): dict(v) for k, v in (kd.get("true_false_answers", {}) or {}).items()},
                        numeric_answers={int(k): str(v) for k, v in (kd.get("numeric_answers", {}) or {}).items()},
                        full_credit_questions={
                            str(sec): [int(x) for x in (vals or []) if str(x).strip().lstrip("-").isdigit()]
                            for sec, vals in (kd.get("full_credit_questions", {}) or {}).items()
                        },
                        invalid_answer_rows={
                            str(sec): {int(k): str(v) for k, v in (vals or {}).items()}
                            for sec, vals in (kd.get("invalid_answer_rows", {}) or {}).items()
                        },
                    ))
                self.answer_keys = repo
                self.imported_exam_codes = sorted(str(k) for k in imported_answer_keys_map.keys())
                self.active_batch_subject_key = answer_key_key
            else:
                # Fallback: load answer keys from subject path or exam/session path if available.
                answer_key_path = str(subject_cfg.get("answer_key_path", "") or (self.session.answer_key_path if self.session else "") or "")
                self.active_batch_subject_key = answer_key_key or self.active_batch_subject_key
                if answer_key_path:
                    pth = Path(answer_key_path)
                    if pth.exists() and pth.suffix.lower() == ".json":
                        try:
                            self.answer_keys = AnswerKeyRepository.load_json(pth)
                            all_codes: set[str] = set()
                            for key_name in self.answer_keys.keys.keys():
                                parts = str(key_name).split("::", 1)
                                if len(parts) == 2 and (not answer_key_key or parts[0] == answer_key_key):
                                    all_codes.add(parts[1])
                            self.imported_exam_codes = sorted(all_codes)
                        except Exception:
                            pass

        if not subject_template_path:
            subject_template_path = self._normalize_template_path(str(getattr(self, "batch_template_path_value", "") or ""))
        if not subject_template_path and hasattr(self, "batch_template_value"):
            subject_template_path = self._normalize_template_path(self.batch_template_value.text().strip())
        if (not scan_folder or scan_folder == "-") and hasattr(self, "batch_scan_folder_value"):
            candidate_folder = str(self.batch_scan_folder_value.text() or "").strip()
            if candidate_folder and candidate_folder != "-":
                scan_folder = candidate_folder

        template_path = subject_template_path or exam_template_path
        if not template_path:
            QMessageBox.warning(self, "Batch Scan", "Chưa cấu hình mẫu giấy để nhận dạng.")
            return
        template_file = Path(template_path)
        if not template_file.exists():
            QMessageBox.warning(self, "Batch Scan", f"Không tìm thấy mẫu giấy\n{template_path}")
            return
        try:
            self.template = Template.load_json(template_file)
        except Exception as exc:
            QMessageBox.warning(self, "Batch Scan", f"Không thể tải mẫu giấy\n{exc}")
            return
        self.template.metadata["recognition_timeout_sec"] = self._recommended_batch_timeout_sec(self.template)

        scan_folder = str(scan_folder or "").strip()
        if not scan_folder:
            QMessageBox.warning(self, "Batch Scan", "Chưa chọn thư mục quét.")
            return
        scan_dir = Path(scan_folder)
        if not scan_dir.exists() or not scan_dir.is_dir():
            QMessageBox.warning(self, "Batch Scan", f"Không tìm thấy thư mục quét\n{scan_folder}")
            return

        file_paths = self._configured_scan_file_paths(subject_cfg or {})

        if not file_paths:
            QMessageBox.warning(self, "Batch Scan", "Không tìm thấy ảnh bài thi trong thư mục quét.")
            return

        subject_key_for_results = self._subject_key_from_cfg(subject_cfg) if subject_cfg else self._resolve_preferred_scoring_subject()
        subject_db_key = self._batch_result_subject_key(subject_key_for_results)
        existing_results = list(self._refresh_scan_results_from_db(subject_key_for_results) or [])
        recognized_paths = {
            str(getattr(item, "image_path", "") or "").strip()
            for item in existing_results
            if str(getattr(item, "image_path", "") or "").strip()
        }
        if file_scope_mode == "new_only":
            file_paths = [path for path in file_paths if str(path).strip() not in recognized_paths]
            if auto_triggered and subject_key_for_results:
                deleted_set = set(self.deleted_scan_images_by_subject.get(subject_db_key, set()) or set())
                if deleted_set:
                    file_paths = [path for path in file_paths if self._result_identity_key(str(path)) not in deleted_set]
            if not file_paths:
                self.scan_results = list(existing_results)
                self._populate_scan_grid_from_results(self.scan_results)
                self._update_batch_scan_scope_summary()
                QMessageBox.information(self, "Batch Scan", "Không còn file mới cần nhận dạng trong phạm vi đã cấu hình.")
                return

        self.scan_list.setRowCount(0)
        self.error_list.clear()
        self.result_preview.clear()
        self.scan_result_preview.setRowCount(0)
        self.manual_edit.clear()
        self.scan_image_preview.setText("Chọn bài thi ở danh sách bên trái")
        self.scan_image_preview.clear_markers()
        self.btn_save_batch_subject.setEnabled(False)
        self.scan_files = [Path(p) for p in file_paths]
        self.scan_blank_questions.clear()
        self.scan_blank_summary.clear()
        self.scan_manual_adjustments.clear()
        self.scan_edit_history.clear()
        self.scan_last_adjustment.clear()
        self.preview_rotation_by_index.clear()
        self.scan_forced_status_by_index.clear()

        self._apply_template_recognition_settings(self.template, sync_mode_selector=False)
        batch_started = time.perf_counter()
        batch_progress_dialog = self._open_batch_progress_screen(len(file_paths), title="Batch Scan - Đang nhận dạng")

        def on_progress(current: int, total: int, image_path: str):
            if self._batch_cancel_requested:
                raise RuntimeError("BATCH_CANCELLED")
            self._update_batch_progress_screen(batch_progress_dialog, current, total, image_path, batch_started)

        self._batch_scan_running = True
        try:
            if file_scope_mode == "all":
                self.database.delete_scan_results_for_subject(subject_db_key)
                self.scan_results = []
            else:
                self.scan_results = list(existing_results)
            duplicate_ids: dict[str, int] = {}
            for existing in self.scan_results:
                sid_existing = (existing.student_id or "").strip()
                if not self._student_id_has_recognition_error(sid_existing):
                    duplicate_ids[sid_existing] = duplicate_ids.get(sid_existing, 0) + 1

            base_count = len(self.scan_results)
            new_results: list[OMRResult] = []
            total = len(file_paths)
            for offset, image_path in enumerate(file_paths):
                on_progress(offset + 1, total, image_path)
                result = self._recognize_image_with_exact_template(
                    str(image_path),
                    template_path,
                    source_tag="batch_scan",
                    allow_retry=False,
                )
                new_results.append(result)
                idx = base_count + offset
                forced_status = ""
                original_meaningful = self._result_has_meaningful_recognition(result)
                original_identity = self._has_valid_identity(result)

                # Keep batch behavior aligned with Template Editor by default (no auto-rotation retry).
                skip_retry_on_poor = self._skip_retry_for_poor_images() and self._result_is_poor_image(result)
                need_retry_180 = self._allow_batch_auto_rotate_retry() and (not skip_retry_on_poor) and ((not original_identity) or (not original_meaningful))
                if need_retry_180:
                    retried, improved = self._try_reprocess_result_rotated_180(result, template_path=template_path, source_tag="batch_scan_retry180")
                    # Accept 180° retry only when quality is strictly improved, otherwise keep original orientation.
                    if improved:
                        result = retried
                        self.preview_rotation_by_index[idx] = (int(self.preview_rotation_by_index.get(idx, 0) or 0) + 180) % 360

                preferred_status = self._preferred_forced_status(result)
                if preferred_status:
                    # Keep raw recognition data (answers + identifiers) and only override status text.
                    forced_status = preferred_status
                elif self._should_force_image_error_status(result):
                    forced_status = "Lỗi file ảnh"

                if forced_status:
                    image_key_for_result = self._result_identity_key(getattr(result, "image_path", ""))
                    if image_key_for_result:
                        self.scan_forced_status_by_index[image_key_for_result] = forced_status

                self._refresh_student_profile_for_result(result)
                result.answer_string = self._build_answer_string_for_result(result, subject_key_for_results)
                self.scan_results.append(self._strip_transient_scan_artifacts(result))
                sid_for_dup = (result.student_id or "").strip()
                if not self._student_id_has_recognition_error(sid_for_dup):
                    duplicate_ids[sid_for_dup] = duplicate_ids.get(sid_for_dup, 0) + 1

                # Keep UI responsive while filling table after recognition reaches 100%.
                if offset % 10 == 0:
                    QApplication.processEvents()
        except RuntimeError as exc:
            if str(exc) != "BATCH_CANCELLED":
                raise
            QMessageBox.information(self, "Batch Scan", "Đã huỷ Batch Scan.")
        finally:
            self._batch_scan_running = False
            self._batch_cancel_requested = False
            self._close_batch_progress_screen(batch_progress_dialog)

        self.scan_results_by_subject[subject_db_key] = list(self.scan_results)
        self._unmark_deleted_scan_images(subject_key_for_results, [str(getattr(x, "image_path", "") or "") for x in new_results])
        self.database.replace_scan_results_for_subject(
            subject_db_key,
            [self._serialize_omr_result(x) for x in self.scan_results],
        )

        forced_status_by_image = {
            str(result.image_path or ""): str(self.scan_forced_status_by_index.get(str(result.image_path or ""), "") or "")
            for result in self.scan_results
        }
        self._populate_scan_grid_from_results(self.scan_results, forced_status_by_image)
        self._finalize_batch_scan_display()
        self.btn_save_batch_subject.setEnabled(False)
        self._update_batch_scan_scope_summary()
        elapsed_sec = max(0.0, float(time.perf_counter() - batch_started))
        total_items = len(file_paths)
        timing_text = f"{elapsed_sec:.1f}s/{total_items} bài"

    def _pick_batch_api_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Chọn file API bài thi",
            "",
            "Data files (*.csv *.txt *.tsv *.xlsx);;All files (*.*)",
        )
        if not path:
            return
        if hasattr(self, "batch_api_file_value"):
            self.batch_api_file_value.setText(path)

    @staticmethod
    def _normalize_mapping_key(text: str) -> str:
        return str(text or "").strip().lower().replace(" ", "").replace("_", "")

    def _load_api_mapping_rows(self, path: Path) -> tuple[list[str], list[dict[str, str]]]:
        ext = path.suffix.lower()
        rows: list[dict[str, str]] = []
        if ext in {".csv", ".txt", ".tsv"}:
            raw = path.read_text(encoding="utf-8-sig", errors="ignore")
            non_empty_lines = [ln for ln in raw.splitlines() if str(ln).strip()]
            first_line = non_empty_lines[0] if non_empty_lines else ""
            if raw.strip():
                force_tab = ext in {".tsv", ".txt"} and ("\t" in first_line)
                if force_tab:
                    dialect = csv.excel_tab
                else:
                    try:
                        dialect = csv.Sniffer().sniff(raw[:2048])
                    except Exception:
                        dialect = csv.excel_tab if "\t" in first_line else csv.excel
            else:
                dialect = csv.excel
            reader = csv.DictReader(raw.splitlines(), dialect=dialect)
            headers = [str(x or "").strip("\ufeff") for x in (reader.fieldnames or []) if str(x or "").strip()]
            # Fallback: if sniffer chose wrong delimiter and collapsed header into one column,
            # try tab-separated parsing explicitly.
            if len(headers) <= 1 and "\t" in first_line:
                reader = csv.DictReader(raw.splitlines(), dialect=csv.excel_tab)
                headers = [str(x or "").strip("\ufeff") for x in (reader.fieldnames or []) if str(x or "").strip()]
            for row in reader:
                rows.append({str(k).strip("\ufeff"): str(v or "") for k, v in (row or {}).items()})
            return headers, rows
        if ext == ".xlsx":
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception as exc:
                raise RuntimeError(f"Thiếu openpyxl để đọc .xlsx: {exc}")
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.values)
            if not values:
                return [], []
            headers = [str(x or "") for x in values[0]]
            for data in values[1:]:
                row = {}
                for idx, key in enumerate(headers):
                    row[str(key)] = str(data[idx] if idx < len(data) and data[idx] is not None else "")
                rows.append(row)
            return [str(h or "") for h in headers if str(h or "").strip()], rows
        raise RuntimeError("Chỉ hỗ trợ .csv/.txt/.tsv/.xlsx")

    def _expected_answer_string_length_for_subject(self, subject_key: str) -> int:
        fetched = self._fetch_answer_keys_for_subject_scoped(subject_key) or {}
        if not fetched:
            return 0
        sample = next(iter(fetched.values()))
        if isinstance(sample, SubjectKey):
            mcq_map = sample.answers if isinstance(sample.answers, dict) else {}
            tf_map = sample.true_false_answers if isinstance(sample.true_false_answers, dict) else {}
            numeric = sample.numeric_answers if isinstance(sample.numeric_answers, dict) else {}
        elif isinstance(sample, dict):
            mcq_map = sample.get("mcq_answers", sample.get("answers", {}))
            tf_map = sample.get("true_false_answers", {})
            numeric = sample.get("numeric_answers", {})
            mcq_map = mcq_map if isinstance(mcq_map, dict) else {}
            tf_map = tf_map if isinstance(tf_map, dict) else {}
            numeric = numeric if isinstance(numeric, dict) else {}
        else:
            mcq_map = {}
            tf_map = {}
            numeric = {}
        mcq_len = len(mcq_map)
        tf_len = 4 * len(tf_map)
        num_len = sum(len(str(v or "")) for v in numeric.values())
        return int(mcq_len + tf_len + num_len)

    def _answer_layout_for_subject(self, subject_key: str) -> tuple[list[int], list[int], list[tuple[int, int]]]:
        fetched = self._fetch_answer_keys_for_subject_scoped(subject_key) or {}
        if not fetched:
            return [], [], []
        sample = next(iter(fetched.values()))
        if isinstance(sample, SubjectKey):
            mcq_map = sample.answers if isinstance(sample.answers, dict) else {}
            tf_map = sample.true_false_answers if isinstance(sample.true_false_answers, dict) else {}
            numeric_map = sample.numeric_answers if isinstance(sample.numeric_answers, dict) else {}
        elif isinstance(sample, dict):
            mcq_map = sample.get("mcq_answers", sample.get("answers", {}))
            tf_map = sample.get("true_false_answers", {})
            numeric_map = sample.get("numeric_answers", {})
            mcq_map = mcq_map if isinstance(mcq_map, dict) else {}
            tf_map = tf_map if isinstance(tf_map, dict) else {}
            numeric_map = numeric_map if isinstance(numeric_map, dict) else {}
        else:
            mcq_map, tf_map, numeric_map = {}, {}, {}

        mcq_questions = sorted(set(int(q) for q in mcq_map.keys() if str(q).strip().lstrip("-").isdigit()))
        tf_questions = sorted(set(int(q) for q in tf_map.keys() if str(q).strip().lstrip("-").isdigit()))
        numeric_questions = sorted(set(int(q) for q in numeric_map.keys() if str(q).strip().lstrip("-").isdigit()))
        numeric_layout = [
            (q, len(str((numeric_map or {}).get(q, (numeric_map or {}).get(str(q), "")) or "")))
            for q in numeric_questions
        ]
        return mcq_questions, tf_questions, numeric_layout

    def _run_batch_scan_from_api_file(self, subject_cfg: dict, file_scope_mode: str, api_file: str) -> None:
        subject_key_for_results = self._subject_key_from_cfg(subject_cfg) if subject_cfg else self._resolve_preferred_scoring_subject()
        scan_folder = str((subject_cfg or {}).get("scan_folder", "") or "").strip()
        if not scan_folder:
            scan_folder = str(getattr(self, "batch_scan_folder_value", QLineEdit("-")).text() if hasattr(self, "batch_scan_folder_value") else "").strip()
        if not scan_folder or scan_folder == "-":
            QMessageBox.warning(self, "API bài thi", "Chưa cấu hình Thư mục bài thi môn.")
            return
        try:
            headers, mapping_rows = self._load_api_mapping_rows(Path(api_file))
        except Exception as exc:
            QMessageBox.warning(self, "API bài thi", f"Không đọc được file mapping:\n{exc}")
            return
        if not mapping_rows:
            QMessageBox.warning(self, "API bài thi", "File mapping không có dữ liệu.")
            return
        if not headers:
            QMessageBox.warning(self, "API bài thi", "Không tìm thấy tiêu đề cột trong file mapping.")
            return

        expected_len = self._expected_answer_string_length_for_subject(subject_key_for_results)
        pick = QDialog(self)
        pick.setWindowTitle("Chọn cột API bài thi")
        pick_lay = QVBoxLayout(pick)
        pick_form = QFormLayout()
        file_col = QComboBox(); file_col.addItems(headers)
        sid_col = QComboBox(); sid_col.addItems(["[Không dùng]"] + headers)
        exam_col = QComboBox(); exam_col.addItems(["[Không dùng]"] + headers)
        answer_col = QComboBox(); answer_col.addItems(["[Không dùng]"] + headers)

        def _find_idx(alias: set[str], combo: QComboBox) -> int:
            for i in range(combo.count()):
                if self._normalize_mapping_key(combo.itemText(i)) in alias:
                    return i
            return 0

        file_col.setCurrentIndex(_find_idx({"filename", "file", "image", "tenfile"}, file_col))
        sid_col.setCurrentIndex(_find_idx({"sdb", "studentid", "student_id", "sobaodanh"}, sid_col))
        exam_col.setCurrentIndex(_find_idx({"made", "examcode", "exam_code", "ma_de"}, exam_col))
        answer_col.setCurrentIndex(_find_idx({"bailam", "answer", "answers", "answerstring"}, answer_col))

        pick_form.addRow("Cột FileName (bắt buộc)", file_col)
        pick_form.addRow("Cột SBD", sid_col)
        pick_form.addRow("Cột mã đề", exam_col)
        pick_form.addRow("Cột bài làm", answer_col)
        pick_lay.addLayout(pick_form)
        pick_buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        pick_buttons.accepted.connect(pick.accept)
        pick_buttons.rejected.connect(pick.reject)
        pick_lay.addWidget(pick_buttons)
        if pick.exec() != QDialog.Accepted:
            return

        selected_file_col = file_col.currentText().strip()
        selected_sid_col = sid_col.currentText().strip()
        selected_exam_col = exam_col.currentText().strip()
        selected_answer_col = answer_col.currentText().strip()
        if not selected_file_col or selected_file_col == "[Không dùng]":
            QMessageBox.warning(self, "API bài thi", "Bắt buộc chọn cột FileName.")
            return

        mcq_questions, tf_questions, numeric_layout = self._answer_layout_for_subject(subject_key_for_results)
        q_counts = (subject_cfg or {}).get("question_counts", {}) if isinstance(subject_cfg or {}, dict) else {}
        if not any(int((q_counts or {}).get(k, 0) or 0) > 0 for k in ["MCQ", "TF", "NUMERIC"]):
            tpl_path = str((subject_cfg or {}).get("template", "") or "").strip()
            if tpl_path:
                q_counts = self._template_question_counts(tpl_path)
        mcq_count = max(0, int((q_counts or {}).get("MCQ", 0) or 0))
        tf_count = max(0, int((q_counts or {}).get("TF", 0) or 0))
        numeric_count = max(0, int((q_counts or {}).get("NUMERIC", 0) or 0))
        if not mcq_questions and mcq_count > 0:
            mcq_questions = [x for x in range(1, mcq_count + 1)]
        next_q = max(mcq_questions) if mcq_questions else 0
        if not tf_questions and tf_count > 0:
            tf_questions = [x for x in range(next_q + 1, next_q + tf_count + 1)]
        next_q = max(tf_questions) if tf_questions else next_q
        if not numeric_layout and numeric_count > 0:
            numeric_layout = [(x, 0) for x in range(next_q + 1, next_q + numeric_count + 1)]
        tf_true_chars = {"Đ", "đ", "D", "d", "T", "t", "1", "Y", "y"}

        configured_answer_keys = self._fetch_answer_keys_for_subject_scoped(subject_key_for_results, subject_cfg) or {}
        imported_answer_keys = self._subject_imported_answer_keys_for_main(subject_cfg or {})
        section_layout_cache: dict[str, tuple[list[int], list[int], list[tuple[int, int]], bool]] = {}

        def _section_layout_from_subject_cfg(exam_code_text: str) -> tuple[list[int], list[int], list[tuple[int, int]], bool]:
            exam_text = str(exam_code_text or "").strip()
            exam_norm = self._normalize_exam_code_text(exam_text)
            cache_key = f"{exam_text}::{exam_norm}"
            cached = section_layout_cache.get(cache_key)
            if cached is not None:
                return cached
            payload = None

            if configured_answer_keys:
                payload = configured_answer_keys.get(exam_text)
                if payload is None:
                    for k, v in configured_answer_keys.items():
                        if self._normalize_exam_code_text(str(k or "")) == exam_norm:
                            payload = v
                            break

            if payload is None:
                payload = imported_answer_keys.get(exam_text) if isinstance(imported_answer_keys, dict) else None
                if payload is None and isinstance(imported_answer_keys, dict):
                    for k, v in imported_answer_keys.items():
                        if self._normalize_exam_code_text(str(k or "")) == exam_norm:
                            payload = v
                            break

            if isinstance(payload, SubjectKey):
                mcq_payload = payload.answers if isinstance(payload.answers, dict) else {}
                tf_payload = payload.true_false_answers if isinstance(payload.true_false_answers, dict) else {}
                numeric_payload = payload.numeric_answers if isinstance(payload.numeric_answers, dict) else {}
            elif isinstance(payload, dict):
                mcq_payload = payload.get("mcq_answers", {}) if isinstance(payload.get("mcq_answers", {}), dict) else {}
                tf_payload = payload.get("true_false_answers", {}) if isinstance(payload.get("true_false_answers", {}), dict) else {}
                numeric_payload = payload.get("numeric_answers", {}) if isinstance(payload.get("numeric_answers", {}), dict) else {}
            else:
                return [], [], [], False

            mcq_q: list[int] = []
            tf_q: list[int] = []
            numeric_layout: list[tuple[int, int]] = []
            for q_raw in mcq_payload.keys():
                try:
                    mcq_q.append(int(q_raw))
                except Exception:
                    pass
            for q_raw in tf_payload.keys():
                try:
                    tf_q.append(int(q_raw))
                except Exception:
                    pass
            for q_raw, ans in numeric_payload.items():
                try:
                    numeric_layout.append((int(q_raw), len(str(ans or ""))))
                except Exception:
                    pass
            parsed = (sorted(set(mcq_q)), sorted(set(tf_q)), sorted(numeric_layout, key=lambda x: int(x[0])), True)
            section_layout_cache[cache_key] = parsed
            return parsed

        def _parse_answer_string(
            raw_answer: str,
            row_mcq_questions: list[int],
            row_tf_questions: list[int],
            row_numeric_layout: list[tuple[int, int]],
        ) -> tuple[dict[int, str], dict[int, dict[str, bool]], dict[int, str], str]:
            raw_text = str(raw_answer or "").strip()
            compact = re.sub(r"\s+", "", raw_text)
            mcq_span = max(0, len(row_mcq_questions))
            tf_span = max(0, len(row_tf_questions) * 4)
            mcq_source = compact[:mcq_span]
            tf_source = compact[mcq_span:mcq_span + tf_span]
            numeric_tail = compact[mcq_span + tf_span:]

            numeric_map: dict[int, str] = {}
            has_fixed_numeric_width = bool(row_numeric_layout) and all(int(expected_len) > 0 for _, expected_len in row_numeric_layout)
            if has_fixed_numeric_width:
                compact_numeric = str(numeric_tail)
                pos = 0
                for q_no, expected_len in row_numeric_layout:
                    token = compact_numeric[pos:pos + int(expected_len)] if pos < len(compact_numeric) else ""
                    numeric_map[int(q_no)] = token.replace("_", "")
                    pos += int(expected_len)
            else:
                # Do not split by comma because decimal answers commonly use comma
                # (e.g. 0,61 / 58,3). Use explicit field separators only.
                numeric_tokens = [tok.strip() for tok in re.split(r"[;|]+", numeric_tail) if tok and tok.strip()]
                if len(numeric_tokens) >= len(row_numeric_layout) and row_numeric_layout:
                    for idx_layout, (q_no, expected_len) in enumerate(row_numeric_layout):
                        token = str(numeric_tokens[idx_layout]) if idx_layout < len(numeric_tokens) else ""
                        if int(expected_len) > 0:
                            numeric_map[int(q_no)] = token[: max(0, int(expected_len))].replace("_", "")
                        else:
                            numeric_map[int(q_no)] = token.replace("_", "")

            mcq_map: dict[int, str] = {}
            for idx_q, q_no in enumerate(row_mcq_questions):
                raw_ch = str(mcq_source[idx_q]).upper() if idx_q < len(mcq_source) else ""
                mcq_map[int(q_no)] = raw_ch if raw_ch and raw_ch != "_" else ""

            tf_map: dict[int, dict[str, bool]] = {}
            for idx_q, q_no in enumerate(row_tf_questions):
                base = idx_q * 4
                tf_row: dict[str, bool] = {}
                for idx_stmt, key in enumerate(["a", "b", "c", "d"]):
                    ch = tf_source[base + idx_stmt] if (base + idx_stmt) < len(tf_source) else ""
                    if ch in tf_true_chars:
                        tf_row[key] = True
                    elif ch in {"S", "s", "0", "N", "n"}:
                        tf_row[key] = False
                tf_map[int(q_no)] = tf_row

            rebuilt_parts: list[str] = []
            for q_no in row_mcq_questions:
                rebuilt_parts.append(str((mcq_map or {}).get(int(q_no), "") or "_")[:1])
            for q_no in row_tf_questions:
                flags = (tf_map or {}).get(int(q_no), {}) or {}
                for key in ["a", "b", "c", "d"]:
                    if key in flags:
                        rebuilt_parts.append("Đ" if bool(flags.get(key)) else "S")
                    else:
                        rebuilt_parts.append("_")
            for q_no, expected_len in row_numeric_layout:
                raw_val = str((numeric_map or {}).get(int(q_no), "") or "")
                if int(expected_len) > 0:
                    token = raw_val[: int(expected_len)]
                    if len(token) < int(expected_len):
                        token = token + ("_" * (int(expected_len) - len(token)))
                    rebuilt_parts.append(token)
                else:
                    rebuilt_parts.append(raw_val)
            rebuilt = "".join(rebuilt_parts)
            return mcq_map, tf_map, numeric_map, rebuilt

        out: list[OMRResult] = []
        for row in mapping_rows:
            def _pick(col_name: str) -> str:
                if not col_name or col_name == "[Không dùng]":
                    return ""
                return str(row.get(col_name, "") or "").strip()

            fname = _pick(selected_file_col)
            if not fname:
                continue
            result = OMRResult(image_path=str(Path(scan_folder) / fname))
            result.student_id = _pick(selected_sid_col)
            result.exam_code = _pick(selected_exam_col)
            raw_answer = _pick(selected_answer_col)
            row_mcq_questions = list(mcq_questions)
            row_tf_questions = list(tf_questions)
            row_numeric_layout = list(numeric_layout)
            cfg_mcq_q, cfg_tf_q, cfg_num_layout, exam_code_valid = _section_layout_from_subject_cfg(result.exam_code)
            if cfg_mcq_q:
                row_mcq_questions = list(cfg_mcq_q)
            if cfg_tf_q:
                row_tf_questions = list(cfg_tf_q)
            if exam_code_valid and cfg_num_layout:
                row_numeric_layout = list(cfg_num_layout)
            # Keep default numeric layout from current subject configuration when exam code
            # is missing/invalid, so Numeric answer string is still cut by configured lengths.
            mcq_map, tf_map, numeric_map, rebuilt_answer = _parse_answer_string(raw_answer, row_mcq_questions, row_tf_questions, row_numeric_layout)
            result.mcq_answers = mcq_map
            result.true_false_answers = tf_map
            result.numeric_answers = numeric_map
            setattr(result, "answer_string_api_mode", True)
            if expected_len > 0:
                result.answer_string = rebuilt_answer[:expected_len]
            else:
                result.answer_string = rebuilt_answer
            out.append(self._strip_transient_scan_artifacts(result))

        if not out:
            QMessageBox.warning(self, "API bài thi", "Không có dòng hợp lệ (cần cột FileName).")
            return

        if file_scope_mode == "all":
            self.database.delete_scan_results_for_subject(self._batch_result_subject_key(subject_key_for_results))
            self.scan_results = []
        else:
            self.scan_results = list(self._refresh_scan_results_from_db(subject_key_for_results) or [])
        by_path = {str(getattr(r, "image_path", "") or ""): r for r in self.scan_results}
        for r in out:
            by_path[str(r.image_path)] = r
        self.scan_results = list(by_path.values())
        self.scan_results_by_subject[self._batch_result_subject_key(subject_key_for_results)] = list(self.scan_results)
        self.database.replace_scan_results_for_subject(self._batch_result_subject_key(subject_key_for_results), [self._serialize_omr_result(x) for x in self.scan_results])
        self._populate_scan_grid_from_results(self.scan_results)
        self._finalize_batch_scan_display()
        self._update_batch_scan_scope_summary()
        QMessageBox.information(self, "API bài thi", f"Đã nạp {len(out)} dòng từ API bài thi.")

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        parsed_parts: list[str] = []
        mcq = self._format_mcq_answers(getattr(result, "mcq_answers", {}) or {})
        tf = self._format_tf_answers(getattr(result, "true_false_answers", {}) or {})
        num = self._format_numeric_answers(getattr(result, "numeric_answers", {}) or {})
        if mcq and mcq != "-":
            parsed_parts.append(f"MCQ: {mcq}")
        if tf and tf != "-":
            parsed_parts.append(f"TF: {tf}")
        if num and num != "-":
            parsed_parts.append(f"NUM: {num}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = blank_map.get(sec, [])
            if vals:
                if sec == "TF":
                    blank_parts.append(f"{sec} trống: {len(vals)}")
                else:
                    blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parsed_parts + blank_parts
        return " | ".join(merged) if merged else "-"

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _scoped_result_copy(self, result):
        scoped = copy.deepcopy(result)
        self._trim_result_answers_to_expected_scope(scoped)
        return scoped

    def _match_answer_key_payload_by_exam_code(self, mapping: dict, exam_code: str):
        if not isinstance(mapping, dict) or not mapping:
            return None
        exam_text = str(exam_code or "").strip()
        normalized_exam_text = self._normalize_exam_code_text(exam_text)
        if exam_text and exam_text in mapping:
            return mapping.get(exam_text)
        if normalized_exam_text:
            for key_text, payload in mapping.items():
                if self._normalize_exam_code_text(str(key_text or "").strip()) == normalized_exam_text:
                    return payload
        return None

    def _question_scope_from_subject_counts(
        self,
        template_expected: dict[str, list[int]],
        subject_cfg: dict | None = None,
        subject_key: str = "",
    ) -> dict[str, list[int]]:
        counts = self._subject_section_question_counts(subject_key) if subject_key else {"MCQ": 0, "TF": 0, "NUMERIC": 0}
        if not any(int(counts.get(sec, 0) or 0) > 0 for sec in ["MCQ", "TF", "NUMERIC"]):
            raw_counts = (subject_cfg or {}).get("question_counts", {}) if isinstance(subject_cfg, dict) else {}
            if isinstance(raw_counts, dict):
                counts = {
                    "MCQ": int(raw_counts.get("MCQ", 0) or 0),
                    "TF": int(raw_counts.get("TF", 0) or 0),
                    "NUMERIC": int(raw_counts.get("NUMERIC", 0) or 0),
                }

        scoped = {"MCQ": [], "TF": [], "NUMERIC": []}
        for sec in ["MCQ", "TF", "NUMERIC"]:
            limit = max(0, int(counts.get(sec, 0) or 0))
            if limit <= 0:
                scoped[sec] = []
                continue
            template_section = [int(q) for q in (template_expected.get(sec, []) or []) if str(q).strip().lstrip("-").isdigit() and int(q) > 0]
            scoped[sec] = list(template_section[:limit]) if template_section else list(range(1, limit + 1))
        return scoped

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        subject_cfg = self._selected_batch_subject_config() or self._resolve_subject_config_for_batch()
        current_subject_key = str(self._current_batch_subject_key() or self.active_batch_subject_key or "").strip()
        answer_subject_key = str(self._answer_key_subject_key(current_subject_key, subject_cfg) or "").strip()
        exam_code = str(getattr(result, "exam_code", "") or "").strip()
        normalized_exam_code = self._normalize_exam_code_text(exam_code)
        key_payload: Any = None
        has_configured_keys = False

        imported_answer_keys = self._subject_imported_answer_keys_for_main(subject_cfg or {})
        if imported_answer_keys:
            has_configured_keys = True
            key_payload = self._match_answer_key_payload_by_exam_code(imported_answer_keys, exam_code)

        if key_payload is None and current_subject_key:
            configured = self._fetch_answer_keys_for_subject_scoped(current_subject_key, subject_cfg) or {}
            if configured:
                has_configured_keys = True
                key_payload = self._match_answer_key_payload_by_exam_code(configured, exam_code)

        if key_payload is None and self.answer_keys is not None:
            subject_candidates: list[str] = []
            for candidate in [current_subject_key, answer_subject_key]:
                candidate_text = str(candidate or "").strip()
                if candidate_text and candidate_text not in subject_candidates:
                    subject_candidates.append(candidate_text)
            for candidate_subject in subject_candidates:
                rows = [
                    row for row in self.answer_keys.keys.values()
                    if str(getattr(row, "subject", "") or "").strip() == candidate_subject
                ]
                if rows:
                    has_configured_keys = True
                for row in rows:
                    row_exam_code = str(getattr(row, "exam_code", "") or "").strip()
                    if exam_code and row_exam_code == exam_code:
                        key_payload = row
                        break
                    if normalized_exam_code and self._normalize_exam_code_text(row_exam_code) == normalized_exam_code:
                        key_payload = row
                        break
                if key_payload is not None:
                    break

        scoped_by_counts = self._question_scope_from_subject_counts(template_expected, subject_cfg, current_subject_key)

        if key_payload is not None:
            full_credit = (key_payload.get("full_credit_questions", {}) if isinstance(key_payload, dict) else getattr(key_payload, "full_credit_questions", {})) or {}
            invalid_rows = (key_payload.get("invalid_answer_rows", {}) if isinstance(key_payload, dict) else getattr(key_payload, "invalid_answer_rows", {})) or {}

            def _extra_for_section(sec: str) -> set[int]:
                extra: set[int] = set()
                for q in list((full_credit.get(sec, []) or [])):
                    try:
                        extra.add(int(q))
                    except Exception:
                        continue
                for q in dict((invalid_rows.get(sec, {}) or {})).keys():
                    try:
                        extra.add(int(q))
                    except Exception:
                        continue
                return extra

            def _sorted_question_keys(payload: object) -> list[int]:
                out: set[int] = set()
                for q in dict(payload or {}).keys():
                    try:
                        out.add(int(q))
                    except Exception:
                        continue
                return sorted(out)

            mcq_map = (key_payload.get("mcq_answers", {}) if isinstance(key_payload, dict) else getattr(key_payload, "answers", {})) or {}
            tf_map = (key_payload.get("true_false_answers", {}) if isinstance(key_payload, dict) else getattr(key_payload, "true_false_answers", {})) or {}
            numeric_map = (key_payload.get("numeric_answers", {}) if isinstance(key_payload, dict) else getattr(key_payload, "numeric_answers", {})) or {}
            key_scope = {
                "MCQ": sorted(set(_sorted_question_keys(mcq_map)) | _extra_for_section("MCQ")),
                "TF": sorted(set(_sorted_question_keys(tf_map)) | _extra_for_section("TF")),
                "NUMERIC": sorted(set(_sorted_question_keys(numeric_map)) | _extra_for_section("NUMERIC")),
            }

            if any(scoped_by_counts.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
                merged_scope: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    allowed = [int(q) for q in (scoped_by_counts.get(sec, []) or [])]
                    if not allowed:
                        merged_scope[sec] = []
                        continue
                    allowed_set = set(allowed)
                    filtered = [int(q) for q in (key_scope.get(sec, []) or []) if int(q) in allowed_set]
                    merged_scope[sec] = filtered or list(allowed)
                return merged_scope

            return key_scope
        if any(scoped_by_counts.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
            return scoped_by_counts

        if has_configured_keys:
            return {"MCQ": [], "TF": [], "NUMERIC": []}

        return {sec: list(vals) for sec, vals in template_expected.items()}

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

        has_exam_code_zone = any(z.zone_type.value == "EXAM_CODE_BLOCK" for z in self.template.zones)
        has_student_id_zone = any(z.zone_type.value == "STUDENT_ID_BLOCK" for z in self.template.zones)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, v in sorted(answers.items(), key=lambda x: int(x[0])):
            value = str(v).strip()
            chunks.append(f"{int(q)}={value if value else '_'}")
        return "; ".join(chunks)

    @staticmethod
    def _format_mcq_answers_with_expected(answers: dict[int, str], expected_questions: list[int]) -> str:
        expected = sorted({int(q) for q in (expected_questions or []) if int(q) > 0})
        if not expected:
            return MainWindow._format_mcq_answers(answers)
        chunks: list[str] = []
        for q in expected:
            ans = str((answers or {}).get(int(q), "") or "").strip().upper()
            chunks.append(f"{int(q)}{ans if ans else '_'}")
        return "; ".join(chunks) if chunks else "-"

    @staticmethod
    def _format_tf_answers_with_expected(answers: dict[int, dict[str, bool]], expected_questions: list[int]) -> str:
        expected = sorted({int(q) for q in (expected_questions or []) if int(q) > 0})
        if not expected:
            return MainWindow._format_tf_answers(answers)
        chunks: list[str] = []
        for q in expected:
            flags = dict((answers or {}).get(int(q), {}) or {})
            marks = "".join(("Đ" if bool(flags.get(k)) else "S") if k in flags else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks) if chunks else "-"

    @staticmethod
    def _format_numeric_answers_with_expected(answers: dict[int, str], expected_questions: list[int]) -> str:
        expected = sorted({int(q) for q in (expected_questions or []) if int(q) > 0})
        if not expected:
            return MainWindow._format_numeric_answers(answers)
        chunks: list[str] = []
        for q in expected:
            value = str((answers or {}).get(int(q), "") or "").strip()
            chunks.append(f"{int(q)}={value if value else '_'}")
        return "; ".join(chunks) if chunks else "-"

    def _build_recognition_content_text(
        self,
        result,
        blank_map: dict[str, list],
        expected_by_section: dict[str, list[int]] | None = None,
    ) -> str:
        parts: list[str] = []
        blank_parts: list[str] = []
        mcq_blank = list((blank_map or {}).get("MCQ", []) or [])
        if mcq_blank:
            blank_parts.append(f"MCQ[{','.join(str(v) for v in mcq_blank)}]")
        tf_blank = list((blank_map or {}).get("TF", []) or [])
        tf_blank_detail = dict(getattr(result, "tf_blank_detail", {}) or {})
        if tf_blank:
            tf_chunks: list[str] = []
            for q in tf_blank:
                missing_count = int(tf_blank_detail.get(int(q), 0) or 0)
                if missing_count > 0:
                    tf_chunks.append(f"{int(q)}({missing_count}/4 ý)")
                else:
                    tf_chunks.append(str(int(q)))
            blank_parts.append(f"TF[{', '.join(tf_chunks)}]")
        numeric_blank = list((blank_map or {}).get("NUMERIC", []) or [])
        if numeric_blank:
            blank_parts.append(f"NUM[{','.join(str(v) for v in numeric_blank)}]")
        if blank_parts:
            parts.append("Trống: " + " | ".join(blank_parts))
        return " | ".join(parts) if parts else ""

    def _normalize_answer_map_to_expected_scope(self, data: dict, expected_questions: list[int], cast_value):
        normalized: dict[int, Any] = {}
        for q_raw, value in (data or {}).items():
            try:
                q_no = int(q_raw)
            except Exception:
                continue
            normalized[q_no] = cast_value(value)

        expected_set = sorted(
            {
                int(q)
                for q in (expected_questions or [])
                if str(q).strip().lstrip("-").isdigit() and int(q) > 0
            }
        )
        if not expected_set:
            return {}
        if not normalized:
            return {}

        actual_set = sorted(set(normalized.keys()))
        overlap = sorted(set(actual_set) & set(expected_set))
        if overlap:
            return {q: normalized[q] for q in expected_set if q in normalized}

        remapped: dict[int, Any] = {}
        for src_q, dst_q in zip(actual_set, expected_set):
            remapped[dst_q] = normalized[src_q]
        return remapped

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        if result is None:
            return

        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(
            code_text
            and "?" not in code_text
            and (
                not avail_codes
                or code_text in avail_codes
                or self._normalize_exam_code_text(code_text) in avail_codes
            )
        )
        if not code_valid:
            return

        subject_key = str(self._current_batch_subject_key() or self.active_batch_subject_key or "").strip()
        if not subject_key:
            return
        if self.template is None:
            return

        expected = self._expected_questions_by_section(result)
        if not any(expected.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
            return

        result.mcq_answers = self._normalize_answer_map_to_expected_scope(
            result.mcq_answers or {},
            list(expected.get("MCQ", []) or []),
            lambda v: str(v or "").strip().upper()[:1],
        )
        result.true_false_answers = self._normalize_answer_map_to_expected_scope(
            result.true_false_answers or {},
            list(expected.get("TF", []) or []),
            lambda v: {
                str(k or "").strip().lower(): bool(flag)
                for k, flag in (dict(v or {}) if isinstance(v, dict) else {}).items()
                if str(k or "").strip().lower() in {"a", "b", "c", "d"}
            },
        )
        result.numeric_answers = self._normalize_answer_map_to_expected_scope(
            result.numeric_answers or {},
            list(expected.get("NUMERIC", []) or []),
            lambda v: str(v or "").strip(),
        )
        result.sync_legacy_aliases()

    def _compute_blank_questions(self, result) -> dict[str, list]:
        expected_by_section = self._expected_questions_by_section(result)
        mcq_payload = {int(q): str(a) for q, a in (dict(getattr(result, "mcq_answers", {}) or {})).items()}
        tf_payload = {int(q): dict(v or {}) for q, v in (dict(getattr(result, "true_false_answers", {}) or {})).items()}
        numeric_payload = {int(q): str(v) for q, v in (dict(getattr(result, "numeric_answers", {}) or {})).items()}

        mcq_expected = sorted(set(int(q) for q in (expected_by_section.get("MCQ", []) or [])))
        tf_expected = sorted(set(int(q) for q in (expected_by_section.get("TF", []) or [])))
        numeric_expected = sorted(set(int(q) for q in (expected_by_section.get("NUMERIC", []) or [])))

        blanks: dict[str, list[int] | list[str]] = {}
        for sec in ["MCQ", "TF", "NUMERIC"]:
            if sec == "MCQ":
                blanks[sec] = [int(q) for q in mcq_expected if not str(mcq_payload.get(int(q), "") or "").strip()]
            elif sec == "NUMERIC":
                blanks[sec] = [int(q) for q in numeric_expected if not str(numeric_payload.get(int(q), "") or "").strip()]
            else:
                blanks[sec] = []
        tf_blank_detail: dict[int, int] = {}
        missing_tf_statements: list[str] = []
        for display_q in tf_expected:
            flags = tf_payload.get(int(display_q), {})
            flags = flags if isinstance(flags, dict) else {}
            missing_count = sum(1 for key in ["a", "b", "c", "d"] if key not in flags)
            if missing_count > 0:
                tf_blank_detail[int(display_q)] = int(missing_count)
                for key in ["a", "b", "c", "d"]:
                    if key not in flags:
                        missing_tf_statements.append(f"{int(display_q)}{key}")
        sec = "TF"
        blanks[sec] = missing_tf_statements

        setattr(result, "tf_blank_detail", tf_blank_detail)
        return {
            "MCQ": list(blanks.get("MCQ", []) or []),
            "TF": list(blanks.get("TF", []) or []),
            "NUMERIC": list(blanks.get("NUMERIC", []) or []),
        }
        messages: list[str] = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            expected_set = set(expected.get(sec, []))
            if not expected_set:
                continue
            actual_set = {int(q) for q in actual_map.get(sec, set())}
            missing = sorted(expected_set - actual_set)
            if missing:
                messages.append(
                    f"thiếu {sec} ({len(expected_set)-len(missing)}/{len(expected_set)}): {','.join(str(v) for v in missing)}"
                )
        return messages

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = {sec: list(vals) for sec, vals in template_expected.items()}
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                full_credit = getattr(key, "full_credit_questions", {}) or {}
                invalid_rows = getattr(key, "invalid_answer_rows", {}) or {}

                def _extra_for_section(sec: str) -> set[int]:
                    extra: set[int] = set()
                    for q in list((full_credit.get(sec, []) or [])):
                        try:
                            extra.add(int(q))
                        except Exception:
                            continue
                    for q in dict((invalid_rows.get(sec, {}) or {})).keys():
                        try:
                            extra.add(int(q))
                        except Exception:
                            continue
                    return extra

                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys()) | _extra_for_section("MCQ")),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys()) | _extra_for_section("TF")),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys()) | _extra_for_section("NUMERIC")),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    key_set = set(key_sections[sec])
                    if key_set:
                        # For display/content scope: always preserve answer-key numbering.
                        expected_by_section[sec] = key_sections[sec]
        if not any(expected_by_section.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
            # fallback only when answer-key scope is unavailable.
            return {sec: list(vals) for sec, vals in template_expected.items()}
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, v in sorted(answers.items(), key=lambda x: int(x[0])):
            value = str(v).strip()
            chunks.append(f"{int(q)}={value if value else '_'}")
        return "; ".join(chunks)

    @staticmethod
    def _format_mcq_answers_with_expected(answers: dict[int, str], expected_questions: list[int]) -> str:
        expected = sorted({int(q) for q in (expected_questions or []) if int(q) > 0})
        if not expected:
            return MainWindow._format_mcq_answers(answers)
        chunks: list[str] = []
        for q in expected:
            ans = str((answers or {}).get(int(q), "") or "").strip().upper()
            chunks.append(f"{int(q)}{ans if ans else '_'}")
        return "; ".join(chunks) if chunks else "-"

    @staticmethod
    def _format_tf_answers_with_expected(answers: dict[int, dict[str, bool]], expected_questions: list[int]) -> str:
        expected = sorted({int(q) for q in (expected_questions or []) if int(q) > 0})
        if not expected:
            return MainWindow._format_tf_answers(answers)
        chunks: list[str] = []
        for q in expected:
            flags = dict((answers or {}).get(int(q), {}) or {})
            marks = "".join(("Đ" if bool(flags.get(k)) else "S") if k in flags else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks) if chunks else "-"

    @staticmethod
    def _format_numeric_answers_with_expected(answers: dict[int, str], expected_questions: list[int]) -> str:
        expected = sorted({int(q) for q in (expected_questions or []) if int(q) > 0})
        if not expected:
            return MainWindow._format_numeric_answers(answers)
        chunks: list[str] = []
        for q in expected:
            value = str((answers or {}).get(int(q), "") or "").strip()
            chunks.append(f"{int(q)}={value if value else '_'}")
        return "; ".join(chunks) if chunks else "-"

    def _build_recognition_content_text(
        self,
        result,
        blank_map: dict[str, list[int]],
        expected_by_section: dict[str, list[int]] | None = None,
    ) -> str:
        parts: list[str] = []
        blank_parts: list[str] = []
        mcq_blank = list((blank_map or {}).get("MCQ", []) or [])
        if mcq_blank:
            blank_parts.append(f"MCQ[{','.join(str(v) for v in mcq_blank)}]")

        tf_blank = list((blank_map or {}).get("TF", []) or [])
        tf_blank_detail = dict(getattr(result, "tf_blank_detail", {}) or {})
        if tf_blank:
            tf_chunks: list[str] = []
            for q in tf_blank:
                missing_count = int(tf_blank_detail.get(int(q), 0) or 0)
                if missing_count > 0:
                    tf_chunks.append(f"{int(q)}({missing_count}/4 ý)")
                else:
                    tf_chunks.append(str(int(q)))
            blank_parts.append(f"TF[{', '.join(tf_chunks)}]")

        numeric_blank = list((blank_map or {}).get("NUMERIC", []) or [])
        if numeric_blank:
            blank_parts.append(f"NUM[{','.join(str(v) for v in numeric_blank)}]")

        if blank_parts:
            parts.append("Trống: " + " | ".join(blank_parts))
        return " | ".join(parts) if parts else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        if result is None:
            return

        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return

        subject_key = str(self._current_batch_subject_key() or self.active_batch_subject_key or "").strip()
        if not subject_key:
            return
        if self.template is None:
            return

        expected = self._expected_questions_by_section(result)
        if not any(expected.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
            return

        def _trim_map_by_expected(data: dict, expected_questions: list[int], cast_value):
            normalized = {int(q): cast_value(v) for q, v in (data or {}).items()}
            expected_set = sorted({int(q) for q in (expected_questions or []) if str(q).strip().lstrip('-').isdigit() and int(q) > 0})
            if not expected_set:
                return normalized
            return {q: normalized[q] for q in expected_set if q in normalized}

        result.mcq_answers = _trim_map_by_expected(result.mcq_answers or {}, list(expected.get("MCQ", []) or []), lambda v: str(v))
        result.true_false_answers = _trim_map_by_expected(result.true_false_answers or {}, list(expected.get("TF", []) or []), lambda v: dict(v or {}))
        result.numeric_answers = _trim_map_by_expected(result.numeric_answers or {}, list(expected.get("NUMERIC", []) or []), lambda v: str(v))

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, v in sorted(answers.items(), key=lambda x: int(x[0])):
            value = str(v).strip()
            chunks.append(f"{int(q)}={value if value else '_'}")
        return "; ".join(chunks)

    def _build_recognition_content_text(
        self,
        result,
        blank_map: dict[str, list[int]],
        expected_by_section: dict[str, list[int]] | None = None,
    ) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})
        expected = expected_by_section or {}
        expected_mcq = list(expected.get("MCQ", []) or [])
        expected_tf = list(expected.get("TF", []) or [])
        expected_numeric = list(expected.get("NUMERIC", []) or [])

        parts: list[str] = []
        blank_parts: list[str] = []
        mcq_blank = list((blank_map or {}).get("MCQ", []) or [])
        if mcq_blank:
            blank_parts.append(f"MCQ[{','.join(str(v) for v in mcq_blank)}]")
        tf_blank = list((blank_map or {}).get("TF", []) or [])
        tf_blank_detail = dict(getattr(result, "tf_blank_detail", {}) or {})
        if tf_blank:
            tf_chunks: list[str] = []
            for q in tf_blank:
                missing_count = int(tf_blank_detail.get(int(q), 0) or 0)
                if missing_count > 0:
                    tf_chunks.append(f"{int(q)}({missing_count}/4 ý)")
                else:
                    tf_chunks.append(str(int(q)))
            blank_parts.append(f"TF[{', '.join(tf_chunks)}]")
        numeric_blank = list((blank_map or {}).get("NUMERIC", []) or [])
        if numeric_blank:
            blank_parts.append(f"NUM[{','.join(str(v) for v in numeric_blank)}]")
        if blank_parts:
            parts.append("Trống: " + " | ".join(blank_parts))
        return " | ".join(parts) if parts else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, v in sorted(answers.items(), key=lambda x: int(x[0])):
            value = str(v).strip()
            chunks.append(f"{int(q)}={value if value else '_'}")
        return "; ".join(chunks)

    def _build_recognition_content_text(
        self,
        result,
        blank_map: dict[str, list[int]],
        expected_by_section: dict[str, list[int]] | None = None,
    ) -> str:
        parts: list[str] = []
        blank_parts: list[str] = []
        mcq_blank = list((blank_map or {}).get("MCQ", []) or [])
        if mcq_blank:
            blank_parts.append(f"MCQ[{','.join(str(v) for v in mcq_blank)}]")
        tf_blank = list((blank_map or {}).get("TF", []) or [])
        tf_blank_detail = dict(getattr(result, "tf_blank_detail", {}) or {})
        if tf_blank:
            tf_chunks: list[str] = []
            for q in tf_blank:
                missing_count = int(tf_blank_detail.get(int(q), 0) or 0)
                if missing_count > 0:
                    tf_chunks.append(f"{int(q)}({missing_count}/4 ý)")
                else:
                    tf_chunks.append(str(int(q)))
            blank_parts.append(f"TF[{', '.join(tf_chunks)}]")
        numeric_blank = list((blank_map or {}).get("NUMERIC", []) or [])
        if numeric_blank:
            blank_parts.append(f"NUM[{','.join(str(v) for v in numeric_blank)}]")
        if blank_parts:
            parts.append("Trống: " + " | ".join(blank_parts))
        return " | ".join(parts) if parts else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        expected_by_section = dict(template_expected)
        subject_key_name = self.active_batch_subject_key
        if not subject_key_name and self.session and self.session.subjects:
            subject_key_name = self.session.subjects[0]
        if self.answer_keys and subject_key_name:
            key = self.answer_keys.get(subject_key_name, (result.exam_code or "").strip())
            if key:
                key_sections = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys())),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys())),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys())),
                }
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    if not key_sections[sec]:
                        continue
                    template_set = set(template_expected.get(sec, []))
                    key_set = set(key_sections[sec])
                    if template_set:
                        overlap = sorted(template_set & key_set)
                        if overlap:
                            expected_by_section[sec] = overlap
                        else:
                            # When numbering between template and answer-key differs,
                            # prioritize answer-key numbering to keep section slicing correct.
                            expected_by_section[sec] = key_sections[sec]
                    else:
                        expected_by_section[sec] = key_sections[sec]
        return expected_by_section

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}={str(v).strip()}" for q, v in sorted(answers.items(), key=lambda x: int(x[0])))

    def _build_recognition_content_text(self, result, blank_map: dict[str, list[int]]) -> str:
        mcq_payload = dict(getattr(result, "mcq_answers", {}) or {})
        tf_payload = dict(getattr(result, "true_false_answers", {}) or {})
        numeric_payload = dict(getattr(result, "numeric_answers", {}) or {})

        parts: list[str] = []
        if mcq_payload:
            parts.append(f"MCQ: {self._format_mcq_answers(mcq_payload)}")
        if tf_payload:
            parts.append(f"TF: {self._format_tf_answers(tf_payload)}")
        if numeric_payload:
            parts.append(f"NUM: {self._format_numeric_answers(numeric_payload)}")

        blank_parts = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            vals = list((blank_map or {}).get(sec, []) or [])
            if vals:
                blank_parts.append(f"{sec} trống: {','.join(str(v) for v in vals)}")
        merged = parts + blank_parts
        return " | ".join(merged) if merged else ""

    def _trim_result_answers_to_expected_scope(self, result) -> None:
        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return
        expected = self._expected_questions_by_section(result)
        if expected.get("MCQ"):
            allow = set(expected["MCQ"])
            result.mcq_answers = {int(q): str(a) for q, a in (result.mcq_answers or {}).items() if int(q) in allow}
        if expected.get("TF"):
            allow = set(expected["TF"])
            result.true_false_answers = {int(q): dict(a) for q, a in (result.true_false_answers or {}).items() if int(q) in allow}
        if expected.get("NUMERIC"):
            allow = set(expected["NUMERIC"])
            result.numeric_answers = {int(q): str(a) for q, a in (result.numeric_answers or {}).items() if int(q) in allow}

    def _expected_questions_by_section(self, result) -> dict[str, list[int]]:
        template_expected: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        if self.template:
            for z in self.template.zones:
                if not z.grid:
                    continue
                count = int(z.grid.question_count or z.grid.rows or 0)
                start = int(z.grid.question_start)
                rng = list(range(start, start + max(0, count)))
                if z.zone_type.value == "MCQ_BLOCK":
                    template_expected["MCQ"].extend(rng)
                elif z.zone_type.value == "TRUE_FALSE_BLOCK":
                    template_expected["TF"].extend(rng)
                elif z.zone_type.value == "NUMERIC_BLOCK":
                    template_expected["NUMERIC"].extend(rng)
            template_expected = {sec: sorted(set(vals)) for sec, vals in template_expected.items()}

        subject_cfg = self._selected_batch_subject_config() or self._resolve_subject_config_for_batch()
        current_subject_key = str(self._current_batch_subject_key() or self.active_batch_subject_key or "").strip()
        answer_subject_key = str(self._answer_key_subject_key(current_subject_key, subject_cfg) or "").strip()
        exam_code = str(getattr(result, "exam_code", "") or "").strip()
        normalized_exam_code = self._normalize_exam_code_text(exam_code)
        key_payload: Any = None
        has_configured_keys = False

        imported_answer_keys = self._subject_imported_answer_keys_for_main(subject_cfg or {})
        if imported_answer_keys:
            has_configured_keys = True
            key_payload = self._match_answer_key_payload_by_exam_code(imported_answer_keys, exam_code)

        if key_payload is None and current_subject_key:
            configured = self._fetch_answer_keys_for_subject_scoped(current_subject_key, subject_cfg) or {}
            if configured:
                has_configured_keys = True
                key_payload = self._match_answer_key_payload_by_exam_code(configured, exam_code)

        if key_payload is None and self.answer_keys is not None:
            subject_candidates: list[str] = []
            for candidate in [current_subject_key, answer_subject_key]:
                candidate_text = str(candidate or "").strip()
                if candidate_text and candidate_text not in subject_candidates:
                    subject_candidates.append(candidate_text)
            for candidate_subject in subject_candidates:
                rows = [
                    row for row in self.answer_keys.keys.values()
                    if str(getattr(row, "subject", "") or "").strip() == candidate_subject
                ]
                if rows:
                    has_configured_keys = True
                for row in rows:
                    row_exam_code = str(getattr(row, "exam_code", "") or "").strip()
                    if exam_code and row_exam_code == exam_code:
                        key_payload = row
                        break
                    if normalized_exam_code and self._normalize_exam_code_text(row_exam_code) == normalized_exam_code:
                        key_payload = row
                        break
                if key_payload is not None:
                    break

        scoped_by_counts = self._question_scope_from_subject_counts(template_expected, subject_cfg, current_subject_key)

        if key_payload is not None:
            full_credit = (key_payload.get("full_credit_questions", {}) if isinstance(key_payload, dict) else getattr(key_payload, "full_credit_questions", {})) or {}
            invalid_rows = (key_payload.get("invalid_answer_rows", {}) if isinstance(key_payload, dict) else getattr(key_payload, "invalid_answer_rows", {})) or {}

            def _extra_for_section(sec: str) -> set[int]:
                extra: set[int] = set()
                for q in list((full_credit.get(sec, []) or [])):
                    try:
                        extra.add(int(q))
                    except Exception:
                        continue
                for q in dict((invalid_rows.get(sec, {}) or {})).keys():
                    try:
                        extra.add(int(q))
                    except Exception:
                        continue
                return extra

            def _sorted_question_keys(payload: object) -> list[int]:
                out: set[int] = set()
                for q in dict(payload or {}).keys():
                    try:
                        out.add(int(q))
                    except Exception:
                        continue
                return sorted(out)

            mcq_map = (key_payload.get("mcq_answers", {}) if isinstance(key_payload, dict) else getattr(key_payload, "answers", {})) or {}
            tf_map = (key_payload.get("true_false_answers", {}) if isinstance(key_payload, dict) else getattr(key_payload, "true_false_answers", {})) or {}
            numeric_map = (key_payload.get("numeric_answers", {}) if isinstance(key_payload, dict) else getattr(key_payload, "numeric_answers", {})) or {}
            key_scope = {
                "MCQ": sorted(set(_sorted_question_keys(mcq_map)) | _extra_for_section("MCQ")),
                "TF": sorted(set(_sorted_question_keys(tf_map)) | _extra_for_section("TF")),
                "NUMERIC": sorted(set(_sorted_question_keys(numeric_map)) | _extra_for_section("NUMERIC")),
            }

            if any(scoped_by_counts.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
                merged_scope: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
                for sec in ["MCQ", "TF", "NUMERIC"]:
                    allowed = [int(q) for q in (scoped_by_counts.get(sec, []) or [])]
                    if not allowed:
                        merged_scope[sec] = []
                        continue
                    allowed_set = set(allowed)
                    filtered = [int(q) for q in (key_scope.get(sec, []) or []) if int(q) in allowed_set]
                    merged_scope[sec] = filtered or list(allowed)
                return merged_scope
            return key_scope

        if any(scoped_by_counts.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
            return scoped_by_counts
        if has_configured_keys:
            return {"MCQ": [], "TF": [], "NUMERIC": []}
        return {sec: list(vals) for sec, vals in template_expected.items()}

    @staticmethod
    def _format_mcq_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        return "; ".join(f"{int(q)}{str(a).strip()}" for q, a in sorted(answers.items(), key=lambda x: int(x[0])))

    @staticmethod
    def _format_tf_answers(answers: dict[int, dict[str, bool]]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, flags in sorted(answers.items(), key=lambda x: int(x[0])):
            marks = "".join(("Đ" if bool((flags or {}).get(k)) else "S") if k in (flags or {}) else "_" for k in ["a", "b", "c", "d"])
            chunks.append(f"{int(q)}{marks}")
        return "; ".join(chunks)

    @staticmethod
    def _format_numeric_answers(answers: dict[int, str]) -> str:
        if not answers:
            return "-"
        chunks: list[str] = []
        for q, v in sorted(answers.items(), key=lambda x: int(x[0])):
            value = str(v).strip()
            chunks.append(f"{int(q)}={value if value else '_'}")
        return "; ".join(chunks)

    def _build_recognition_content_text(
        self,
        result,
        blank_map: dict[str, list[int]],
        expected_by_section: dict[str, list[int]] | None = None,
    ) -> str:
        expected = expected_by_section or self._expected_questions_by_section(result)
        lines: list[str] = []

        mcq_expected = set(int(q) for q in (expected.get("MCQ", []) or []) if str(q).strip())
        mcq_blank = [
            int(v)
            for v in list((blank_map or {}).get("MCQ", []) or [])
            if str(v).strip() and int(v) in mcq_expected
        ]
        if mcq_blank:
            lines.append(f"MCQ trống: {', '.join(str(v) for v in mcq_blank)}")

        tf_expected = sorted(int(q) for q in (expected.get("TF", []) or []) if str(q).strip())
        tf_blank_tokens = [str(v).strip().lower() for v in list((blank_map or {}).get("TF", []) or []) if str(v).strip()]
        tf_blank_set = {
            int(token[:-1])
            for token in tf_blank_tokens
            if len(token) >= 2 and token[:-1].lstrip("-").isdigit() and token[-1] in {"a", "b", "c", "d"}
        }
        tf_answers = {
            int(q): dict(v or {})
            for q, v in (dict(getattr(result, "true_false_answers", {}) or {})).items()
            if str(q).strip()
        }
        missing_tf_statements: list[str] = []
        for display_q in tf_expected:
            if display_q not in tf_blank_set:
                continue
            flags = tf_answers.get(int(display_q), {}) if isinstance(tf_answers.get(int(display_q), {}), dict) else {}
            for key in ["a", "b", "c", "d"]:
                if key not in flags:
                    missing_tf_statements.append(f"{int(display_q)}{key}")
        if missing_tf_statements:
            lines.append("TF trống: " + ", ".join(missing_tf_statements))

        numeric_expected = set(int(q) for q in (expected.get("NUMERIC", []) or []) if str(q).strip())
        numeric_blank = [
            int(v)
            for v in list((blank_map or {}).get("NUMERIC", []) or [])
            if str(v).strip() and int(v) in numeric_expected
        ]
        if numeric_blank:
            lines.append(f"Num trống: {', '.join(str(v) for v in numeric_blank)}")

        return "\n".join(lines)
    def _trim_result_answers_to_expected_scope(self, result) -> None:
        if result is None:
            return

        code_text = str(getattr(result, "exam_code", "") or "").strip()
        avail_codes = self._available_exam_codes()
        code_valid = bool(code_text and "?" not in code_text and (not avail_codes or code_text in avail_codes or self._normalize_exam_code_text(code_text) in avail_codes))
        if not code_valid:
            return

        subject_key = str(self._current_batch_subject_key() or self.active_batch_subject_key or "").strip()
        if not subject_key:
            return
        if self.template is None:
            return

        expected = self._expected_questions_by_section(result)
        if not any(expected.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
            return

        def _trim_map_by_expected(data: dict, expected_questions: list[int], cast_value):
            normalized = {int(q): cast_value(v) for q, v in (data or {}).items()}
            expected_set = sorted({int(q) for q in (expected_questions or []) if str(q).strip().lstrip('-').isdigit() and int(q) > 0})
            if not expected_set:
                return normalized
            return {q: normalized[q] for q in expected_set if q in normalized}

        result.mcq_answers = _trim_map_by_expected(result.mcq_answers or {}, list(expected.get("MCQ", []) or []), lambda v: str(v))
        result.true_false_answers = _trim_map_by_expected(result.true_false_answers or {}, list(expected.get("TF", []) or []), lambda v: dict(v or {}))
        result.numeric_answers = _trim_map_by_expected(result.numeric_answers or {}, list(expected.get("NUMERIC", []) or []), lambda v: str(v))

    def _count_mismatch_status_parts(self, result) -> list[str]:
        expected = self._expected_questions_by_section(result)
        actual_map = {
            "MCQ": set((result.mcq_answers or {}).keys()),
            "TF": set((result.true_false_answers or {}).keys()),
            "NUMERIC": set((result.numeric_answers or {}).keys()),
        }
        messages: list[str] = []
        for sec in ["MCQ", "TF", "NUMERIC"]:
            expected_set = set(expected.get(sec, []))
            if not expected_set:
                continue
            actual_set = {int(q) for q in actual_map.get(sec, set())}
            missing = sorted(expected_set - actual_set)
            if missing:
                messages.append(
                    f"thiếu {sec} ({len(expected_set)-len(missing)}/{len(expected_set)}): {','.join(str(v) for v in missing)}"
                )
        return messages

    def _retrim_batch_results_to_answer_key_scope(self) -> None:
        if not self.scan_results:
            return
        for idx, res in enumerate(self.scan_results):
            scoped = self._scoped_result_copy(res)
            blank_map = self._compute_blank_questions(scoped)
            expected = self._expected_questions_by_section(scoped)
            self.scan_blank_summary[idx] = blank_map
            self.scan_blank_questions[idx] = blank_map.get("MCQ", [])
            if idx < self.scan_list.rowCount():
                self.scan_list.setItem(idx, self.SCAN_COL_CONTENT, QTableWidgetItem(self._build_recognition_content_text(scoped, blank_map, expected)))
                sid_item = self.scan_list.item(idx, self.SCAN_COL_STUDENT_ID)
                if sid_item:
                    sid_item.setData(Qt.UserRole + 2, self._short_recognition_text_for_result(scoped))
        self._refresh_all_statuses()
        current = self.scan_list.currentRow() if hasattr(self, "scan_list") else -1
        if 0 <= current < len(self.scan_results):
            self._update_scan_preview(current)
            self._load_selected_result_for_correction()

    def _apply_scan_filter(self) -> None:
        def _normalize(text: str) -> str:
            return " ".join(str(text or "").strip().lower().split())

        value = _normalize(self.search_value.text())
        col = self._scan_filter_column_from_combo_index(self.filter_column.currentIndex())
        for i in range(self.scan_list.rowCount()):
            status_text = str(self.scan_list.item(i, self.SCAN_COL_STATUS).text() if self.scan_list.item(i, self.SCAN_COL_STATUS) else "").strip().lower()
            is_edited = "đã sửa" in status_text
            status_ok = True
            if self.batch_status_filter_mode == "error":
                status_ok = (not is_edited) and bool(status_text and status_text != "ok")
            elif self.batch_status_filter_mode == "duplicate":
                status_ok = (not is_edited) and ("trùng sbd" in status_text or "duplicate" in status_text)
            elif self.batch_status_filter_mode == "wrong_code":
                status_ok = (not is_edited) and (("mã đề" in status_text) and ("sai" in status_text or "không" in status_text or "?" in status_text))
            elif self.batch_status_filter_mode == "edited":
                status_ok = is_edited
            if not value:
                self.scan_list.setRowHidden(i, not status_ok)
                continue
            if col is None:
                searchable = []
                for j in range(0, self.SCAN_COL_STATUS + 1):
                    item = self.scan_list.item(i, j)
                    searchable.append(_normalize(item.text() if item else ""))
                cell = " | ".join(searchable)
            else:
                item = self.scan_list.item(i, col)
                cell = _normalize(item.text() if item else "")
            self.scan_list.setRowHidden(i, (value not in cell) or (not status_ok))
        self._update_batch_scan_bottom_status_text()

    def _handle_batch_status_filter_link(self, link: str) -> None:
        self.batch_status_filter_mode = str(link or "all").strip() or "all"
        self._apply_scan_filter()

    def _on_scan_header_clicked(self, section: int) -> None:
        combo_index = self._scan_filter_combo_index_from_header_section(section)
        if combo_index is None:
            return
        if 0 <= combo_index < self.filter_column.count():
            self.filter_column.setCurrentIndex(combo_index)
        self._apply_scan_filter()

    @staticmethod
    def _scan_filter_column_from_combo_index(combo_index: int) -> int | None:
        mapping = {
            0: None,  # Tất cả
            1: 0,     # STT
            2: 1,     # STUDENT ID
            3: 2,     # Phòng thi
            4: 3,     # Mã đề
            5: 4,     # Họ tên
            6: 5,     # Ngày sinh
            7: 6,     # Nội dung
            8: 7,     # Status
        }
        return mapping.get(int(combo_index), None)

    @staticmethod
    def _scan_filter_combo_index_from_header_section(section: int) -> int | None:
        mapping = {
            0: 1,  # STT
            1: 2,  # STUDENT ID
            2: 3,  # Phòng thi
            3: 4,  # Mã đề
            4: 5,  # Họ tên
            5: 6,  # Ngày sinh
            6: 7,  # Nội dung
            7: 8,  # Status
        }
        return mapping.get(int(section), None)

    def eventFilter(self, obj, event):
        if hasattr(self, "scan_image_scroll") and obj == self.scan_image_scroll.viewport() and event.type() == QEvent.Wheel:
            if event.modifiers() & Qt.ControlModifier:
                if event.angleDelta().y() > 0:
                    self._zoom_preview_in()
                else:
                    self._zoom_preview_out()
                return True
        if hasattr(self, "scan_image_scroll") and obj == self.scan_image_scroll.viewport():
            if event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton and not self.preview_source_pixmap.isNull():
                self.preview_drag_active = True
                self.preview_last_pos = event.position().toPoint()
                self.scan_image_scroll.viewport().setCursor(Qt.ClosedHandCursor)
                return True
            if event.type() == QEvent.MouseMove and self.preview_drag_active and self.preview_last_pos is not None:
                pos = event.position().toPoint()
                dx = pos.x() - self.preview_last_pos.x()
                dy = pos.y() - self.preview_last_pos.y()
                self.preview_last_pos = pos
                self.scan_image_scroll.horizontalScrollBar().setValue(
                    self.scan_image_scroll.horizontalScrollBar().value() - dx
                )
                self.scan_image_scroll.verticalScrollBar().setValue(
                    self.scan_image_scroll.verticalScrollBar().value() - dy
                )
                return True
            if event.type() == QEvent.MouseButtonRelease and event.button() == Qt.LeftButton:
                self.preview_drag_active = False
                self.preview_last_pos = None
                self.scan_image_scroll.viewport().unsetCursor()
                return True
        return super().eventFilter(obj, event)

    def _selected_scan_row_index(self) -> int:
        idx = self.scan_list.currentRow()
        if idx < 0 or idx >= self.scan_list.rowCount():
            return -1
        return idx

    def _rotate_selected_scan(self, degrees: int) -> None:
        row_idx = self._selected_scan_row_index()
        if row_idx < 0:
            return
        current = int(self.preview_rotation_by_index.get(row_idx, 0) or 0)
        self.preview_rotation_by_index[row_idx] = (current + int(degrees)) % 360
        if row_idx < len(self.scan_results):
            self._update_scan_preview(row_idx)
        else:
            self._update_scan_preview_from_saved_row(row_idx)


    def _rebuild_error_list(self) -> None:
        self.error_list.clear()
        for result in self.scan_results:
            rec_errors = list(getattr(result, "recognition_errors", [])) or list(getattr(result, "errors", []))
            for issue in result.issues:
                self.error_list.addItem(f"{Path(result.image_path).name}: {issue.code} - {issue.message}")
            for err in rec_errors:
                self.error_list.addItem(f"{Path(result.image_path).name}: RECOGNITION - {err}")

    @staticmethod
    def _student_sort_token(student_id: str) -> tuple[int, int, object]:
        sid = str(student_id or "").strip()
        if not sid or sid == "-":
            return (1, 0, "")
        if sid.isdigit():
            return (0, 0, int(sid))
        return (0, 1, sid.casefold())

    def _row_sort_bucket(self, status: str, blank_map: dict[str, list[int]]) -> int:
        status_text = str(status or "").strip()
        has_error = bool(status_text and status_text != "OK")
        has_blank = any(bool(blank_map.get(sec, [])) for sec in ["MCQ", "TF", "NUMERIC"])
        if has_error:
            return 0
        if has_blank:
            return 1
        return 2

    def _populate_scan_grid_from_results(
        self,
        results: list[OMRResult],
        forced_status_by_image: dict[str, str] | None = None,
        skip_expensive_checks: bool = False,
        preserve_selection_image_path: str = "",
        refresh_statuses: bool = False,
        rebuild_error_list: bool = False,
    ) -> None:
        # scan_list columns: 0 stt, 1 sid, 2 exam_room, 3 exam_code, 4 full_name, 5 birth_date, 6 content, 7 status, 8 actions
        forced_status_by_image = forced_status_by_image or {}
        duplicate_ids: dict[str, int] = {}
        subject_scope: tuple[set[str], set[str]] | None = None
        available_exam_codes: set[str] | None = None
        if not skip_expensive_checks:
            for res in results:
                sid = str(getattr(res, "student_id", "") or "").strip()
                if not self._student_id_has_recognition_error(sid):
                    duplicate_ids[sid] = duplicate_ids.get(sid, 0) + 1
            subject_scope = self._subject_student_room_scope()
            available_exam_codes = self._available_exam_codes()

        row_views: list[dict[str, object]] = []
        for result in results:
            self._refresh_student_profile_for_result(result)
            scoped = self._scoped_result_copy(result)
            sid = str(result.student_id or "").strip()
            exam_code_text = str(result.exam_code or "").strip()
            image_path = str(result.image_path or "")
            subject_cfg = self._selected_batch_subject_config() or {}
            room_text = ""
            if sid:
                room_text = str(self._subject_room_for_student_id(sid, subject_cfg) or "").strip()
            if not room_text:
                room_text = str(getattr(result, "exam_room", "") or "").strip()
            if not room_text:
                profile = self._student_profile_by_id(sid) if sid else {}
                room_text = str((profile or {}).get("exam_room", "") or "").strip()
            setattr(result, "exam_room", room_text)
            forced_status = str(forced_status_by_image.get(image_path, "") or "")
            status_override = ""
            if forced_status:
                status_override = forced_status
            elif skip_expensive_checks:
                status_override = str(getattr(result, "cached_status", "") or "OK")
            payload = self._build_scan_row_payload_from_result(
                result,
                row_idx=None,
                duplicate_count=duplicate_ids.get(sid, 0),
                subject_scope=None if skip_expensive_checks else subject_scope,
                available_exam_codes=None if skip_expensive_checks else available_exam_codes,
                forced_status=status_override,
            )
            setattr(result, "cached_forced_status", str(forced_status or getattr(result, "cached_forced_status", "") or ""))
            payload["forced_status"] = str(forced_status or payload.get("forced_status", "") or "")
            row_views.append(payload)

        row_views.sort(
            key=lambda item: (
                self._row_sort_bucket(str(item["status"]), item["blank_map"]),
                self._student_sort_token(str(item["student_id"])),
                str(item["exam_code"] or "").casefold(),
                Path(str(item["image_path"] or "")).name.casefold(),
            )
        )

        self.scan_results = [item["result"] for item in row_views]
        self.scan_edit_history = {}
        self.scan_last_adjustment = {}
        self.scan_manual_adjustments = {}
        for res in self.scan_results:
            image_key = self._result_identity_key(getattr(res, "image_path", ""))
            if not image_key:
                continue
            history = [str(x) for x in (getattr(res, "edit_history", []) or []) if str(x or "").strip()]
            if history:
                self.scan_edit_history[image_key] = list(history)
                self.scan_last_adjustment[image_key] = str(getattr(res, "last_adjustment", "") or history[-1])
                setattr(res, "manually_edited", True)
                if not str(getattr(res, "cached_forced_status", "") or "").strip():
                    setattr(res, "cached_forced_status", "Đã sửa")
                self.scan_forced_status_by_index[image_key] = "Đã sửa"
            manual_items = [str(x) for x in (getattr(res, "manual_adjustments", []) or []) if str(x or "").strip()]
            if manual_items:
                self.scan_manual_adjustments[image_key] = list(sorted(set(manual_items)))
        subject_key = self._current_batch_subject_key()
        if subject_key:
            self.scan_results_by_subject[self._batch_result_subject_key(subject_key)] = list(self.scan_results)
        self.scan_blank_questions = {idx: list(item["blank_map"].get("MCQ", [])) for idx, item in enumerate(row_views)}
        self.scan_blank_summary = {idx: dict(item["blank_map"]) for idx, item in enumerate(row_views)}
        self.scan_forced_status_by_index = {
            self._result_identity_key(str(item.get("image_path", "") or "")): str(item["forced_status"] or "")
            for item in row_views
            if self._result_identity_key(str(item.get("image_path", "") or "")) and str(item["forced_status"] or "")
        }

        scan_list = self.scan_list
        selected_image = str(preserve_selection_image_path or "").strip()
        if not selected_image and 0 <= scan_list.currentRow() < scan_list.rowCount():
            current_item = scan_list.item(scan_list.currentRow(), self.SCAN_COL_STUDENT_ID)
            selected_image = str(current_item.data(Qt.UserRole) if current_item else "").strip()

        self._begin_scan_grid_update()
        try:
            scan_list.setRowCount(len(row_views))
            for idx, item in enumerate(row_views):
                payload = {
                    "student_id": item.get("student_id", ""),
                    "exam_room": item.get("exam_room", ""),
                    "exam_code": item.get("exam_code", ""),
                    "full_name": item.get("full_name", "-"),
                    "birth_date": item.get("birth_date", "-"),
                    "content": item.get("content", ""),
                    "status": item.get("status", "OK"),
                    "recognized_short": item.get("recognized_short", ""),
                    "image_path": item.get("image_path", ""),
                    "serialized_result": item.get("serialized_result", {}),
                }
                self._apply_scan_row_payload_to_grid(idx, payload, skip_actions=skip_expensive_checks)
            scan_list.resizeRowsToContents()
            for fit_col in [self.SCAN_COL_STT, self.SCAN_COL_STUDENT_ID, self.SCAN_COL_EXAM_ROOM, self.SCAN_COL_EXAM_CODE, self.SCAN_COL_FULL_NAME, self.SCAN_COL_BIRTH_DATE, self.SCAN_COL_ACTIONS]:
                scan_list.resizeColumnToContents(fit_col)
        finally:
            self._end_scan_grid_update()

        if refresh_statuses:
            self._refresh_all_statuses()
        if rebuild_error_list:
            self._rebuild_error_list()

        if selected_image:
            for row in range(scan_list.rowCount()):
                cell = scan_list.item(row, self.SCAN_COL_STUDENT_ID)
                if str(cell.data(Qt.UserRole) if cell else "").strip() == selected_image:
                    scan_list.setCurrentCell(row, self.SCAN_COL_STUDENT_ID)
                    scan_list.selectRow(row)
                    break

    def _finalize_batch_scan_display(self, refresh_statuses: bool = True) -> None:
        if not hasattr(self, "scan_list"):
            return
        if refresh_statuses:
            self._refresh_all_statuses()
        self._rebuild_error_list()
        self._apply_scan_filter()
        if self.scan_list.rowCount() <= 0:
            if hasattr(self, "scan_result_preview"):
                self.scan_result_preview.setRowCount(0)
            if hasattr(self, "result_preview"):
                self.result_preview.clear()
            if hasattr(self, "manual_edit"):
                self.manual_edit.clear()
            return
        target_row = -1
        for row in range(self.scan_list.rowCount()):
            if not self.scan_list.isRowHidden(row):
                target_row = row
                break
        if target_row < 0:
            target_row = 0
        self.scan_list.setCurrentCell(target_row, self.SCAN_COL_STUDENT_ID)
        self.scan_list.selectRow(target_row)
        self._on_scan_selected()
        self._refresh_ribbon_action_states()

    def _build_scan_row_payload_from_result(
        self,
        result,
        row_idx: int | None = None,
        duplicate_count: int = 1,
        subject_scope: tuple[set[str], set[str]] | None = None,
        available_exam_codes: set[str] | None = None,
        forced_status: str = "",
    ) -> dict:
        scoped_result = self._scoped_result_copy(result)
        blank_map = self._compute_blank_questions(scoped_result)
        expected_by_section = self._expected_questions_by_section(scoped_result)
        sid = str(getattr(result, "student_id", "") or "").strip()
        exam_code_text = str(getattr(result, "exam_code", "") or "").strip()
        room_text = str(self._subject_room_for_student_id(sid) or "").strip() if sid else ""
        if not room_text:
            room_text = str(getattr(result, "exam_room", "") or "").strip()
        if not room_text and sid:
            profile = self._student_profile_by_id(sid)
            room_text = str((profile or {}).get("exam_room", "") or "").strip()
        setattr(result, "exam_room", room_text)

        status_parts = self._status_parts_for_result(
            result,
            duplicate_count,
            subject_scope=subject_scope,
            available_exam_codes=available_exam_codes,
        )
        has_edit_history = bool([str(x) for x in (getattr(result, "edit_history", []) or []) if str(x or "").strip()])
        is_manual_edited = bool(getattr(result, "manually_edited", False)) or has_edit_history
        forced_status_text = str(forced_status or "").strip()
        status_parts_text = ", ".join(status_parts) if status_parts else ""
        if is_manual_edited:
            if status_parts_text:
                status_text = f"Đã sửa ({status_parts_text})"
            elif forced_status_text and forced_status_text not in {"Đã sửa", "OK"}:
                status_text = f"Đã sửa ({forced_status_text})"
            else:
                status_text = "Đã sửa"
            effective_forced_status = "Đã sửa"
        else:
            effective_forced_status = forced_status_text
            status_text = effective_forced_status or status_parts_text or "OK"
        manual_content_override = str(getattr(result, "manual_content_override", "") or "").strip()
        content_text = self._build_recognition_content_text(scoped_result, blank_map, expected_by_section)
        recognized_short = self._short_recognition_text_for_result(scoped_result)

        if row_idx is not None and row_idx >= 0:
            self.scan_blank_summary[row_idx] = dict(blank_map)
            self.scan_blank_questions[row_idx] = list(blank_map.get("MCQ", []))
            image_key_for_row = self._row_image_key(row_idx) or self._result_identity_key(getattr(result, "image_path", ""))
            if effective_forced_status:
                if image_key_for_row:
                    self.scan_forced_status_by_index[image_key_for_row] = effective_forced_status
            elif image_key_for_row and image_key_for_row in self.scan_forced_status_by_index:
                self.scan_forced_status_by_index.pop(image_key_for_row, None)

        setattr(result, "cached_status", status_text)
        setattr(result, "cached_content", content_text)
        setattr(result, "cached_recognized_short", recognized_short)
        setattr(result, "cached_blank_summary", dict(blank_map))
        setattr(result, "manual_content_override", manual_content_override)
        setattr(result, "cached_forced_status", effective_forced_status)

        return {
            "result": result,
            "image_path": str(getattr(result, "image_path", "") or ""),
            "student_id": sid,
            "exam_room": room_text,
            "exam_code": exam_code_text,
            "full_name": str(getattr(result, "full_name", "") or "-"),
            "birth_date": str(getattr(result, "birth_date", "") or "-"),
            "content": content_text,
            "manual_content_override": manual_content_override,
            "status": status_text,
            "recognized_short": recognized_short,
            "forced_status": effective_forced_status,
            "manually_edited": is_manual_edited,
            "edit_history": [str(x) for x in (getattr(result, "edit_history", []) or []) if str(x or "").strip()],
            "last_adjustment": str(getattr(result, "last_adjustment", "") or ""),
            "blank_map": dict(blank_map),
            "serialized_result": self._serialize_omr_result(result),
            "recognized_template_path": str(getattr(result, "recognized_template_path", "") or ""),
            "recognized_alignment_profile": str(getattr(result, "recognized_alignment_profile", "") or ""),
            "recognized_fill_threshold": float(getattr(result, "recognized_fill_threshold", 0.45) or 0.45),
            "recognized_empty_threshold": float(getattr(result, "recognized_empty_threshold", 0.20) or 0.20),
            "recognized_certainty_margin": float(getattr(result, "recognized_certainty_margin", 0.08) or 0.08),
        }

    def _apply_scan_row_payload_to_grid(self, row_idx: int, payload: dict, *, skip_actions: bool = False) -> None:
        if row_idx < 0:
            return
        if row_idx >= self.scan_list.rowCount():
            self.scan_list.setRowCount(row_idx + 1)
        self.scan_list.setItem(row_idx, self.SCAN_COL_STT, QTableWidgetItem(str(row_idx + 1)))
        sid_item = QTableWidgetItem(str(payload.get("student_id", "") or "-"))
        sid_item.setData(Qt.UserRole, str(payload.get("image_path", "") or ""))
        sid_item.setData(Qt.UserRole + 1, str(payload.get("exam_code", "") or ""))
        sid_item.setData(Qt.UserRole + 2, str(payload.get("recognized_short", "") or ""))
        sid_item.setData(Qt.UserRole + 10, dict(payload.get("serialized_result", {}) or {}))
        sid_item.setData(Qt.UserRole + 11, str(payload.get("manual_content_override", "") or ""))
        sid_item.setData(Qt.UserRole + 12, dict(payload or {}))
        self.scan_list.setItem(row_idx, self.SCAN_COL_STUDENT_ID, sid_item)
        self.scan_list.setItem(row_idx, self.SCAN_COL_EXAM_ROOM, QTableWidgetItem(str(payload.get("exam_room", "") or "-")))
        self.scan_list.setItem(row_idx, self.SCAN_COL_EXAM_CODE, QTableWidgetItem(str(payload.get("exam_code", "") or "-")))
        self.scan_list.setItem(row_idx, self.SCAN_COL_FULL_NAME, QTableWidgetItem(str(payload.get("full_name", "") or "-")))
        self.scan_list.setItem(row_idx, self.SCAN_COL_BIRTH_DATE, QTableWidgetItem(str(payload.get("birth_date", "") or "-")))
        content_text = str(payload.get("content", "") or "")
        content_item = QTableWidgetItem(content_text)
        content_item.setToolTip(content_text)
        self.scan_list.setItem(row_idx, self.SCAN_COL_CONTENT, content_item)
        full_status = str(payload.get("status", "") or "OK")
        status_item = QTableWidgetItem(self._compact_status_text(full_status, max_len=150))
        status_item.setToolTip(full_status)
        if full_status != "OK":
            status_item.setForeground(Qt.red)
        self.scan_list.setItem(row_idx, self.SCAN_COL_STATUS, status_item)
        if skip_actions:
            self.scan_list.setItem(row_idx, self.SCAN_COL_ACTIONS, QTableWidgetItem("..."))
        else:
            self._set_scan_action_widget(row_idx)

    def _update_scan_row_from_result(self, idx: int, result) -> None:
        # scan_list columns: 0 stt, 1 sid, 2 exam_room, 3 exam_code, 4 full_name, 5 birth_date, 6 content, 7 status, 8 actions
        target_idx = int(idx)
        image_key = self._result_identity_key(getattr(result, "image_path", ""))
        if image_key:
            mapped_idx = self._row_index_by_image_path(image_key)
            if mapped_idx >= 0:
                target_idx = mapped_idx
        if target_idx < 0 or target_idx >= self.scan_list.rowCount():
            return
        self._refresh_student_profile_for_result(result)
        payload = self._build_scan_row_payload_from_result(result, row_idx=target_idx)
        self._apply_scan_row_payload_to_grid(target_idx, payload)

    def _current_scan_results_snapshot(self) -> list[OMRResult]:
        # scan_list columns: 0 stt, 1 sid, 2 exam_room, 3 exam_code, 4 full_name, 5 birth_date, 6 content, 7 status, 8 actions
        base = list(self.scan_results or [])
        if not hasattr(self, "scan_list"):
            return base
        row_count = self.scan_list.rowCount()
        if row_count <= 0:
            return base

        base_by_image: dict[str, list[OMRResult]] = {}
        for src in base:
            key = self._result_identity_key(getattr(src, "image_path", ""))
            if not key:
                continue
            base_by_image.setdefault(key, []).append(src)

        out: list[OMRResult] = []
        for idx in range(row_count):
            sid_item = self.scan_list.item(idx, self.SCAN_COL_STUDENT_ID)
            sid_text = str(sid_item.text() if sid_item else "").strip()
            image_path = str(sid_item.data(Qt.UserRole) if sid_item else "").strip()
            serialized_payload = sid_item.data(Qt.UserRole + 10) if sid_item else None
            result = None
            if isinstance(serialized_payload, dict) and serialized_payload:
                try:
                    result = self._deserialize_omr_result(serialized_payload)
                except Exception:
                    result = None
            if result is None:
                matched_base = None
                image_key = self._result_identity_key(image_path)
                if image_key and image_key in base_by_image and base_by_image[image_key]:
                    matched_base = base_by_image[image_key].pop(0)
                if matched_base is not None:
                    result = self._lightweight_result_copy(matched_base)
                elif image_key:
                    fallback = self._build_result_from_saved_table_row(idx)
                    result = fallback if fallback is not None else OMRResult(image_path=image_path)
                elif idx < len(base):
                    result = self._lightweight_result_copy(base[idx])
                else:
                    fallback = self._build_result_from_saved_table_row(idx)
                    result = fallback if fallback is not None else OMRResult(image_path="")

            result.student_id = "" if sid_text in {"", "-"} else sid_text
            result.exam_code = str(sid_item.data(Qt.UserRole + 1) if sid_item else "").strip()
            if image_path:
                result.image_path = image_path
            setattr(result, "exam_room", str(self.scan_list.item(idx, self.SCAN_COL_EXAM_ROOM).text() if self.scan_list.item(idx, self.SCAN_COL_EXAM_ROOM) else ""))
            setattr(result, "full_name", str(self.scan_list.item(idx, self.SCAN_COL_FULL_NAME).text() if self.scan_list.item(idx, self.SCAN_COL_FULL_NAME) else ""))
            setattr(result, "birth_date", str(self.scan_list.item(idx, self.SCAN_COL_BIRTH_DATE).text() if self.scan_list.item(idx, self.SCAN_COL_BIRTH_DATE) else ""))
            setattr(result, "manual_content_override", str(sid_item.data(Qt.UserRole + 11) if sid_item else "").strip())
            result.answer_string = self._normalize_non_api_answer_string(result)
            payload = self._build_scan_row_payload_from_result(result, row_idx=idx)
            setattr(result, "cached_content", str(payload.get("content", "") or ""))
            setattr(result, "cached_status", str(payload.get("status", "") or ""))
            setattr(result, "cached_recognized_short", str(payload.get("recognized_short", "") or ""))
            self._debug_scan_result_state("current_scan_results_snapshot", result)
            out.append(result)
        return out

    @staticmethod
    def _result_has_recognition_payload(result: OMRResult | None) -> bool:
        if result is None:
            return False
        if bool(str(getattr(result, "answer_string", "") or "").strip()):
            return True
        if bool(getattr(result, "mcq_answers", {}) or {}):
            return True
        if bool(getattr(result, "true_false_answers", {}) or {}):
            return True
        if bool(getattr(result, "numeric_answers", {}) or {}):
            return True
        if bool(getattr(result, "recognition_errors", []) or []):
            return True
        return bool(getattr(result, "issues", []) or [])

    def _restore_full_result_for_row(self, row_idx: int) -> OMRResult | None:
        if row_idx < 0 or row_idx >= self.scan_list.rowCount():
            return None
        sid_item = self.scan_list.item(row_idx, self.SCAN_COL_STUDENT_ID)
        image_path = str(sid_item.data(Qt.UserRole) if sid_item else "")
        sid_text = str(sid_item.text() if sid_item else "").strip()
        exam_code = str(sid_item.data(Qt.UserRole + 1) if sid_item else "").strip()
        if sid_text == "-":
            sid_text = ""

        def _match_result_pool(pool: list[OMRResult]) -> OMRResult | None:
            sid_matches: list[OMRResult] = []
            for cand in (pool or []):
                cand_img = str(getattr(cand, "image_path", "") or "").strip()
                cand_sid = str(getattr(cand, "student_id", "") or "").strip()
                cand_code = str(getattr(cand, "exam_code", "") or "").strip()
                if image_path and cand_img == image_path:
                    return cand
                if (not image_path) and sid_text and cand_sid == sid_text and (not exam_code or cand_code == exam_code):
                    sid_matches.append(cand)
            if len(sid_matches) == 1:
                return sid_matches[0]
            return None

        if 0 <= row_idx < len(self.scan_results):
            direct = self.scan_results[row_idx]
            if self._result_has_recognition_payload(direct):
                return self._lightweight_result_copy(direct)

        subject_key = self._current_batch_subject_key()
        subject_pool = list(self.scan_results_by_subject.get(self._batch_result_subject_key(subject_key), []) or [])
        matched = _match_result_pool(subject_pool)
        if matched is not None and self._result_has_recognition_payload(matched):
            return self._lightweight_result_copy(matched)

        db_rows = self.database.fetch_scan_results_for_subject(self._batch_result_subject_key(subject_key))
        db_pool: list[OMRResult] = []
        for item in db_rows:
            try:
                db_pool.append(self._deserialize_omr_result(item))
            except Exception:
                continue
        matched = _match_result_pool(db_pool)
        if matched is not None and self._result_has_recognition_payload(matched):
            return self._lightweight_result_copy(matched)

        return self._build_result_from_saved_table_row(row_idx)
    def _build_result_from_saved_table_row(self, idx: int) -> OMRResult | None:
        # scan_list columns: 0 stt, 1 sid, 2 exam_room, 3 exam_code, 4 full_name, 5 birth_date, 6 content, 7 status, 8 actions
        if idx < 0 or idx >= self.scan_list.rowCount():
            return None
        sid_item = self.scan_list.item(idx, self.SCAN_COL_STUDENT_ID)
        image_path = str(sid_item.data(Qt.UserRole) if sid_item else "")
        if not image_path:
            return None
        student_id = str(sid_item.text() if sid_item else "").strip()
        if student_id == "-":
            student_id = ""
        exam_code = str(sid_item.data(Qt.UserRole + 1) if sid_item else "").strip()
        result = OMRResult(image_path=image_path, student_id=student_id, exam_code=exam_code)
        serialized_payload = sid_item.data(Qt.UserRole + 10) if sid_item else None
        row_payload = sid_item.data(Qt.UserRole + 12) if sid_item else None
        if isinstance(serialized_payload, dict) and serialized_payload:
            try:
                restored = self._deserialize_omr_result(serialized_payload)
                result = restored
            except Exception:
                pass
        result = self._hydrate_result_from_saved_row_payload(result, row_payload) or result
        result.image_path = image_path or str(getattr(result, "image_path", "") or "")
        result.student_id = student_id
        result.exam_code = exam_code
        room_text = str(self.scan_list.item(idx, self.SCAN_COL_EXAM_ROOM).text() if self.scan_list.item(idx, self.SCAN_COL_EXAM_ROOM) else "").strip()
        if not room_text and student_id:
            room_text = str(self._subject_room_for_student_id(student_id) or "").strip()
        result.exam_room = room_text
        result.full_name = str(self.scan_list.item(idx, self.SCAN_COL_FULL_NAME).text() if self.scan_list.item(idx, self.SCAN_COL_FULL_NAME) else "")
        result.birth_date = str(self.scan_list.item(idx, self.SCAN_COL_BIRTH_DATE).text() if self.scan_list.item(idx, self.SCAN_COL_BIRTH_DATE) else "")
        manual_content_override = str(sid_item.data(Qt.UserRole + 11) if sid_item else "").strip()
        setattr(result, "manual_content_override", manual_content_override)
        result.answer_string = self._normalize_non_api_answer_string(result)
        payload = self._build_scan_row_payload_from_result(result, row_idx=idx)
        setattr(result, "cached_content", str(payload.get("content", "") or ""))
        setattr(result, "cached_status", str(payload.get("status", "") or ""))
        setattr(result, "cached_recognized_short", str(payload.get("recognized_short", "") or ""))
        self._debug_scan_result_state("build_result_from_saved_table_row", result)
        result.sync_legacy_aliases()
        return result

    def _ensure_template_for_selected_subject(self) -> bool:
        cfg = self._selected_batch_subject_config() or self._resolve_subject_config_for_batch()
        template_path = self._resolve_template_path_for_subject(cfg)
        if not template_path:
            return False
        pth = Path(template_path)
        if not pth.exists():
            return False
        active_template_path = str(getattr(self, "_active_template_path", "") or "").strip()
        desired_template_path = str(pth.resolve())
        try:
            if (self.template is None) or (active_template_path != desired_template_path):
                self.template = Template.load_json(pth)
                setattr(self, "_active_template_path", desired_template_path)
            self._apply_template_recognition_settings(self.template, sync_mode_selector=False)
            return True
        except Exception:
            return False

    @staticmethod
    def _aligned_image_to_qpixmap(image) -> QPixmap:
        try:
            if image is None:
                return QPixmap()
            h, w, ch = image.shape
            if h <= 0 or w <= 0 or ch < 3:
                return QPixmap()
            rgb = image[:, :, :3][:, :, ::-1].copy()
            qimg = QImage(rgb.data, w, h, 3 * w, QImage.Format_RGB888)
            return QPixmap.fromImage(qimg.copy())
        except Exception:
            return QPixmap()

    def _recognition_overlay_positions_for_result(self, result: OMRResult) -> list[dict[str, float]]:
        if self.template is None or self.preview_source_pixmap.isNull():
            return []
        states_by_zone = getattr(result, "bubble_states_by_zone", None)
        if not isinstance(states_by_zone, dict) or not states_by_zone:
            return []
        tpl_w = max(1.0, float(self.template.width))
        tpl_h = max(1.0, float(self.template.height))
        img_w = max(1.0, float(self.preview_source_pixmap.width()))
        img_h = max(1.0, float(self.preview_source_pixmap.height()))
        sx, sy = img_w / tpl_w, img_h / tpl_h
        markers: list[dict[str, float]] = []
        for z in self.template.zones:
            g = z.grid
            if not g or not g.bubble_positions:
                continue
            states = list(states_by_zone.get(z.id, []) or [])
            if not states:
                continue
            for i, pos in enumerate(g.bubble_positions):
                if i >= len(states) or not bool(states[i]):
                    continue
                bx, by = pos
                markers.append({"zone_id": z.id, "x": float(bx) * sx, "y": float(by) * sy})
        return markers

    def _set_scan_result_at_row(self, idx: int, result: OMRResult) -> None:
        if idx < 0:
            return
        result_image = self._result_identity_key(getattr(result, "image_path", ""))
        if result_image:
            for i, existing in enumerate(self.scan_results):
                if self._result_identity_key(getattr(existing, "image_path", "")) == result_image:
                    if any(j != i and self.scan_results[j] is result for j in range(len(self.scan_results))):
                        result = self._lightweight_result_copy(result)
                    self.scan_results[i] = result
                    return
        if idx == len(self.scan_results):
            self.scan_results.append(self._lightweight_result_copy(result))
            return
        if idx > len(self.scan_results):
            return
        if any(i != idx and self.scan_results[i] is result for i in range(len(self.scan_results))):
            result = self._lightweight_result_copy(result)
        self.scan_results[idx] = result

    def _rerecognize_selected_scan(self) -> None:
        idx = self._selected_scan_row_index()
        if idx < 0:
            QMessageBox.warning(self, "Nhận dạng lại", "Chọn một bài thi trong danh sách bên trái trước.")
            return
        if not self._ensure_template_for_selected_subject() or not self.template:
            QMessageBox.warning(self, "Nhận dạng lại", "Chưa có template khả dụng (theo môn hoặc theo kỳ thi).")
            return

        old_result = self.scan_results[idx] if idx < len(self.scan_results) else self._build_result_from_saved_table_row(idx)
        if old_result is None:
            QMessageBox.warning(self, "Nhận dạng lại", "Không tìm thấy dữ liệu dòng đang chọn để nhận dạng lại.")
            return
        subject_key = self._current_batch_subject_key()
        has_existing_score = bool(subject_key and self._scan_has_existing_score(subject_key, old_result))
        confirm_message = "Bạn có chắc muốn nhận dạng lại bài thi đã chọn?"
        if has_existing_score:
            confirm_message = (
                "Bài thi này đã có điểm. Nhận dạng lại ảnh đã chọn có thể làm thay đổi điểm đã tính.\n\n"
                "Bạn có muốn tiếp tục không?"
            )
        confirm_rerun = QMessageBox.question(
            self,
            "Nhận dạng lại ảnh chọn",
            confirm_message,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm_rerun != QMessageBox.Yes:
            return
        image_path = str(old_result.image_path or "").strip()
        if not image_path or not Path(image_path).exists():
            QMessageBox.warning(self, "Nhận dạng lại", f"Không tìm thấy ảnh để nhận dạng lại:\n{image_path or '-'}")
            return

        process_path = image_path
        rotation = int(self.preview_rotation_by_index.get(idx, 0) or 0) % 360
        temp_rotated_path = None
        if rotation:
            pix = QPixmap(image_path)
            if pix.isNull():
                QMessageBox.warning(self, "Nhận dạng lại", "Không thể mở ảnh để xoay tạm thời trước khi nhận dạng lại.")
                return
            rotated = pix.transformed(QTransform().rotate(float(rotation)), Qt.SmoothTransformation)
            temp_rotated_path = str(Path(image_path).with_name(f".{Path(image_path).stem}_tmp_rerun_{rotation}.png"))
            if not rotated.save(temp_rotated_path):
                QMessageBox.warning(self, "Nhận dạng lại", "Không thể tạo ảnh xoay tạm thời để nhận dạng lại.")
                return
            process_path = temp_rotated_path

        recognized_template_path = str(getattr(old_result, "recognized_template_path", "") or "").strip()
        template_path = recognized_template_path or self._resolve_template_path_for_subject(self._selected_batch_subject_config() or self._resolve_subject_config_for_batch())
        wait_dlg = self._open_wait_progress("Đang nhận dạng lại ảnh đã chọn...")
        try:
            new_result = self._recognize_image_with_exact_template(process_path, template_path, source_tag="rerecognize_selected", allow_retry=False)
        finally:
            if temp_rotated_path:
                try:
                    Path(temp_rotated_path).unlink(missing_ok=True)
                except Exception:
                    pass
            self._close_wait_progress(wait_dlg)
        new_result.image_path = image_path
        setattr(new_result, "manually_edited", bool(getattr(old_result, "manually_edited", False)))
        setattr(new_result, "cached_forced_status", str(getattr(old_result, "cached_forced_status", "") or ""))
        image_key = self._result_identity_key(image_path)
        prior_history = [str(x) for x in (getattr(old_result, "edit_history", []) or []) if str(x or "").strip()]
        if image_key:
            prior_history = [str(x) for x in (self.scan_edit_history.get(image_key, []) or []) if str(x or "").strip()] or prior_history
        setattr(new_result, "edit_history", list(prior_history))
        setattr(new_result, "last_adjustment", str(getattr(old_result, "last_adjustment", "") or ""))
        setattr(new_result, "manual_adjustments", [str(x) for x in (getattr(old_result, "manual_adjustments", []) or []) if str(x or "").strip()])

        sid = (new_result.student_id or "").strip()
        profile = self._student_profile_by_id(sid)
        if profile.get("name"):
            setattr(new_result, "full_name", profile.get("name"))
        if profile.get("birth_date"):
            setattr(new_result, "birth_date", self._format_birth_date_mmddyyyy(profile.get("birth_date")))
        if profile.get("class_name"):
            setattr(new_result, "class_name", profile.get("class_name"))
        if profile.get("exam_room"):
            setattr(new_result, "exam_room", profile.get("exam_room"))
        scoped_new = self._scoped_result_copy(new_result)
        blank_map = self._compute_blank_questions(scoped_new)
        rec_errors = list(getattr(new_result, "recognition_errors", [])) or list(getattr(new_result, "errors", []))
        old_sid_for_score = str(getattr(old_result, "student_id", "") or "").strip()
        new_sid_for_score = str(getattr(new_result, "student_id", "") or "").strip()

        if rec_errors:
            # Failed/partial re-recognition must restart workflow like a newly recognized file:
            # clear manual-edit traces and reset forced-status cache.
            setattr(new_result, "manually_edited", False)
            setattr(new_result, "cached_forced_status", "")
            setattr(new_result, "cached_status", "")
            setattr(new_result, "edit_history", [])
            setattr(new_result, "manual_adjustments", [])
            setattr(new_result, "last_adjustment", "")
            if image_key:
                self.scan_edit_history.pop(image_key, None)
                self.scan_manual_adjustments.pop(image_key, None)
                self.scan_last_adjustment.pop(image_key, None)
                self.scan_forced_status_by_index.pop(image_key, None)
            # If re-recognition has errors, scoring must be reset and follow full scoring flow again.
            self._invalidate_scoring_for_student_ids(
                [sid for sid in [old_sid_for_score, new_sid_for_score] if sid],
                subject_key=subject_key,
                reason="rerecognize_with_errors_reset",
            )

        self._set_scan_result_at_row(idx, new_result)
        subject_key = self._current_batch_subject_key()
        if subject_key:
            self.scan_results_by_subject[self._batch_result_subject_key(subject_key)] = list(self.scan_results)
        self.scan_blank_questions[idx] = blank_map.get("MCQ", [])
        self.scan_blank_summary[idx] = blank_map
        self._update_scan_row_from_result(idx, new_result)
        self._refresh_all_statuses()
        self._rebuild_error_list()
        self._update_scan_preview(idx)
        self._sync_correction_detail_panel(new_result, rebuild_editor=False)
        if prior_history and not rec_errors:
            old_sid = str(getattr(old_result, "student_id", "") or "").strip() or "-"
            old_code = str(getattr(old_result, "exam_code", "") or "").strip() or "-"
            new_sid = str(getattr(new_result, "student_id", "") or "").strip() or "-"
            new_code = str(getattr(new_result, "exam_code", "") or "").strip() or "-"
            self._record_adjustment(
                idx,
                [f"Nhận dạng lại ảnh chọn: STUDENT ID {old_sid} -> {new_sid}; Mã đề {old_code} -> {new_code}"],
                "rerecognize_selected",
            )
            self._persist_single_scan_result_to_db(new_result, note="rerecognize_selected_with_history")
            self._refresh_all_statuses()
        self._persist_single_scan_result_to_db(new_result, note="rerecognize_selected_scan")
        self.btn_save_batch_subject.setEnabled(False)

        summary = (
            f"Ảnh: {Path(image_path).name}\n"
            f"STUDENT ID: {new_result.student_id or '-'}\n"
            f"Họ tên: {str(getattr(new_result, 'full_name', '') or '-')}\n"
            f"Ngày sinh: {str(getattr(new_result, 'birth_date', '') or '-')}\n"
            f"Mã đề: {new_result.exam_code or '-'}\n"
            f"Nhận dạng ngắn: {self._compact_value(self._short_recognition_text_for_result(scoped_new), 180)}\n"
            f"Số lỗi nhận dạng: {len(rec_errors)}"
        )
        QMessageBox.information(self, "Kết quả nhận dạng lại", summary)

    def _scan_has_existing_score(self, subject_key: str, result: OMRResult) -> bool:
        if not subject_key or result is None:
            return False
        sid = str(getattr(result, "student_id", "") or "").strip()
        exam_code = str(getattr(result, "exam_code", "") or "").strip()

        # DB-first: kiểm tra trực tiếp trong scores đã persist.
        try:
            db_rows = list(self.database.fetch_scores_for_subject(subject_key) or [])
        except Exception:
            db_rows = []
        for row in db_rows:
            row_sid = str((row or {}).get("student_id", "") or "").strip()
            row_exam = str((row or {}).get("exam_code", "") or "").strip()
            if sid and sid == row_sid:
                return True
            if exam_code and row_exam and exam_code == row_exam and ((not sid) or (not row_sid)):
                return True
        return False
    def _render_preview_pixmap(self) -> None:
        if self.preview_source_pixmap.isNull():
            self.scan_image_preview.setPixmap(QPixmap())
            return
        src_size = self.preview_source_pixmap.size()
        target_w = max(1, int(src_size.width() * self.preview_zoom_factor))
        target_h = max(1, int(src_size.height() * self.preview_zoom_factor))
        scaled = self.preview_source_pixmap.scaled(
            target_w,
            target_h,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        self.scan_image_preview.setPixmap(scaled)
        self.scan_image_preview.resize(scaled.size())
        self.scan_image_preview.adjustSize()

    def _zoom_preview_in(self) -> None:
        self.preview_zoom_factor = min(4.0, self.preview_zoom_factor + 0.1)
        self._render_preview_pixmap()
        self.btn_zoom_reset.setText(f"{int(self.preview_zoom_factor*100)}%")

    def _zoom_preview_out(self) -> None:
        self.preview_zoom_factor = max(0.3, self.preview_zoom_factor - 0.1)
        self._render_preview_pixmap()
        self.btn_zoom_reset.setText(f"{int(self.preview_zoom_factor*100)}%")

    def _zoom_preview_reset(self) -> None:
        self.preview_zoom_factor = 0.3
        self._render_preview_pixmap()
        self.btn_zoom_reset.setText("30%")

    @staticmethod
    def _compact_value(value, limit: int = 120) -> str:
        text = str(value)
        return text if len(text) <= limit else text[:limit] + "..."

    def _update_scan_preview_from_saved_row(self, row: int) -> None:
        sid = self.scan_list.item(row, self.SCAN_COL_STUDENT_ID).text() if self.scan_list.item(row, self.SCAN_COL_STUDENT_ID) else "-"
        exam_code_cell = self.scan_list.item(row, self.SCAN_COL_EXAM_CODE).text() if self.scan_list.item(row, self.SCAN_COL_EXAM_CODE) else "-"
        content = self.scan_list.item(row, self.SCAN_COL_CONTENT).text() if self.scan_list.item(row, self.SCAN_COL_CONTENT) else "-"
        status = self.scan_list.item(row, self.SCAN_COL_STATUS).text() if self.scan_list.item(row, self.SCAN_COL_STATUS) else "-"
        img_path = ""
        exam_code = ""
        item0 = self.scan_list.item(row, self.SCAN_COL_STUDENT_ID)
        if item0:
            img_path = str(item0.data(Qt.UserRole) or "")
            exam_code = str(item0.data(Qt.UserRole + 1) or "")

        pix = QPixmap(img_path) if img_path else QPixmap()
        if pix.isNull():
            self.preview_source_pixmap = QPixmap()
            self.scan_image_preview.setPixmap(QPixmap())
            self.scan_image_preview.setText("Không có ảnh tương ứng cho dòng đã lưu")
            self.scan_image_preview.clear_markers()
            self.btn_zoom_reset.setText("30%")
        else:
            rotation = int(self.preview_rotation_by_index.get(row, 0) or 0) % 360
            if rotation:
                pix = pix.transformed(QTransform().rotate(float(rotation)), Qt.SmoothTransformation)
            self.preview_source_pixmap = pix
            self._render_preview_pixmap()
            self.btn_zoom_reset.setText(f"{int(self.preview_zoom_factor*100)}%")

        tmp_result = self._restore_full_result_for_row(row)
        if tmp_result is not None:
            self.scan_image_preview.set_overlay_markers(self._recognition_overlay_positions_for_result(tmp_result))
        else:
            self.scan_image_preview.clear_markers()

        rows = [
            ("File ảnh", Path(str(img_path or "")).name or "-"),
            ("STUDENT ID", sid),
            ("Mã đề", exam_code or exam_code_cell or "-"),
            ("Nội dung", self._compact_value(content, 220)),
            ("Status", status),
        ]
        self.scan_result_preview.setRowCount(0)
        for r, (k, v) in enumerate(rows):
            self.scan_result_preview.insertRow(r)
            self.scan_result_preview.setItem(r, 0, QTableWidgetItem(str(k)))
            self.scan_result_preview.setItem(r, 1, QTableWidgetItem(str(v)))

    def _on_scan_selected(self) -> None:
        if self._scan_grid_loading:
            return
        index = self.scan_list.currentRow()
        if index < 0:
            return
        self._ensure_scan_action_widget(index)
        result = self._restore_full_result_for_row(index)
        if result is not None:
            if index >= len(self.scan_results):
                self._set_scan_result_at_row(index, result)
            self._update_scan_preview(index)
            self._load_selected_result_for_correction()
            return
        self._update_scan_preview_from_saved_row(index)

    @staticmethod
    def _normalize_exam_code_text(code: str) -> str:
        c = str(code or "").strip()
        if not c:
            return ""
        if c.isdigit():
            # Treat purely numeric exam codes with/without leading zeros as equivalent.
            c2 = c.lstrip("0")
            return c2 if c2 else "0"
        return c

    def _available_exam_codes(self, subject_key: str = "") -> set[str]:
        out: set[str] = set()
        subject = str(subject_key or self._current_batch_subject_key() or "").strip()
        if subject:
            cfg = self._subject_config_by_subject_key(subject) or {}
            imported_cfg = cfg.get("imported_answer_keys", {}) if isinstance(cfg.get("imported_answer_keys", {}), dict) else {}
            for raw in imported_cfg.keys():
                code_text = str(raw).strip()
                if not code_text:
                    continue
                out.add(code_text)
                out.add(self._normalize_exam_code_text(code_text))
            try:
                fetched = self._fetch_answer_keys_for_subject_scoped(subject)
            except Exception:
                fetched = {}
            for raw in (fetched or {}).keys():
                code_text = str(raw).strip()
                if not code_text:
                    continue
                out.add(code_text)
                out.add(self._normalize_exam_code_text(code_text))
            if self.answer_keys is not None:
                for key_obj in self.answer_keys.keys.values():
                    key_subject = str(getattr(key_obj, "subject", "") or "").strip()
                    code_text = str(getattr(key_obj, "exam_code", "") or "").strip()
                    if key_subject != subject or not code_text:
                        continue
                    out.add(code_text)
                    out.add(self._normalize_exam_code_text(code_text))
        if not out:
            for x in (self.imported_exam_codes or []):
                raw = str(x).strip()
                if not raw:
                    continue
                out.add(raw)
                out.add(self._normalize_exam_code_text(raw))
        return {v for v in out if v}

    def _subject_answer_key_for_result(self, result, subject_key: str = ""):
        subject = str(subject_key or self._current_batch_subject_key() or "").strip()
        exam_code = str(getattr(result, "exam_code", "") or "").strip()
        if not subject or not exam_code:
            return None
        if self.answer_keys is not None:
            key = self.answer_keys.get(subject, exam_code)
            if key is not None:
                return key
            normalized = self._normalize_exam_code_text(exam_code)
            for candidate in (self.imported_exam_codes or []):
                candidate_text = str(candidate).strip()
                if candidate_text and self._normalize_exam_code_text(candidate_text) == normalized:
                    key = self.answer_keys.get(subject, candidate_text)
                    if key is not None:
                        return key
        fetched = self._fetch_answer_keys_for_subject_scoped(subject)
        if exam_code in fetched:
            return fetched[exam_code]
        normalized = self._normalize_exam_code_text(exam_code)
        for candidate_text, candidate_key in fetched.items():
            if self._normalize_exam_code_text(candidate_text) == normalized:
                return candidate_key
        return None

    @staticmethod
    def _answer_string_from_maps(
        mcq_answers: dict[int, str],
        tf_answers: dict[int, dict[str, bool]],
        numeric_answers: dict[int, str],
        answer_key,
        *,
        use_semicolon: bool = True,
    ) -> str:
        if answer_key is None:
            return ""

        def _question_numbers(valid_map, invalid_map, fallback_map=None) -> list[int]:
            primary_nums = set()
            for src in [valid_map or {}, invalid_map or {}]:
                for key in src.keys():
                    if str(key).strip().lstrip("-").isdigit():
                        primary_nums.add(int(key))
            if primary_nums:
                return sorted(primary_nums)
            nums = set()
            for key in (fallback_map or {}).keys():
                if str(key).strip().lstrip("-").isdigit():
                    nums.add(int(key))
            return sorted(nums)

        invalid_rows = getattr(answer_key, "invalid_answer_rows", {}) or {}
        parts: list[str] = []
        for q_no in _question_numbers(
            getattr(answer_key, "answers", {}) or {},
            (invalid_rows.get("MCQ", {}) or {}),
            mcq_answers or {},
        ):
            value = str((mcq_answers or {}).get(q_no, "") or "").strip().upper()[:1]
            parts.append(value or "_")
        for q_no in _question_numbers(
            getattr(answer_key, "true_false_answers", {}) or {},
            (invalid_rows.get("TF", {}) or {}),
            tf_answers or {},
        ):
            flags = (tf_answers or {}).get(q_no, {}) or {}
            for key in ["a", "b", "c", "d"]:
                parts.append("Đ" if key in flags and bool(flags.get(key)) else ("S" if key in flags else "_"))
        numeric_valid = getattr(answer_key, "numeric_answers", {}) or {}
        numeric_invalid = (invalid_rows.get("NUMERIC", {}) or {})
        for q_no in _question_numbers(
            numeric_valid,
            numeric_invalid,
            numeric_answers or {},
        ):
            raw_key = str((getattr(answer_key, "numeric_answers", {}) or {}).get(q_no, ((invalid_rows.get("NUMERIC", {}) or {}).get(q_no, ""))) or "")
            student_text_full = str((numeric_answers or {}).get(q_no, "") or "").strip().replace(" ", "").lstrip("+").replace(".", ",")
            normalized_key = str(raw_key).strip().replace(" ", "").lstrip("+").replace(".", ",")
            if normalized_key:
                width = len(normalized_key)
                student_text = student_text_full[:width]
                if len(student_text) < width:
                    student_text = student_text + ("_" * (width - len(student_text)))
                parts.append(student_text)
            else:
                parts.append(student_text_full or "_")
        return ";".join(parts) if use_semicolon else "".join(parts)

    def _build_answer_string_for_result(self, result, subject_key: str = "") -> str:
        scoped_result = self._scoped_result_copy(result)
        key = self._subject_answer_key_for_result(scoped_result, subject_key)
        expected_by_section = self._expected_questions_by_section(scoped_result)
        if key is not None and any(expected_by_section.get(sec, []) for sec in ["MCQ", "TF", "NUMERIC"]):
            invalid_rows = getattr(key, "invalid_answer_rows", {}) or {}

            def _filter_question_map(payload: object, expected_questions: list[int]) -> dict[int, Any]:
                expected_set = {int(q) for q in (expected_questions or [])}
                filtered: dict[int, Any] = {}
                for q_raw, value in dict(payload or {}).items():
                    try:
                        q_no = int(q_raw)
                    except Exception:
                        continue
                    if q_no in expected_set:
                        filtered[q_no] = value
                return filtered

            scoped_key = type("ScopedAnswerKey", (), {})()
            setattr(scoped_key, "answers", _filter_question_map(getattr(key, "answers", {}) or {}, expected_by_section.get("MCQ", [])))
            setattr(scoped_key, "true_false_answers", _filter_question_map(getattr(key, "true_false_answers", {}) or {}, expected_by_section.get("TF", [])))
            setattr(scoped_key, "numeric_answers", _filter_question_map(getattr(key, "numeric_answers", {}) or {}, expected_by_section.get("NUMERIC", [])))
            setattr(
                scoped_key,
                "invalid_answer_rows",
                {
                    "MCQ": _filter_question_map((invalid_rows.get("MCQ", {}) or {}), expected_by_section.get("MCQ", [])),
                    "TF": _filter_question_map((invalid_rows.get("TF", {}) or {}), expected_by_section.get("TF", [])),
                    "NUMERIC": _filter_question_map((invalid_rows.get("NUMERIC", {}) or {}), expected_by_section.get("NUMERIC", [])),
                },
            )
            key = scoped_key
        use_semicolon = not bool(getattr(scoped_result, "answer_string_api_mode", False))
        return self._answer_string_from_maps(
            scoped_result.mcq_answers or {},
            scoped_result.true_false_answers or {},
            scoped_result.numeric_answers or {},
            key,
            use_semicolon=use_semicolon,
        )

    def _short_recognition_text_for_result(self, result) -> str:
        scoped_result = self._scoped_result_copy(result)
        parts: list[str] = []
        mcq = self._format_mcq_answers(scoped_result.mcq_answers or {})
        tf = self._format_tf_answers(scoped_result.true_false_answers or {})
        num = self._format_numeric_answers(scoped_result.numeric_answers or {})
        if mcq and mcq != "-":
            parts.append(f"MCQ: {mcq}")
        if tf and tf != "-":
            parts.append(f"TF: {tf}")
        if num and num != "-":
            parts.append(f"NUM: {num}")
        return " | ".join(parts) if parts else "-"

    def _status_parts_for_row(
        self,
        sid: str,
        exam_code_text: str,
        duplicate_count: int,
        subject_scope: tuple[set[str], set[str]] | None = None,
        available_exam_codes: set[str] | None = None,
        result: OMRResult | None = None,
    ) -> list[str]:
        sid_text = str(sid or "").strip()
        has_duplicate = False
        status_parts: list[str] = []

        if self._student_id_has_recognition_error(sid_text):
            status_parts.append("SBD không nhận dạng")
            has_duplicate = duplicate_count > 1
        else:
            has_duplicate = duplicate_count > 1
            all_sids, _room_sids = subject_scope if subject_scope is not None else self._subject_student_room_scope()
            sid_norm = self._normalized_student_id_for_match(sid_text)
            profile = self._resolve_student_profile_for_status(sid_text)
            cfg = self._selected_batch_subject_config() or {}

            if not self._has_valid_student_reference(sid_text):
                status_parts.append("SBD không có trong danh sách")
            elif all_sids and sid_norm not in all_sids:
                status_parts.append("SBD không có trong danh sách")
            else:
                room_text = self._subject_room_for_student_id(sid_text, cfg)
                if self._is_missing_room_for_status(room_text):
                    status_parts.append("Thiếu phòng thi")
                else:
                    mapping_by_room = self._normalized_exam_room_mapping_by_room(cfg)
                    if mapping_by_room:
                        resolved_norm = self._normalized_room_for_match(room_text)
                        resolved_candidates = [
                            room for room in mapping_by_room.keys()
                            if self._normalized_room_for_match(room) == resolved_norm
                        ]
                        if resolved_candidates:
                            resolved_room_sids: set[str] = set()
                            for room in resolved_candidates:
                                resolved_room_sids.update(mapping_by_room.get(room, set()))
                            if sid_norm not in resolved_room_sids:
                                status_parts.append("SBD không thuộc phòng thi môn")
                if self._is_missing_name_for_status(str(profile.get("name", "") or "")):
                    status_parts.append("Thiếu họ tên")

        if not self._is_valid_exam_code_for_subject(exam_code_text, available_exam_codes=available_exam_codes):
            status_parts.append("Mã đề không hợp lệ")
        if has_duplicate:
            status_parts.append("Trùng SBD")

        if result is not None:
            rec_errors = list(getattr(result, "recognition_errors", [])) or list(getattr(result, "errors", []))
            issue_codes = [str(getattr(issue, "code", "") or "").strip().upper() for issue in (getattr(result, "issues", []) or [])]
            rec_error_codes = [str(err or "").strip().upper() for err in rec_errors if str(err or "").strip()]
            if rec_errors or issue_codes:
                if any(code in {"ANCHOR_MISSING", "ANCHOR_FAIL", "SCANNER_LOCK_FAIL"} for code in issue_codes) or any("ANCHOR" in code or "SCANNER_LOCK_FAIL" in code for code in rec_error_codes):
                    status_parts.append("Lỗi nhận dạng anchor")
                elif any(code in {"POOR_IDENTIFIER_ZONE", "STUDENT_ID_FAST_FAIL", "EXAM_CODE_FAST_FAIL"} for code in issue_codes) or any(code in {"POOR_IDENTIFIER_ZONE", "STUDENT_ID_FAST_FAIL", "EXAM_CODE_FAST_FAIL"} or "HEADER" in code for code in rec_error_codes):
                    status_parts.append("Lỗi nhận dạng vùng header")
                else:
                    status_parts.append("Lỗi nhận dạng")
        return list(dict.fromkeys([x for x in status_parts if str(x or "").strip()]))

    def _normalized_exam_room_mapping_by_room(self, cfg: dict) -> dict[str, set[str]]:
        normalized: dict[str, set[str]] = {}
        mapping_by_room = cfg.get("exam_room_sbd_mapping_by_room", {}) if isinstance(cfg.get("exam_room_sbd_mapping_by_room", {}), dict) else {}
        for room_key, vals in mapping_by_room.items():
            room_text = str(room_key or "").strip()
            if not room_text:
                continue
            raw_vals: list[str] = []
            if isinstance(vals, (list, tuple, set)):
                raw_vals.extend(str(x).strip() for x in vals if str(x).strip())
            elif isinstance(vals, str):
                raw_vals.extend(x.strip() for x in vals.replace(";", ",").replace("\n", ",").split(",") if x.strip())
            normalized_vals = {self._normalized_student_id_for_match(x) for x in raw_vals if x}
            if normalized_vals:
                normalized[room_text] = normalized_vals
        return normalized

    def _subject_student_room_scope(self) -> tuple[set[str], set[str]]:
        all_sids: set[str] = set()
        for s in (self.session.students if self.session else []):
            sid = str(getattr(s, "student_id", "") or "").strip()
            if sid:
                all_sids.add(self._normalized_student_id_for_match(sid))
        cfg = self._selected_batch_subject_config() or {}
        room_name = str(cfg.get("exam_room_name", "") or "").strip()
        room_name_norm = self._normalized_room_for_match(room_name)
        mapping_text = str(cfg.get("exam_room_sbd_mapping", "") or "").strip()
        mapping_by_room = self._normalized_exam_room_mapping_by_room(cfg)
        room_sids: set[str] = set()
        if mapping_by_room:
            matched_rooms = [room for room in mapping_by_room.keys() if self._normalized_room_for_match(room) == room_name_norm] if room_name else []
            if not matched_rooms and len(mapping_by_room) == 1:
                matched_rooms = [next(iter(mapping_by_room.keys()))]
            if matched_rooms:
                for room in matched_rooms:
                    room_sids.update(mapping_by_room.get(room, set()))
            elif not room_name:
                # exam_room_name trống: dùng union của toàn bộ mapping để tránh false negative.
                for values in mapping_by_room.values():
                    room_sids.update(values)
        elif mapping_text:
            chunks = [x.strip() for x in mapping_text.replace(";", ",").replace("\n", ",").split(",")]
            room_sids = {self._normalized_student_id_for_match(x) for x in chunks if x}
        elif room_name:
            for s in (self.session.students if self.session else []):
                sid = str(getattr(s, "student_id", "") or "").strip()
                extra = getattr(s, "extra", {}) or {}
                exam_room = str(extra.get("exam_room", "") or "").strip()
                if sid and exam_room and self._normalized_room_for_match(exam_room) == room_name_norm:
                    room_sids.add(self._normalized_student_id_for_match(sid))
        return all_sids, room_sids

    def _subject_room_for_student_id(self, sid: str, subject_cfg: dict | None = None) -> str:
        sid_norm = self._normalized_student_id_for_match(sid)
        if not sid_norm:
            return ""
        cfg = subject_cfg if isinstance(subject_cfg, dict) else (self._selected_batch_subject_config() or {})
        mapping_by_room = self._normalized_exam_room_mapping_by_room(cfg)
        if mapping_by_room:
            matched_rooms = [room for room, sids in mapping_by_room.items() if sid_norm in sids]
            if matched_rooms:
                preferred_room = str(cfg.get("exam_room_name", "") or "").strip()
                preferred_norm = self._normalized_room_for_match(preferred_room)
                if preferred_norm:
                    for room in matched_rooms:
                        if self._normalized_room_for_match(room) == preferred_norm:
                            return str(room or "").strip()
                picked = sorted(matched_rooms, key=lambda x: self._normalized_room_for_match(x))[0]
                return str(picked or "").strip()

        room_name = str(cfg.get("exam_room_name", "") or "").strip()
        mapping_text = str(cfg.get("exam_room_sbd_mapping", "") or "").strip()
        if room_name and mapping_text:
            chunks = [x.strip() for x in mapping_text.replace(";", ",").replace("\n", ",").split(",") if x.strip()]
            if sid_norm in {self._normalized_student_id_for_match(x) for x in chunks if x}:
                return room_name

        prof = self._student_profile_by_id(sid)
        fallback_room = str(prof.get("exam_room", "") or "")
        return fallback_room

    @staticmethod
    def _student_id_has_recognition_error(student_id: str) -> bool:
        sid = str(student_id or "").strip()
        return sid in {"", "-"} or ("?" in sid)

    @staticmethod
    def _is_missing_name_for_status(name_text: str) -> bool:
        return str(name_text or "").strip() in {"", "-"}

    def _resolve_student_profile_for_status(self, student_id: str) -> dict:
        sid = str(student_id or "").strip()
        if not sid:
            return {}
        profile = self._student_profile_by_id(sid)
        return dict(profile) if isinstance(profile, dict) else {}

    def _has_valid_student_reference(self, student_id: str) -> bool:
        profile = self._resolve_student_profile_for_status(student_id)
        return bool(profile) and not self._is_missing_name_for_status(str(profile.get("name", "") or ""))

    def _is_missing_room_for_status(self, room_text: str) -> bool:
        normalized = self._normalized_room_for_match(str(room_text or ""))
        return normalized in {
            "",
            self._normalized_room_for_match("-"),
            self._normalized_room_for_match("Không rõ phòng"),
            self._normalized_room_for_match("[Không rõ phòng]"),
        }

    def _is_valid_exam_code_for_subject(self, exam_code_text: str, available_exam_codes: set[str] | None = None) -> bool:
        code = str(exam_code_text or "").strip()
        if not code or code == "-" or "?" in code:
            return False
        valid_codes = available_exam_codes if available_exam_codes is not None else self._available_exam_codes()
        if not valid_codes:
            return True
        norm_code = self._normalize_exam_code_text(code)
        return code in valid_codes or norm_code in valid_codes

    @staticmethod
    def _name_missing(name_text: str) -> bool:
        name = str(name_text or "").strip()
        return name in {"", "-"}

    def _status_parts_for_result(
        self,
        result,
        duplicate_count: int,
        subject_scope: tuple[set[str], set[str]] | None = None,
        available_exam_codes: set[str] | None = None,
    ) -> list[str]:
        sid = str(getattr(result, "student_id", "") or "").strip()
        exam_code_text = str(getattr(result, "exam_code", "") or "").strip()
        return self._status_parts_for_row(
            sid,
            exam_code_text,
            duplicate_count,
            subject_scope=subject_scope,
            available_exam_codes=available_exam_codes,
            result=result,
        )

    def _status_text_for_row(
        self,
        idx: int,
        duplicate_count_map: dict[str, int] | None = None,
        subject_scope: tuple[set[str], set[str]] | None = None,
        available_exam_codes: set[str] | None = None,
    ) -> str:
        if idx < 0 or idx >= len(self.scan_results):
            return "OK"
        res = self.scan_results[idx]
        sid = (res.student_id or "").strip()
        if self._student_id_has_recognition_error(sid):
            dup = 0
        elif duplicate_count_map is not None:
            dup = int(duplicate_count_map.get(sid, 0) or 0)
        else:
            dup = sum(
                1
                for r in self.scan_results
                if not self._student_id_has_recognition_error((r.student_id or "").strip()) and (r.student_id or "").strip() == sid
            )
        status_parts = self._status_parts_for_result(
            res,
            dup,
            subject_scope=subject_scope,
            available_exam_codes=available_exam_codes,
        )
        return ", ".join(status_parts) if status_parts else "OK"

    def _current_batch_subject_key(self) -> str:
        cfg = self._selected_batch_subject_config()
        if cfg:
            return self._subject_key_from_cfg(cfg)
        return str(self.active_batch_subject_key or "").strip() or self._resolve_preferred_scoring_subject()

    def _invalidate_scoring_for_student_ids(self, student_ids: list[str], subject_key: str = "", reason: str = "") -> int:
        subject = str(subject_key or self._current_batch_subject_key() or "").strip()
        if not subject:
            return 0
        subject_scores = dict(self.scoring_results_by_subject.get(subject, {}))
        if not subject_scores:
            return 0
        changed = 0
        for sid in student_ids:
            sid_key = str(sid or "").strip()
            if not sid_key:
                continue
            if sid_key in subject_scores:
                subject_scores.pop(sid_key, None)
                changed += 1
        if changed <= 0:
            return 0
        self.scoring_results_by_subject[subject] = subject_scores
        # Đồng bộ DB ngay khi invalidate để tránh resurrect dữ liệu chấm cũ
        # khi chuyển màn hình hoặc tải lại phiên.
        try:
            current_mode = str((self._subject_config_by_subject_key(subject) or {}).get("scoring_last_mode", "") or "Tính lại toàn bộ")
            self._persist_scoring_results_for_subject(
                subject,
                list(subject_scores.values()),
                current_mode,
                reason or "invalidate_scoring_records",
                mark_saved=False,
            )
        except Exception:
            pass
        try:
            self.database.log_change(
                "scoring_results",
                subject,
                "invalidate_scoring_records",
                "",
                f"removed={changed}; student_ids={','.join(sorted({str(s or '').strip() for s in student_ids if str(s or '').strip()}))}",
                reason or "invalidate_scoring_records",
            )
        except Exception:
            pass
        self._persist_runtime_session_state_quietly()
        return changed

    def _record_adjustment(self, idx: int, details: list[str], source: str) -> None:
        if not details:
            return
        message = f"({source}) " + "; ".join(details)
        image_key = self._row_image_key(idx)
        if (not image_key) and 0 <= idx < len(self.scan_results):
            image_key = self._result_identity_key(getattr(self.scan_results[idx], "image_path", ""))
        if not image_key:
            return
        history = self.scan_edit_history.setdefault(image_key, [])
        history.append(message)
        self.scan_last_adjustment[image_key] = message
        self.scan_manual_adjustments[image_key] = sorted(set(self.scan_manual_adjustments.get(image_key, []) + details))
        target_result = self._result_by_image_path(image_key)
        if target_result is not None:
            setattr(target_result, "edit_history", list(history))
            setattr(target_result, "last_adjustment", message)
            setattr(target_result, "manual_adjustments", list(self.scan_manual_adjustments.get(image_key, [])))
            setattr(target_result, "manually_edited", True)
            setattr(target_result, "cached_forced_status", "Đã sửa")
            setattr(target_result, "cached_status", "Đã sửa")
        self.scan_forced_status_by_index[image_key] = "Đã sửa"
        current_subject = str(self._current_batch_subject_key() or "").strip()
        if current_subject:
            self.scan_results_by_subject[self._batch_result_subject_key(current_subject)] = list(self.scan_results or [])
        self.database.log_change("scan_results", image_key, source, "", message, source)

    @staticmethod
    def _result_identity_key(image_path: str) -> str:
        return str(image_path or "").strip()

    def _row_image_key(self, row_idx: int) -> str:
        if row_idx < 0 or (not hasattr(self, "scan_list")) or row_idx >= self.scan_list.rowCount():
            return ""
        sid_item = self.scan_list.item(row_idx, self.SCAN_COL_STUDENT_ID)
        return self._result_identity_key(str(sid_item.data(Qt.UserRole) if sid_item else ""))

    def _row_index_by_image_path(self, image_path: str) -> int:
        key = self._result_identity_key(image_path)
        if not key or not hasattr(self, "scan_list"):
            return -1
        for row in range(self.scan_list.rowCount()):
            sid_item = self.scan_list.item(row, self.SCAN_COL_STUDENT_ID)
            row_image = self._result_identity_key(str(sid_item.data(Qt.UserRole) if sid_item else ""))
            if row_image == key:
                return row
        return -1

    def _result_by_image_path(self, image_path: str) -> OMRResult | None:
        image_key = self._result_identity_key(image_path)
        if not image_key:
            return None
        for result in (self.scan_results or []):
            if self._result_identity_key(getattr(result, "image_path", "")) == image_key:
                return result
        return None

    def _mark_result_manually_edited(self, result: OMRResult, row_idx: int | None = None) -> None:
        setattr(result, "manually_edited", True)
        setattr(result, "cached_forced_status", "Đã sửa")
        setattr(result, "cached_status", "Đã sửa")
        image_key = self._result_identity_key(getattr(result, "image_path", ""))
        resolved_row = row_idx if row_idx is not None and row_idx >= 0 else self._row_index_by_image_path(image_key)
        if image_key:
            self.scan_forced_status_by_index[image_key] = "Đã sửa"
        if resolved_row is not None and resolved_row >= 0 and hasattr(self, "scan_list") and resolved_row < self.scan_list.rowCount():
            status_item = self.scan_list.item(resolved_row, self.SCAN_COL_STATUS)
            if status_item is None:
                status_item = QTableWidgetItem("Đã sửa")
                self.scan_list.setItem(resolved_row, self.SCAN_COL_STATUS, status_item)
            else:
                status_item.setText("Đã sửa")
            status_item.setToolTip("Đã sửa")
            status_item.setForeground(Qt.red)

    def _persist_scan_results_to_db(self, subject_key: str) -> None:
        scoped_key = self._batch_result_subject_key(subject_key)
        active_subject = str(self._current_batch_subject_key() or "").strip()
        if active_subject and self._batch_result_subject_key(active_subject) == scoped_key and self.scan_results:
            source_rows = list(self.scan_results)
            self.scan_results_by_subject[scoped_key] = list(self.scan_results)
        else:
            source_rows = list(self.scan_results_by_subject.get(scoped_key, self.scan_results) or [])
        for result in source_rows:
            result.answer_string = self._normalize_non_api_answer_string(result, subject_key)
        rows = [self._serialize_omr_result(x) for x in source_rows]
        self.database.replace_scan_results_for_subject(scoped_key, rows)
        self.database.log_change("scan_results", scoped_key, "replace_subject_rows", "", f"{len(rows)} rows", "batch_save")
        self._refresh_scan_results_from_db(subject_key)

    def _normalize_non_api_answer_string(self, result: OMRResult, subject_key: str = "") -> str:
        if result is None:
            return ""
        answer_text = str(getattr(result, "answer_string", "") or "").strip()
        if bool(getattr(result, "answer_string_api_mode", False)):
            return answer_text
        if answer_text:
            return answer_text
        subject = str(subject_key or self._current_batch_subject_key() or self.active_batch_subject_key or "").strip()
        rebuilt = str(self._build_answer_string_for_result(result, subject) or "").strip()
        if rebuilt:
            setattr(result, "answer_string", rebuilt)
        return rebuilt

    @staticmethod
    def _debug_scan_result_state(tag: str, result: OMRResult | None) -> None:
        if result is None:
            return

    def _update_working_batch_state_single_row(self, subject_key: str, result: OMRResult, row_payload: dict, row_idx: int) -> None:
        # DB-only mode: không duy trì working cache theo môn.
        return
    def _persist_single_scan_result_to_db(self, result: OMRResult, note: str = "") -> None:
        subject_key = str(self._current_batch_subject_key() or "").strip()
        image_path = str(getattr(result, "image_path", "") or "").strip()
        if not subject_key or not image_path:
            return
        result.answer_string = self._normalize_non_api_answer_string(result, subject_key)
        row_idx = self._row_index_by_image_path(image_path)
        serialized = self._serialize_omr_result(result)
        self._debug_scan_result_state("persist_single_before_db", result)

        subject_db_key = self._batch_result_subject_key(subject_key)
        updated = False
        try:
            self.database.update_scan_result_payload(
                subject_db_key,
                image_path,
                serialized,
                note=note,
            )
            db_rows = list(self.database.fetch_scan_results_for_subject(subject_db_key) or [])
            for payload in db_rows:
                if not isinstance(payload, dict):
                    continue
                if str(payload.get("image_path", "") or "").strip() != image_path:
                    continue
                payload_history = [str(x) for x in (payload.get("edit_history", []) or []) if str(x or "").strip()]
                payload_forced = str(payload.get("cached_forced_status", payload.get("forced_status", "")) or "").strip()
                if bool(payload.get("manually_edited", False)) or bool(payload_history) or payload_forced == "Đã sửa":
                    updated = True
                    break
        except Exception:
            updated = False

        if not updated:
            try:
                db_rows = list(self.database.fetch_scan_results_for_subject(subject_db_key) or [])
            except Exception:
                db_rows = []
            replaced = False
            for i, payload in enumerate(list(db_rows)):
                if isinstance(payload, dict) and str(payload.get("image_path", "") or "").strip() == image_path:
                    db_rows[i] = dict(serialized)
                    replaced = True
                    break
            if not replaced:
                db_rows.append(dict(serialized))
            try:
                self.database.replace_scan_results_for_subject(subject_db_key, db_rows)
                try:
                    self.database.log_change("scan_results", image_path, "persist_single_fallback", "", note or "inline_edit", note or "inline_edit")
                except Exception:
                    pass
            except Exception:
                pass

        self.scan_results_by_subject[subject_db_key] = list(self.scan_results or [])
        self._update_single_subject_saved_snapshot(subject_key, result, {}, row_idx)
    def _update_single_subject_saved_snapshot(self, subject_key: str, result: OMRResult, row_payload: dict, row_idx: int) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        cfg = self._subject_config_by_subject_key(subject)
        if not isinstance(cfg, dict):
            return

        scoped_key = self._batch_result_subject_key(subject)
        result_count = len(self.scan_results_by_subject.get(scoped_key, []) or [])
        if result_count <= 0 and subject == str(self._current_batch_subject_key() or "").strip() and hasattr(self, "scan_list"):
            result_count = self.scan_list.rowCount()

        cfg["batch_saved"] = bool(result_count)
        cfg["batch_saved_at"] = datetime.now().isoformat(timespec="seconds") if result_count else ""
        cfg["batch_result_count"] = result_count
        cfg["batch_saved_rows"] = []
        cfg["batch_saved_preview"] = []
        cfg["batch_saved_results"] = []

        if self.session and isinstance(self.session.config, dict):
            subject_cfgs = self.session.config.get("subject_configs", [])
            if isinstance(subject_cfgs, list):
                for i, item in enumerate(subject_cfgs):
                    if item is cfg:
                        subject_cfgs[i] = cfg
                        break
                    if isinstance(item, dict) and self._subject_key_from_cfg(item) == subject:
                        subject_cfgs[i].update(cfg)
                        break
        if isinstance(self.batch_editor_return_payload, dict):
            subject_cfgs = self.batch_editor_return_payload.get("subject_configs", [])
            if isinstance(subject_cfgs, list):
                for i, item in enumerate(subject_cfgs):
                    if isinstance(item, dict) and self._subject_key_from_cfg(item) == subject:
                        subject_cfgs[i].update(cfg)
                        break

        self.batch_working_state_by_subject.pop(self._batch_runtime_key(subject), None)
    def _sync_subject_batch_snapshot_after_inline_edit(self, subject_key: str = "") -> None:
        # Inline edit đã lưu trực tiếp vào DB; chỉ cập nhật metadata nhẹ trong runtime.
        # Không save lại toàn bộ session để tránh làm chậm luồng sửa bài.
        subject = str(subject_key or self._current_batch_subject_key() or "").strip()
        if not subject:
            return
        cfg = self._subject_config_by_subject_key(subject)
        if not isinstance(cfg, dict):
            return
        result_count = len(self.scan_results_by_subject.get(self._batch_result_subject_key(subject), []) or [])
        if result_count <= 0 and subject == str(self._current_batch_subject_key() or "").strip() and hasattr(self, "scan_list"):
            result_count = self.scan_list.rowCount()
        cfg["batch_saved"] = bool(result_count)
        cfg["batch_result_count"] = result_count
        if not cfg.get("batch_saved_at") and result_count:
            cfg["batch_saved_at"] = datetime.now().isoformat(timespec="seconds")
        cfg["batch_saved_rows"] = []
        cfg["batch_saved_preview"] = []
        cfg["batch_saved_results"] = []
        self.batch_working_state_by_subject.pop(self._batch_runtime_key(subject), None)
    def _refresh_all_statuses(self) -> None:
        duplicate_count_map: dict[str, int] = {}
        canonical_by_image: dict[str, OMRResult] = {}
        for res in self.scan_results:
            canonical_by_image[self._result_identity_key(getattr(res, "image_path", ""))] = res
        for res in canonical_by_image.values():
            sid = str(getattr(res, "student_id", "") or "").strip()
            if not self._student_id_has_recognition_error(sid):
                duplicate_count_map[sid] = duplicate_count_map.get(sid, 0) + 1
        subject_scope = self._subject_student_room_scope()
        available_exam_codes = self._available_exam_codes()
        for row_idx in range(self.scan_list.rowCount()):
            self._refresh_row_status(
                row_idx,
                duplicate_count_map=duplicate_count_map,
                subject_scope=subject_scope,
                available_exam_codes=available_exam_codes,
            )

    def _on_scan_cell_clicked(self, row: int, col: int) -> None:
        if row < 0:
            return
        # Show edit history from Status / Actions columns for quick access in all status scenarios.
        if col not in {self.SCAN_COL_STATUS, self.SCAN_COL_ACTIONS}:
            return
        image_key = self._row_image_key(row)
        history = list(self.scan_edit_history.get(image_key, []) or [])
        if (not history) and image_key:
            result = self._result_by_image_path(image_key)
            history = [str(x) for x in (getattr(result, "edit_history", []) or []) if str(x or "").strip()] if result is not None else []
            if history:
                self.scan_edit_history[image_key] = list(history)
                self.scan_last_adjustment[image_key] = str(getattr(result, "last_adjustment", "") or history[-1])
        if not history:
            sid_item = self.scan_list.item(row, self.SCAN_COL_STUDENT_ID)
            row_payload = dict(sid_item.data(Qt.UserRole + 12) or {}) if sid_item is not None else {}
            history = [str(x) for x in (row_payload.get("edit_history", []) or []) if str(x or "").strip()]
            if history and image_key:
                self.scan_edit_history[image_key] = list(history)
                self.scan_last_adjustment[image_key] = str(row_payload.get("last_adjustment", "") or history[-1])
        if not history:
            QMessageBox.information(self, "Lịch sử sửa", "Chưa có lịch sử điều chỉnh trong Status cho bài thi này.")
            return
        latest = self.scan_last_adjustment.get(image_key, history[-1])
        QMessageBox.information(
            self,
            "Lịch sử sửa bài",
            "Điều chỉnh gần nhất:\n"
            + latest
            + "\n\nToàn bộ lịch sử:\n"
            + "\n".join(history),
        )

    def _status_text_for_saved_table_row(self, row_idx: int) -> str:
        sid_item = self.scan_list.item(row_idx, self.SCAN_COL_STUDENT_ID)
        payload = dict(sid_item.data(Qt.UserRole + 12) or {}) if sid_item is not None else {}
        payload_status = str(payload.get("status", "") or "").strip()
        payload_history = [str(x) for x in (payload.get("edit_history", []) or []) if str(x or "").strip()]
        payload_is_edited = bool(payload.get("manually_edited", False)) or bool(payload_history) or ("đã sửa" in payload_status.lower())
        sid = (sid_item.text().strip() if sid_item else "")
        exam_code_text = str(sid_item.data(Qt.UserRole + 1) if sid_item else "").strip()
        dup = 0
        if not self._student_id_has_recognition_error(sid):
            unique_pairs: set[tuple[str, str]] = set()
            for result in (self._current_scan_results_snapshot() or []):
                sid_val = str(getattr(result, "student_id", "") or "").strip()
                img_val = self._result_identity_key(getattr(result, "image_path", ""))
                if sid_val and (not self._student_id_has_recognition_error(sid_val)):
                    unique_pairs.add((img_val, sid_val))
            dup = sum(1 for _img, sid_val in unique_pairs if sid_val == sid)
        status_parts = self._status_parts_for_row("" if self._student_id_has_recognition_error(sid) else sid, exam_code_text, dup)
        status_parts_text = ", ".join(status_parts) if status_parts else ""
        if payload_is_edited:
            if status_parts_text:
                return f"Đã sửa ({status_parts_text})"
            return "Đã sửa"
        return status_parts_text or "OK"

    @staticmethod
    def _compact_status_text(status_text: str, max_len: int = 150) -> str:
        text = str(status_text or "").strip()
        if len(text) <= max_len:
            return text
        return text[: max(0, max_len - 3)].rstrip() + "..."

    def _refresh_row_status(
        self,
        idx: int,
        duplicate_count_map: dict[str, int] | None = None,
        subject_scope: tuple[set[str], set[str]] | None = None,
        available_exam_codes: set[str] | None = None,
    ) -> None:
        if idx < 0 or idx >= self.scan_list.rowCount():
            return
        image_key = self._row_image_key(idx)
        forced_status = self.scan_forced_status_by_index.get(image_key, "")
        row_result = self._result_by_image_path(image_key) if image_key else None
        if row_result is None and idx < len(self.scan_results):
            row_result = self.scan_results[idx]
        if not forced_status and row_result is not None:
            row_history = [str(x) for x in (getattr(row_result, "edit_history", []) or []) if str(x or "").strip()]
            if bool(getattr(row_result, "manually_edited", False)) or bool(row_history):
                forced_status = "Đã sửa"
                row_image = self._result_identity_key(getattr(row_result, "image_path", "")) or image_key
                if row_image:
                    self.scan_forced_status_by_index[row_image] = forced_status
        if row_result is not None:
            sid_text = str(getattr(row_result, "student_id", "") or "").strip()
            duplicate_count = 0
            if duplicate_count_map is not None and not self._student_id_has_recognition_error(sid_text):
                duplicate_count = int(duplicate_count_map.get(sid_text, 0) or 0)
            payload = self._build_scan_row_payload_from_result(
                row_result,
                row_idx=idx,
                duplicate_count=duplicate_count,
                subject_scope=subject_scope,
                available_exam_codes=available_exam_codes,
                forced_status=forced_status,
            )
            sid_item = self.scan_list.item(idx, self.SCAN_COL_STUDENT_ID)
            if sid_item is not None:
                sid_item.setData(Qt.UserRole + 1, str(payload.get("exam_code", "") or ""))
                sid_item.setData(Qt.UserRole + 2, str(payload.get("recognized_short", "") or ""))
                sid_item.setData(Qt.UserRole + 10, dict(payload.get("serialized_result", {}) or {}))
                sid_item.setData(Qt.UserRole + 11, str(payload.get("manual_content_override", "") or ""))
                sid_item.setData(Qt.UserRole + 12, dict(payload or {}))
            content_text = str(payload.get("content", "") or "")
            content_item = QTableWidgetItem(content_text)
            content_item.setToolTip(content_text)
            self.scan_list.setItem(idx, self.SCAN_COL_CONTENT, content_item)
            full_status = str(payload.get("status", "") or "OK")
        else:
            full_status = str(self._status_text_for_saved_table_row(idx) or "OK")
        display_status = self._compact_status_text(full_status, max_len=150)
        status_item = QTableWidgetItem(display_status)
        status_item.setToolTip(full_status)
        if full_status != "OK":
            status_item.setForeground(Qt.red)
        self.scan_list.setItem(idx, self.SCAN_COL_STATUS, status_item)

    def _update_scan_preview(self, index: int) -> None:
        if index < 0 or index >= len(self.scan_results):
            return
        result = self.scan_results[index]
        img_path = Path(result.image_path)
        aligned_pix = self._aligned_image_to_qpixmap(getattr(result, "aligned_image", None))
        pix = aligned_pix if not aligned_pix.isNull() else QPixmap(str(img_path))
        if pix.isNull():
            self.preview_source_pixmap = QPixmap()
            self.scan_image_preview.setText(f"Cannot load image: {img_path.name}")
            self.scan_image_preview.clear_markers()
        else:
            rotation = int(self.preview_rotation_by_index.get(index, 0) or 0) % 360
            if rotation:
                pix = pix.transformed(QTransform().rotate(float(rotation)), Qt.SmoothTransformation)
            self.preview_source_pixmap = pix
            self._render_preview_pixmap()
            self.scan_image_preview.set_overlay_markers(self._recognition_overlay_positions_for_result(result))

        preview_result = self._scoped_result_copy(self._lightweight_result_copy(result))
        section_counts = self._subject_section_question_counts(self._current_batch_subject_key())
        expected_actual = self._expected_questions_by_section(preview_result)

        def _build_display_mapping(actual_questions: list[int], limit: int) -> tuple[list[int], dict[int, int], dict[int, int]]:
            actual_sorted = sorted(set(int(q) for q in (actual_questions or [])))
            if limit > 0:
                actual_sorted = actual_sorted[:limit]
                if not actual_sorted:
                    actual_sorted = list(range(1, limit + 1))
            display_questions = list(range(1, len(actual_sorted) + 1))
            actual_to_display = {int(a): int(d) for d, a in zip(display_questions, actual_sorted)}
            display_to_actual = {int(d): int(a) for d, a in zip(display_questions, actual_sorted)}
            return display_questions, actual_to_display, display_to_actual

        expected_display: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        actual_to_display: dict[str, dict[int, int]] = {"MCQ": {}, "TF": {}, "NUMERIC": {}}
        for sec in ["MCQ", "TF", "NUMERIC"]:
            sec_limit = max(0, int(section_counts.get(sec, 0) or 0))
            display_qs, map_actual_to_display, _map_display_to_actual = _build_display_mapping(
                list(expected_actual.get(sec, []) or []),
                sec_limit,
            )
            expected_display[sec] = list(display_qs)
            actual_to_display[sec] = dict(map_actual_to_display)

        mcq_display = {
            int(actual_to_display["MCQ"].get(int(q), int(q))): str(v)
            for q, v in (preview_result.mcq_answers or {}).items()
        }
        tf_display = {
            int(actual_to_display["TF"].get(int(q), int(q))): dict(v or {})
            for q, v in (preview_result.true_false_answers or {}).items()
        }
        num_display = {
            int(actual_to_display["NUMERIC"].get(int(q), int(q))): str(v)
            for q, v in (preview_result.numeric_answers or {}).items()
        }
        blank_map = self.scan_blank_summary.get(index) or self._compute_blank_questions(self._lightweight_result_copy(result))

        mcq_text = self._compact_value(self._format_mcq_answers_with_expected(mcq_display, expected_display.get("MCQ", [])), 220)
        tf_text = self._compact_value(self._format_tf_answers_with_expected(tf_display, expected_display.get("TF", [])), 220)
        num_text = self._compact_value(self._format_numeric_answers_with_expected(num_display, expected_display.get("NUMERIC", [])), 220)
        rows = [
            ("File ảnh", img_path.name),
            ("STUDENT ID", result.student_id or "-"),
            ("Exam code", result.exam_code or "-"),
        ]
        if section_counts.get("MCQ", 0) > 0:
            rows.extend([
                ("MCQ", mcq_text),
                ("MCQ không tô", ", ".join(str(x) for x in blank_map.get("MCQ", [])) or "-"),
            ])
        if section_counts.get("TF", 0) > 0:
            rows.extend([
                ("TF", tf_text),
                ("TF không tô", ", ".join(str(x) for x in blank_map.get("TF", [])) or "-"),
            ])
        if section_counts.get("NUMERIC", 0) > 0:
            rows.extend([
                ("NUM", num_text),
                ("NUMERIC không tô", ", ".join(str(x) for x in blank_map.get("NUMERIC", [])) or "-"),
            ])
        self.scan_result_preview.setRowCount(0)
        for r, (k, v) in enumerate(rows):
            self.scan_result_preview.insertRow(r)
            self.scan_result_preview.setItem(r, 0, QTableWidgetItem(str(k)))
            self.scan_result_preview.setItem(r, 1, QTableWidgetItem(str(v)))

    def _sync_correction_detail_panel(self, res: OMRResult, rebuild_editor: bool = False) -> None:
        subject_key = self._current_batch_subject_key()
        self._load_exam_code_correction_options(subject_key, str(res.exam_code or "").strip())
        self._load_student_correction_options(str(res.student_id or "").strip())
        scoped = self._scoped_result_copy(res)
        if rebuild_editor:
            self._build_visual_answer_editor(scoped)
        payload = {
            "student_id": res.student_id,
            "exam_code": res.exam_code,
            "answer_string": str(getattr(scoped, "answer_string", "") or self._build_answer_string_for_result(scoped, subject_key)),
            "mcq_answers": scoped.mcq_answers,
            "true_false_answers": scoped.true_false_answers,
            "numeric_answers": scoped.numeric_answers,
            "issues": [{"code": i.code, "message": i.message, "zone_id": i.zone_id} for i in res.issues],
            "recognition_errors": list(getattr(res, "recognition_errors", [])) or list(getattr(res, "errors", [])),
        }
        self.result_preview.setPlainText(json.dumps(payload, ensure_ascii=False, indent=2))
        self.manual_edit.setPlainText(
            json.dumps(
                {
                    "student_id": res.student_id,
                    "exam_code": res.exam_code,
                    "answer_string": payload["answer_string"],
                    "mcq_answers": scoped.mcq_answers,
                    "true_false_answers": scoped.true_false_answers,
                    "numeric_answers": scoped.numeric_answers,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        self.correction_ui_loading = False

    def _load_selected_result_for_correction(self) -> None:
        idx = self.scan_list.currentRow()
        if idx < 0:
            return
        res = self._restore_full_result_for_row(idx)
        if res is None:
            return
        self._trim_result_answers_to_expected_scope(res)
        self._set_scan_result_at_row(idx, res)
        self.correction_ui_loading = True
        self._sync_correction_detail_panel(res, rebuild_editor=True)
        self.correction_ui_loading = False

    def _open_edit_selected_scan(self, *_args) -> None:
        idx = self.scan_list.currentRow()
        if idx < 0:
            QMessageBox.warning(self, "No selection", "Chọn bài thi cần sửa trước.")
            return
        restored_for_edit = self._restore_full_result_for_row(idx)
        if restored_for_edit is not None and idx >= len(self.scan_results):
            self._set_scan_result_at_row(idx, restored_for_edit)
        if idx >= len(self.scan_results):
            sid_item_existing = self.scan_list.item(idx, self.SCAN_COL_STUDENT_ID)
            sid = sid_item_existing.text() if sid_item_existing else "-"
            content = ""
            if restored_for_edit is not None:
                scoped_for_edit = self._scoped_result_copy(restored_for_edit)
                content = str(self._build_answer_string_for_result(scoped_for_edit, self._current_batch_subject_key()) or "").strip()
                if not content:
                    content = str(getattr(scoped_for_edit, "cached_recognized_short", "") or "").strip()
                if not content:
                    content = str(self._short_recognition_text_for_result(scoped_for_edit) or "").strip()
                if not content:
                    content = str(getattr(restored_for_edit, "manual_content_override", "") or "").strip()
            if not content:
                content = str(sid_item_existing.data(Qt.UserRole + 2) if sid_item_existing else "").strip()
            if not content:
                content = self.scan_list.item(idx, self.SCAN_COL_CONTENT).text() if self.scan_list.item(idx, self.SCAN_COL_CONTENT) else "-"
            exam_code = str(sid_item_existing.data(Qt.UserRole + 1) if sid_item_existing else "").strip()
            if not exam_code:
                for r in range(self.scan_result_preview.rowCount()):
                    k = self.scan_result_preview.item(r, 0)
                    v = self.scan_result_preview.item(r, 1)
                    if k and v and k.text().strip().lower() in {"exam code", "mã đề"}:
                        exam_code = v.text().strip()
                        break
            dlg = QDialog(self)
            dlg.setWindowTitle("Sửa bài thi đã lưu")
            lay = QVBoxLayout(dlg)
            form = QFormLayout()
            inp_sid = QLineEdit(sid)
            inp_code = QComboBox()
            inp_code.setEditable(True)
            inp_code.setInsertPolicy(QComboBox.NoInsert)
            subject_cfg = self._selected_batch_subject_config() or {}
            subject_key = self._current_batch_subject_key()
            valid_codes: set[str] = set()
            imported = subject_cfg.get("imported_answer_keys", {}) if isinstance(subject_cfg.get("imported_answer_keys", {}), dict) else {}
            valid_codes.update(str(x).strip() for x in imported.keys() if str(x).strip())
            try:
                valid_codes.update(str(x).strip() for x in self._fetch_answer_keys_for_subject_scoped(subject_key).keys() if str(x).strip())
            except Exception:
                pass
            if self.answer_keys:
                valid_codes.update(
                    str(item.exam_code).strip()
                    for item in self.answer_keys.keys.values()
                    if str(getattr(item, "subject", "") or "").strip() == str(subject_key or "").strip() and str(getattr(item, "exam_code", "") or "").strip()
                )
            if exam_code:
                valid_codes.add(exam_code)
            if valid_codes:
                for code in sorted(valid_codes):
                    inp_code.addItem(code, code)
            else:
                inp_code.addItem("-", "-")
            if exam_code:
                idx_code = inp_code.findData(exam_code)
                if idx_code >= 0:
                    inp_code.setCurrentIndex(idx_code)
                else:
                    inp_code.setEditText(exam_code)
            form.addRow("Student ID", inp_sid)
            form.addRow("Exam Code", inp_code)
            lay.addLayout(form)
            buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
            buttons.accepted.connect(dlg.accept)
            buttons.rejected.connect(dlg.reject)
            lay.addWidget(buttons)
            if dlg.exec() != QDialog.Accepted:
                return
            old_item = self.scan_list.item(idx, self.SCAN_COL_STUDENT_ID)
            old_sid = old_item.text().strip() if old_item else ""
            old_img = str(old_item.data(Qt.UserRole) if old_item else "")
            old_exam_code = str(old_item.data(Qt.UserRole + 1) if old_item else "").strip()
            old_recognized_short = str(old_item.data(Qt.UserRole + 2) if old_item else "")
            new_sid_text = inp_sid.text().strip()
            new_exam_code = str(inp_code.currentData() or inp_code.currentText() or "").strip() or old_exam_code
            sid_item = QTableWidgetItem(new_sid_text or "-")
            sid_item.setData(Qt.UserRole, old_img)
            sid_item.setData(Qt.UserRole + 1, new_exam_code)
            sid_item.setData(Qt.UserRole + 2, old_recognized_short)
            self.scan_list.setItem(idx, self.SCAN_COL_STUDENT_ID, sid_item)
            self.scan_list.setItem(idx, self.SCAN_COL_EXAM_CODE, QTableWidgetItem(new_exam_code or "-"))
            changes: list[str] = []
            if new_sid_text != old_sid:
                changes.append(f"student_id: '{old_sid}' -> '{new_sid_text}'")
            if new_exam_code != old_exam_code:
                changes.append(f"exam_code: '{old_exam_code}' -> '{new_exam_code}'")
            rebuilt = self._build_result_from_saved_table_row(idx)
            if rebuilt is not None:
                setattr(rebuilt, "manual_content_override", "")
                if changes:
                    self._mark_result_manually_edited(rebuilt, idx)
                self._refresh_student_profile_for_result(rebuilt, idx)
                self._set_scan_result_at_row(idx, rebuilt)
                subject_key_now = self._current_batch_subject_key()
                if subject_key_now:
                    self.scan_results_by_subject[self._batch_result_subject_key(subject_key_now)] = list(self.scan_results)
                self._update_scan_row_from_result(idx, rebuilt)
                if changes:
                    self._record_adjustment(idx, changes, "saved_row_edit")
                    self._persist_single_scan_result_to_db(rebuilt, note="saved_row_edit")
                else:
                    self._refresh_row_status(idx)
            else:
                self._refresh_row_status(idx)
            for r in range(self.scan_result_preview.rowCount()):
                k = self.scan_result_preview.item(r, 0)
                if k and k.text().strip().lower() in {"exam code", "mã đề"}:
                    self.scan_result_preview.setItem(r, 1, QTableWidgetItem(new_exam_code or "-"))
                    break
            if self.scan_list.currentRow() == idx:
                self._update_scan_preview_from_saved_row(idx)
            self.btn_save_batch_subject.setEnabled(False)
            invalidated = self._invalidate_scoring_for_student_ids(
                [old_sid, inp_sid.text().strip()],
                reason="saved_row_edit",
            )
            if invalidated > 0:
                QMessageBox.information(
                    self,
                    "Tính điểm",
                    f"Đã đánh dấu {invalidated} bản ghi cần chấm lại do sửa bài. Vui lòng chạy lại Tính điểm.",
                )
            return
        dialog_state = {"index": idx, "loading": False}
        dialog_requires_full_status_refresh = {"value": False}
        default_zoom_factor = 0.34

        dlg = QDialog(None)
        dlg.setWindowTitle(f"Sửa bài thi: {Path(self.scan_results[dialog_state['index']].image_path).name}")
        dlg.setModal(True)
        dlg.setWindowFlag(Qt.Window, True)
        screen_geom = QApplication.primaryScreen().availableGeometry() if QApplication.primaryScreen() else self.geometry()
        dlg.setGeometry(screen_geom)
        dlg.setMinimumSize(screen_geom.size())
        dlg.setMaximumSize(screen_geom.size())
        QTimer.singleShot(0, lambda: (dlg.raise_(), dlg.showMaximized(), dlg.activateWindow()))
        lay = QHBoxLayout(dlg)
        splitter = QSplitter(Qt.Horizontal)
        left = QWidget()
        left_lay = QVBoxLayout(left)
        form = QFormLayout()

        inp_sid = QLineEdit()

        inp_code = QComboBox()
        inp_code.setEditable(True)
        inp_code.setInsertPolicy(QComboBox.NoInsert)

        mcq_label = QLabel("MCQ")
        mcq_hint = QLabel("Nhập đáp án trực tiếp vào từng ô MCQ")
        tf_label = QLabel("True / False")
        numeric_label = QLabel("Numeric")
        mcq_host = QWidget()
        mcq_host_lay = QVBoxLayout(mcq_host)
        mcq_host_lay.setContentsMargins(0, 0, 0, 0)
        tf_host = QWidget()
        tf_host_lay = QVBoxLayout(tf_host)
        tf_host_lay.setContentsMargins(0, 0, 0, 0)
        numeric_host = QWidget()
        numeric_host_lay = QVBoxLayout(numeric_host)
        numeric_host_lay.setContentsMargins(0, 0, 0, 0)

        form.addRow("Student ID", inp_sid)
        form.addRow("Exam Code", inp_code)
        left_lay.addLayout(form)
        answer_grid = QTableWidget(0, 3)
        answer_grid.setHorizontalHeaderLabels(["Phần", "Câu", "Bài làm"])
        answer_grid.verticalHeader().setVisible(False)
        answer_grid.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        answer_grid.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        answer_grid.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        _orig_answer_grid_key_release = answer_grid.keyReleaseEvent

        def _answer_grid_key_release(event) -> None:
            _orig_answer_grid_key_release(event)
            if answer_grid.currentColumn() != 2:
                return
            sec_item = answer_grid.item(answer_grid.currentRow(), 0)
            section_text = str(sec_item.text() if sec_item else "").strip().upper()
            cur_item = answer_grid.currentItem()
            if cur_item is not None:
                txt = str(cur_item.text() or "")
                up_txt = txt.upper()
                if txt != up_txt:
                    answer_grid.blockSignals(True)
                    cur_item.setText(up_txt)
                    answer_grid.blockSignals(False)
            key_code = int(getattr(event, "key", lambda: 0)() or 0)
            move_next = False
            if section_text == "MCQ":
                move_next = (65 <= key_code <= 90) or (48 <= key_code <= 57)
            elif section_text in {"TF", "NUMERIC"}:
                move_next = key_code in {Qt.Key_Return, Qt.Key_Enter, Qt.Key_Tab}
            if move_next:
                row_local = answer_grid.currentRow()
                if 0 <= row_local < answer_grid.rowCount() - 1:
                    answer_grid.setCurrentCell(row_local + 1, 2)
                elif row_local == answer_grid.rowCount() - 1:
                    answer_grid.setCurrentCell(row_local, 2)

        answer_grid.keyReleaseEvent = _answer_grid_key_release  # type: ignore[method-assign]

        def _normalize_answer_grid_cell_text(item: QTableWidgetItem) -> None:
            if item is None or item.column() != 2:
                return
            txt = str(item.text() or "")
            up_txt = txt.upper()
            if txt != up_txt:
                answer_grid.blockSignals(True)
                item.setText(up_txt)
                answer_grid.blockSignals(False)

        answer_grid.itemChanged.connect(_normalize_answer_grid_cell_text)
        left_lay.addWidget(answer_grid, 1)
        left_lay.addWidget(mcq_label)
        left_lay.addWidget(mcq_host)
        mcq_hint_row = QWidget()
        row_mcq = QHBoxLayout(mcq_hint_row)
        row_mcq.setContentsMargins(0, 0, 0, 0)
        row_mcq.addWidget(mcq_hint)
        row_mcq.addStretch()
        left_lay.addWidget(mcq_hint_row)
        left_lay.addWidget(tf_label)
        left_lay.addWidget(tf_host)
        tf_actions_row = QWidget()
        tf_actions = QHBoxLayout(tf_actions_row)
        tf_actions.setContentsMargins(0, 0, 0, 0)
        btn_add_tf = QPushButton("Thêm dòng TF")
        btn_del_tf = QPushButton("Xoá dòng chọn")
        tf_actions.addWidget(btn_add_tf)
        tf_actions.addWidget(btn_del_tf)
        tf_actions.addStretch()
        left_lay.addWidget(tf_actions_row)
        left_lay.addWidget(numeric_label)
        left_lay.addWidget(numeric_host)
        num_actions_row = QWidget()
        num_actions = QHBoxLayout(num_actions_row)
        num_actions.setContentsMargins(0, 0, 0, 0)
        btn_add_num = QPushButton("Thêm dòng Numeric")
        btn_del_num = QPushButton("Xoá dòng chọn")
        num_actions.addWidget(btn_add_num)
        num_actions.addWidget(btn_del_num)
        num_actions.addStretch()
        left_lay.addWidget(num_actions_row)

        button_row = QHBoxLayout()
        btn_prev = QPushButton("← Bài trước")
        btn_next = QPushButton("Bài tiếp →")
        btn_save = QPushButton("Save")
        btn_close = QPushButton("Close")
        button_row.addWidget(btn_prev)
        button_row.addWidget(btn_next)
        button_row.addStretch()
        button_row.addWidget(btn_save)
        button_row.addWidget(btn_close)
        left_lay.addLayout(button_row)

        splitter.addWidget(left)
        right = QWidget()
        right_lay = QVBoxLayout(right)
        right_lay.addWidget(QLabel("Ảnh bài làm"))
        zoom_row = QHBoxLayout()
        btn_prev_top = QPushButton("←")
        btn_zoom_out_dlg = QPushButton("-")
        btn_zoom_reset_dlg = QPushButton("34%")
        btn_zoom_in_dlg = QPushButton("+")
        btn_next_top = QPushButton("→")
        for btn in [btn_prev_top, btn_zoom_out_dlg, btn_zoom_reset_dlg, btn_zoom_in_dlg, btn_next_top]:
            btn.setMaximumWidth(60)
        zoom_row.addWidget(btn_prev_top)
        zoom_row.addWidget(btn_zoom_out_dlg)
        zoom_row.addWidget(btn_zoom_reset_dlg)
        zoom_row.addWidget(btn_zoom_in_dlg)
        zoom_row.addWidget(btn_next_top)
        zoom_row.addStretch()
        right_lay.addLayout(zoom_row)
        preview = QLabel()
        preview.setAlignment(Qt.AlignCenter)
        preview.setText("Đang tải ảnh bài làm...")
        preview_scroll = QScrollArea()
        preview_scroll.setWidgetResizable(False)
        preview_scroll.setAlignment(Qt.AlignCenter)
        preview_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        preview_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        preview_scroll.setFrameShape(QFrame.NoFrame)
        preview_scroll.setWidget(preview)
        right_lay.addWidget(preview_scroll, 1)
        splitter.addWidget(right)
        splitter.setChildrenCollapsible(False)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)
        lay.addWidget(splitter)
        QTimer.singleShot(0, lambda s=splitter: s.setSizes([max(1, s.width() // 2), max(1, s.width() // 2)]))

        editor_refs: dict[str, object] = {"mcq_edits": {}, "table_tf": None, "table_num": None, "answer_grid": answer_grid}
        expected_questions_state: dict[str, list[int]] = {"MCQ": [], "TF": [], "NUMERIC": []}
        question_mapping_state: dict[str, dict[str, dict[int, int]]] = {
            "MCQ": {"display_to_actual": {}, "actual_to_display": {}},
            "TF": {"display_to_actual": {}, "actual_to_display": {}},
            "NUMERIC": {"display_to_actual": {}, "actual_to_display": {}},
        }
        preview_state: dict[str, object] = {"pix": QPixmap(), "image_name": "-", "zoom": default_zoom_factor}
        loaded_snapshots: dict[str, dict[str, object]] = {}
        dialog_saved_images: set[str] = set()

        def _question_numbers(values) -> list[int]:
            out = {int(q) for q in (values or {}).keys() if str(q).strip().lstrip('-').isdigit()}
            return sorted(out)

        def _current_result() -> OMRResult:
            return self.scan_results[dialog_state["index"]]

        def _current_result_image_key() -> str:
            return self._result_identity_key(getattr(_current_result(), "image_path", ""))

        def _snapshot_from_result(result: OMRResult) -> dict[str, object]:
            return {
                "image_path": self._result_identity_key(getattr(result, "image_path", "")),
                "student_id": str(result.student_id or "").strip(),
                "exam_code": str(result.exam_code or "").strip(),
                "mcq_answers": {int(q): str(v) for q, v in (result.mcq_answers or {}).items()},
                "true_false_answers": {int(q): dict(v or {}) for q, v in (result.true_false_answers or {}).items()},
                "numeric_answers": {int(q): str(v) for q, v in (result.numeric_answers or {}).items()},
            }

        def _set_combo_to_value(combo: QComboBox, value: str) -> None:
            idx_found = combo.findData(value)
            if idx_found < 0:
                idx_found = combo.findText(value)
            if idx_found >= 0:
                combo.setCurrentIndex(idx_found)
            else:
                combo.setEditText(value)

        student_label_to_sid = dict(getattr(self, "_student_option_sid_map", {}) or {})
        valid_student_id_set = set(getattr(self, "_student_option_sid_set", set()) or set())
        sid_pattern = re.compile(r"^\[([^\]]+)\]")

        def _normalize_student_id_input(raw_text: str) -> str:
            text = str(raw_text or "").strip()
            if not text:
                return ""
            if text in student_label_to_sid:
                return str(student_label_to_sid[text] or "").strip()
            match = sid_pattern.match(text)
            if match:
                return str(match.group(1) or "").strip()
            return text

        def _configured_exam_codes_for_subject_cfg(subject_cfg: dict | None) -> list[str]:
            codes: set[str] = set()
            cfg = subject_cfg if isinstance(subject_cfg, dict) else {}
            imported = cfg.get("imported_answer_keys", {}) if isinstance(cfg.get("imported_answer_keys", {}), dict) else {}
            codes.update(str(x).strip() for x in imported.keys() if str(x).strip())
            return sorted(codes)

        def _valid_exam_codes(subject_key: str, current_code: str = "", subject_cfg: dict | None = None) -> list[str]:
            codes: set[str] = set()
            subject = str(subject_key or "").strip()
            if not subject:
                return []
            codes.update(_configured_exam_codes_for_subject_cfg(subject_cfg))
            try:
                codes.update(str(x).strip() for x in self._fetch_answer_keys_for_subject_scoped(subject).keys() if str(x).strip())
            except Exception:
                pass
            if self.answer_keys:
                codes.update(
                    str(item.exam_code).strip()
                    for item in self.answer_keys.keys.values()
                    if str(getattr(item, "subject", "") or "").strip() == subject and str(getattr(item, "exam_code", "") or "").strip()
                )
            if str(current_code or "").strip():
                codes.add(str(current_code or "").strip())
            return sorted(codes)

        def _populate_exam_code_combo(combo: QComboBox, subject_key: str, current_code: str) -> None:
            combo.blockSignals(True)
            combo.clear()
            subject_cfg_local = self._selected_batch_subject_config() or {}
            valid_codes = _valid_exam_codes(subject_key, current_code, subject_cfg_local)
            selected_code = str(current_code or "").strip()
            selected_is_valid = bool(selected_code and selected_code in valid_codes)
            if not selected_is_valid:
                combo.addItem("-", "-")
            for code in valid_codes:
                combo.addItem(code, code)
            if combo.count() == 0:
                combo.addItem("-", "-")
            if selected_is_valid:
                _set_combo_to_value(combo, selected_code)
            else:
                combo.setCurrentIndex(0)
            combo.blockSignals(False)

        def _answer_key_for_exam_code(exam_code_text: str):
            subject = str(self._current_batch_subject_key() or self.active_batch_subject_key or "").strip()
            code = str(exam_code_text or "").strip()
            if not subject or not self.answer_keys:
                return None
            if code and code != "-" and "?" not in code:
                key = self.answer_keys.get_flexible(subject, code)
                if key is not None:
                    return key
            return None

        def _expected_questions_for_dialog(result: OMRResult, exam_code_text: str, data_snapshot: dict[str, object]) -> dict[str, list[int]]:
            configured_counts = self._subject_section_question_counts(self._current_batch_subject_key())
            default_by_config = {
                sec: list(range(1, max(0, int(configured_counts.get(sec, 0) or 0)) + 1))
                for sec in ["MCQ", "TF", "NUMERIC"]
            }
            mapping_payload = {
                sec: {"display_to_actual": {}, "actual_to_display": {}}
                for sec in ["MCQ", "TF", "NUMERIC"]
            }
            key = _answer_key_for_exam_code(exam_code_text)
            if key is not None:
                full_credit = getattr(key, "full_credit_questions", {}) or {}
                invalid_rows = getattr(key, "invalid_answer_rows", {}) or {}

                def _extra_for_section(sec: str) -> set[int]:
                    extra: set[int] = set()
                    for q in list((full_credit.get(sec, []) or [])):
                        try:
                            extra.add(int(q))
                        except Exception:
                            continue
                    for q in dict((invalid_rows.get(sec, {}) or {})).keys():
                        try:
                            extra.add(int(q))
                        except Exception:
                            continue
                    return extra

                expected = {
                    "MCQ": sorted(set(int(q) for q in (key.answers or {}).keys()) | _extra_for_section("MCQ")) or list(default_by_config["MCQ"]),
                    "TF": sorted(set(int(q) for q in (key.true_false_answers or {}).keys()) | _extra_for_section("TF")) or list(default_by_config["TF"]),
                    "NUMERIC": sorted(set(int(q) for q in (key.numeric_answers or {}).keys()) | _extra_for_section("NUMERIC")) or list(default_by_config["NUMERIC"]),
                }
            else:
                expected = {
                    "MCQ": list(default_by_config["MCQ"]),
                    "TF": list(default_by_config["TF"]),
                    "NUMERIC": list(default_by_config["NUMERIC"]),
                }
            for sec in ["MCQ", "TF", "NUMERIC"]:
                limit = max(0, int(configured_counts.get(sec, 0) or 0))
                if limit <= 0:
                    expected[sec] = []
                else:
                    actual_questions = sorted(expected.get(sec, []))[:limit]
                    if not actual_questions:
                        display_questions = list(range(1, limit + 1))
                        actual_questions = list(display_questions)
                    else:
                        contiguous_local = actual_questions == list(range(1, len(actual_questions) + 1))
                        display_questions = list(actual_questions) if contiguous_local else list(range(1, len(actual_questions) + 1))
                    expected[sec] = list(display_questions)
                    mapping_payload[sec]["display_to_actual"] = {
                        int(display_q): int(actual_q)
                        for display_q, actual_q in zip(display_questions, actual_questions)
                    }
                    mapping_payload[sec]["actual_to_display"] = {
                        int(actual_q): int(display_q)
                        for display_q, actual_q in zip(display_questions, actual_questions)
                    }
            for sec in ["MCQ", "TF", "NUMERIC"]:
                question_mapping_state[sec]["display_to_actual"] = dict(mapping_payload[sec]["display_to_actual"])
                question_mapping_state[sec]["actual_to_display"] = dict(mapping_payload[sec]["actual_to_display"])
            return expected

        def _clear_layout(layout_obj: QVBoxLayout) -> None:
            while layout_obj.count():
                item = layout_obj.takeAt(0)
                widget = item.widget()
                child_layout = item.layout()
                if widget is not None:
                    widget.deleteLater()
                elif child_layout is not None:
                    while child_layout.count():
                        sub = child_layout.takeAt(0)
                        if sub.widget() is not None:
                            sub.widget().deleteLater()

        def _build_pair_table(question_numbers: list[int], data: dict[int, str], value_placeholder: str = "") -> QTableWidget:
            table = QTableWidget(0, 2)
            table.setHorizontalHeaderLabels(["Câu", "Giá trị"])
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            rows = list(question_numbers or [])
            for q in rows:
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QTableWidgetItem(str(int(q))))
                item_v = QTableWidgetItem(str((data or {}).get(int(q), "") or ""))
                if value_placeholder:
                    item_v.setToolTip(value_placeholder)
                table.setItem(r, 1, item_v)
            return table

        def _build_mcq_grid(question_numbers: list[int], data: dict[int, str]) -> tuple[QWidget, dict[int, QLineEdit]]:
            box = QWidget()
            grid = QGridLayout(box)
            grid.setContentsMargins(0, 0, 0, 0)
            grid.setHorizontalSpacing(8)
            grid.setVerticalSpacing(6)
            edits: dict[int, QLineEdit] = {}
            questions = list(question_numbers or [])
            cols = 8
            for idx_q, q_no in enumerate(questions):
                row = (idx_q // cols) * 2
                col = idx_q % cols
                lbl = QLabel(str(q_no))
                lbl.setAlignment(Qt.AlignCenter)
                edit = QLineEdit(str((data or {}).get(int(q_no), "") or ""))
                edit.setMaxLength(1)
                edit.setMaximumWidth(52)
                edit.setAlignment(Qt.AlignCenter)
                edits[int(q_no)] = edit
                grid.addWidget(lbl, row, col)
                grid.addWidget(edit, row + 1, col)
            grid.setColumnStretch(cols, 1)
            return box, edits

        def _build_tf_table(question_numbers: list[int], data: dict[int, dict[str, bool]]) -> QTableWidget:
            table = QTableWidget(0, 5)
            labels = ["a", "b", "c", "d"]
            table.setHorizontalHeaderLabels(["Câu", *[s.upper() for s in labels]])
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
            for c in range(1, 5):
                table.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
            rows = list(question_numbers or [])
            for q in rows:
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QTableWidgetItem(str(int(q))))
                flags = dict((data or {}).get(int(q), {}) or {})
                for i, key in enumerate(labels, start=1):
                    cb = QComboBox()
                    cb.addItem("-", None)
                    cb.addItem("Đúng", True)
                    cb.addItem("Sai", False)
                    val = flags.get(key, None)
                    idx_found = 0
                    for k in range(cb.count()):
                        if cb.itemData(k) is val:
                            idx_found = k
                            break
                    cb.setCurrentIndex(idx_found)
                    table.setCellWidget(r, i, cb)
            return table

        def _collect_editor_snapshot(validate: bool = True) -> dict[str, object]:
            result = _current_result()
            student_id_text = _normalize_student_id_input(inp_sid.text())
            exam_code_text = str(inp_code.currentData() or inp_code.currentText() or "").strip()
            if exam_code_text == "-":
                exam_code_text = ""
            if validate:
                valid_exam_codes = _valid_exam_codes(self._current_batch_subject_key(), exam_code_text)
                if valid_student_id_set and student_id_text and student_id_text not in valid_student_id_set:
                    raise ValueError(f"Student ID '{student_id_text}' không có trong danh sách học sinh hợp lệ của ca thi.")
                if valid_exam_codes and exam_code_text not in {"", "-"} and exam_code_text not in valid_exam_codes:
                    raise ValueError(f"Exam code '{exam_code_text}' không có đáp án hợp lệ cho môn hiện tại.")
            snapshot = {
                "image_path": self._result_identity_key(getattr(result, "image_path", "")),
                "student_id": student_id_text,
                "exam_code": exam_code_text,
                "mcq_answers": {},
                "true_false_answers": {},
                "numeric_answers": {},
            }
            answer_grid_local = editor_refs.get("answer_grid")
            if isinstance(answer_grid_local, QTableWidget):
                for r in range(answer_grid_local.rowCount()):
                    sec_item = answer_grid_local.item(r, 0)
                    q_item = answer_grid_local.item(r, 1)
                    v_item = answer_grid_local.item(r, 2)
                    sec = str(sec_item.text() if sec_item else "").strip().upper()
                    q_text = str(q_item.text() if q_item else "").strip()
                    v_text = str(v_item.text() if v_item else "").strip()
                    if not q_text.lstrip('-').isdigit():
                        continue
                    display_q = int(q_text)
                    if sec == "MCQ":
                        actual_q = int(question_mapping_state.get("MCQ", {}).get("display_to_actual", {}).get(display_q, display_q))
                        if v_text:
                            snapshot["mcq_answers"][actual_q] = str(v_text).upper()[:1]
                    elif sec == "TF":
                        actual_q = int(question_mapping_state.get("TF", {}).get("display_to_actual", {}).get(display_q, display_q))
                        compact = "".join(ch for ch in str(v_text).upper() if ch in {"Đ", "D", "S", "_", "T", "F", "1", "0"})
                        flags: dict[str, bool] = {}
                        for i, ch in enumerate(compact[:4]):
                            key_flag = ["a", "b", "c", "d"][i]
                            if ch == "_":
                                continue
                            flags[key_flag] = ch in {"Đ", "D", "T", "1"}
                        if flags:
                            snapshot["true_false_answers"][actual_q] = flags
                    elif sec == "NUMERIC":
                        actual_q = int(question_mapping_state.get("NUMERIC", {}).get("display_to_actual", {}).get(display_q, display_q))
                        snapshot["numeric_answers"][actual_q] = v_text
                return snapshot
            mcq_edits = editor_refs.get("mcq_edits", {}) or {}
            for q_no, edit in mcq_edits.items():
                v_text = str(edit.text() if edit else "").strip().upper()[:1]
                if v_text:
                    actual_q = int(question_mapping_state.get("MCQ", {}).get("display_to_actual", {}).get(int(q_no), int(q_no)))
                    snapshot["mcq_answers"][actual_q] = v_text

            table_num = editor_refs.get("table_num")
            if isinstance(table_num, QTableWidget):
                for r in range(table_num.rowCount()):
                    q_item = table_num.item(r, 0)
                    v_item = table_num.item(r, 1)
                    q_text = str(q_item.text() if q_item else "").strip()
                    v_text = str(v_item.text() if v_item else "").strip()
                    if not q_text and not v_text:
                        continue
                    if validate and not q_text.lstrip('-').isdigit():
                        raise ValueError(f"Numeric dòng {r+1}: Câu phải là số nguyên.")
                    if not q_text.lstrip('-').isdigit() or not v_text:
                        continue
                    display_q = int(q_text)
                    actual_q = int(question_mapping_state.get("NUMERIC", {}).get("display_to_actual", {}).get(display_q, display_q))
                    snapshot["numeric_answers"][actual_q] = v_text

            table_tf = editor_refs.get("table_tf")
            if isinstance(table_tf, QTableWidget):
                labels = ["a", "b", "c", "d"]
                for r in range(table_tf.rowCount()):
                    q_item = table_tf.item(r, 0)
                    q_text = str(q_item.text() if q_item else "").strip()
                    if not q_text:
                        continue
                    if validate and not q_text.lstrip('-').isdigit():
                        raise ValueError(f"TF dòng {r+1}: Câu phải là số nguyên.")
                    if not q_text.lstrip('-').isdigit():
                        continue
                    display_q = int(q_text)
                    q = int(question_mapping_state.get("TF", {}).get("display_to_actual", {}).get(display_q, display_q))
                    flags: dict[str, bool] = {}
                    for i, key in enumerate(labels, start=1):
                        cb = table_tf.cellWidget(r, i)
                        if not isinstance(cb, QComboBox):
                            continue
                        val = cb.currentData()
                        if isinstance(val, bool):
                            flags[key] = val
                    if flags:
                        snapshot["true_false_answers"][q] = flags
            return snapshot

        def _remove_selected_row(table_key: str) -> None:
            table = editor_refs.get(table_key)
            if isinstance(table, QTableWidget):
                row = table.currentRow()
                if row >= 0:
                    table.removeRow(row)

        def _add_pair_row() -> None:
            table = editor_refs.get("table_num")
            if isinstance(table, QTableWidget):
                numeric_limit = len(expected_questions_state.get("NUMERIC", []) or [])
                if numeric_limit > 0 and table.rowCount() >= numeric_limit:
                    return
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QTableWidgetItem(""))
                table.setItem(r, 1, QTableWidgetItem(""))

        def _add_tf_row() -> None:
            table = editor_refs.get("table_tf")
            if isinstance(table, QTableWidget):
                tf_limit = len(expected_questions_state.get("TF", []) or [])
                if tf_limit > 0 and table.rowCount() >= tf_limit:
                    return
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QTableWidgetItem(""))
                for i in range(1, 5):
                    cb = QComboBox()
                    cb.addItem("-", None)
                    cb.addItem("Đúng", True)
                    cb.addItem("Sai", False)
                    table.setCellWidget(r, i, cb)

        btn_add_tf.clicked.connect(_add_tf_row)
        btn_del_tf.clicked.connect(lambda: _remove_selected_row("table_tf"))
        btn_add_num.clicked.connect(_add_pair_row)
        btn_del_num.clicked.connect(lambda: _remove_selected_row("table_num"))

        def _render_preview_for_result(result: OMRResult) -> None:
            image_path = str(getattr(result, "image_path", "") or "").strip()
            image_name = Path(image_path).name if image_path else "-"
            aligned_pix = self._aligned_image_to_qpixmap(getattr(result, "aligned_image", None))
            if not aligned_pix.isNull():
                pix = aligned_pix
            elif hasattr(self, "preview_source_pixmap") and not self.preview_source_pixmap.isNull() and self.scan_list.currentRow() == dialog_state["index"]:
                pix = self.preview_source_pixmap
            else:
                pix = QPixmap(image_path) if image_path else QPixmap()
            preview_state["pix"] = pix
            preview_state["image_name"] = image_name
            preview_state["zoom"] = default_zoom_factor
            _apply_preview_zoom()

        def _apply_preview_zoom() -> None:
            base_pix = preview_state.get("pix", QPixmap())
            image_name = str(preview_state.get("image_name", "-"))
            factor = float(preview_state.get("zoom", default_zoom_factor) or default_zoom_factor)
            if isinstance(base_pix, QPixmap) and not base_pix.isNull():
                scaled = base_pix.scaled(base_pix.size() * factor, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                preview.setText("")
                preview.setPixmap(scaled)
                preview.resize(scaled.size())
                preview.setMinimumSize(scaled.size())
                preview.setMaximumSize(scaled.size())
                preview.adjustSize()
                preview_scroll.ensureVisible(0, 0)
            else:
                preview.setPixmap(QPixmap())
                preview.setMinimumSize(preview_scroll.viewport().size())
                preview.setMaximumSize(16777215, 16777215)
                preview.setText(f"Không thể tải ảnh bài làm tương ứng.\n{image_name}")
                preview.resize(preview_scroll.viewport().size())
            btn_zoom_reset_dlg.setText(f"{int(round(factor * 100))}%")

        btn_zoom_out_dlg.clicked.connect(lambda: (preview_state.__setitem__("zoom", max(0.1, float(preview_state.get("zoom", default_zoom_factor)) / 1.2)), _apply_preview_zoom()))
        btn_zoom_in_dlg.clicked.connect(lambda: (preview_state.__setitem__("zoom", min(5.0, float(preview_state.get("zoom", default_zoom_factor)) * 1.2)), _apply_preview_zoom()))
        btn_zoom_reset_dlg.clicked.connect(lambda: (preview_state.__setitem__("zoom", default_zoom_factor), _apply_preview_zoom()))

        def _refresh_editor_widgets(data_snapshot: dict[str, object]) -> None:
            result = _current_result()
            expected = _expected_questions_for_dialog(result, str(data_snapshot.get("exam_code", "") or ""), data_snapshot)
            expected_questions_state["MCQ"] = list(expected.get("MCQ", []) or [])
            expected_questions_state["TF"] = list(expected.get("TF", []) or [])
            expected_questions_state["NUMERIC"] = list(expected.get("NUMERIC", []) or [])
            section_limits = self._subject_section_question_counts(self._current_batch_subject_key())
            mcq_map_actual_to_display = question_mapping_state.get("MCQ", {}).get("actual_to_display", {}) or {}
            mcq_data = data_snapshot.get("mcq_answers", {}) or {}
            mcq_limit = int(section_limits.get("MCQ", 0) or 0)
            if mcq_limit <= 0:
                mcq_questions = []
                mcq_data_display = {}
            else:
                mcq_data_display_keys = {
                    int(mcq_map_actual_to_display.get(int(q), int(q)))
                    for q in mcq_data.keys()
                }
                mcq_questions = sorted(set(expected.get("MCQ", []) or []).union(mcq_data_display_keys))
                mcq_questions = list(mcq_questions[:mcq_limit])
                allowed_mcq_questions = set(mcq_questions)
                mcq_data_display = {
                    int(mcq_map_actual_to_display.get(int(q), int(q))): str(v)
                    for q, v in mcq_data.items()
                    if int(mcq_map_actual_to_display.get(int(q), int(q))) in allowed_mcq_questions
                }

            tf_data = data_snapshot.get("true_false_answers", {}) or {}
            tf_map_actual_to_display = question_mapping_state.get("TF", {}).get("actual_to_display", {}) or {}
            tf_limit = int(section_limits.get("TF", 0) or 0)
            if tf_limit <= 0:
                tf_questions = []
                tf_data = {}
            else:
                tf_data_display_keys = {
                    int(tf_map_actual_to_display.get(int(q), int(q)))
                    for q in tf_data.keys()
                }
                tf_questions = sorted(set(expected.get("TF", []) or []).union(tf_data_display_keys))
                tf_questions = list(tf_questions[:tf_limit])
                allowed_tf_questions = set(tf_questions)
                tf_data = {
                    int(tf_map_actual_to_display.get(int(q), int(q))): dict(v or {})
                    for q, v in tf_data.items()
                    if int(tf_map_actual_to_display.get(int(q), int(q))) in allowed_tf_questions
                }

            numeric_data = data_snapshot.get("numeric_answers", {}) or {}
            numeric_map_actual_to_display = question_mapping_state.get("NUMERIC", {}).get("actual_to_display", {}) or {}
            numeric_limit = int(section_limits.get("NUMERIC", 0) or 0)
            if numeric_limit <= 0:
                numeric_questions = []
                numeric_data = {}
            else:
                numeric_data_display_keys = {
                    int(numeric_map_actual_to_display.get(int(q), int(q)))
                    for q in numeric_data.keys()
                }
                numeric_questions = sorted(set(expected.get("NUMERIC", []) or []).union(numeric_data_display_keys))
                numeric_questions = list(numeric_questions[:numeric_limit])
                allowed_numeric_questions = set(numeric_questions)
                numeric_data = {
                    int(numeric_map_actual_to_display.get(int(q), int(q))): str(v)
                    for q, v in numeric_data.items()
                    if int(numeric_map_actual_to_display.get(int(q), int(q))) in allowed_numeric_questions
                }
            grid_local = editor_refs.get("answer_grid")
            if isinstance(grid_local, QTableWidget):
                grid_local.blockSignals(True)
                grid_local.setRowCount(0)
                for q in mcq_questions:
                    r = grid_local.rowCount()
                    grid_local.insertRow(r)
                    grid_local.setItem(r, 0, QTableWidgetItem("MCQ"))
                    grid_local.setItem(r, 1, QTableWidgetItem(str(int(q))))
                    grid_local.setItem(r, 2, QTableWidgetItem(str(mcq_data_display.get(int(q), "") or "")))
                for q in tf_questions:
                    flags = dict(tf_data.get(int(q), {}) or {})
                    token = "".join(("Đ" if bool(flags.get(k)) else ("S" if k in flags else "_")) for k in ["a", "b", "c", "d"])
                    r = grid_local.rowCount()
                    grid_local.insertRow(r)
                    grid_local.setItem(r, 0, QTableWidgetItem("TF"))
                    grid_local.setItem(r, 1, QTableWidgetItem(str(int(q))))
                    grid_local.setItem(r, 2, QTableWidgetItem(token))
                for q in numeric_questions:
                    r = grid_local.rowCount()
                    grid_local.insertRow(r)
                    grid_local.setItem(r, 0, QTableWidgetItem("NUMERIC"))
                    grid_local.setItem(r, 1, QTableWidgetItem(str(int(q))))
                    grid_local.setItem(r, 2, QTableWidgetItem(str(numeric_data.get(int(q), "") or "")))
                for r in range(grid_local.rowCount()):
                    sec_item = grid_local.item(r, 0)
                    q_item = grid_local.item(r, 1)
                    if sec_item:
                        sec_item.setFlags(sec_item.flags() & ~Qt.ItemIsEditable)
                    if q_item:
                        q_item.setFlags(q_item.flags() & ~Qt.ItemIsEditable)
                grid_local.blockSignals(False)
            mcq_label.setVisible(False)
            mcq_host.setVisible(False)
            mcq_hint_row.setVisible(False)
            tf_label.setVisible(False)
            tf_host.setVisible(False)
            tf_actions_row.setVisible(False)
            numeric_label.setVisible(False)
            numeric_host.setVisible(False)
            num_actions_row.setVisible(False)

        def _apply_changes(save_feedback: bool = False) -> bool:
            idx_local = dialog_state["index"]
            current_ref = _current_result()
            if any(i != idx_local and self.scan_results[i] is current_ref for i in range(len(self.scan_results))):
                self.scan_results[idx_local] = self._lightweight_result_copy(current_ref)
            result = _current_result()
            try:
                snapshot = _collect_editor_snapshot(validate=True)
            except Exception as exc:
                QMessageBox.warning(self, "Dữ liệu không hợp lệ", str(exc))
                return False

            old_sid_for_score = str(result.student_id or "").strip()
            changes: list[str] = []
            new_sid = str(snapshot["student_id"] or "").strip()
            new_code = str(snapshot["exam_code"] or "").strip()
            new_mcq_answers = {int(q): str(v) for q, v in (snapshot.get("mcq_answers", {}) or {}).items()}
            new_tf_answers = {int(q): dict(v or {}) for q, v in (snapshot.get("true_false_answers", {}) or {}).items()}
            new_numeric_answers = {int(q): str(v) for q, v in (snapshot.get("numeric_answers", {}) or {}).items()}

            if new_sid != (result.student_id or ""):
                changes.append(f"student_id: '{result.student_id or ''}' -> '{new_sid}'")
                result.student_id = new_sid
                dialog_requires_full_status_refresh["value"] = True
            if new_code != (result.exam_code or ""):
                changes.append(f"exam_code: '{result.exam_code or ''}' -> '{new_code}'")
                result.exam_code = new_code
                dialog_requires_full_status_refresh["value"] = True
            if new_mcq_answers != (result.mcq_answers or {}):
                result.mcq_answers = new_mcq_answers
                changes.append("mcq_answers updated")
            if new_tf_answers != (result.true_false_answers or {}):
                result.true_false_answers = new_tf_answers
                changes.append("true_false_answers updated")
            if new_numeric_answers != (result.numeric_answers or {}):
                result.numeric_answers = new_numeric_answers
                changes.append("numeric_answers updated")

            if not changes:
                return True

            # Clear legacy manual-content text so Batch Scan "Nội dung" always reflects
            # the current structured answers edited in this dialog.
            setattr(result, "manual_content_override", "")
            self._mark_result_manually_edited(result, idx_local)
            self._refresh_student_profile_for_result(result)
            scoped = self._scoped_result_copy(result)
            self.scan_blank_summary[idx_local] = self._compute_blank_questions(scoped)
            self._update_scan_row_from_result(idx_local, result)
            self._record_adjustment(idx_local, changes, "dialog_edit")
            self._persist_single_scan_result_to_db(result, note="dialog_edit")
            subject_key_now = self._current_batch_subject_key()
            if subject_key_now:
                self.scan_results_by_subject[self._batch_result_subject_key(subject_key_now)] = list(self.scan_results)
            dialog_saved_images.add(_current_result_image_key())
            self.btn_save_batch_subject.setEnabled(False)
            invalidated = self._invalidate_scoring_for_student_ids([old_sid_for_score, str(result.student_id or "").strip()], reason="dialog_edit")
            self._update_batch_scan_bottom_status_text()
            loaded_snapshots[_current_result_image_key()] = _snapshot_from_result(result)
            if save_feedback:
                notices = ["Đã lưu thay đổi cho bài hiện tại."]
                if invalidated > 0:
                    notices.append(f"Đã đánh dấu {invalidated} bản ghi cần chấm lại. Vui lòng chạy lại Tính điểm.")
                QMessageBox.information(self, "Sửa bài thi", "\n".join(notices))
            return True

        def _load_result_into_dialog(new_index: int, preserve_snapshot: dict[str, object] | None = None) -> None:
            if new_index < 0 or new_index >= len(self.scan_results):
                return
            dialog_state["loading"] = True
            dialog_state["index"] = new_index
            self.scan_list.setCurrentCell(new_index, self.SCAN_COL_STUDENT_ID)
            result = self.scan_results[new_index]
            image_key = self._result_identity_key(getattr(result, "image_path", ""))
            dlg.setWindowTitle(f"Sửa bài thi: {Path(result.image_path).name}")
            self._load_student_correction_options(str(result.student_id or "").strip())
            inp_sid.blockSignals(True)
            sid_value = str((preserve_snapshot or {}).get("student_id", result.student_id or "")).strip()
            inp_sid.setText(sid_value)
            sid_labels = list(getattr(self, "_student_option_labels_cache", []) or [])
            sid_completer = QCompleter(sid_labels, inp_sid)
            sid_completer.setCaseSensitivity(Qt.CaseInsensitive)
            sid_completer.setFilterMode(Qt.MatchContains)
            sid_completer.activated.connect(lambda text: inp_sid.setText(_normalize_student_id_input(str(text))))
            inp_sid.setCompleter(sid_completer)
            inp_sid.blockSignals(False)

            subject_key = self._current_batch_subject_key()
            current_code = str((preserve_snapshot or {}).get("exam_code", result.exam_code or ""))
            _populate_exam_code_combo(inp_code, subject_key, current_code)

            data_snapshot = preserve_snapshot or loaded_snapshots.get(image_key) or _snapshot_from_result(result)
            loaded_snapshots[image_key] = {
                "image_path": image_key,
                "student_id": str(data_snapshot.get("student_id", "") or "").strip(),
                "exam_code": str(data_snapshot.get("exam_code", "") or "").strip(),
                "mcq_answers": {int(q): str(v) for q, v in (data_snapshot.get("mcq_answers", {}) or {}).items()},
                "true_false_answers": {int(q): dict(v or {}) for q, v in (data_snapshot.get("true_false_answers", {}) or {}).items()},
                "numeric_answers": {int(q): str(v) for q, v in (data_snapshot.get("numeric_answers", {}) or {}).items()},
            }
            _refresh_editor_widgets(loaded_snapshots[image_key])
            _render_preview_for_result(result)
            btn_prev.setEnabled(new_index > 0)
            btn_prev_top.setEnabled(new_index > 0)
            btn_next.setEnabled(new_index < len(self.scan_results) - 1)
            btn_next_top.setEnabled(new_index < len(self.scan_results) - 1)
            dialog_state["loading"] = False

        def _rebuild_for_exam_code_change() -> None:
            if dialog_state["loading"]:
                return
            try:
                snapshot = _collect_editor_snapshot(validate=False)
            except Exception:
                snapshot = _snapshot_from_result(_current_result())
            image_key = _current_result_image_key()
            loaded_snapshots[image_key] = {
                "image_path": image_key,
                "student_id": str(snapshot.get("student_id", "") or "").strip(),
                "exam_code": str(snapshot.get("exam_code", "") or "").strip(),
                "mcq_answers": {int(q): str(v) for q, v in (snapshot.get("mcq_answers", {}) or {}).items()},
                "true_false_answers": {int(q): dict(v or {}) for q, v in (snapshot.get("true_false_answers", {}) or {}).items()},
                "numeric_answers": {int(q): str(v) for q, v in (snapshot.get("numeric_answers", {}) or {}).items()},
            }
            _refresh_editor_widgets(loaded_snapshots[image_key])

        def _navigate(offset: int) -> None:
            target = dialog_state["index"] + offset
            if target < 0 or target >= len(self.scan_results):
                return
            if not _apply_changes(save_feedback=False):
                return
            _load_result_into_dialog(target)

        def _has_unsaved_changes() -> bool:
            try:
                snapshot = _collect_editor_snapshot(validate=False)
            except Exception:
                return False
            baseline = loaded_snapshots.get(_current_result_image_key(), _snapshot_from_result(_current_result()))
            return snapshot != baseline

        def _request_close() -> None:
            if _has_unsaved_changes():
                msg = QMessageBox(self)
                msg.setWindowTitle("Sửa bài thi")
                msg.setText("Bài hiện tại có thay đổi chưa lưu.")
                msg.setInformativeText("Bạn muốn lưu trước khi đóng cửa sổ sửa bài?")
                msg.setStandardButtons(QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
                msg.setDefaultButton(QMessageBox.Save)
                choice = msg.exec()
                if choice == QMessageBox.Cancel:
                    return
                if choice == QMessageBox.Save and not _apply_changes(save_feedback=False):
                    return
            if dialog_saved_images:
                for saved_image in sorted(dialog_saved_images):
                    saved_idx = self._row_index_by_image_path(saved_image)
                    if 0 <= saved_idx < len(self.scan_results):
                        saved_result = self.scan_results[saved_idx]
                        self._refresh_student_profile_for_result(saved_result, saved_idx)
                        scoped_saved = self._scoped_result_copy(saved_result)
                        self.scan_blank_summary[saved_idx] = self._compute_blank_questions(scoped_saved)
                        self._update_scan_row_from_result(saved_idx, saved_result)
                subject_key_now = self._current_batch_subject_key()
                if subject_key_now:
                    self.scan_results_by_subject[self._batch_result_subject_key(subject_key_now)] = list(self.scan_results)
                if dialog_requires_full_status_refresh["value"]:
                    self._refresh_all_statuses()
                else:
                    for saved_image in sorted(dialog_saved_images):
                        saved_idx = self._row_index_by_image_path(saved_image)
                        if 0 <= saved_idx < self.scan_list.rowCount():
                            self._refresh_row_status(saved_idx)
                self._update_batch_scan_bottom_status_text()
                current_idx = dialog_state["index"]
                if 0 <= current_idx < len(self.scan_results):
                    self.scan_list.setCurrentCell(current_idx, self.SCAN_COL_STUDENT_ID)
                    self._update_scan_preview(current_idx)
                    self._sync_correction_detail_panel(self.scan_results[current_idx], rebuild_editor=False)
            dlg.accept()

        inp_code.currentIndexChanged.connect(lambda _=0: _rebuild_for_exam_code_change())
        inp_code.editTextChanged.connect(lambda _text="": _rebuild_for_exam_code_change())
        btn_prev.clicked.connect(lambda: _navigate(-1))
        btn_prev_top.clicked.connect(lambda: _navigate(-1))
        btn_next.clicked.connect(lambda: _navigate(1))
        btn_next_top.clicked.connect(lambda: _navigate(1))
        btn_save.clicked.connect(lambda: _apply_changes(save_feedback=True))
        btn_close.clicked.connect(_request_close)
        dlg.reject = _request_close

        _load_result_into_dialog(dialog_state["index"])
        dlg.exec()
    def apply_manual_correction(self) -> None:
        idx = self.scan_list.currentRow()
        if idx < 0 or idx >= len(self.scan_results):
            QMessageBox.warning(self, "No selection", "Select a scanned result first.")
            return
        txt = self.manual_edit.toPlainText().strip()
        if not txt:
            return
        try:
            patch = json.loads(txt)
        except Exception as exc:
            QMessageBox.warning(self, "Invalid JSON", f"Cannot parse manual correction:\n{exc}")
            return

        res = self.scan_results[idx]
        if any(i != idx and self.scan_results[i] is res for i in range(len(self.scan_results))):
            res = self._lightweight_result_copy(res)
            self.scan_results[idx] = res
        old_sid_for_score = str(res.student_id or "").strip()
        changes: list[str] = []
        if "student_id" in patch:
            new_sid = str(patch["student_id"])
            if new_sid != (res.student_id or ""):
                changes.append(f"student_id: '{res.student_id or ''}' -> '{new_sid}'")
            res.student_id = new_sid
        if "exam_code" in patch:
            new_code = str(patch["exam_code"])
            if new_code != (res.exam_code or ""):
                changes.append(f"exam_code: '{res.exam_code or ''}' -> '{new_code}'")
            res.exam_code = new_code
        if isinstance(patch.get("mcq_answers"), dict):
            new_mcq_answers = {int(k): str(v) for k, v in patch["mcq_answers"].items()}
            if new_mcq_answers != (res.mcq_answers or {}):
                res.mcq_answers = new_mcq_answers
                changes.append("mcq_answers updated")
        if isinstance(patch.get("numeric_answers"), dict):
            new_numeric_answers = {int(k): str(v) for k, v in patch["numeric_answers"].items()}
            if new_numeric_answers != (res.numeric_answers or {}):
                res.numeric_answers = new_numeric_answers
                changes.append("numeric_answers updated")
        if isinstance(patch.get("true_false_answers"), dict):
            new_tf_answers = patch["true_false_answers"]
            if new_tf_answers != (res.true_false_answers or {}):
                res.true_false_answers = new_tf_answers
                changes.append("true_false_answers updated")

        sid = (res.student_id or "").strip() or "-"
        sid_item = QTableWidgetItem(sid)
        sid_item.setData(Qt.UserRole, str(res.image_path))
        sid_item.setData(Qt.UserRole + 1, res.exam_code or "")
        sid_item.setData(Qt.UserRole + 2, self._short_recognition_text_for_result(res))
        self.scan_list.setItem(idx, self.SCAN_COL_STUDENT_ID, sid_item)
        if changes:
            self._mark_result_manually_edited(res, idx)
            self._refresh_student_profile_for_result(res)
            scoped = self._scoped_result_copy(res)
            self.scan_blank_summary[idx] = self._compute_blank_questions(scoped)
            self._update_scan_row_from_result(idx, res)
            self._record_adjustment(idx, changes, "manual_json")
            self._persist_single_scan_result_to_db(res, note="manual_json")
            self.btn_save_batch_subject.setEnabled(False)
            invalidated = self._invalidate_scoring_for_student_ids(
                [old_sid_for_score, str(res.student_id or "").strip()],
                reason="manual_json",
            )
            if invalidated > 0:
                QMessageBox.information(
                    self,
                    "Tính điểm",
                    f"Đã đánh dấu {invalidated} bản ghi cần chấm lại do sửa bài. Vui lòng chạy lại Tính điểm.",
                )
        self._refresh_all_statuses()
        self._update_scan_preview(idx)
        self._load_selected_result_for_correction()
        QMessageBox.information(self, "Correction", "Manual correction applied to selected scan.")

    def _finalize_scoring_rows(
        self,
        subject: str,
        rows: list,
        prev_subject_scores: dict[str, dict],
        edited_student_ids: set[str],
        mode_text: str,
        note: str,
        missing: int,
        failed_scans: list[dict[str, str]],
        smart_scored_scans: list[dict[str, str]],
    ) -> list:
        self.score_rows = rows
        rows.sort(key=lambda r: (0 if str(getattr(r, "scoring_note", "") or "").startswith("Lỗi") else 1, str(getattr(r, "student_id", "") or "")))
        success_count = 0
        error_count = 0
        edited_count = 0
        pending_count = 0
        self.score_preview_table.setSortingEnabled(False)
        self.score_preview_table.setUpdatesEnabled(False)
        try:
            self.score_preview_table.setRowCount(0)
            for i, r in enumerate(rows):
                sid_key = str(r.student_id or "").strip()
                existing_payload = prev_subject_scores.get(sid_key, {}) if sid_key else {}
                recheck_score = existing_payload.get("recheck_score", "") if isinstance(existing_payload, dict) else ""
                final_score_display = recheck_score if recheck_score not in {"", None} else r.score
                class_name = str(getattr(r, "class_name", "") or existing_payload.get("class_name", "") or "-") if isinstance(existing_payload, dict) else str(getattr(r, "class_name", "") or "-")
                birth_date = str(getattr(r, "birth_date", "") or existing_payload.get("birth_date", "") or "-") if isinstance(existing_payload, dict) else str(getattr(r, "birth_date", "") or "-")
                status_text = str(getattr(r, "scoring_note", "") or "OK")
                if sid_key in edited_student_ids:
                    status_text = "Đã sửa"
                elif isinstance(existing_payload, dict) and existing_payload.get("status") == "Đã sửa":
                    status_text = "Đã sửa"

                self.score_preview_table.insertRow(i)
                self.score_preview_table.setItem(i, 0, QTableWidgetItem(r.student_id or "-"))
                self.score_preview_table.setItem(i, 1, QTableWidgetItem(r.name or "-"))
                self.score_preview_table.setItem(i, 2, QTableWidgetItem(class_name))
                self.score_preview_table.setItem(i, 3, QTableWidgetItem(birth_date))
                self.score_preview_table.setItem(i, 4, QTableWidgetItem(r.exam_code))
                self.score_preview_table.setItem(i, 5, QTableWidgetItem(str(getattr(r, "mcq_correct", 0))))
                tf_statement_count = self._tf_statement_correct_count(str(getattr(r, "tf_compare", "") or ""))
                self.score_preview_table.setItem(i, 6, QTableWidgetItem(str(tf_statement_count)))
                self.score_preview_table.setItem(i, 7, QTableWidgetItem(str(getattr(r, "numeric_correct", 0))))
                self.score_preview_table.setItem(i, 8, QTableWidgetItem(str(r.correct)))
                self.score_preview_table.setItem(i, 9, QTableWidgetItem(str(r.wrong)))
                self.score_preview_table.setItem(i, 10, QTableWidgetItem(str(r.blank)))
                self.score_preview_table.setItem(i, 11, QTableWidgetItem(str(final_score_display)))
                status_item = QTableWidgetItem(status_text)
                if status_text != "OK":
                    status_item.setForeground(QColor("red"))
                category = "success"
                if status_text.startswith("Lỗi"):
                    category = "error"
                elif status_text == "Đã sửa":
                    category = "edited"
                elif status_text in {"Chưa chấm", "Cần chấm lại"}:
                    category = "pending"
                status_item.setData(Qt.UserRole, category)
                self.score_preview_table.setItem(i, 12, status_item)
                if status_text.startswith("Lỗi"):
                    for col in range(self.score_preview_table.columnCount()):
                        item = self.score_preview_table.item(i, col)
                        if item:
                            item.setBackground(QColor(255, 225, 225))
                    error_count += 1
                else:
                    success_count += 1
                if status_text == "Đã sửa":
                    edited_count += 1
                if status_text in {"Chưa chấm", "Cần chấm lại"}:
                    pending_count += 1
        finally:
            self.score_preview_table.setUpdatesEnabled(True)
            self.score_preview_table.viewport().update()

        phase = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "subject": subject,
            "mode": mode_text,
            "count": len(rows),
            "missing": missing,
            "failed_count": len(failed_scans),
            "success_count": len(rows),
            "failed_scans": list(failed_scans),
            "smart_scoring_count": len(smart_scored_scans),
            "smart_scored_scans": list(smart_scored_scans),
            "note": note,
        }
        phase_marker = f"{phase['timestamp']}::{subject}::{mode_text}"

        subject_scores = dict(prev_subject_scores)
        for r in rows:
            sid_key = (r.student_id or "").strip()
            if sid_key:
                old_payload = (prev_subject_scores.get(sid_key, {}) or {})
                profile = self._student_profile_by_id(sid_key)
                row_name = str(profile.get("name", "") or getattr(r, "name", "") or old_payload.get("name", "") or "")
                row_class_name = str(profile.get("class_name", "") or getattr(r, "class_name", "") or old_payload.get("class_name", "") or "-")
                row_birth_date = str(profile.get("birth_date", "") or getattr(r, "birth_date", "") or old_payload.get("birth_date", "") or "-")
                row_exam_room = str(
                    self._subject_room_for_student_id(sid_key, self._subject_config_by_subject_key(subject) or {})
                    or getattr(r, "exam_room", "")
                    or old_payload.get("exam_room", "")
                    or profile.get("exam_room", "")
                    or ""
                )
                row_status = "Đã sửa" if (sid_key in edited_student_ids or old_payload.get("status") == "Đã sửa") else str(getattr(r, "scoring_note", "") or "OK")
                subject_scores[sid_key] = {
                    "student_id": r.student_id,
                    "name": row_name,
                    "subject": r.subject,
                    "class_name": row_class_name,
                    "birth_date": row_birth_date,
                    "exam_room": row_exam_room,
                    "exam_code": r.exam_code,
                    "mcq_correct": getattr(r, "mcq_correct", 0),
                    "tf_correct": getattr(r, "tf_correct", 0),
                    "numeric_correct": getattr(r, "numeric_correct", 0),
                    "tf_compare": getattr(r, "tf_compare", ""),
                    "numeric_compare": getattr(r, "numeric_compare", ""),
                    "correct": r.correct,
                    "wrong": r.wrong,
                    "blank": r.blank,
                    "score": r.score,
                    "recheck_score": (prev_subject_scores.get(sid_key, {}) or {}).get("recheck_score", ""),
                    "baithiphuctra": (prev_subject_scores.get(sid_key, {}) or {}).get("baithiphuctra", ""),
                    "phase": phase_marker,
                    "phase_timestamp": phase["timestamp"],
                    "phase_mode": mode_text,
                    "note": str(getattr(r, "scoring_note", "") or ""),
                    "status": row_status,
                }

        self.scoring_results_by_subject[subject] = subject_scores
        self._persist_scoring_results(subject, phase=phase)
        # DB is canonical source: reload full payload (especially after "Chỉ tính bài chưa có điểm")
        # so the grid always shows newest complete data from storage.
        self.scoring_results_by_subject.pop(subject, None)
        self._load_cached_scoring_results_for_subject(subject)
        refreshed_map = self._load_scoring_results_for_subject_from_storage(subject)
        self.score_rows = [self.scoring_engine.score_result_from_dict(dict(x)) for x in refreshed_map.values() if isinstance(x, dict)]
        self._update_scoring_status_bar(success_count, error_count, edited_count, pending_count)
        self._apply_scoring_filter()
        self._current_scoring_subject = subject
        self._scoring_dirty_subjects.discard(subject)
        self._persist_runtime_session_state_quietly()
        return rows

    def calculate_scores(self, subject_key: str = "", mode: str = "Tính lại toàn bộ", note: str = "") -> list:
        subject = (subject_key or self._resolve_preferred_scoring_subject() or "General").strip()
        subject_cfg = self._subject_config_by_subject_key(subject) or {}
        is_direct_score_subject = self._subject_uses_direct_score_import(subject_cfg)
        self._apply_subject_section_visibility(subject)

        mode_text = (mode or ("Nhập điểm trực tiếp" if is_direct_score_subject else "Tính lại toàn bộ")).strip()
        prev_subject_scores = self.scoring_results_by_subject.get(subject, {})
        rows = []
        missing = 0
        failed_scans: list[dict[str, str]] = []
        smart_scored_scans: list[dict[str, str]] = []
        edited_student_ids: set[str] = set()

        if is_direct_score_subject:
            imported_payloads = self._build_direct_score_payloads_for_subject(subject)
            if not imported_payloads:
                silent_missing_notes = {"auto_refresh_subject_change", "auto_refresh_open_scoring"}
                if str(note or "").strip() not in silent_missing_notes:
                    QMessageBox.warning(self, "Missing data", "Môn tự luận chưa có dữ liệu import điểm trực tiếp.")
                return []
            for payload in imported_payloads:
                sid_key = str(payload.get("student_id", "") or "").strip()
                if mode_text == "Chỉ tính bài chưa có điểm" and sid_key and sid_key in prev_subject_scores:
                    continue
                if sid_key and bool(payload.get("manually_edited", False)):
                    edited_student_ids.add(sid_key)
                row_obj = self.scoring_engine.score_result_from_dict(dict(payload))
                setattr(row_obj, "scoring_note", str(payload.get("note", "") or "Nhập điểm trực tiếp"))
                setattr(row_obj, "exam_room", str(payload.get("exam_room", "") or ""))
                rows.append(row_obj)
            return self._finalize_scoring_rows(
                subject,
                rows,
                prev_subject_scores,
                edited_student_ids,
                mode_text,
                note or "essay_direct_import",
                missing,
                failed_scans,
                smart_scored_scans,
            )

        # Chấm điểm chỉ đọc từ DB; không lấy fallback từ grid/config/runtime snapshot.
        subject_scans = list(self._refresh_scan_results_from_db(subject) or [])
        self.scan_results_by_subject[self._batch_result_subject_key(subject)] = list(subject_scans)

        if not subject_scans:
            silent_missing_notes = {"auto_refresh_subject_change", "auto_refresh_open_scoring"}
            if str(note or "").strip() not in silent_missing_notes:
                QMessageBox.warning(self, "Missing data", "Môn này chưa có dữ liệu Batch Scan để tính điểm.")
            return []

        self._ensure_answer_keys_for_subject(subject)
        if not self.answer_keys:
            QMessageBox.warning(self, "Missing data", "Không tìm thấy đáp án cho môn đã chọn. Vui lòng kiểm tra cấu hình môn.")
            return []

        all_subject_keys: list[SubjectKey] = []
        if self.answer_keys is not None:
            for key_obj in self.answer_keys.keys.values():
                if isinstance(key_obj, SubjectKey) and str(getattr(key_obj, "subject", "") or "").strip() == subject:
                    all_subject_keys.append(key_obj)
        fetched_keys = self._fetch_answer_keys_for_subject_scoped(subject)
        for exam_code, key_obj in (fetched_keys or {}).items():
            if not isinstance(key_obj, SubjectKey):
                continue
            if any(str(k.exam_code or "").strip() == str(exam_code or "").strip() for k in all_subject_keys):
                continue
            all_subject_keys.append(key_obj)

        for scan_row in (subject_scans or []):
            sid_key = str(getattr(scan_row, "student_id", "") or "").strip()
            if not sid_key:
                continue
            has_history = bool([str(x) for x in (getattr(scan_row, "edit_history", []) or []) if str(x or "").strip()])
            if bool(getattr(scan_row, "manually_edited", False)) or has_history:
                edited_student_ids.add(sid_key)

        for scan in subject_scans:
            sid = (scan.student_id or "").strip()
            profile = self._student_profile_by_id(sid)
            if profile.get("name") and not str(getattr(scan, "full_name", "") or "").strip():
                setattr(scan, "full_name", profile.get("name"))
            if profile.get("birth_date") and not str(getattr(scan, "birth_date", "") or "").strip():
                setattr(scan, "birth_date", profile.get("birth_date"))
            if profile.get("class_name") and not str(getattr(scan, "class_name", "") or "").strip():
                setattr(scan, "class_name", profile.get("class_name"))

            if mode_text == "Chỉ tính bài chưa có điểm" and sid and sid in prev_subject_scores:
                continue

            full_name = str(getattr(scan, "full_name", "") or profile.get("name", "") or "").strip()
            class_name = str(getattr(scan, "class_name", "") or profile.get("class_name", "") or "").strip()
            birth_date = str(getattr(scan, "birth_date", "") or profile.get("birth_date", "") or "").strip()

            if not sid or not full_name or not class_name or not birth_date:
                err_msg = "Lỗi thông tin thí sinh: thiếu SBD/Họ tên/Lớp/Ngày sinh"
                error_row = self.scoring_engine.score_result_from_dict({
                    "student_id": sid or "-",
                    "name": full_name or "-",
                    "subject": subject,
                    "exam_code": str(getattr(scan, "exam_code", "") or ""),
                    "mcq_correct": 0,
                    "tf_correct": 0,
                    "numeric_correct": 0,
                    "correct": 0,
                    "wrong": 0,
                    "blank": 0,
                    "score": 0.0,
                    "class_name": class_name or "-",
                    "birth_date": birth_date or "-",
                })
                setattr(error_row, "scoring_note", err_msg)
                rows.append(error_row)
                failed_scans.append({"file": str(getattr(scan, "image_path", "") or "-"), "reason": err_msg})
                continue

            key = self.answer_keys.get_flexible(subject, scan.exam_code)
            if not key:
                exam_code_text = str(getattr(scan, "exam_code", "") or "").strip()
                has_student_identity = bool(sid and (profile.get("name") or profile.get("class_name") or profile.get("birth_date")))
                allow_smart = (not exam_code_text) or ("?" in exam_code_text) or (bool(exam_code_text) and has_student_identity)
                best_row = None
                best_key_code = ""
                if allow_smart:
                    for candidate_key in all_subject_keys:
                        try:
                            cand_row = self.scoring_engine.score(
                                scan,
                                candidate_key,
                                student_name=str(getattr(scan, "full_name", "") or ""),
                                subject_config=subject_cfg,
                            )
                        except Exception:
                            continue
                        if best_row is None or float(getattr(cand_row, "score", 0.0) or 0.0) > float(getattr(best_row, "score", 0.0) or 0.0):
                            best_row = cand_row
                            best_key_code = str(getattr(candidate_key, "exam_code", "") or "").strip()
                if allow_smart and best_row is not None:
                    setattr(best_row, "scoring_note", "Chấm thông minh")
                    setattr(best_row, "class_name", str(getattr(scan, "class_name", "") or profile.get("class_name", "") or ""))
                    setattr(best_row, "birth_date", str(getattr(scan, "birth_date", "") or profile.get("birth_date", "") or ""))
                    if best_key_code:
                        try:
                            setattr(scan, "exam_code", best_key_code)
                            self.database.update_scan_result_payload(
                                self._batch_result_subject_key(subject),
                                str(getattr(scan, "image_path", "") or ""),
                                self._serialize_omr_result(scan),
                                note="smart_scoring_exam_code_pick",
                            )
                        except Exception:
                            pass
                    rows.append(best_row)
                    smart_scored_scans.append({
                        "file": str(getattr(scan, "image_path", "") or "-"),
                        "student_id": sid or "-",
                        "picked_exam_code": best_key_code or "-",
                    })
                    continue

                err_msg = f"Không tìm thấy đáp án phù hợp cho mã đề '{exam_code_text or '-'}'"
                missing += 1
                error_row = self.scoring_engine.score_result_from_dict({
                    "student_id": sid or "-",
                    "name": str(getattr(scan, "full_name", "") or profile.get("name", "") or "-"),
                    "subject": subject,
                    "exam_code": exam_code_text,
                    "mcq_correct": 0,
                    "tf_correct": 0,
                    "numeric_correct": 0,
                    "correct": 0,
                    "wrong": 0,
                    "blank": 0,
                    "score": 0.0,
                    "class_name": str(getattr(scan, "class_name", "") or profile.get("class_name", "") or "-"),
                    "birth_date": str(getattr(scan, "birth_date", "") or profile.get("birth_date", "") or "-"),
                })
                setattr(error_row, "scoring_note", err_msg)
                rows.append(error_row)
                failed_scans.append({
                    "file": str(getattr(scan, "image_path", "") or "-"),
                    "reason": err_msg,
                })
                continue

            try:
                scored = self.scoring_engine.score(
                    scan,
                    key,
                    student_name=str(getattr(scan, "full_name", "") or ""),
                    subject_config=subject_cfg,
                )
                setattr(scored, "class_name", str(getattr(scan, "class_name", "") or profile.get("class_name", "") or ""))
                setattr(scored, "birth_date", str(getattr(scan, "birth_date", "") or profile.get("birth_date", "") or ""))
                rows.append(scored)
            except Exception as exc:
                err_msg = f"Lỗi chấm điểm: {exc}"
                error_row = self.scoring_engine.score_result_from_dict({
                    "student_id": sid or "-",
                    "name": str(getattr(scan, "full_name", "") or profile.get("name", "") or "-"),
                    "subject": subject,
                    "exam_code": str(getattr(scan, "exam_code", "") or ""),
                    "mcq_correct": 0,
                    "tf_correct": 0,
                    "numeric_correct": 0,
                    "correct": 0,
                    "wrong": 0,
                    "blank": 0,
                    "score": 0.0,
                    "class_name": str(getattr(scan, "class_name", "") or profile.get("class_name", "") or "-"),
                    "birth_date": str(getattr(scan, "birth_date", "") or profile.get("birth_date", "") or "-"),
                })
                setattr(error_row, "scoring_note", err_msg)
                rows.append(error_row)
                failed_scans.append({
                    "file": str(getattr(scan, "image_path", "") or "-"),
                    "reason": err_msg,
                })

        return self._finalize_scoring_rows(
            subject,
            rows,
            prev_subject_scores,
            edited_student_ids,
            mode_text,
            note,
            missing,
            failed_scans,
            smart_scored_scans,
        )
    def action_open_recheck(self) -> None:
        return open_recheck_dialog(self)

    def _export_student_subject_matrix_excel(self, output_path: Path) -> None:
        from openpyxl import Workbook

        for _label, key in self._iter_export_subjects():
            self._load_scoring_results_for_subject_from_storage(key, force_db=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "scores_by_student"

        subjects = sorted(set(str(k or "").strip() for k in (self.scoring_results_by_subject or {}).keys() if str(k or "").strip()))
        header = ["Student ID", "Name"] + subjects
        ws.append(header)

        base_students: list[tuple[str, str]] = []
        seen: set[str] = set()
        if self.session:
            for st in (self.session.students or []):
                sid = str(getattr(st, "student_id", "") or "").strip()
                name = str(getattr(st, "name", "") or "").strip()
                if not sid or sid in seen:
                    continue
                seen.add(sid)
                base_students.append((sid, name))

        for subject in subjects:
            for sid, row in (self.scoring_results_by_subject.get(subject, {}) or {}).items():
                sid_text = str(sid or "").strip()
                if not sid_text or sid_text in seen:
                    continue
                seen.add(sid_text)
                base_students.append((sid_text, str((row or {}).get("name", "") or "")))

        for sid, name in base_students:
            vals = [sid, name]
            for subject in subjects:
                row = (self.scoring_results_by_subject.get(subject, {}) or {}).get(sid, {}) or {}
                vals.append(row.get("score", ""))
            ws.append(vals)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[col_letter].width = min(40, max(12, max_len + 2))

        wb.save(output_path)

    def export_results(self) -> None:
        rows = self.score_rows or self.calculate_scores()
        if not rows:
            return

        output_dir = QFileDialog.getExistingDirectory(self, "Select export folder")
        if not output_dir:
            return
        self.scoring_engine.export_csv(rows, Path(output_dir) / "results.csv")
        self.scoring_engine.export_json(rows, Path(output_dir) / "results.json")
        self.scoring_engine.export_xml(rows, Path(output_dir) / "results.xml")
        self.scoring_engine.export_excel(rows, Path(output_dir) / "results.xlsx")
        try:
            self._export_student_subject_matrix_excel(Path(output_dir) / "scores_by_student.xlsx")
            QMessageBox.information(self, "Export", "Exported CSV, JSON, XML, XLSX, scores_by_student.xlsx.")
        except Exception as exc:
            QMessageBox.warning(self, "Export", f"Đã export CSV/JSON/XML/XLSX nhưng không tạo được scores_by_student.xlsx:\n{exc}")

    def _export_subject_scores(self, subject_key: str) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        rows = self._ensure_export_score_rows_for_subject(subject)
        if not rows:
            QMessageBox.information(self, "Xuất điểm môn", "Môn này chưa có dữ liệu điểm.")
            return
        cfg = self._subject_config_by_subject_key(subject) or {}
        suggested = self._safe_sheet_name(str(cfg.get("name", "") or "subject_scores"), fallback="subject_scores")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Xuất điểm môn",
            f"{suggested}.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)",
        )
        if not path:
            return
        out = Path(path)
        headers = self._subject_score_export_headers()
        normalized = self._build_subject_score_export_rows(subject)
        if out.suffix.lower() == ".csv":
            with out.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(normalized)
        else:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            from openpyxl.styles import Border, Font, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "scores"
            ws.append(headers)
            for row in normalized:
                ws.append([row.get(key, "") for key in headers])
            border_side = Side(style="thin", color="000000")
            base_font = Font(name="Times New Roman", size=12)
            for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row_cells:
                    cell.font = base_font
                    cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            for cell in ws[1]:
                cell.font = Font(name="Times New Roman", size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            name_col = headers.index("Họ tên") + 1 if "Họ tên" in headers else -1
            if name_col > 0:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=name_col).alignment = Alignment(horizontal="left", vertical="center")
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    max_len = max(max_len, len(str(cell.value or "")))
                ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))
            wb.save(out)
        QMessageBox.information(self, "Xuất điểm môn", f"Đã xuất dữ liệu:\n{out}")

    def _subject_score_export_headers(self) -> list[str]:
        return [
            "STT",
            "SBD",
            "Phòng thi",
            "Họ tên",
            "Ngày sinh",
            "Lớp",
            "Mã đề",
            "Đúng",
            "Sai",
            "Bỏ trống",
            "Điểm",
            "Trạng thái",
            "Ghi chú",
        ]

    def _build_subject_score_export_rows(self, subject: str) -> list[dict[str, object]]:
        rows = self._ensure_export_score_rows_for_subject(subject)
        student_meta = self._student_meta_by_sid()
        room_by_sid: dict[str, str] = {}
        for scan_item in self._scan_rows_for_subject(subject):
            sid_scan = str(getattr(scan_item, "student_id", "") or "").strip()
            room_scan = str(getattr(scan_item, "exam_room", "") or "").strip()
            if sid_scan and room_scan and sid_scan not in room_by_sid:
                room_by_sid[sid_scan] = room_scan
        normalized: list[dict[str, object]] = []
        for row in rows:
            sid = str(row.get("student_id", "") or "").strip()
            meta = student_meta.get(sid, {})
            status_text = str(row.get("status", "") or "").strip()
            status_fold = status_text.casefold()
            has_error = bool(status_fold) and status_fold not in {"ok", "đã sửa"}
            score_value = row.get("recheck_score", "")
            if score_value in {"", None}:
                score_value = row.get("score", "")
            normalized.append({
                "_has_error": has_error,
                "SBD": sid,
                "Phòng thi": str(row.get("exam_room", "") or room_by_sid.get(sid, "") or meta.get("exam_room", "")),
                "Họ tên": str(row.get("name", "") or meta.get("name", "")),
                "Ngày sinh": self._format_birth_date_for_export(row.get("birth_date", "") or meta.get("birth_date", "")),
                "Lớp": str(row.get("class_name", "") or meta.get("class_name", "")),
                "Mã đề": str(row.get("exam_code", "") or ""),
                "Đúng": row.get("correct", 0),
                "Sai": row.get("wrong", 0),
                "Bỏ trống": row.get("blank", 0),
                "Điểm": "" if score_value in {"", None} else score_value,
                "Trạng thái": status_text,
                "Ghi chú": str(row.get("note", "") or ""),
            })
        for idx, row in enumerate(normalized, start=1):
            row["STT"] = idx
        return normalized

    def _export_subject_score_matrix(self) -> None:
        subjects = self._iter_export_subjects()
        if not subjects:
            QMessageBox.information(self, "Xuất điểm các môn", "Chưa có môn nào để xuất.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Xuất điểm các môn", "subject_scores_matrix.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "scores"
        base_headers, score_headers, sorted_rows = self._build_subject_score_matrix_rows(subjects)
        self._write_subject_score_matrix_sheet(ws, base_headers, score_headers, sorted_rows)
        wb.save(path)
        QMessageBox.information(self, "Xuất điểm các môn", f"Đã xuất dữ liệu:\n{path}")

    def _short_subject_label_for_export(self, subject_key: str, fallback_label: str) -> str:
        cfg = self._subject_config_by_subject_key(subject_key) or {}
        raw = str(cfg.get("name", "") or fallback_label or subject_key).strip()
        if not raw:
            raw = str(subject_key or "").strip()
        compact = re.sub(r"\s*[-–]\s*kỳ thi\b.*$", "", raw, flags=re.IGNORECASE).strip()
        compact = re.sub(r"\s*[-–]\s*khối\b.*$", "", compact, flags=re.IGNORECASE).strip()
        return compact or raw

    def _build_subject_score_matrix_rows(self, subjects: list[tuple[str, str]]) -> tuple[list[str], list[str], list[dict[str, object]]]:
        base_headers = ["STT", "SBD", "Họ tên", "Ngày sinh", "Lớp"]
        score_headers = [self._short_subject_label_for_export(key, label) for label, key in subjects]
        student_rows: dict[str, dict[str, object]] = {}
        for idx, (label, key) in enumerate(subjects):
            score_header = score_headers[idx]
            for row in self._build_subject_score_export_rows(key):
                sid = str(row.get("SBD", "") or "").strip()
                if not sid:
                    continue
                rec = student_rows.setdefault(sid, {
                    "SBD": sid,
                    "Họ tên": row.get("Họ tên", ""),
                    "Ngày sinh": self._format_birth_date_for_export(row.get("Ngày sinh", "")),
                    "Lớp": row.get("Lớp", ""),
                })
                if not rec.get("Họ tên") and row.get("Họ tên"):
                    rec["Họ tên"] = row.get("Họ tên", "")
                if not rec.get("Ngày sinh") and row.get("Ngày sinh"):
                    rec["Ngày sinh"] = self._format_birth_date_for_export(row.get("Ngày sinh", ""))
                if not rec.get("Lớp") and row.get("Lớp"):
                    rec["Lớp"] = row.get("Lớp", "")
                rec[score_header] = row.get("Điểm", "")
        sorted_rows = sorted(student_rows.values(), key=lambda x: str(x.get("SBD", "")))
        return base_headers, score_headers, sorted_rows

    def _write_subject_score_matrix_sheet(
        self,
        ws,
        base_headers: list[str],
        score_headers: list[str],
        rows: list[dict[str, object]],
    ) -> None:
        from openpyxl.styles import Alignment
        from openpyxl.styles import Border, Font, Side
        from openpyxl.utils import get_column_letter

        ws.append(base_headers + score_headers)
        for idx, row in enumerate(rows, start=1):
            row["STT"] = idx
            ws.append([row.get(col, "") for col in base_headers] + [row.get(col, "") for col in score_headers])

        border_side = Side(style="thin", color="000000")
        base_font = Font(name="Times New Roman", size=12)
        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row_cells:
                cell.font = base_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        for cell in ws[1]:
            cell.font = Font(name="Times New Roman", size=12, bold=True)
        name_col = base_headers.index("Họ tên") + 1 if "Họ tên" in base_headers else -1
        if name_col > 0:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=name_col).alignment = Alignment(horizontal="left", vertical="center")

        score_col_width = 136 / 7
        score_col_indexes = set(range(len(base_headers) + 1, len(base_headers) + len(score_headers) + 1))
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx in score_col_indexes:
                ws.column_dimensions[col_letter].width = score_col_width
                continue
            max_len = 0
            for cell in ws[col_letter]:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

    @staticmethod
    def _trim_subject_headers_without_scores(rows: list[dict[str, object]], score_headers: list[str]) -> list[str]:
        kept_subject_headers: list[str] = []
        for header in score_headers:
            has_score = False
            for row in rows:
                if header not in row:
                    continue
                value = row.get(header, None)
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                has_score = True
                break
            if has_score:
                kept_subject_headers.append(header)
        return kept_subject_headers

    def _build_class_matrix_rows(
        self,
        class_name: str,
        score_rows: list[dict[str, object]],
        score_headers: list[str],
    ) -> list[dict[str, object]]:
        score_by_sid: dict[str, dict[str, object]] = {}
        for row in score_rows:
            sid = str(row.get("SBD", "") or "").strip()
            if sid:
                score_by_sid[sid] = row
        out_rows: list[dict[str, object]] = []
        seen: set[str] = set()
        if self.session:
            for st in (self.session.students or []):
                sid = str(getattr(st, "student_id", "") or "").strip()
                st_class = str(getattr(st, "class_name", "") or "").strip()
                if not sid or st_class != class_name:
                    continue
                rec: dict[str, object] = {
                    "SBD": sid,
                    "Họ tên": str(getattr(st, "name", "") or ""),
                    "Ngày sinh": self._format_birth_date_for_export(getattr(st, "birth_date", "") or ""),
                    "Lớp": st_class,
                }
                score_row = score_by_sid.get(sid, {})
                for header in score_headers:
                    if header in score_row:
                        rec[header] = score_row.get(header, "")
                out_rows.append(rec)
                seen.add(sid)
        for row in score_rows:
            sid = str(row.get("SBD", "") or "").strip()
            st_class = str(row.get("Lớp", "") or "").strip()
            if not sid or sid in seen or st_class != class_name:
                continue
            rec = {
                "SBD": sid,
                "Họ tên": row.get("Họ tên", ""),
                "Ngày sinh": self._format_birth_date_for_export(row.get("Ngày sinh", "")),
                "Lớp": st_class,
            }
            for header in score_headers:
                if header in row:
                    rec[header] = row.get(header, "")
            out_rows.append(rec)
        out_rows.sort(key=lambda x: str(x.get("SBD", "")))
        return out_rows

    def _class_options_for_export(self) -> list[str]:
        class_names: list[str] = []
        seen: set[str] = set()
        if self.session:
            for st in (self.session.students or []):
                class_name = str(getattr(st, "class_name", "") or "").strip()
                if class_name and class_name not in seen:
                    seen.add(class_name)
                    class_names.append(class_name)
        if class_names:
            return sorted(class_names)
        subjects = self._iter_export_subjects()
        _base_headers, _score_headers, rows = self._build_subject_score_matrix_rows(subjects)
        for row in rows:
            class_name = str(row.get("Lớp", "") or "").strip()
            if class_name and class_name not in seen:
                seen.add(class_name)
                class_names.append(class_name)
        return sorted(class_names)

    def _export_class_subject_scores(self) -> None:
        subjects = self._iter_export_subjects()
        if not subjects:
            QMessageBox.information(self, "Xuất điểm lớp", "Chưa có môn nào để xuất.")
            return
        class_options = self._class_options_for_export()
        if not class_options:
            QMessageBox.information(self, "Xuất điểm lớp", "Chưa có dữ liệu lớp để xuất.")
            return
        picked_class, ok = QInputDialog.getItem(self, "Xuất điểm lớp", "Chọn lớp cần xuất:", class_options, 0, False)
        if not ok:
            return
        class_name = str(picked_class or "").strip()
        if not class_name:
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Xuất điểm lớp",
            f"{self._safe_sheet_name(class_name, fallback='class_scores')}.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return
        from openpyxl import Workbook

        base_headers, score_headers, rows = self._build_subject_score_matrix_rows(subjects)
        class_rows = self._build_class_matrix_rows(class_name, rows, score_headers)
        if not class_rows:
            QMessageBox.information(self, "Xuất điểm lớp", "Lớp đã chọn chưa có dữ liệu học sinh để xuất.")
            return
        score_headers = self._trim_subject_headers_without_scores(class_rows, score_headers)
        wb = Workbook()
        ws = wb.active
        ws.title = self._safe_sheet_name(class_name, fallback="class_scores")
        self._write_subject_score_matrix_sheet(ws, base_headers, score_headers, class_rows)
        wb.save(path)
        QMessageBox.information(self, "Xuất điểm lớp", f"Đã xuất dữ liệu:\n{path}")

    def _export_all_classes_subject_scores(self) -> None:
        subjects = self._iter_export_subjects()
        if not subjects:
            QMessageBox.information(self, "Xuất điểm các lớp", "Chưa có môn nào để xuất.")
            return
        class_options = self._class_options_for_export()
        if not class_options:
            QMessageBox.information(self, "Xuất điểm các lớp", "Chưa có dữ liệu lớp để xuất.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Xuất điểm các lớp", "all_class_subject_scores.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook

        base_headers, score_headers, rows = self._build_subject_score_matrix_rows(subjects)
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        exported_count = 0
        for class_name in class_options:
            ws = wb.create_sheet(self._safe_sheet_name(class_name, fallback="class_scores"))
            class_rows = self._build_class_matrix_rows(class_name, rows, score_headers)
            if not class_rows:
                wb.remove(ws)
                continue
            exported_count += 1
            class_score_headers = self._trim_subject_headers_without_scores(class_rows, score_headers)
            self._write_subject_score_matrix_sheet(ws, base_headers, class_score_headers, class_rows)
        if exported_count <= 0:
            QMessageBox.information(self, "Xuất điểm các lớp", "Chưa có dữ liệu học sinh để xuất.")
            return
        wb.save(path)
        QMessageBox.information(self, "Xuất điểm các lớp", f"Đã xuất dữ liệu:\n{path}")

    def _export_all_subject_scores(self) -> None:
        subjects = self._iter_export_subjects()
        if not subjects:
            QMessageBox.information(self, "Xuất điểm chi tiết các môn", "Chưa có môn nào để xuất.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Xuất điểm chi tiết các môn", "all_subject_scores.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "summary"
        ws_summary.append(["subject_key", "subject_label", "student_count", "avg_score", "min_score", "max_score"])
        for idx, (label, key) in enumerate(subjects):
            rows = self._ensure_export_score_rows_for_subject(key)
            sheet_name = self._safe_sheet_name(label, fallback=f"subject_{idx + 1}")
            ws = wb.create_sheet(sheet_name)
            headers = ["student_id", "name", "exam_code", "correct", "wrong", "blank", "score", "status", "note"]
            ws.append(headers)
            scores: list[float] = []
            for row in rows:
                ws.append([
                    str(row.get("student_id", "") or ""),
                    str(row.get("name", "") or ""),
                    str(row.get("exam_code", "") or ""),
                    row.get("correct", 0),
                    row.get("wrong", 0),
                    row.get("blank", 0),
                    row.get("score", 0),
                    str(row.get("status", "") or ""),
                    str(row.get("note", "") or ""),
                ])
                score_value = self._score_value_for_statistics(row)
                if score_value is not None:
                    scores.append(score_value)
            if scores:
                ws_summary.append([key, label, len(scores), round(sum(scores) / len(scores), 4), min(scores), max(scores)])
            else:
                ws_summary.append([key, label, 0, 0, 0, 0])
        wb.save(path)
        QMessageBox.information(self, "Xuất điểm chi tiết các môn", f"Đã xuất dữ liệu:\n{path}")

    @staticmethod
    def _safe_file_component(value: str, fallback: str = "unknown") -> str:
        text = str(value or "").strip()
        if not text:
            return fallback
        normalized = unicodedata.normalize("NFKD", text)
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r"[\\/:*?\"<>|]+", "_", normalized)
        normalized = re.sub(r"\s+", "_", normalized).strip("._ ")
        return normalized or fallback

    def _export_return_by_class(self) -> None:
        if not self.session:
            QMessageBox.information(self, "Trả bài theo lớp", "Chưa có kỳ thi hiện tại.")
            return
        class_options = self._class_options_for_export()
        if not class_options:
            QMessageBox.information(self, "Trả bài theo lớp", "Chưa có dữ liệu lớp.")
            return
        picked_class, ok = QInputDialog.getItem(self, "Trả bài theo lớp", "Chọn lớp cần trả bài:", class_options, 0, False)
        if not ok:
            return
        class_name = str(picked_class or "").strip()
        if not class_name:
            return
        output_root = QFileDialog.getExistingDirectory(self, "Chọn thư mục lưu trả bài theo lớp")
        if not output_root:
            return

        student_meta = self._student_meta_by_sid()
        target_class_fold = class_name.casefold()
        class_dir = Path(output_root) / self._safe_file_component(class_name, fallback="class")
        class_dir.mkdir(parents=True, exist_ok=True)

        copied_count = 0
        missing_image_count = 0
        skipped_count = 0
        for subject_label, subject_key in self._iter_export_subjects():
            rows = self._scan_rows_for_subject(subject_key)
            if not rows:
                continue
            subject_dir = class_dir / self._safe_file_component(subject_label, fallback=self._safe_sheet_name(subject_key, fallback="subject"))
            subject_dir.mkdir(parents=True, exist_ok=True)
            used_names: set[str] = set()
            for row in rows:
                image_path = str(getattr(row, "image_path", "") or "").strip()
                sid = str(getattr(row, "student_id", "") or "").strip()
                meta = student_meta.get(sid, {})
                row_class = str(getattr(row, "class_name", "") or meta.get("class_name", "") or "").strip()
                if row_class.casefold() != target_class_fold:
                    continue
                if not image_path:
                    skipped_count += 1
                    continue
                source_path = Path(image_path)
                if not source_path.exists() or not source_path.is_file():
                    missing_image_count += 1
                    continue
                student_name = str(getattr(row, "full_name", "") or meta.get("name", "") or "").strip()
                base_name = f"{self._safe_file_component(sid, fallback='SBD')}_{self._safe_file_component(student_name, fallback='ho_ten')}"
                final_name = base_name
                ext = source_path.suffix
                dup_idx = 2
                while final_name in used_names or (subject_dir / f"{final_name}{ext}").exists():
                    final_name = f"{base_name}_{dup_idx}"
                    dup_idx += 1
                used_names.add(final_name)
                shutil.copy2(source_path, subject_dir / f"{final_name}{ext}")
                copied_count += 1

        QMessageBox.information(
            self,
            "Trả bài theo lớp",
            f"Đã xử lý xong.\n- Lớp: {class_name}\n- File đã copy: {copied_count}\n- File thiếu/không tồn tại: {missing_image_count}\n- Bản ghi bỏ qua (không có đường dẫn ảnh): {skipped_count}\n- Thư mục đích: {class_dir}",
        )

    def _export_subject_api_payload(self, subject_key: str) -> None:
        subject = str(subject_key or "").strip()
        if not subject:
            return
        rows = self._scan_rows_for_subject(subject)
        if not rows:
            cfg = self._subject_config_by_subject_key(subject) or {}
            if self._subject_uses_direct_score_import(cfg):
                for payload in self._build_direct_score_payloads_for_subject(subject):
                    sid = str(payload.get("student_id", "") or "").strip()
                    rows.append(OMRResult(
                        image_path="",
                        student_id=sid,
                        exam_code=str(payload.get("exam_code", "") or ""),
                        answer_string="",
                    ))
                    if rows:
                        row_obj = rows[-1]
                        setattr(row_obj, "full_name", str(payload.get("name", "") or ""))
                        setattr(row_obj, "class_name", str(payload.get("class_name", "") or ""))
                        setattr(row_obj, "exam_room", str(payload.get("exam_room", "") or ""))
        if not rows:
            QMessageBox.information(self, "Xuất API bài làm", "Môn này chưa có dữ liệu bài làm.")
            return
        cfg = self._subject_config_by_subject_key(subject) or {}
        suggested = self._safe_sheet_name(str(cfg.get("name", "") or "subject_api"), fallback="subject_api")
        path, _ = QFileDialog.getSaveFileName(self, "Xuất API bài làm", f"{suggested}_api.csv", "CSV (*.csv)")
        if not path:
            return
        headers = ["STT", "student_id", "name", "class_name", "exam_room", "subject_key", "exam_code", "answer_string"]
        student_meta = self._student_meta_by_sid()
        sid_display_by_normalized: dict[str, str] = {}
        for st in (self.session.students if self.session else []):
            sid_origin = str(getattr(st, "student_id", "") or "").strip()
            sid_norm = self._normalized_student_id_for_match(sid_origin)
            if sid_origin and sid_norm and sid_norm not in sid_display_by_normalized:
                sid_display_by_normalized[sid_norm] = sid_origin

        def _sorted_question_numbers(payload: object) -> list[int]:
            out: list[int] = []
            if isinstance(payload, dict):
                for key in payload.keys():
                    key_text = str(key or "").strip()
                    if key_text.lstrip("-").isdigit():
                        out.append(int(key_text))
            return sorted(set(out))

        def _ordered_question_numbers(valid_map: object, invalid_map: object, fallback_map: object) -> list[int]:
            primary_nums: set[int] = set()
            for src in [valid_map or {}, invalid_map or {}]:
                if not isinstance(src, dict):
                    continue
                for key in src.keys():
                    key_text = str(key or "").strip()
                    if key_text.lstrip("-").isdigit():
                        primary_nums.add(int(key_text))
            if primary_nums:
                return sorted(primary_nums)
            return _sorted_question_numbers(fallback_map)

        def _answer_string_for_api(result) -> str:
            manual_text = str(getattr(result, "manual_content_override", "") or "").strip()
            if manual_text:
                return manual_text.replace(",", ";")

            answer_key = self._subject_answer_key_for_result(result, subject)
            mcq_map = dict(getattr(result, "mcq_answers", {}) or {})
            tf_map = dict(getattr(result, "true_false_answers", {}) or {})
            numeric_map = dict(getattr(result, "numeric_answers", {}) or {})

            invalid_rows = getattr(answer_key, "invalid_answer_rows", {}) or {} if answer_key is not None else {}
            mcq_qs = _ordered_question_numbers(
                getattr(answer_key, "answers", {}) or {} if answer_key is not None else {},
                (invalid_rows.get("MCQ", {}) or {}),
                mcq_map,
            )
            tf_qs = _ordered_question_numbers(
                getattr(answer_key, "true_false_answers", {}) or {} if answer_key is not None else {},
                (invalid_rows.get("TF", {}) or {}),
                tf_map,
            )
            numeric_qs = _ordered_question_numbers(
                getattr(answer_key, "numeric_answers", {}) or {} if answer_key is not None else {},
                (invalid_rows.get("NUMERIC", {}) or {}),
                numeric_map,
            )

            parts: list[str] = []
            for q_no in mcq_qs:
                value = str(mcq_map.get(q_no, "") or "").strip().upper()[:1]
                parts.append(value or "_")
            for q_no in tf_qs:
                flags = dict(tf_map.get(q_no, {}) or {})
                token = "".join(
                    "Đ" if key in flags and bool(flags.get(key)) else ("S" if key in flags else "_")
                    for key in ["a", "b", "c", "d"]
                )
                parts.append(token or "____")
            for q_no in numeric_qs:
                value = str(numeric_map.get(q_no, "") or "").strip().replace(" ", "").lstrip("+").replace(".", ",")
                parts.append(value or "_")
            if parts:
                return ";".join(parts)

            fallback = str(getattr(result, "answer_string", "") or "").strip()
            return fallback.replace(",", ";") if fallback else ""

        with Path(path).open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for idx, result in enumerate(rows, start=1):
                sid_raw = str(getattr(result, "student_id", "") or "").strip()
                sid_display = sid_display_by_normalized.get(self._normalized_student_id_for_match(sid_raw), sid_raw)
                meta = student_meta.get(sid_display, student_meta.get(sid_raw, {}))
                sid_export = f'="{sid_display}"' if sid_display and sid_display.isdigit() else sid_display
                exam_room_text = str(self._subject_room_for_student_id(sid_display or sid_raw, cfg) or "").strip()
                if not exam_room_text:
                    exam_room_text = str(getattr(result, "exam_room", "") or meta.get("exam_room", ""))
                writer.writerow({
                    "STT": idx,
                    "student_id": sid_export,
                    "name": str(getattr(result, "full_name", "") or meta.get("name", "")),
                    "class_name": str(getattr(result, "class_name", "") or meta.get("class_name", "")),
                    "exam_room": exam_room_text,
                    "subject_key": subject,
                    "exam_code": str(getattr(result, "exam_code", "") or ""),
                    "answer_string": _answer_string_for_api(result),
                })
        QMessageBox.information(self, "Xuất API bài làm", f"Đã xuất API bài làm:\n{path}")

    def _export_score_range_report(self, ranges: list[tuple[float, float]]) -> None:
        path, _ = QFileDialog.getSaveFileName(self, "Báo cáo thống kê khoảng điểm", "score_range_report.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "score_ranges"
        ws.append(["subject_key", "subject_label", "range", "count", "ratio_percent"])
        subjects = self._iter_export_subjects()
        for label, key in subjects:
            rows = self._ensure_export_score_rows_for_subject(key)
            values = []
            for row in rows:
                score_value = self._score_value_for_statistics(row)
                if score_value is None:
                    continue
                values.append(score_value)
            total = len(values)
            if total <= 0:
                for start, end in ranges:
                    ws.append([key, label, f"{start:g}-{end:g}", 0, 0.0])
                continue
            for idx, (start, end) in enumerate(ranges):
                if idx == len(ranges) - 1:
                    cnt = sum(1 for x in values if start <= x <= end)
                else:
                    cnt = sum(1 for x in values if start <= x < end)
                ratio = round((cnt * 100.0 / total), 2)
                ws.append([key, label, f"{start:g}-{end:g}", cnt, ratio])
        wb.save(path)
        QMessageBox.information(self, "Báo cáo thống kê khoảng điểm", f"Đã xuất báo cáo:\n{path}")

    def _export_class_report(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, "Báo cáo thống kê theo lớp", "class_score_report.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "class_stats"
        ws.append(["subject_key", "subject_label", "class_name", "student_count", "avg_score", "min_score", "max_score", "pass_rate_percent"])
        subjects = self._iter_export_subjects()
        student_meta = self._student_meta_by_sid()
        for label, key in subjects:
            rows = self._ensure_export_score_rows_for_subject(key)
            grouped: dict[str, list[float]] = {}
            for row in rows:
                sid = str(row.get("student_id", "") or "").strip()
                class_name = str(row.get("class_name", "") or "").strip() or str((student_meta.get(sid, {}) or {}).get("class_name", "") or "").strip() or "Chưa phân lớp"
                score_value = self._score_value_for_statistics(row)
                if score_value is None:
                    continue
                grouped.setdefault(class_name, []).append(score_value)
            for class_name, values in sorted(grouped.items(), key=lambda x: x[0]):
                count = len(values)
                if count <= 0:
                    continue
                avg_score = round(sum(values) / count, 4)
                pass_rate = round((sum(1 for x in values if x >= 5.0) * 100.0 / count), 2)
                ws.append([key, label, class_name, count, avg_score, min(values), max(values), pass_rate])
        wb.save(path)
        QMessageBox.information(self, "Báo cáo thống kê theo lớp", f"Đã xuất báo cáo:\n{path}")

    def _export_management_report(self) -> None:
        default_ranges = [(0.0, 2.0), (2.0, 4.0), (4.0, 6.0), (6.0, 8.0), (8.0, 10.0)]
        path, _ = QFileDialog.getSaveFileName(self, "Báo cáo tổng hợp quản lý", "management_report.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook
        wb = Workbook()
        ws_overview = wb.active
        ws_overview.title = "overview"
        ws_overview.append(["subject_key", "subject_label", "student_count", "avg_score", "std_like", "pass_rate_percent"])
        ws_ranges = wb.create_sheet("score_ranges")
        ws_ranges.append(["subject_key", "subject_label", "range", "count", "ratio_percent"])
        ws_classes = wb.create_sheet("class_stats")
        ws_classes.append(["subject_key", "subject_label", "class_name", "student_count", "avg_score"])
        subjects = self._iter_export_subjects()
        student_meta = self._student_meta_by_sid()
        for label, key in subjects:
            rows = self._ensure_export_score_rows_for_subject(key)
            scores: list[float] = []
            class_bucket: dict[str, list[float]] = {}
            for row in rows:
                val = self._score_value_for_statistics(row)
                if val is None:
                    continue
                scores.append(val)
                sid = str(row.get("student_id", "") or "").strip()
                class_name = str(row.get("class_name", "") or "").strip() or str((student_meta.get(sid, {}) or {}).get("class_name", "") or "").strip() or "Chưa phân lớp"
                class_bucket.setdefault(class_name, []).append(val)
            count = len(scores)
            if count <= 0:
                ws_overview.append([key, label, 0, 0.0, 0.0, 0.0])
                for start, end in default_ranges:
                    ws_ranges.append([key, label, f"{start:g}-{end:g}", 0, 0.0])
                continue
            avg_score = sum(scores) / count
            variance = sum((x - avg_score) ** 2 for x in scores) / count
            pass_rate = (sum(1 for x in scores if x >= 5.0) * 100.0 / count)
            ws_overview.append([key, label, count, round(avg_score, 4), round(variance ** 0.5, 4), round(pass_rate, 2)])
            for idx, (start, end) in enumerate(default_ranges):
                if idx == len(default_ranges) - 1:
                    c = sum(1 for x in scores if start <= x <= end)
                else:
                    c = sum(1 for x in scores if start <= x < end)
                ws_ranges.append([key, label, f"{start:g}-{end:g}", c, round(c * 100.0 / count, 2)])
            for class_name, vals in sorted(class_bucket.items(), key=lambda x: x[0]):
                ws_classes.append([key, label, class_name, len(vals), round(sum(vals) / len(vals), 4)])
        wb.save(path)
        QMessageBox.information(self, "Báo cáo tổng hợp quản lý", f"Đã xuất báo cáo:\n{path}")

    def _refresh_session_info(self) -> None:
        if not self.session:
            return
        cfg = self.session.config or {}
        self.session_info.setPlainText(
            f"Exam: {self.session.exam_name}\nDate: {self.session.exam_date}\n"
            f"Subjects: {', '.join(self.session.subjects)}\n"
            f"Template: {self.session.template_path}\n"
            f"AnswerKey: {self.session.answer_key_path}\n"
            f"Scan mode: {cfg.get('scan_mode', '-')}\n"
            f"Scan root: {cfg.get('scan_root', '-')}\n"
            f"Paper parts: {cfg.get('paper_part_count', '-')}\n"
            f"Students: {len(self.session.students)}"
        )
        codes = ", ".join(self.imported_exam_codes) if self.imported_exam_codes else "-"
        self.exam_code_preview.setText(f"Mã đề trên phiếu trả lời mẫu: {codes}")

    def closeEvent(self, event) -> None:  # type: ignore[override]
        if self._has_pending_unsaved_work():
            choice = self._prompt_save_changes_word_style(
                "Thoát ứng dụng",
                "Bạn muốn lưu thay đổi trước khi thoát không?",
            )
            if choice == "cancel":
                event.ignore()
                return
            if choice == "save" and not self._save_current_work():
                event.ignore()
                return
        event.accept()



def run() -> None:
    bootstrap_application_db()
    app = QApplication([])
    window = MainWindow()
    window.showMaximized()
    app.exec()


if __name__ == "__main__":
    run()
