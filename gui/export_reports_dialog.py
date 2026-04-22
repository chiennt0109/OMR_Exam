from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from statistics import mean, median

from PySide6.QtCore import Qt, QTimer, QRectF
from PySide6.QtGui import QColor, QFont, QPageSize, QPainter, QPdfWriter, QPen
from PySide6.QtWidgets import (
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)


@dataclass
class ReportTable:
    headers: list[str]
    rows: list[list[object]]
    grouped_rows: dict[str, list[list[object]]] | None = None


class ExportReportsDialog(QDialog):
    REPORT_SUBJECT_DIST = "Phổ điểm theo môn"
    REPORT_SUBJECT_STATS = "Chỉ số theo môn"
    REPORT_COMBO_RANK = "Bảng điểm theo tổ hợp"
    REPORT_COMBO_DIST = "Phổ điểm tổ hợp"
    REPORT_CLASS_SUMMARY = "Tổng hợp theo lớp"
    REPORT_ABSENT_EXAM = "Thống kê học sinh vắng thi"
    ABSENT_GROUP_BY_CLASS = "Theo lớp"
    ABSENT_GROUP_BY_SUBJECT = "Theo môn"

    def __init__(self, parent_window) -> None:
        super().__init__(parent_window)
        self.main_window = parent_window
        self.setWindowTitle("Báo cáo thống kê")
        self.setWindowFlag(Qt.Window, True)
        self.setWindowFlag(Qt.WindowMinimizeButtonHint, True)
        self.setWindowFlag(Qt.WindowMaximizeButtonHint, True)
        self.setWindowFlag(Qt.WindowCloseButtonHint, True)

        self.combo_defs: list[tuple[str, list[str]]] = []
        self._subject_pairs_cache: list[tuple[str, str]] | None = None
        self._student_profile_cache: dict[str, dict[str, str]] | None = None
        self._score_rows_cache: dict[str, list[dict]] = {}
        self._subject_score_cache: dict[str, dict[str, float]] = {}

        self.report_list = QListWidget()
        for name in [
            self.REPORT_SUBJECT_DIST,
            self.REPORT_SUBJECT_STATS,
            self.REPORT_COMBO_RANK,
            self.REPORT_COMBO_DIST,
            self.REPORT_CLASS_SUMMARY,
            self.REPORT_ABSENT_EXAM,
        ]:
            self.report_list.addItem(name)

        self.exam_label = QLabel(self._current_exam_text())
        self.title_label = QLabel(self.REPORT_SUBJECT_DIST)
        self.title_label.setStyleSheet("font-size: 16px; font-weight: 600;")
        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)

        self.class_combo = QComboBox()
        self.class_combo.addItems(self._collect_class_options())
        self.absent_group_combo = QComboBox()
        self.absent_group_combo.addItems([self.ABSENT_GROUP_BY_CLASS, self.ABSENT_GROUP_BY_SUBJECT])

        self.combo_name_edit = QLineEdit()
        self.combo_name_edit.setPlaceholderText("Tên tổ hợp (VD: A00)")
        subject_options = [label for label, _ in self._collect_subject_pairs()]
        self.combo_sub1 = QComboBox(); self.combo_sub1.addItems(subject_options)
        self.combo_sub2 = QComboBox(); self.combo_sub2.addItems(subject_options)
        self.combo_sub3 = QComboBox(); self.combo_sub3.addItems(subject_options)
        self.combo_list = QListWidget()
        self.btn_add_combo = QPushButton("Thêm tổ hợp")
        self.btn_remove_combo = QPushButton("Xóa tổ hợp")

        self.right_widget = QWidget()
        self.right_form = QFormLayout(self.right_widget)
        self.row_class = QWidget()
        row_class_layout = QVBoxLayout(self.row_class)
        row_class_layout.setContentsMargins(0, 0, 0, 0)
        row_class_layout.addWidget(self.class_combo)
        self.row_absent_group = QWidget()
        row_absent_group_layout = QVBoxLayout(self.row_absent_group)
        row_absent_group_layout.setContentsMargins(0, 0, 0, 0)
        row_absent_group_layout.addWidget(self.absent_group_combo)
        self.row_combo = QWidget()
        combo_layout = QVBoxLayout(self.row_combo)
        combo_layout.setContentsMargins(0, 0, 0, 0)
        combo_layout.addWidget(self.combo_name_edit)
        combo_layout.addWidget(self.combo_sub1)
        combo_layout.addWidget(self.combo_sub2)
        combo_layout.addWidget(self.combo_sub3)
        row_btns = QHBoxLayout()
        row_btns.addWidget(self.btn_add_combo)
        row_btns.addWidget(self.btn_remove_combo)
        combo_layout.addLayout(row_btns)
        combo_layout.addWidget(self.combo_list)

        self.right_form.addRow("Kỳ thi hiện tại", self.exam_label)
        self.right_form.addRow("Lớp", self.row_class)
        self.right_form.addRow("Nhóm vắng thi", self.row_absent_group)
        self.right_form.addRow("Chọn tổ hợp", self.row_combo)

        center = QWidget()
        center_layout = QVBoxLayout(center)
        center_layout.addWidget(self.title_label)
        center_layout.addWidget(self.preview_table)

        self.btn_preview = QPushButton("Xem trước")
        self.btn_export_excel = QPushButton("Xuất Excel")
        self.btn_export_pdf = QPushButton("Xuất PDF")
        self.btn_close = QPushButton("Đóng")
        bottom_ribbon = QHBoxLayout()
        bottom_ribbon.addStretch(1)
        bottom_ribbon.addWidget(self.btn_preview)
        bottom_ribbon.addWidget(self.btn_export_excel)
        bottom_ribbon.addWidget(self.btn_export_pdf)
        bottom_ribbon.addWidget(self.btn_close)

        layout = QGridLayout(self)
        layout.addWidget(self.report_list, 0, 0)
        layout.addWidget(center, 0, 1)
        layout.addWidget(self.right_widget, 0, 2)
        layout.addLayout(bottom_ribbon, 1, 0, 1, 3)
        layout.setColumnStretch(0, 2)
        layout.setColumnStretch(1, 7)
        layout.setColumnStretch(2, 3)

        self.report_list.currentTextChanged.connect(self._on_report_changed)
        self.btn_add_combo.clicked.connect(self._add_combo)
        self.btn_remove_combo.clicked.connect(self._remove_combo)
        self.absent_group_combo.currentTextChanged.connect(lambda _text: self.preview_report())
        self.btn_preview.clicked.connect(self.preview_report)
        self.btn_export_excel.clicked.connect(self.export_excel)
        self.btn_export_pdf.clicked.connect(self.export_pdf)
        self.btn_close.clicked.connect(self.close)

        self._last_report: ReportTable | None = None
        self.report_list.setCurrentRow(0)
        self._load_combo_defs_from_db()
        self._ensure_default_combo()
        self.preview_report()
        QTimer.singleShot(0, self.showMaximized)

    def _current_exam_text(self) -> str:
        session_id = str(getattr(self.main_window, "current_session_id", "") or "").strip()
        return session_id or "Chưa có kỳ thi"

    def _collect_subject_pairs(self) -> list[tuple[str, str]]:
        if self._subject_pairs_cache is not None:
            return list(self._subject_pairs_cache)
        out: list[tuple[str, str]] = []
        for _label, key in self.main_window._iter_export_subjects():
            cfg = self.main_window._subject_config_by_subject_key(key) or {}
            raw_name = str(cfg.get("name", "") or key).strip()
            clean = raw_name
            clean = clean.replace("Môn:", "").strip()
            clean = clean.split("- Kỳ thi", 1)[0].strip()
            clean = clean.split("- Khối", 1)[0].strip()
            out.append((clean or str(key), key))
        seen: set[str] = set()
        dedup: list[tuple[str, str]] = []
        for label, key in out:
            final = label
            if final in seen:
                final = f"{label} ({key})"
            seen.add(final)
            dedup.append((final, key))
        self._subject_pairs_cache = dedup
        return list(dedup)

    def _score_rows_for_subject_cached(self, subject_key: str) -> list[dict]:
        key = str(subject_key or "").strip()
        if not key:
            return []
        if key not in self._score_rows_cache:
            rows = list(self.main_window._ensure_export_score_rows_for_subject(key) or [])
            source_ids: set[str] = set()
            source_count = 0
            if hasattr(self.main_window, "_scoring_source_student_ids"):
                try:
                    source_ids, source_count = self.main_window._scoring_source_student_ids(key)
                except Exception:
                    source_ids, source_count = set(), 0
            if source_count > 0:
                filtered: list[dict] = []
                for row in rows:
                    sid = str((row or {}).get("student_id", "") or "").strip()
                    status_text = str((row or {}).get("status", "") or (row or {}).get("note", "") or "").strip()
                    if sid and sid in source_ids:
                        filtered.append(dict(row))
                        continue
                    if not sid and status_text.startswith("Lỗi"):
                        filtered.append(dict(row))
                rows = filtered
            self._score_rows_cache[key] = rows
        return list(self._score_rows_cache.get(key, []))

    def _collect_class_options(self) -> list[str]:
        vals = ["Tất cả"]
        classes = self._session_class_order()
        extra_classes = sorted(
            set(
                str(profile.get("class_name", "") or "").strip()
                for profile in self._student_profile_map().values()
                if str(profile.get("class_name", "") or "").strip() and str(profile.get("class_name", "") or "").strip() not in classes
            )
        )
        classes.extend(extra_classes)
        vals.extend(classes)
        return vals

    def _session_class_order(self) -> list[str]:
        ordered: list[str] = []
        seen: set[str] = set()
        if self.main_window.session:
            for st in (self.main_window.session.students or []):
                cls = self._student_extra_text(st, ["class_name", "class", "Lớp", "lop", "Mã lớp", "ma_lop"])
                if not cls or cls in seen:
                    continue
                seen.add(cls)
                ordered.append(cls)
        return ordered

    @staticmethod
    def _student_extra_text(student: object, keys: list[str]) -> str:
        if not keys:
            return ""
        meta = getattr(student, "extra", {}) or {}
        if not isinstance(meta, dict):
            meta = {}
        lowered = {str(k).strip().lower(): v for k, v in meta.items()}
        for key in keys:
            key_text = str(key or "").strip()
            if not key_text:
                continue
            if hasattr(student, key_text):
                raw = getattr(student, key_text, "")
                if str(raw or "").strip():
                    return str(raw).strip()
            raw = meta.get(key_text, None)
            if str(raw or "").strip():
                return str(raw).strip()
            raw = lowered.get(key_text.lower(), None)
            if str(raw or "").strip():
                return str(raw).strip()
        return ""

    def _student_profile_map(self) -> dict[str, dict[str, str]]:
        if self._student_profile_cache is not None:
            return dict(self._student_profile_cache)
        out: dict[str, dict[str, str]] = {}
        has_session_students = False
        if self.main_window.session:
            for st in (self.main_window.session.students or []):
                sid = str(getattr(st, "student_id", "") or "").strip()
                if not sid:
                    continue
                has_session_students = True
                out[sid] = {
                    "name": str(getattr(st, "name", "") or ""),
                    "birth_date": self.main_window._format_birth_date_for_export(
                        self._student_extra_text(st, ["birth_date", "dob", "Ngày sinh", "ngay_sinh"])
                    ),
                    "class_name": self._student_extra_text(st, ["class_name", "class", "Lớp", "lop", "Mã lớp", "ma_lop"]),
                    "exam_room": self._student_extra_text(st, ["exam_room", "room", "Phòng thi", "phong_thi"]),
                }
        for _label, key in self._collect_subject_pairs():
            for row in self._score_rows_for_subject_cached(key):
                sid = str(row.get("student_id", "") or "").strip()
                if not sid:
                    continue
                if has_session_students and sid not in out:
                    # Lớp phải theo danh sách import ban đầu của kỳ thi; không tạo mới từ dữ liệu phát sinh.
                    continue
                rec = out.setdefault(
                    sid,
                    {
                        "name": str(row.get("name", "") or ""),
                        "birth_date": self.main_window._format_birth_date_for_export(row.get("birth_date", "") or ""),
                        "class_name": str(row.get("class_name", "") or ""),
                        "exam_room": str(row.get("exam_room", "") or ""),
                    },
                )
                if not rec.get("name"):
                    rec["name"] = str(row.get("name", "") or "")
                if not rec.get("birth_date"):
                    rec["birth_date"] = self.main_window._format_birth_date_for_export(row.get("birth_date", "") or "")
                if not rec.get("class_name"):
                    rec["class_name"] = str(row.get("class_name", "") or "")
                if not rec.get("exam_room"):
                    rec["exam_room"] = str(row.get("exam_room", "") or "")
        self._student_profile_cache = out
        return dict(out)

    def _ensure_default_combo(self) -> None:
        if self.combo_defs:
            return
        subjects = [label for label, _ in self._collect_subject_pairs()]
        if len(subjects) >= 3:
            self.combo_defs.append(("Mặc định", subjects[:3]))
            self._refresh_combo_list(select_index=0)
            self._save_combo_defs_to_db()

    def _combo_state_key(self) -> str:
        session_id = str(getattr(self.main_window, "current_session_id", "") or "").strip() or "global"
        return f"report_combo_defs:{session_id}"

    def _normalize_combo_defs(self, raw_value: object) -> list[tuple[str, list[str]]]:
        subject_labels = {label for label, _ in self._collect_subject_pairs()}
        normalized: list[tuple[str, list[str]]] = []
        seen: set[tuple[str, str, str, str]] = set()
        if not isinstance(raw_value, list):
            return normalized
        for item in raw_value:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name", "") or "").strip()
            subjects_raw = item.get("subjects", [])
            if not isinstance(subjects_raw, list):
                continue
            subjects = [str(s or "").strip() for s in subjects_raw[:3]]
            if len(subjects) != 3 or len(set(subjects)) < 3:
                continue
            if not all(sub in subject_labels for sub in subjects):
                continue
            key = (name.casefold(), subjects[0], subjects[1], subjects[2])
            if key in seen:
                continue
            seen.add(key)
            normalized.append((name or f"Tổ hợp {len(normalized) + 1}", subjects))
        return normalized

    def _load_combo_defs_from_db(self) -> None:
        db = getattr(self.main_window, "database", None)
        if db is None:
            return
        payload = db.get_app_state(self._combo_state_key(), [])
        self.combo_defs = self._normalize_combo_defs(payload)
        if self.combo_defs:
            self._refresh_combo_list(select_index=0)

    def _save_combo_defs_to_db(self) -> None:
        db = getattr(self.main_window, "database", None)
        if db is None:
            return
        payload = [{"name": name, "subjects": list(subjects)} for name, subjects in self.combo_defs]
        db.set_app_state(self._combo_state_key(), payload)

    def _selected_combo(self) -> tuple[str, list[str]] | None:
        if not self.combo_defs:
            return None
        row = self.combo_list.currentRow()
        if 0 <= row < len(self.combo_defs):
            return self.combo_defs[row]
        return self.combo_defs[0]

    def _add_combo(self) -> None:
        name = self.combo_name_edit.text().strip() or f"Tổ hợp {len(self.combo_defs) + 1}"
        subjects = [
            self.combo_sub1.currentText().strip(),
            self.combo_sub2.currentText().strip(),
            self.combo_sub3.currentText().strip(),
        ]
        if len(set(subjects)) < 3:
            QMessageBox.warning(self, "Tổ hợp", "3 môn trong tổ hợp phải khác nhau.")
            return
        self.combo_defs.append((name, subjects))
        self._refresh_combo_list(select_index=len(self.combo_defs) - 1)
        self._last_report = None
        self._save_combo_defs_to_db()

    def _remove_combo(self) -> None:
        row = self.combo_list.currentRow()
        if row < 0 or row >= len(self.combo_defs):
            return
        self.combo_defs.pop(row)
        next_row = min(row, len(self.combo_defs) - 1)
        self._refresh_combo_list(select_index=next_row)
        self._last_report = None
        self._save_combo_defs_to_db()

    def _refresh_combo_list(self, select_index: int | None = None) -> None:
        old_row = self.combo_list.currentRow()
        self.combo_list.clear()
        for name, subjects in self.combo_defs:
            self.combo_list.addItem(f"{name}: {subjects[0]}, {subjects[1]}, {subjects[2]}")
        if self.combo_defs:
            target = old_row if select_index is None else select_index
            target = max(0, min(target, len(self.combo_defs) - 1))
            self.combo_list.setCurrentRow(target)

    def _on_report_changed(self, text: str) -> None:
        self.title_label.setText(text or "Báo cáo thống kê")
        needs_filters = text in {self.REPORT_COMBO_RANK, self.REPORT_COMBO_DIST, self.REPORT_CLASS_SUMMARY, self.REPORT_ABSENT_EXAM}
        self.right_widget.setVisible(needs_filters)
        self.row_class.setVisible(text == self.REPORT_CLASS_SUMMARY)
        self.row_absent_group.setVisible(text == self.REPORT_ABSENT_EXAM)
        self.row_combo.setVisible(text in {self.REPORT_COMBO_RANK, self.REPORT_COMBO_DIST, self.REPORT_CLASS_SUMMARY})

    @staticmethod
    def _is_missing_room_text(room_text: str) -> bool:
        normalized = str(room_text or "").strip().casefold()
        return normalized in {"", "-", "không rõ phòng", "[không rõ phòng]"}

    def _subject_mapping_room_by_sid(self, subject_cfg: dict, profiles: dict[str, dict[str, str]]) -> dict[str, str]:
        cfg = subject_cfg if isinstance(subject_cfg, dict) else {}
        room_by_sid: dict[str, str] = {}
        normalized_to_sid: dict[str, str] = {}

        normalize_sid = getattr(self.main_window, "_normalized_student_id_for_match", None)
        if callable(normalize_sid):
            for sid in profiles.keys():
                sid_text = str(sid or "").strip()
                sid_norm = str(normalize_sid(sid_text) or "").strip()
                if sid_text and sid_norm and sid_norm not in normalized_to_sid:
                    normalized_to_sid[sid_norm] = sid_text
        else:
            for sid in profiles.keys():
                sid_text = str(sid or "").strip()
                if sid_text and sid_text not in normalized_to_sid:
                    normalized_to_sid[sid_text] = sid_text

        preferred_room = str(cfg.get("exam_room_name", "") or "").strip()
        normalize_room = getattr(self.main_window, "_normalized_room_for_match", None)
        preferred_room_norm = str(normalize_room(preferred_room) or "").strip() if callable(normalize_room) else preferred_room.casefold()

        mapping_by_room_func = getattr(self.main_window, "_normalized_exam_room_mapping_by_room", None)
        mapping_by_room = mapping_by_room_func(cfg) if callable(mapping_by_room_func) else {}
        if isinstance(mapping_by_room, dict) and mapping_by_room:
            for room_name, sid_set in mapping_by_room.items():
                room_text = str(room_name or "").strip()
                if not room_text or not isinstance(sid_set, set):
                    continue
                for sid_norm in sid_set:
                    sid_key = normalized_to_sid.get(str(sid_norm or "").strip(), "")
                    if not sid_key:
                        continue
                    current = room_by_sid.get(sid_key, "")
                    if not current:
                        room_by_sid[sid_key] = room_text
                        continue
                    current_norm = str(normalize_room(current) or "").strip() if callable(normalize_room) else current.casefold()
                    next_norm = str(normalize_room(room_text) or "").strip() if callable(normalize_room) else room_text.casefold()
                    if preferred_room_norm and next_norm == preferred_room_norm and current_norm != preferred_room_norm:
                        room_by_sid[sid_key] = room_text
            return room_by_sid

        mapping_text = str(cfg.get("exam_room_sbd_mapping", "") or "").strip()
        if preferred_room and mapping_text:
            chunks = [x.strip() for x in mapping_text.replace(";", ",").replace("\n", ",").split(",") if x.strip()]
            for token in chunks:
                sid_norm = str(normalize_sid(token) or "").strip() if callable(normalize_sid) else token
                sid_key = normalized_to_sid.get(sid_norm, "")
                if sid_key:
                    room_by_sid[sid_key] = preferred_room
        return room_by_sid

    def build_absent_exam_report(self) -> ReportTable:
        subjects = self._collect_subject_pairs()
        profiles = self._student_profile_map()
        absent_items: list[dict[str, str]] = []
        for subject_label, subject_key in subjects:
            cfg = self.main_window._subject_config_by_subject_key(subject_key) or {}
            mapped_room_by_sid = self._subject_mapping_room_by_sid(cfg, profiles)
            if not mapped_room_by_sid:
                continue
            source_ids: set[str] = set()
            if hasattr(self.main_window, "_scoring_source_student_ids"):
                try:
                    source_ids, _count = self.main_window._scoring_source_student_ids(subject_key)
                except Exception:
                    source_ids = set()
            normalize_sid = getattr(self.main_window, "_normalized_student_id_for_match", None)
            source_sid_normalized = {
                str(normalize_sid(sid) or "").strip() if callable(normalize_sid) else str(sid or "").strip()
                for sid in source_ids
                if str(sid or "").strip()
            }
            for sid_text, room_text in mapped_room_by_sid.items():
                profile = profiles.get(sid_text, {})
                missing_room = self._is_missing_room_text(room_text)
                if hasattr(self.main_window, "_is_missing_room_for_status"):
                    try:
                        missing_room = bool(self.main_window._is_missing_room_for_status(room_text))
                    except Exception:
                        missing_room = self._is_missing_room_text(room_text)
                if missing_room:
                    continue
                sid_norm = str(normalize_sid(sid_text) or "").strip() if callable(normalize_sid) else sid_text
                if sid_norm in source_sid_normalized:
                    continue
                absent_items.append({
                    "student_id": sid_text,
                    "name": str(profile.get("name", "") or "").strip(),
                    "class_name": str(profile.get("class_name", "") or "").strip() or "(Không lớp)",
                    "subject_name": subject_label,
                    "exam_room": room_text,
                })

        group_mode = self.absent_group_combo.currentText().strip()
        group_by_class = group_mode != self.ABSENT_GROUP_BY_SUBJECT
        headers = ["Nhóm", "STT", "SBD", "Họ tên", "Lớp", "Môn", "Phòng thi"]
        grouped: dict[str, list[list[object]]] = {}
        for item in absent_items:
            key = item["class_name"] if group_by_class else item["subject_name"]
            grouped.setdefault(key, []).append([
                key,
                0,
                item["student_id"],
                item["name"],
                item["class_name"],
                item["subject_name"],
                item["exam_room"],
            ])

        if group_by_class:
            class_order = self._session_class_order()
            ordered_groups = [cls for cls in class_order if cls in grouped]
            ordered_groups.extend(sorted(cls for cls in grouped if cls not in class_order))
        else:
            subject_order = [label for label, _ in subjects]
            ordered_groups = [subject for subject in subject_order if subject in grouped]
            ordered_groups.extend(sorted(subject for subject in grouped if subject not in subject_order))

        all_rows: list[list[object]] = []
        for key in ordered_groups:
            rows = grouped.get(key, [])
            rows.sort(key=lambda row: (str(row[2]), str(row[5])))
            for idx, row in enumerate(rows, start=1):
                row[1] = idx
            all_rows.extend(rows)
        return ReportTable(headers, all_rows, grouped_rows=grouped)

    def _subject_score_by_sid(self, subject_key: str) -> dict[str, float]:
        key = str(subject_key or "").strip()
        if not key:
            return {}
        if key in self._subject_score_cache:
            return dict(self._subject_score_cache[key])
        out: dict[str, float] = {}
        for row in self._score_rows_for_subject_cached(key):
            sid = str(row.get("student_id", "") or "").strip()
            if not sid:
                continue
            score_value = self.main_window._score_value_for_statistics(row)
            if score_value is None:
                continue
            out[sid] = score_value
        self._subject_score_cache[key] = out
        return dict(out)

    def build_subject_distribution_report(self) -> ReportTable:
        headers = ["Mã môn", "Tổng bài", "0–<=1", "1–<2", "2–<3", "3–<4", "4–<5", "5–<6", "6–<7", "7–<8", "8–<9", "9–<10", "=10"]
        rows: list[list[object]] = []
        bins = [(0, 1), (1, 2), (2, 3), (3, 4), (4, 5), (5, 6), (6, 7), (7, 8), (8, 9), (9, 10)]
        for label, key in self._collect_subject_pairs():
            subject_rows = self._score_rows_for_subject_cached(key)
            scores: list[float] = []
            for row in subject_rows:
                score_value = self.main_window._score_value_for_statistics(row)
                if score_value is not None:
                    scores.append(score_value)
            counts = [0] * len(bins)
            perfect = 0
            for score in scores:
                if abs(score - 10.0) < 1e-9:
                    perfect += 1
                    continue
                for idx, (lo, hi) in enumerate(bins):
                    if idx == 0 and lo <= score <= hi:
                        counts[idx] += 1
                        break
                    if lo <= score < hi:
                        counts[idx] += 1
                        break
            rows.append([label, len(scores), *counts, perfect])
        return ReportTable(headers, rows)

    def build_subject_stats_report(self) -> ReportTable:
        headers = ["Tên môn", "Điểm trung bình", "Điểm trung vị", "Điểm cao nhất", "Điểm thấp nhất"]
        rows: list[list[object]] = []
        for label, key in self._collect_subject_pairs():
            scores: list[float] = []
            for row in self._score_rows_for_subject_cached(key):
                score_value = self.main_window._score_value_for_statistics(row)
                if score_value is not None:
                    scores.append(score_value)
            if scores:
                rows.append([label, round(mean(scores), 4), round(median(scores), 4), max(scores), min(scores)])
            else:
                rows.append([label, "", "", "", ""])
        return ReportTable(headers, rows)

    def _combo_score_maps(self, subjects: list[str]) -> list[dict[str, float]]:
        label_to_key = {label: key for label, key in self._collect_subject_pairs()}
        return [self._subject_score_by_sid(label_to_key.get(sub, "")) for sub in subjects]

    def _build_combo_ranking_rows(self, subjects: list[str]) -> list[list[object]]:
        score_maps = self._combo_score_maps(subjects)
        rows: list[list[object]] = []
        profiles = self._student_profile_map()
        for sid, profile in profiles.items():
            if not sid:
                continue
            vals = [score_maps[i].get(sid, None) for i in range(3)]
            if not all(isinstance(v, (int, float)) for v in vals):
                continue
            total = round(sum(float(v) for v in vals), 4)
            full_name = str(profile.get("name", "") or "").strip()
            parts = full_name.split()
            ho_dem = " ".join(parts[:-1]) if len(parts) > 1 else ""
            ten = parts[-1] if parts else ""
            rows.append([
                0,
                ho_dem,
                ten,
                str(profile.get("birth_date", "") or ""),
                str(profile.get("class_name", "") or ""),
                *vals,
                total,
                "",
            ])
        rows.sort(key=lambda item: float(item[8]) if isinstance(item[8], (int, float)) else -1.0, reverse=True)
        for idx, row in enumerate(rows, start=1):
            row[0] = idx
            row[9] = idx
        return rows

    def build_combo_ranking_report(self) -> ReportTable:
        selected = self._selected_combo()
        if not selected:
            return ReportTable(["STT", "Họ đệm", "Tên", "Ngày sinh", "Mã lớp", "Môn 1", "Môn 2", "Môn 3", "Tổng điểm", "Xếp hạng"], [])
        grouped: dict[str, list[list[object]]] = {}
        for combo_name, subjects in self.combo_defs:
            grouped[combo_name] = self._build_combo_ranking_rows(subjects)
        sel_name, sel_subjects = selected
        headers = ["STT", "Họ đệm", "Tên", "Ngày sinh", "Mã lớp", sel_subjects[0], sel_subjects[1], sel_subjects[2], "Tổng điểm", "Xếp hạng"]
        return ReportTable(headers, grouped.get(sel_name, []), grouped_rows=grouped)

    def build_combo_distribution_report(self) -> ReportTable:
        headers = ["Khoảng điểm", "Tổng cộng"] + [name for name, _ in self.combo_defs]
        if not self.combo_defs:
            return ReportTable(headers, [])
        bins: list[tuple[float, float]] = []
        cur = 27.0
        while cur < 30.0:
            nxt = round(cur + 0.25, 2)
            bins.append((round(cur, 2), nxt))
            cur = nxt
        bins.append((30.0, 30.0))

        profiles = self._student_profile_map()
        combo_totals: list[list[float]] = []
        for _name, subjects in self.combo_defs:
            maps = self._combo_score_maps(subjects)
            totals: list[float] = []
            for sid in profiles.keys():
                vals = [maps[i].get(sid, None) for i in range(3)]
                if all(isinstance(v, (int, float)) for v in vals):
                    totals.append(float(vals[0]) + float(vals[1]) + float(vals[2]))
            combo_totals.append(totals)

        rows: list[list[object]] = []
        for lo, hi in bins:
            label = f"{lo:.2f}-{hi:.2f}" if lo != hi else "=30.00"
            counts: list[int] = []
            for totals in combo_totals:
                count = 0
                for total in totals:
                    if lo == hi and abs(total - 30.0) < 1e-9:
                        count += 1
                    elif lo <= total < hi:
                        count += 1
                counts.append(count)
            rows.append([label, sum(counts), *counts])
        return ReportTable(headers, rows)

    def build_class_summary_report(self) -> ReportTable:
        subjects = self._collect_subject_pairs()
        score_by_subject = {label: self._subject_score_by_sid(key) for label, key in subjects}
        combo_score_maps = {
            combo_name: self._combo_score_maps(combo_subjects)
            for combo_name, combo_subjects in self.combo_defs
        }
        headers = ["STT", "SBD", "Họ tên", "Ngày sinh", "Mã lớp", "Phòng thi"] + [label for label, _ in subjects] + [name for name, _ in self.combo_defs]
        grouped: dict[str, list[list[object]]] = {}
        profiles = self._student_profile_map()
        for sid, profile in profiles.items():
            if not sid:
                continue
            cls = str(profile.get("class_name", "") or "").strip() or "(Không lớp)"
            row = [0, sid, str(profile.get("name", "") or ""), str(profile.get("birth_date", "") or ""), cls, str(profile.get("exam_room", "") or "")]
            for label, _ in subjects:
                row.append(score_by_subject.get(label, {}).get(sid, ""))
            for combo_name, _combo_subjects in self.combo_defs:
                maps = combo_score_maps.get(combo_name, [])
                vals = [maps[i].get(sid, None) for i in range(3)] if len(maps) >= 3 else [None, None, None]
                row.append(round(sum(float(v) for v in vals), 4) if all(isinstance(v, (int, float)) for v in vals) else "")
            grouped.setdefault(cls, []).append(row)

        selected_class = self.class_combo.currentText().strip()
        if selected_class and selected_class != "Tất cả":
            grouped = {selected_class: grouped.get(selected_class, [])}

        class_order = self._session_class_order()
        ordered_classes = [cls for cls in class_order if cls in grouped]
        ordered_classes.extend(sorted(cls for cls in grouped if cls not in class_order))
        all_rows: list[list[object]] = []
        for cls in ordered_classes:
            rows = grouped[cls]
            rows.sort(key=lambda item: str(item[1]))
            for idx, row in enumerate(rows, start=1):
                row[0] = idx
            if rows:
                avg_row = ["", "", "", "", "Trung bình lớp", ""]
                for col_idx in range(6, len(headers)):
                    vals = [float(r[col_idx]) for r in rows if isinstance(r[col_idx], (int, float))]
                    avg_row.append(round(mean(vals), 4) if vals else "")
                rows.append(avg_row)
            all_rows.extend(rows)
        return ReportTable(headers, all_rows, grouped)

    @staticmethod
    def _compact_columns_for_rows(headers: list[str], rows: list[list[object]], fixed_cols: int = 6) -> tuple[list[str], list[list[object]]]:
        if not headers:
            return headers, rows
        keep_indices = list(range(min(fixed_cols, len(headers))))
        for col_idx in range(fixed_cols, len(headers)):
            has_score = any(isinstance((row[col_idx] if col_idx < len(row) else ""), (int, float)) for row in rows)
            if has_score:
                keep_indices.append(col_idx)
        compact_headers = [headers[i] for i in keep_indices]
        compact_rows: list[list[object]] = []
        for row in rows:
            compact_rows.append([row[i] if i < len(row) else "" for i in keep_indices])
        return compact_headers, compact_rows

    def _build_report(self) -> ReportTable:
        name = self.report_list.currentItem().text() if self.report_list.currentItem() else self.REPORT_SUBJECT_DIST
        if name == self.REPORT_SUBJECT_DIST:
            return self.build_subject_distribution_report()
        if name == self.REPORT_SUBJECT_STATS:
            return self.build_subject_stats_report()
        if name == self.REPORT_COMBO_RANK:
            return self.build_combo_ranking_report()
        if name == self.REPORT_COMBO_DIST:
            return self.build_combo_distribution_report()
        if name == self.REPORT_ABSENT_EXAM:
            return self.build_absent_exam_report()
        return self.build_class_summary_report()

    def _render_table(self, table: ReportTable) -> None:
        self.preview_table.clear()
        self.preview_table.setColumnCount(len(table.headers))
        self.preview_table.setHorizontalHeaderLabels(table.headers)
        self.preview_table.setRowCount(len(table.rows))
        for r_idx, row in enumerate(table.rows):
            for c_idx, value in enumerate(row):
                item = QTableWidgetItem("" if value is None else str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.preview_table.setItem(r_idx, c_idx, item)
        self.preview_table.resizeColumnsToContents()

    def preview_report(self) -> None:
        if not str(getattr(self.main_window, "current_session_id", "") or "").strip():
            QMessageBox.information(self, "Báo cáo thống kê", "Vui lòng chọn kỳ thi hiện tại trước khi xem báo cáo.")
            return
        report = self._build_report()
        self._last_report = report
        self._render_table(report)

    def export_excel(self) -> None:
        if self._last_report is None:
            self.preview_report()
        if self._last_report is None:
            return
        path, _ = QFileDialog.getSaveFileName(self, "Xuất Excel", "report.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        def _apply_name_alignment(ws_obj, headers: list[str]) -> None:
            if "Họ tên" not in headers:
                return
            name_col = headers.index("Họ tên") + 1
            for row_idx in range(2, ws_obj.max_row + 1):
                ws_obj.cell(row=row_idx, column=name_col).alignment = Alignment(horizontal="left", vertical="center")

        wb = Workbook()
        report_name = self.report_list.currentItem().text() if self.report_list.currentItem() else "report"
        if report_name == self.REPORT_COMBO_RANK and self._last_report.grouped_rows:
            if wb.active:
                wb.remove(wb.active)
            title_font = Font(bold=True, size=15, color="0B4EA2")
            header_font = Font(bold=True, color="FFFFFF")
            body_font = Font(size=11, color="1F2D3D")
            bold_body_font = Font(size=11, color="1F2D3D", bold=True)
            title_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            header_fill = PatternFill(fill_type="solid", fgColor="1677E5")
            body_fill = PatternFill(fill_type="solid", fgColor="EFEFEF")
            thin = Side(border_style="thin", color="D0D0D0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for combo_name, rows in self._last_report.grouped_rows.items():
                combo_def = next(((n, s) for n, s in self.combo_defs if n == combo_name), None)
                if combo_def is None:
                    continue
                _, subjects = combo_def
                headers = ["STT", "HỌ ĐỆM", "TÊN", "NGÀY SINH", "MÃ LỚP", subjects[0].upper(), subjects[1].upper(), subjects[2].upper(), "TỔNG ĐIỂM", "XẾP HẠNG"]
                ws = wb.create_sheet(self.main_window._safe_sheet_name(combo_name, fallback="to_hop"))
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                title_cell = ws.cell(row=1, column=1, value=f"THỐNG KÊ TỔ HỢP: {combo_name.upper()}")
                title_cell.font = title_font
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = title_fill
                ws.row_dimensions[1].height = 30
                for c_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=3, column=c_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border
                for r_offset, row in enumerate(rows, start=4):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws.cell(row=r_offset, column=c_idx, value=value)
                        cell.fill = body_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = bold_body_font if c_idx == 9 else body_font
                widths = [8, 26, 16, 16, 14, 10, 10, 10, 16, 14]
                for idx, width in enumerate(widths, start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = width
            wb.save(Path(path))
            QMessageBox.information(self, "Báo cáo", f"Đã xuất Excel:\n{path}")
            return
        if report_name == self.REPORT_CLASS_SUMMARY and self._last_report.grouped_rows:
            if wb.active:
                wb.remove(wb.active)
            title_font = Font(bold=True, size=15, color="0B4EA2")
            header_font = Font(bold=True, color="FFFFFF")
            body_font = Font(size=11, color="1F2D3D")
            title_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            header_fill = PatternFill(fill_type="solid", fgColor="1677E5")
            body_fill = PatternFill(fill_type="solid", fgColor="EFEFEF")
            thin = Side(border_style="thin", color="D0D0D0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cls, rows in self._last_report.grouped_rows.items():
                class_headers, class_rows = self._compact_columns_for_rows(self._last_report.headers, rows, fixed_cols=6)
                ws = wb.create_sheet(self.main_window._safe_sheet_name(cls, fallback="class"))
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(class_headers)))
                title_cell = ws.cell(row=1, column=1, value=f"TỔNG HỢP THEO LỚP: {cls.upper()}")
                title_cell.font = title_font
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = title_fill
                ws.row_dimensions[1].height = 30
                for c_idx, header in enumerate(class_headers, start=1):
                    cell = ws.cell(row=3, column=c_idx, value=header.upper())
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border
                for r_offset, row in enumerate(class_rows, start=4):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws.cell(row=r_offset, column=c_idx, value=value)
                        cell.fill = body_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = body_font
                for col_idx in range(1, len(class_headers) + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 14
                if len(class_headers) >= 3:
                    ws.column_dimensions["C"].width = 28
                if len(class_headers) >= 4:
                    ws.column_dimensions["D"].width = 16
                if len(class_headers) >= 5:
                    ws.column_dimensions["E"].width = 14
                if len(class_headers) >= 6:
                    ws.column_dimensions["F"].width = 14
                _apply_name_alignment(ws, class_headers)
                avg_row = ["", "", "", "", "Trung bình lớp", ""]
                for col_idx in range(6, len(class_headers)):
                    vals = [float(r[col_idx]) for r in class_rows if col_idx < len(r) and isinstance(r[col_idx], (int, float))]
                    avg_row.append(round(mean(vals), 4) if vals else "")
                ws.append(avg_row)
                for c_idx in range(1, len(class_headers) + 1):
                    cell = ws.cell(row=ws.max_row, column=c_idx)
                    cell.fill = body_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(size=11, color="1F2D3D", bold=True)
                _apply_name_alignment(ws, class_headers)
        else:
            ws = wb.active
            ws.title = "report"
            ws.append(self._last_report.headers)
            for row in self._last_report.rows:
                ws.append(row)
            _apply_name_alignment(ws, self._last_report.headers)
        wb.save(Path(path))
        QMessageBox.information(self, "Báo cáo", f"Đã xuất Excel:\n{path}")

    def export_pdf(self) -> None:
        if self._last_report is None:
            self.preview_report()
        if self._last_report is None:
            return
        path, _ = QFileDialog.getSaveFileName(self, "Xuất PDF", "report.pdf", "PDF (*.pdf)")
        if not path:
            return
        writer = QPdfWriter(path)
        writer.setPageSize(QPageSize(QPageSize.A4))
        painter = QPainter(writer)
        report_name = self.report_list.currentItem().text() if self.report_list.currentItem() else "report"
        if report_name == self.REPORT_COMBO_RANK and self._last_report.grouped_rows:
            page_w = writer.width()
            page_h = writer.height()
            margin = 36
            y = margin
            cols_ratio = [0.06, 0.23, 0.11, 0.12, 0.11, 0.08, 0.07, 0.07, 0.13, 0.12]
            table_w = page_w - margin * 2
            col_w = [table_w * ratio for ratio in cols_ratio]
            row_h = 26
            title_font = QFont("Arial", 18, QFont.Bold)
            header_font = QFont("Arial", 12, QFont.Bold)
            body_font = QFont("Arial", 11)
            body_bold_font = QFont("Arial", 11, QFont.Bold)
            blue = QColor("#1677E5")
            white = QColor("#FFFFFF")
            text_blue = QColor("#0B4EA2")
            gray = QColor("#EFEFEF")
            line_color = QColor("#D0D0D0")
            for combo_name, rows in self._last_report.grouped_rows.items():
                combo_def = next(((n, s) for n, s in self.combo_defs if n == combo_name), None)
                if combo_def is None:
                    continue
                _, subjects = combo_def
                headers = ["STT", "HỌ ĐỆM", "TÊN", "NGÀY SINH", "MÃ LỚP", subjects[0].upper(), subjects[1].upper(), subjects[2].upper(), "TỔNG ĐIỂM", "XẾP HẠNG"]
                required_h = 70 + row_h + max(1, len(rows)) * row_h + 20
                if y + required_h > page_h - margin:
                    writer.newPage()
                    y = margin
                painter.setFont(title_font)
                painter.setPen(QPen(text_blue))
                painter.drawText(QRectF(margin, y, table_w, 34), int(Qt.AlignCenter), f"THỐNG KÊ TỔ HỢP: {combo_name.upper()}")
                y += 40
                painter.setFont(header_font)
                x = margin
                for idx, head in enumerate(headers):
                    rect = QRectF(x, y, col_w[idx], row_h)
                    painter.fillRect(rect, blue)
                    painter.setPen(QPen(white))
                    painter.drawText(rect.adjusted(8, 0, -2, 0), int(Qt.AlignVCenter | Qt.AlignLeft), head)
                    painter.setPen(QPen(line_color))
                    painter.drawRect(rect)
                    x += col_w[idx]
                y += row_h
                for row in rows:
                    x = margin
                    for idx, value in enumerate(row):
                        rect = QRectF(x, y, col_w[idx], row_h)
                        painter.fillRect(rect, gray)
                        painter.setPen(QPen(line_color))
                        painter.drawRect(rect)
                        painter.setPen(QPen(QColor("#1F2D3D")))
                        painter.setFont(body_bold_font if idx == 8 else body_font)
                        painter.drawText(rect, int(Qt.AlignCenter), "" if value is None else str(value))
                        x += col_w[idx]
                    y += row_h
                y += 20
            painter.end()
            QMessageBox.information(self, "Báo cáo", f"Đã xuất PDF:\n{path}")
            return
        if report_name == self.REPORT_CLASS_SUMMARY and self._last_report.grouped_rows:
            page_w = writer.width()
            page_h = writer.height()
            margin = 36
            y = margin
            title_font = QFont("Arial", 16, QFont.Bold)
            header_font = QFont("Arial", 10, QFont.Bold)
            body_font = QFont("Arial", 9)
            blue = QColor("#1677E5")
            white = QColor("#FFFFFF")
            text_blue = QColor("#0B4EA2")
            gray = QColor("#EFEFEF")
            line_color = QColor("#D0D0D0")
            for cls, rows in self._last_report.grouped_rows.items():
                class_headers, class_rows = self._compact_columns_for_rows(self._last_report.headers, rows, fixed_cols=6)
                if not class_headers:
                    continue
                col_count = len(class_headers)
                table_w = page_w - margin * 2
                col_w = [table_w / col_count] * col_count
                row_h = 22
                avg_row = ["", "", "", "", "Trung bình lớp", ""]
                for col_idx in range(6, len(class_headers)):
                    vals = [float(r[col_idx]) for r in class_rows if col_idx < len(r) and isinstance(r[col_idx], (int, float))]
                    avg_row.append(round(mean(vals), 4) if vals else "")
                display_rows = class_rows + [avg_row]
                required_h = 60 + row_h + max(1, len(display_rows)) * row_h + 20
                if y + required_h > page_h - margin:
                    writer.newPage()
                    y = margin
                painter.setFont(title_font)
                painter.setPen(QPen(text_blue))
                painter.drawText(QRectF(margin, y, table_w, 30), int(Qt.AlignCenter), f"TỔNG HỢP THEO LỚP: {cls.upper()}")
                y += 34
                painter.setFont(header_font)
                x = margin
                for idx, head in enumerate(class_headers):
                    rect = QRectF(x, y, col_w[idx], row_h)
                    painter.fillRect(rect, blue)
                    painter.setPen(QPen(white))
                    painter.drawText(rect.adjusted(6, 0, -2, 0), int(Qt.AlignVCenter | Qt.AlignLeft), str(head).upper())
                    painter.setPen(QPen(line_color))
                    painter.drawRect(rect)
                    x += col_w[idx]
                y += row_h
                for row_idx, row in enumerate(display_rows):
                    x = margin
                    for idx, value in enumerate(row):
                        rect = QRectF(x, y, col_w[idx], row_h)
                        painter.fillRect(rect, gray)
                        painter.setPen(QPen(line_color))
                        painter.drawRect(rect)
                        painter.setPen(QPen(QColor("#1F2D3D")))
                        painter.setFont(QFont("Arial", 9, QFont.Bold) if row_idx == len(display_rows) - 1 else body_font)
                        painter.drawText(rect, int(Qt.AlignCenter), "" if value is None else str(value))
                        x += col_w[idx]
                    y += row_h
                y += 16
            painter.end()
            QMessageBox.information(self, "Báo cáo", f"Đã xuất PDF:\n{path}")
            return
        x = 40
        y = 60
        line_h = 22
        header_line = " | ".join(self._last_report.headers)
        painter.drawText(x, y, header_line)
        y += line_h
        for row in self._last_report.rows:
            painter.drawText(x, y, " | ".join("" if v is None else str(v) for v in row))
            y += line_h
            if y > writer.height() - 60:
                writer.newPage()
                y = 60
                painter.drawText(x, y, header_line)
                y += line_h
        painter.end()
        QMessageBox.information(self, "Báo cáo", f"Đã xuất PDF:\n{path}")
